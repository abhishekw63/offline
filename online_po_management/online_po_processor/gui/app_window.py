"""
gui.app_window
==============

Main Tkinter window — ``OnlinePOApp``.

Layout (520×720 px — bumped in v1.5.0 to accommodate D365 + Email buttons)::

    ┌─────────────────────────────────────────┐
    │         Online PO Processor              │  ← title
    │  Marketplace PO → ERP Sales Order Import │
    │                                          │
    │ Marketplace: [Myntra ▼]  Margin: [70]%   │  ← mkt + margin row
    │                                          │
    │ ┌─ Input Files ─────────────────────┐   │
    │ │ Items Master:    ✓ Items March... │   │
    │ │                  Updated: …        │   │
    │ │ Ship-To Mapping: ✓ Ship to B2B... │   │
    │ │                  Updated: …        │   │
    │ │ Marketplace PO:  Not selected      │   │
    │ └────────────────────────────────────┘   │
    │                                          │
    │        [▶ Generate SO]                   │  ← primary action
    │        [📂 Open Last Output]             │
    │        [📋 Download PO Template]         │
    │        [📁 Update Bundled Files]         │
    │        [📤 Export D365 Package]          │  ← NEW v1.5.0
    │        [📧 Send Email Report]            │  ← NEW v1.5.0
    │                                          │
    │ Status: ...                              │
    │ ┌─ Log ──────────────────────────────┐   │
    │ │ [time] message                     │   │
    │ └────────────────────────────────────┘   │
    └─────────────────────────────────────────┘

Responsibilities
----------------
* Wire the UI and state together.
* Auto-load bundled master/mapping on startup.
* Route user actions to the engine/exporter/template-writer.
* Surface progress in the Log panel and Status line.
* Gate D365 + Email actions on a successful SO generation (both need
  the ``ProcessingResult`` produced by ``generate()``).

The class is intentionally "procedural inside a class" — it holds Tk
widget references plus a few StringVars and path strings. Business
logic lives in ``engine`` / ``exporter`` / ``emailer`` modules; this
file is the thin layer on top.
"""

from __future__ import annotations
import logging
import os
import shutil
import time
import tkinter as tk
from pathlib import Path
from tkinter import filedialog, messagebox, ttk
from typing import List, Optional, Tuple

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from online_po_processor.config.constants import (
    BUNDLED_DATA_FOLDER, BUNDLED_MAPPING_NAME, BUNDLED_MASTER_NAME,
)
from online_po_processor.config.email_config import get_email_config
from online_po_processor.config.marketplaces import (
    DEFAULT_WAREHOUSE, MARKETPLACE_CONFIGS, MARKETPLACE_NAMES,
    WAREHOUSE_CODES, WAREHOUSE_DISPLAY_NAMES,
)
from online_po_processor.config.paths import (
    get_bundled_data_folder, get_bundled_mapping_path,
    get_bundled_master_path, get_update_timestamp, record_update,
)
from online_po_processor.data.mapping_loader import MappingLoader
from online_po_processor.data.master_loader import MasterLoader
from online_po_processor.data.models import ProcessingResult
from online_po_processor.emailer import EmailSender
from online_po_processor.engine.marketplace_engine import MarketplaceEngine
from online_po_processor.exporter.d365_exporter import D365Exporter
from online_po_processor.exporter.so_exporter import SOExporter
from online_po_processor.gui._file_row import build_file_row
from online_po_processor.gui._update_dialog import UpdateDialog
from online_po_processor.utils.platform_open import open_file


class OnlinePOApp:
    """GUI for Online Marketplace PO → SO generation."""

    # ── Construction ───────────────────────────────────────────────────

    def __init__(self) -> None:
        self.root = tk.Tk()
        self.root.title("Online PO Processor — Marketplace SO Generator")
        self.root.geometry("520x720")
        self.root.resizable(False, False)

        # ── File paths (None until picked or auto-loaded) ───────────────
        self.master_path: Optional[str] = None
        self.mapping_path: Optional[str] = None
        self.po_path: Optional[str] = None
        # v1.7.0: multi-file upload support (currently Reliance only).
        # Always contains a list — empty until user picks files, then
        # one-element list for single-select marketplaces, or N-element
        # list for batch-capable ones (i.e. those whose config has
        # ``pre_process`` set).
        self.po_paths: List[str] = []

        # ── Output tracking ─────────────────────────────────────────────
        self.last_output: Optional[Path] = None

        # v1.5.0: stash the full ProcessingResult from the last successful
        # generate() so the D365 and Email actions can reuse it without
        # re-running the engine. Reset to None on every new generate()
        # attempt; set to a real result only if the engine produced rows.
        self.last_result: Optional[ProcessingResult] = None

        # Track whether master/mapping came from the bundled folder (vs
        # user-picked). Used so the GUI can show "(auto-loaded)" and the
        # "Update Bundled Files" flow knows what's in use.
        self.master_is_bundled: bool = False
        self.mapping_is_bundled: bool = False

        # ── Engine-side state ───────────────────────────────────────────
        # MappingLoader is held on the app because it gets re-loaded each
        # run (when marketplace changes). MasterLoader is created fresh
        # in generate() — no state to carry between runs.
        self.mapping_loader = MappingLoader()
        self.exporter = SOExporter()

        # v1.5.0: D365 template filler — created once, reused across
        # multiple "Export D365 Package" clicks. Stateless so one
        # instance is fine.
        self.d365_exporter = D365Exporter()

        # ── Widget references populated by _build_ui ────────────────────
        self.marketplace_var: tk.StringVar
        self.marketplace_dropdown: ttk.Combobox
        self.margin_var: tk.StringVar
        self.margin_entry: tk.Entry
        self.master_var: tk.StringVar
        self.master_ts_var: tk.StringVar
        self.mapping_var: tk.StringVar
        self.mapping_ts_var: tk.StringVar
        self.po_var: tk.StringVar
        self.open_btn: tk.Button
        # v1.5.0: D365 + email buttons start disabled; enabled after a
        # successful Generate SO run (same UX as open_btn).
        self.d365_btn: tk.Button
        self.email_btn: tk.Button
        self.status_var: tk.StringVar
        self.status_label: tk.Label
        self.log_text: tk.Text

        self._build_ui()

        # Auto-load AFTER the UI exists, so we can log and update
        # picker labels in one go.
        self._auto_load_bundled_files()

    # ── UI construction ────────────────────────────────────────────────

    def _build_ui(self) -> None:
        """Build the Tk widget tree."""

        # ── Title ───────────────────────────────────────────────────────
        tk.Label(
            self.root, text="Online PO Processor",
            font=("Arial", 14, "bold"),
        ).pack(pady=(12, 2))

        tk.Label(
            self.root, text="Marketplace PO → ERP Sales Order Import",
            font=("Arial", 9), fg='gray',
        ).pack(pady=(0, 10))

        # ── Marketplace selector + Margin input ─────────────────────────
        mkt_frame = tk.Frame(self.root)
        mkt_frame.pack(fill='x', padx=20, pady=(0, 8))

        tk.Label(
            mkt_frame, text="Marketplace:", font=("Arial", 10, "bold"),
        ).pack(side='left')

        self.marketplace_var = tk.StringVar(
            value=MARKETPLACE_NAMES[0] if MARKETPLACE_NAMES else ''
        )
        self.marketplace_dropdown = ttk.Combobox(
            mkt_frame, textvariable=self.marketplace_var,
            values=MARKETPLACE_NAMES, state='readonly', width=20,
        )
        self.marketplace_dropdown.pack(side='left', padx=8)
        self.marketplace_dropdown.bind(
            '<<ComboboxSelected>>', self._on_marketplace_change,
        )

        # Margin % — user can override per run (pre-filled from config)
        tk.Label(
            mkt_frame, text="Margin:", font=("Arial", 10, "bold"),
        ).pack(side='left', padx=(12, 0))
        self.margin_var = tk.StringVar(value=str(self._get_default_margin()))
        self.margin_entry = tk.Entry(
            mkt_frame, textvariable=self.margin_var, width=5,
            font=("Arial", 10), justify='center',
        )
        self.margin_entry.pack(side='left', padx=4)
        tk.Label(mkt_frame, text="%", font=("Arial", 10)).pack(side='left')
        tk.Label(
            mkt_frame, text="(Landing Cost)", font=("Arial", 8), fg='gray',
        ).pack(side='left', padx=4)

        # ── v1.9.0: Warehouse selector (own row — v1.9.2) ───────────────
        # Lives on its own row below Marketplace/Margin so the widgets
        # don't get clipped off the right edge of the 520px window.
        # Lets the user pick which RENEE warehouse fulfills this batch.
        # The friendly code (AHD/BLR/...) maps to the ERP location
        # string (PICK/DS_BL_OFF1/...) in config.WAREHOUSE_CODES and
        # gets stamped on every D365 Sales Header col K + Sales Line
        # col F. Default is AHD because that's the primary warehouse
        # for most batches today. Adding a warehouse = one line in
        # WAREHOUSE_CODES; no UI changes needed.
        wh_frame = tk.Frame(self.root)
        wh_frame.pack(fill='x', padx=20, pady=(0, 8))

        tk.Label(
            wh_frame, text="Warehouse:", font=("Arial", 10, "bold"),
        ).pack(side='left')
        self.warehouse_var = tk.StringVar(value=DEFAULT_WAREHOUSE)
        self.warehouse_combo = ttk.Combobox(
            wh_frame, textvariable=self.warehouse_var,
            values=WAREHOUSE_DISPLAY_NAMES, state='readonly', width=8,
            font=("Arial", 10),
        )
        self.warehouse_combo.pack(side='left', padx=8)
        # Live hint beside the dropdown — shows the ERP code for the
        # currently-selected warehouse (e.g. 'AHD → PICK'). Kept small
        # and gray so it reads as meta-info, not a primary field.
        self._warehouse_hint_var = tk.StringVar(
            value=f'→ {WAREHOUSE_CODES[DEFAULT_WAREHOUSE]}'
        )
        tk.Label(
            wh_frame, textvariable=self._warehouse_hint_var,
            font=("Arial", 9), fg='gray',
        ).pack(side='left', padx=4)
        self.warehouse_combo.bind(
            '<<ComboboxSelected>>', self._on_warehouse_change,
        )

        # ── File selectors ──────────────────────────────────────────────
        files_frame = tk.LabelFrame(
            self.root, text="Input Files", font=("Arial", 10, "bold"),
            padx=10, pady=8,
        )
        files_frame.pack(fill='x', padx=20, pady=(0, 8))

        # Items Master (with timestamp sub-line)
        self.master_var = tk.StringVar(value="Not selected")
        self.master_ts_var = tk.StringVar(value="")
        build_file_row(
            files_frame, "Items Master:", self.master_var,
            self._select_master, ts_var=self.master_ts_var,
        )

        # Ship-To Mapping (with timestamp sub-line)
        self.mapping_var = tk.StringVar(value="Not selected")
        self.mapping_ts_var = tk.StringVar(value="")
        build_file_row(
            files_frame, "Ship-To Mapping:", self.mapping_var,
            self._select_mapping, ts_var=self.mapping_ts_var,
        )

        # Marketplace PO (no timestamp sub-line — per-run input)
        self.po_var = tk.StringVar(value="Not selected")
        build_file_row(
            files_frame, "Marketplace PO:", self.po_var, self._select_po,
        )

        # ── Action buttons ──────────────────────────────────────────────
        btn_frame = tk.Frame(self.root)
        btn_frame.pack(pady=8)

        tk.Button(
            btn_frame, text="▶  Generate SO", width=20,
            font=("Arial", 10, "bold"),
            bg="#00C853", fg='white', command=self.generate,
        ).pack(pady=4)

        self.open_btn = tk.Button(
            btn_frame, text="📂  Open Last Output", width=20,
            state=tk.DISABLED, command=self.open_last,
        )
        self.open_btn.pack(pady=4)

        tk.Button(
            btn_frame, text="📋  Download PO Template", width=20,
            command=self._download_template,
        ).pack(pady=4)

        tk.Button(
            btn_frame, text="📁  Update Bundled Files", width=20,
            command=self._update_bundled_files,
        ).pack(pady=4)

        # ── v1.5.0: D365 Package Export ─────────────────────────────────
        # Starts disabled — becomes active only after a successful
        # Generate SO run (because both need ``self.last_result``).
        self.d365_btn = tk.Button(
            btn_frame, text="📤  Export D365 Package", width=20,
            state=tk.DISABLED, command=self._export_d365,
        )
        self.d365_btn.pack(pady=4)

        # ── v1.5.0: Email Report ────────────────────────────────────────
        # Same gating rule as the D365 button.
        self.email_btn = tk.Button(
            btn_frame, text="📧  Send Email Report", width=20,
            state=tk.DISABLED, command=self._send_email,
        )
        self.email_btn.pack(pady=4)

        # ── Status line ─────────────────────────────────────────────────
        self.status_var = tk.StringVar(
            value="Status: Waiting — select files and generate"
        )
        self.status_label = tk.Label(
            self.root, textvariable=self.status_var,
            font=("Arial", 10), fg='gray', wraplength=460,
        )
        self.status_label.pack(pady=6)

        # ── Log panel ───────────────────────────────────────────────────
        log_frame = tk.LabelFrame(self.root, text="Log", font=("Arial", 9))
        log_frame.pack(fill='both', expand=True, padx=20, pady=(0, 12))

        scroll = ttk.Scrollbar(log_frame, orient='vertical')
        scroll.pack(side='right', fill='y')

        self.log_text = tk.Text(
            log_frame, height=6, font=("Consolas", 9),
            state='disabled', wrap='word',
            yscrollcommand=scroll.set,
        )
        self.log_text.pack(fill='both', expand=True)
        scroll.config(command=self.log_text.yview)

    # ── Logging helpers ────────────────────────────────────────────────

    def _log(self, msg: str) -> None:
        """Append a timestamped message to the log panel."""
        self.log_text.config(state='normal')
        ts = time.strftime("%H:%M:%S")
        self.log_text.insert('end', f"[{ts}] {msg}\n")
        self.log_text.see('end')
        self.log_text.config(state='disabled')

    # ── Margin helpers ─────────────────────────────────────────────────

    def _get_default_margin(self) -> int:
        """Default margin % for the currently selected marketplace."""
        mkt = (self.marketplace_var.get()
               if hasattr(self, 'marketplace_var') else '')
        if mkt and mkt in MARKETPLACE_CONFIGS:
            return MARKETPLACE_CONFIGS[mkt].get('default_margin', 70)
        return 70

    def _on_marketplace_change(self, _event=None) -> None:
        """Reset margin to the newly-selected marketplace's default."""
        margin = self._get_default_margin()
        self.margin_var.set(str(margin))
        self._log(f"Marketplace changed to {self.marketplace_var.get()}, "
                  f"margin set to {margin}%")

    def _on_warehouse_change(self, _event=None) -> None:
        """
        Sync the gray hint label next to the warehouse dropdown so the
        user can see which ERP code their selection maps to.

        Example: switching dropdown from ``AHD`` to ``BLR`` updates
        the trailing hint from ``→ PICK`` to ``→ DS_BL_OFF1``. Useful
        when someone asks "why did BLR land on that weird code?" —
        the answer is visible right next to the dropdown instead of
        buried in config.
        """
        wh = self.warehouse_var.get()
        code = WAREHOUSE_CODES.get(wh, wh)
        self._warehouse_hint_var.set(f'→ {code}')
        self._log(f"Warehouse changed to {wh} (ERP code: {code})")

    def _get_margin(self) -> float:
        """
        Current margin as a decimal (e.g. ``70`` → ``0.70``).

        Falls back to the marketplace default if the input field is
        empty or invalid. Valid range: 1..100 (inclusive).
        """
        try:
            val = float(self.margin_var.get().strip())
            if val <= 0 or val > 100:
                raise ValueError
            return val / 100.0
        except (ValueError, AttributeError):
            default = self._get_default_margin()
            self._log(f"Invalid margin input, using default {default}%")
            return default / 100.0

    # ── Bundled-file handling ─────────────────────────────────────────
    #
    # Items Master and Ship-To Mapping live in ``Calculation Data/``.
    # Startup auto-loads them so the user doesn't re-pick every run.
    # The "Update Bundled Files" button replaces what's in that folder.

    def _auto_load_bundled_files(self) -> None:
        """
        Look for Items Master + Ship-To Mapping in ``Calculation Data/``
        and pre-populate the picker labels if found.

        Does not abort startup on missing files — logs a hint and leaves
        the pickers in their default "Not selected" state.
        """
        master_p = get_bundled_master_path()
        mapping_p = get_bundled_mapping_path()

        if master_p:
            self.master_path = str(master_p)
            self.master_is_bundled = True
            self.master_var.set(f"✓ {master_p.name} (auto-loaded)")
            self._refresh_ts_label(self.master_ts_var, master_p.name)
            self._log(f"Auto-loaded master from "
                      f"{BUNDLED_DATA_FOLDER}/{master_p.name}")
        else:
            self._log(f"No bundled master at "
                      f"{BUNDLED_DATA_FOLDER}/{BUNDLED_MASTER_NAME} "
                      f"— pick one manually or use 'Update Bundled Files'")

        if mapping_p:
            self.mapping_path = str(mapping_p)
            self.mapping_is_bundled = True
            self.mapping_var.set(f"✓ {mapping_p.name} (auto-loaded)")
            self._refresh_ts_label(self.mapping_ts_var, mapping_p.name)
            self._log(f"Auto-loaded mapping from "
                      f"{BUNDLED_DATA_FOLDER}/{mapping_p.name}")
        else:
            self._log(f"No bundled mapping at "
                      f"{BUNDLED_DATA_FOLDER}/{BUNDLED_MAPPING_NAME} "
                      f"— pick one manually or use 'Update Bundled Files'")

    def _update_bundled_files(self) -> None:
        """
        Replace the bundled master and/or mapping in ``Calculation Data/``.

        Workflow:

        1. Ask which file(s) to update via :class:`UpdateDialog`.
        2. For each chosen kind, open a file picker and copy the picked
           file into the bundled folder under the canonical name.
        3. Refresh in-memory paths and picker labels.
        4. Log the outcome.
        """
        target_folder = get_bundled_data_folder(create=True)

        dialog = UpdateDialog(self.root, folder=target_folder)
        choice = dialog.show()
        if choice is None:
            return  # user cancelled

        updated_any = False

        if choice in ('master', 'both'):
            updated_any |= self._do_update_one_bundled(
                kind_label='Items Master',
                source_title='Select new Items Master file to bundle',
                target_path=target_folder / BUNDLED_MASTER_NAME,
                on_success=self._refresh_master_after_update,
            )

        if choice in ('mapping', 'both'):
            updated_any |= self._do_update_one_bundled(
                kind_label='Ship-To Mapping',
                source_title='Select new Ship-To Mapping file to bundle',
                target_path=target_folder / BUNDLED_MAPPING_NAME,
                on_success=self._refresh_mapping_after_update,
            )

        if updated_any:
            messagebox.showinfo(
                "Bundled Files Updated",
                f"Bundled files updated in:\n{target_folder}\n\n"
                f"Future runs will auto-load the new version.",
            )

    def _do_update_one_bundled(self, kind_label: str, source_title: str,
                                target_path: Path, on_success) -> bool:
        """
        Prompt for a source file and copy it to ``target_path``.

        Args:
            kind_label:   Display label (used in log/dialog text).
            source_title: Title of the file-picker dialog.
            target_path:  Destination (e.g.
                          ``Calculation Data/Items March.xlsx``).
            on_success:   Callback to refresh GUI state after a
                          successful copy.

        Returns:
            True if a copy was performed, False if the user cancelled
            or the copy failed.
        """
        src = filedialog.askopenfilename(
            title=source_title,
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
        )
        if not src:
            self._log(f"Update cancelled for {kind_label}")
            return False

        try:
            shutil.copy2(src, str(target_path))
            # Stamp history BEFORE refresh so the sub-line shows the new
            # timestamp immediately.
            record_update(target_path.name)
            self._log(f"Bundled {kind_label} updated → {target_path}")
            on_success()
            return True
        except Exception as e:  # noqa: BLE001 — surface ANY copy error
            self._log(f"ERROR copying {kind_label}: {e}")
            messagebox.showerror(
                "Update Failed",
                f"Could not copy {kind_label}:\n{e}",
            )
            return False

    def _refresh_ts_label(self, ts_var: tk.StringVar, filename: str) -> None:
        """
        Refresh a timestamp StringVar from the in-app update history.

        Sets to ``"Updated: <date>"`` when there's a record, empty
        string otherwise (which renders as a blank sub-line).
        """
        ts = get_update_timestamp(filename)
        ts_var.set(f"Updated: {ts}" if ts else "")

    def _refresh_master_after_update(self) -> None:
        """Re-point in-memory master to the freshly bundled file."""
        p = get_bundled_master_path()
        if p:
            self.master_path = str(p)
            self.master_is_bundled = True
            self.master_var.set(f"✓ {p.name} (auto-loaded)")
            self._refresh_ts_label(self.master_ts_var, p.name)

    def _refresh_mapping_after_update(self) -> None:
        """Re-point in-memory mapping to the freshly bundled file."""
        p = get_bundled_mapping_path()
        if p:
            self.mapping_path = str(p)
            self.mapping_is_bundled = True
            self.mapping_var.set(f"✓ {p.name} (auto-loaded)")
            self._refresh_ts_label(self.mapping_ts_var, p.name)

    # ── Manual file pickers ────────────────────────────────────────────

    def _select_master(self) -> None:
        """
        Manually pick an Items Master file.

        Marks master as user-picked (not bundled); the bundled file in
        ``Calculation Data/`` is NOT touched — use "Update Bundled Files"
        for that.
        """
        path = filedialog.askopenfilename(
            title="Select Items Master file",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
        )
        if path:
            self.master_path = path
            self.master_is_bundled = False
            self.master_var.set(os.path.basename(path))
            # Clear the bundled timestamp — manual picks aren't tracked.
            self.master_ts_var.set("")
            self._log(f"Master (manual override): {os.path.basename(path)}")

    def _select_mapping(self) -> None:
        """
        Manually pick a Ship-To B2B mapping file. Bundled file untouched.
        """
        path = filedialog.askopenfilename(
            title="Select Mapping File (Ship-To B2B)",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
        )
        if path:
            self.mapping_path = path
            self.mapping_is_bundled = False
            self.mapping_var.set(os.path.basename(path))
            self.mapping_ts_var.set("")
            self._log(f"Mapping (manual override): {os.path.basename(path)}")

    def _select_po(self) -> None:
        """
        Pick the marketplace PO/punch file(s) for this run.

        v1.7.0: For Reliance, the dialog allows multi-file selection
        because each Reliance PO arrives as its own .xlsx file and
        users routinely process a batch of 5-10 POs at once. Other
        marketplaces stay single-select (Blink/Myntra/RK already
        consolidate multiple POs inside one file so there's no
        reason to complicate their upload flow).

        Result of the dialog is stored in two attributes:
            * ``self.po_path``  — first (or only) file path. Used by
              all the single-file display/log paths.
            * ``self.po_paths`` — list of all selected files.
              ``generate()`` uses this list when the marketplace
              config has ``pre_process`` set, which is the marker
              for "this marketplace supports multi-file upload".
        """
        marketplace = self.marketplace_var.get()
        supports_multi = bool(
            marketplace in MARKETPLACE_CONFIGS
            and MARKETPLACE_CONFIGS[marketplace].get('pre_process')
        )

        if supports_multi:
            # Multi-select — returns a tuple (possibly empty).
            paths = filedialog.askopenfilenames(
                title=f"Select {marketplace} PO File(s) — "
                      f"pick one or many",
                filetypes=[("Excel files", "*.xlsx"),
                           ("All files", "*.*")],
            )
            if not paths:
                return  # user cancelled
            self.po_paths = list(paths)
            self.po_path = self.po_paths[0]
            n = len(self.po_paths)
            if n == 1:
                self.po_var.set(os.path.basename(self.po_path))
                self._log(f"PO file: {os.path.basename(self.po_path)}")
            else:
                self.po_var.set(f"{n} files selected")
                self._log(f"PO files: {n} selected for batch upload")
                for p in self.po_paths:
                    self._log(f"  • {os.path.basename(p)}")
        else:
            # Single-select — existing behavior.
            path = filedialog.askopenfilename(
                title="Select Marketplace PO File",
                filetypes=[("Excel files", "*.xlsx"),
                           ("All files", "*.*")],
            )
            if path:
                self.po_path = path
                self.po_paths = [path]
                self.po_var.set(os.path.basename(path))
                self._log(f"PO file: {os.path.basename(path)}")

    # ── Main processing flow ───────────────────────────────────────────

    def generate(self) -> None:
        """Main action: load mapping → parse PO → generate output."""
        marketplace = self.marketplace_var.get()
        if not marketplace or marketplace not in MARKETPLACE_CONFIGS:
            messagebox.showwarning(
                "No Marketplace", "Please select a marketplace.",
            )
            return

        if not self.mapping_path:
            messagebox.showwarning(
                "No Mapping", "Please select the Ship-To mapping file.",
            )
            return

        if not self.po_path:
            messagebox.showwarning(
                "No PO File", "Please select the marketplace PO file.",
            )
            return

        config = MARKETPLACE_CONFIGS[marketplace]
        margin_pct = self._get_margin()
        start_time = time.time()

        self.status_var.set("Processing...")
        self.status_label.config(fg='blue')
        self.root.update()

        self._log(f"Marketplace: {marketplace} | "
                  f"Margin: {int(margin_pct * 100)}%")

        # ── Load mapping for this marketplace ───────────────────────────
        self._log(f"Loading mapping for '{marketplace}'...")
        warnings: List[Tuple[str, str, str]] = []
        loc_count = self.mapping_loader.load(
            self.mapping_path, config['party_name'], warnings,
        )

        if loc_count == 0:
            self._log("ERROR: No mapping locations found!")
            for _, _, msg in warnings:
                self._log(f"  {msg}")
            self.status_var.set("Failed — mapping load error")
            self.status_label.config(fg='red')
            return

        self._log(f"Loaded {loc_count} locations for {marketplace}")

        # v1.5.0: clear any previous run's stashed result BEFORE we do
        # anything else. If the new run fails anywhere below, the D365
        # and Email buttons will disable themselves — we don't want
        # them acting on stale data from the previous successful run.
        self.last_result = None
        self.d365_btn.config(state=tk.DISABLED)
        self.email_btn.config(state=tk.DISABLED)

        # ── Load Items_March (master) ───────────────────────────────────
        master_loader: Optional[MasterLoader] = None
        if self.master_path:
            self._log("Loading Items_March for validation...")
            master_loader = MasterLoader()
            try:
                item_count = master_loader.load(self.master_path)
                self._log(f"Loaded {item_count:,} items from master")
            except Exception as e:  # noqa: BLE001
                self._log(f"WARNING: Master load failed: {e} "
                          f"— skipping validation")
                master_loader = None

        # ── Engine run ──────────────────────────────────────────────────
        # v1.7.0: route to process_multi when the marketplace supports
        # batch upload (currently Reliance only, signalled by the
        # ``pre_process`` config key). For single-file marketplaces
        # we continue to call process() directly so there's no
        # behavioural change for Blink/Myntra/RK.
        engine = MarketplaceEngine(self.mapping_loader, master=master_loader)

        supports_multi = bool(config.get('pre_process'))
        if supports_multi and len(self.po_paths) > 1:
            self._log(
                f"Batch processing {len(self.po_paths)} "
                f"{marketplace} files..."
            )
            result = engine.process_multi(
                self.po_paths, config, margin_pct=margin_pct,
            )
        else:
            self._log(f"Processing {os.path.basename(self.po_path)}...")
            result = engine.process(
                self.po_path, config, margin_pct=margin_pct,
            )
        result.margin_pct = margin_pct  # redundant but explicit

        # v1.9.0: stamp the GUI's warehouse selection onto the result
        # so D365Exporter uses the right Location Code (col K + F)
        # and the Summary footer + email banner can show which
        # warehouse fulfilled this batch.
        selected_wh = self.warehouse_var.get()
        result.warehouse_display = selected_wh
        result.warehouse_code = WAREHOUSE_CODES.get(selected_wh, 'PICK')
        self._log(
            f"Warehouse: {selected_wh} → ERP code {result.warehouse_code}"
        )

        if not result.rows:
            self._log("ERROR: No valid rows extracted!")
            for _, _, msg in result.warnings:
                self._log(f"  WARNING: {msg}")
            self.status_var.set("Failed — no data extracted")
            self.status_label.config(fg='red')
            return

        # ── Log summary ─────────────────────────────────────────────────
        unique_pos = {r.po_number for r in result.rows}
        total_qty = sum(r.qty for r in result.rows)

        self._log(f"Extracted: {len(result.rows)} items, "
                  f"{len(unique_pos)} PO(s), {total_qty} total qty")
        if result.warnings:
            self._log(f"Warnings: {len(result.warnings)}")
            for po, _loc, msg in result.warnings[:5]:
                self._log(f"  [{po}] {msg}")
            if len(result.warnings) > 5:
                self._log(f"  ... and {len(result.warnings) - 5} more "
                          f"(see Warnings sheet)")

        # ── Export ──────────────────────────────────────────────────────
        self._log("Writing output...")
        output_path = self.exporter.export(result)

        elapsed = time.time() - start_time

        if output_path:
            self.last_output = output_path
            self.open_btn.config(state=tk.NORMAL)

            # v1.5.0: stash the result for the D365 + Email actions and
            # record elapsed time on the result itself (used by the
            # email report footer). Enable the downstream buttons.
            result.elapsed_seconds = elapsed
            self.last_result = result
            self.d365_btn.config(state=tk.NORMAL)
            self.email_btn.config(state=tk.NORMAL)

            status_msg = (f"Done — {len(result.rows)} items, "
                          f"{len(unique_pos)} PO(s), "
                          f"{total_qty} qty | {elapsed:.2f}s")
            if result.warnings:
                status_msg += f" | {len(result.warnings)} warning(s)"
                self.status_label.config(fg='orange')
            else:
                self.status_label.config(fg='darkgreen')

            self.status_var.set(status_msg)
            self._log(f"Saved: {output_path}")

            answer = messagebox.askyesno(
                "SO Generated",
                f"Sales Order generated successfully!\n\n"
                f"Marketplace : {marketplace}\n"
                f"PO(s)       : {len(unique_pos)}\n"
                f"Items       : {len(result.rows)}\n"
                f"Total Qty   : {total_qty}\n"
                f"Warnings    : {len(result.warnings)}\n"
                f"Time        : {elapsed:.2f}s\n\n"
                f"Do you want to open the output file?",
            )
            if answer:
                open_file(output_path)
        else:
            self.status_var.set("Failed — no output generated")
            self.status_label.config(fg='red')

    def open_last(self) -> None:
        """Open the last generated output file in the default app."""
        if self.last_output and self.last_output.exists():
            open_file(self.last_output)
        else:
            messagebox.showwarning("Not Found", "Output file not found.")

    # ── v1.5.0: D365 Package Export ──────────────────────────────────────

    def _export_d365(self) -> None:
        """
        Fill the D365 sample template with the last generated result.

        Flow:
            1. Guard: ``self.last_result`` must be populated (button is
               disabled when it isn't, so this is belt-and-suspenders).
            2. Warn the user about any PO(s) whose Ship-To mapping
               failed — those rows will be written with empty Location
               Code in the D365 output, which D365 may or may not
               accept depending on the company's location defaults.
            3. Prompt the user to pick the D365 template file.
            4. Delegate the actual fill to :class:`D365Exporter`.
            5. Offer to open the resulting file.

        The output file is written to the same ``output/`` folder as
        the main SO export, derived from the original PO file's
        directory — so both artefacts end up side by side.
        """
        result = self.last_result
        if result is None or not result.rows:
            messagebox.showwarning(
                "No Data",
                "Generate an SO successfully first before exporting to "
                "D365.",
            )
            return

        # ── Step 2: warn about unmapped POs ─────────────────────────────
        # A PO is "unmapped" if any of its rows came back with
        # ``mapped=False`` from the engine — meaning the facility name
        # didn't match any entry in Ship-To B2B for this marketplace.
        unmapped_pos = sorted({
            r.po_number for r in result.rows if not r.mapped
        })

        if unmapped_pos:
            preview = "\n".join(f"  • {p}" for p in unmapped_pos[:10])
            more = (
                f"\n  … and {len(unmapped_pos) - 10} more"
                if len(unmapped_pos) > 10 else ""
            )
            proceed = messagebox.askyesno(
                "⚠️ Unmapped Ship-To Locations",
                f"{len(unmapped_pos)} PO(s) have no Ship-To mapping "
                f"and will export with EMPTY Location Code:\n\n"
                f"{preview}{more}\n\n"
                f"D365 import may reject or warn on these rows.\n\n"
                f"Continue with export anyway?",
            )
            if not proceed:
                self._log("D365 export cancelled by user (unmapped POs).")
                return

        # ── Step 3: pick the D365 template ──────────────────────────────
        template_path = filedialog.askopenfilename(
            title="Select D365 Sample Package Template",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
        )
        if not template_path:
            return  # user cancelled the dialog

        # ── Step 4: delegate to exporter ────────────────────────────────
        # Output directory mirrors where the main SO workbook landed:
        # ``<punch_dir>/output/``. That keeps every artefact from one
        # run in a single predictable place.
        punch_dir = Path(result.input_file_path).parent
        output_dir = punch_dir / 'output'

        self._log(f"D365: filling template {os.path.basename(template_path)}...")
        self.status_var.set("D365 export in progress...")
        self.status_label.config(fg='blue')
        self.root.update()

        try:
            d365_path = self.d365_exporter.export(
                result, template_path, output_dir,
            )
        except Exception as e:  # noqa: BLE001
            # Defensive: D365Exporter swallows most errors and returns
            # None, but a truly unexpected bug shouldn't crash the GUI.
            logging.exception("D365 export crashed unexpectedly")
            self.status_var.set("D365 export failed")
            self.status_label.config(fg='red')
            messagebox.showerror(
                "D365 Export Failed",
                f"An unexpected error occurred:\n\n{e}",
            )
            return

        if d365_path is None:
            self.status_var.set("D365 export failed")
            self.status_label.config(fg='red')
            self._log("D365 export returned no file — check log for cause.")
            messagebox.showerror(
                "D365 Export Failed",
                "Could not produce the D365 import file. Check the log "
                "panel for details.",
            )
            return

        # ── Step 5: success popup ───────────────────────────────────────
        self._log(f"D365 file saved → {d365_path}")
        self.status_var.set(f"D365 export done — {d365_path.name}")
        self.status_label.config(fg='darkgreen')

        po_count = len({r.po_number for r in result.rows})
        item_count = len(result.rows)

        answer = messagebox.askyesno(
            "D365 Package Exported",
            f"D365 import file created successfully!\n\n"
            f"File  : {d365_path.name}\n"
            f"POs   : {po_count}\n"
            f"Items : {item_count}\n\n"
            f"Open the file now?",
        )
        if answer:
            open_file(d365_path)

    # ── v1.5.0: Email Report ─────────────────────────────────────────────

    def _send_email(self) -> None:
        """
        Send the HTML report email for the last generated result.

        Flow:
            1. Guard on ``self.last_result`` (defense-in-depth — the
               button is disabled when there's nothing to send).
            2. Freeze the UI with a "Sending..." status while SMTP is
               in flight.
            3. Ask :class:`EmailSender` to build + send the report.
            4. Show a success info box or a clear error box based on
               the return value.

        We don't disable the button during send — the root.update()
        freeze plus the status label are sufficient feedback, and
        an accidental double-click is handled fine by the stateless
        SMTP layer (it'll just send twice).
        """
        result = self.last_result
        if result is None or not result.rows:
            messagebox.showwarning(
                "No Data",
                "Generate an SO successfully first before emailing the "
                "report.",
            )
            return

        # ── Load effective config (defaults + optional JSON overrides)
        # Reloaded every send so edits to ``email_config.json`` take
        # effect without restarting the app.
        email_config = get_email_config()

        # ── Show "sending" state ────────────────────────────────────────
        self._log(
            f"Email: sending to {email_config['DEFAULT_RECIPIENT']}..."
        )
        self.status_var.set("Sending email...")
        self.status_label.config(fg='blue')
        self.root.update()

        # ── Dispatch ────────────────────────────────────────────────────
        ok, err = EmailSender.send(result, email_config)

        if ok:
            self.status_var.set("Email sent ✓")
            self.status_label.config(fg='darkgreen')
            self._log(
                f"Email sent OK → {email_config['DEFAULT_RECIPIENT']}"
                f" + {len(email_config.get('CC_RECIPIENTS', []))} CC"
            )

            cc_list = email_config.get('CC_RECIPIENTS', []) or []
            cc_display = ', '.join(cc_list) if cc_list else 'none'

            messagebox.showinfo(
                "Email Sent",
                f"Report sent successfully!\n\n"
                f"To : {email_config['DEFAULT_RECIPIENT']}\n"
                f"CC : {cc_display}",
            )
        else:
            self.status_var.set("Email failed")
            self.status_label.config(fg='red')
            self._log(f"Email failed: {err}")
            messagebox.showerror("Email Failed", err)

    # ── PO template download ──────────────────────────────────────────

    def _download_template(self) -> None:
        """
        Generate a blank PO template for the selected marketplace.

        Headers are colour-coded so the user knows which columns to
        actually fill in:

        * **BLUE** (``#1A237E``) — Required. Script fails without these.
        * **GREEN** (``#1B5E20``) — Validation. Used for price check +
          master lookup.
        * **GREY** (``#9E9E9E``) — Not read by script. Kept only to
          mirror the marketplace's native file format.

        Includes a 3-row legend below the header explaining each colour
        plus a final orange instruction line.

        The required set adapts to ``item_resolution``:
          * ``from_column`` → item_col is the required identifier.
          * ``from_ean`` → ean_col is the required identifier
            (promoted from GREEN to BLUE).
        """
        marketplace = self.marketplace_var.get()
        if not marketplace or marketplace not in MARKETPLACE_CONFIGS:
            messagebox.showwarning(
                "No Marketplace", "Please select a marketplace first.",
            )
            return

        config = MARKETPLACE_CONFIGS[marketplace]

        save_path = filedialog.asksaveasfilename(
            title=f"Save {marketplace} PO Template",
            defaultextension=".xlsx",
            initialfile=f"{marketplace}_PO_Template.xlsx",
            filetypes=[("Excel files", "*.xlsx")],
        )
        if not save_path:
            return

        try:
            self._write_template_workbook(save_path, marketplace, config)
            self._log(f"{marketplace} template saved → {save_path}")
            messagebox.showinfo(
                "Template Saved",
                f"{marketplace} PO template saved to:\n{save_path}\n\n"
                f"Header colours:\n"
                f"  • Blue  = Required (must fill)\n"
                f"  • Green = Validation (recommended)\n"
                f"  • Grey  = Not read by script",
            )
        except Exception as e:  # noqa: BLE001
            self._log(f"Template save failed: {e}")
            messagebox.showerror(
                "Error", f"Failed to save template:\n{e}",
            )

    @staticmethod
    def _write_template_workbook(save_path: str, marketplace: str,
                                   config: dict) -> None:
        """
        Actually build + save the PO template workbook.

        Split out of :meth:`_download_template` so the wrapping
        save-dialog / success-messagebox flow stays short and readable.
        """
        wb = Workbook()
        ws = wb.active
        ws.title = f'{marketplace} PO'

        # ── Determine required vs validation vs unused cols ─────────────
        # v1.5.5: column-config values can be either a scalar string
        # (most marketplaces) or a list of accepted aliases (Myntra's
        # ``po_col = ['PO', 'PO Number']``). For the template we need
        # ALL possible rendered header names to color-code correctly:
        #   * each alias itself (in case someone customises the
        #     template_headers to use one of the literal names)
        #   * the slash-joined combined label (e.g. 'PO/PO Number')
        #     which is what we put in template_headers so the user
        #     sees both options and knows to rename before uploading.
        item_resolution = config.get('item_resolution', 'from_column')

        def _normalize(val):
            """Turn a config column value into the set of names that,
            if seen as a template header, should be coloured as this
            column. Handles both scalar strings and list aliases."""
            if val is None:
                return set()
            if isinstance(val, list):
                names = set(val)
                # Also include the 'A/B' slash label so templates that
                # combine aliases as a single column header (the
                # recommended pattern for user-facing templates) get
                # coloured correctly.
                names.add('/'.join(val))
                return names
            return {val}

        required_cols: set = set()
        required_cols |= _normalize(config.get('po_col'))
        required_cols |= _normalize(config.get('loc_col'))
        required_cols |= _normalize(config.get('qty_col'))

        if item_resolution == 'from_ean':
            # EAN is the required identifier when Item No is resolved
            # from it — promote from GREEN to BLUE.
            required_cols |= _normalize(config.get('ean_col'))
        else:  # 'from_column'
            required_cols |= _normalize(config.get('item_col'))

        validation_cols: set = set()
        ean_names = _normalize(config.get('ean_col'))
        if ean_names and not (ean_names & required_cols):
            validation_cols |= ean_names
        validation_cols |= _normalize(config.get('fob_col'))

        # ── Build column list ───────────────────────────────────────────
        # Prefer the full marketplace template_headers; fall back to a
        # minimal list built from required + validation columns.
        # v1.5.5: list-valued column configs collapse to their
        # slash-joined combined label here ('A/B' form), matching the
        # convention used in the canonical template_headers.
        def _primary_label(val):
            if val is None:
                return None
            if isinstance(val, list):
                return '/'.join(val)
            return val

        headers = config.get('template_headers')
        if not headers:
            headers = [
                _primary_label(config.get('po_col')),
                _primary_label(config.get('loc_col')),
            ]
            if (item_resolution == 'from_column'
                    and config.get('item_col')):
                headers.append(_primary_label(config.get('item_col')))
            headers.append(_primary_label(config.get('qty_col')))
            ean_label = _primary_label(config.get('ean_col'))
            if ean_label and ean_label not in headers:
                headers.append(ean_label)
            fob_label = _primary_label(config.get('fob_col'))
            if fob_label and fob_label not in headers:
                headers.append(fob_label)
            # Filter out None in case any of the above weren't
            # configured — shouldn't happen for a valid marketplace
            # config but belt-and-braces.
            headers = [h for h in headers if h]

        # ── Styles per role ─────────────────────────────────────────────
        required_fill = PatternFill('solid', fgColor='1A237E')   # blue
        validation_fill = PatternFill('solid', fgColor='1B5E20')  # green
        unused_fill = PatternFill('solid', fgColor='9E9E9E')       # grey

        hdr_font_white = Font(bold=True, color='FFFFFF',
                               name='Aptos Display', size=11)
        hdr_font_dim = Font(bold=True, color='EEEEEE',
                             name='Aptos Display', size=11, italic=True)

        # ── Header row (colour-coded) ───────────────────────────────────
        for c, h in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=c, value=h)
            if h in required_cols:
                cell.fill = required_fill
                cell.font = hdr_font_white
            elif h in validation_cols:
                cell.fill = validation_fill
                cell.font = hdr_font_white
            else:
                cell.fill = unused_fill
                cell.font = hdr_font_dim
            cell.alignment = Alignment(horizontal='center', vertical='center')
            ws.column_dimensions[get_column_letter(c)].width = max(
                len(h) + 4, 12,
            )

        # ── Legend rows (3-5) ───────────────────────────────────────────
        # v1.5.5: for list-valued column configs, ``required_cols``
        # contains BOTH every alias AND the slash-joined combined
        # label. When we list the required/validation column NAMES
        # in the legend we only want the user-facing labels — the
        # ones that actually appear in the header row. So we filter
        # to the intersection with ``headers``.
        header_set = set(headers)
        required_labels = sorted(required_cols & header_set)
        validation_labels = sorted(validation_cols & header_set)

        legend_row = 3
        legend_items = [
            ('1A237E', 'FFFFFF', 'REQUIRED',
             f'Script fails without these — fill them in: '
             f'{", ".join(required_labels)}'),
            ('1B5E20', 'FFFFFF', 'VALIDATION',
             f'Used for price check & master lookup: '
             f'{", ".join(validation_labels) or "(none)"}'),
            ('9E9E9E', 'FFFFFF', 'NOT READ',
             'Optional — kept only to match marketplace file format; '
             'can stay blank'),
        ]
        for fg, fc, label, desc in legend_items:
            tag = ws.cell(row=legend_row, column=1, value=label)
            tag.fill = PatternFill('solid', fgColor=fg)
            tag.font = Font(bold=True, color=fc,
                             name='Aptos Display', size=10)
            tag.alignment = Alignment(horizontal='center')

            desc_cell = ws.cell(row=legend_row, column=2, value=desc)
            desc_cell.font = Font(name='Aptos Display', size=10,
                                    color='333333', italic=True)
            ws.merge_cells(start_row=legend_row, start_column=2,
                            end_row=legend_row,
                            end_column=min(8, len(headers)))
            legend_row += 1

        # v1.5.5: extra instruction row for marketplaces that use
        # slash-joined column labels (e.g. 'PO/PO Number'). Tells the
        # user to rename the header to one of the options before
        # uploading, so the engine's alias matcher can find it.
        slash_labels = [h for h in headers if '/' in h]
        if slash_labels:
            msg = (
                "⚠ Rename '" + "', '".join(slash_labels) + "' "
                "column(s) to ONE of the listed options (e.g. 'PO' OR "
                "'PO Number') before uploading. The script accepts "
                "either name but the header must be a single choice."
            )
            rename_cell = ws.cell(
                row=legend_row, column=1, value=msg,
            )
            rename_cell.font = Font(
                name='Aptos Display', size=10,
                color='C62828', italic=True, bold=True,
            )
            ws.merge_cells(
                start_row=legend_row, start_column=1,
                end_row=legend_row,
                end_column=min(8, len(headers)),
            )
            legend_row += 1

        # ── Final orange instruction row ────────────────────────────────
        ws.cell(
            row=legend_row + 1, column=1,
            value=(f'← {marketplace} PO template. Fill data rows below '
                   f'the header. Only the BLUE & GREEN columns are read '
                   f'by the script.'),
        ).font = Font(name='Aptos Display', size=10,
                      color='FF6600', italic=True)

        ws.freeze_panes = 'A2'

        # ── v1.8.1: 'How this works' reference sheet ────────────────────
        # Adds a second sheet to the template workbook that documents
        # exactly what the engine does with this marketplace's data:
        # which columns are consulted, what the match mode is
        # (exact / case-insensitive / alias list), what each column
        # contributes to the output (lookup / calculation / reference
        # only). The goal is that opening the template answers the
        # question "which of these columns actually matter?" without
        # having to read Python code.
        OnlinePOApp._append_how_this_works_sheet(wb, marketplace, config)

        wb.save(save_path)

    @staticmethod
    def _append_how_this_works_sheet(wb, marketplace: str,
                                       config: dict) -> None:
        """
        Append a readable reference sheet explaining the engine's
        behavior for this marketplace.

        Layout::

            Row 1    : title banner
            Row 2    : blank
            Row 3    : "OVERVIEW" subheading
            Row 4-5  : paragraph covering sheet selection, compare
                       basis, HSN check, pre-processing, case-
                       insensitivity — whichever apply
            Row 6    : blank
            Row 7    : per-column reference table header
            Row 8+   : one row per column with 4 fields:
                         Column header in file
                         Required / Optional / Not read
                         Match mode (Exact / Case-insensitive /
                           Alias list / Synthetic)
                         What the engine does with it
            Row N    : blank
            Row N+1  : footer note about version + generation date

        The table is deterministic — derived from the same config
        dict the engine uses — so any config change auto-updates the
        docs on the next template download.
        """
        from datetime import datetime as _dt
        from openpyxl.styles import Alignment as _Align

        ref_ws = wb.create_sheet('How this works')

        # Colors: match the main template's scheme so users recognise
        # them instantly.
        BLUE = PatternFill('solid', fgColor='001A237E')     # Required
        GREEN = PatternFill('solid', fgColor='FF2E7D32')    # Validation
        GREY = PatternFill('solid', fgColor='FFBDBDBD')     # Not read
        SYN = PatternFill('solid', fgColor='FF6A1B9A')      # Synthetic
        HEADER_FILL = PatternFill('solid', fgColor='FF37474F')
        WHITE_FONT = Font(color='FFFFFFFF', bold=True,
                           name='Aptos Display', size=11)
        TITLE_FONT = Font(color='FF1A237E', bold=True,
                           name='Aptos Display', size=16)
        SECTION_FONT = Font(color='FF1A237E', bold=True,
                             name='Aptos Display', size=12)
        BODY_FONT = Font(name='Aptos Display', size=10)

        # ── Row 1: title ────────────────────────────────────────────────
        ref_ws.cell(row=1, column=1,
                     value=f'{marketplace} — How this template is read')
        ref_ws.cell(row=1, column=1).font = TITLE_FONT
        ref_ws.merge_cells('A1:D1')

        # ── Row 3: overview heading + paragraph ─────────────────────────
        ref_ws.cell(row=3, column=1, value='OVERVIEW').font = SECTION_FONT

        # Build the overview paragraph dynamically based on what the
        # config opts into.
        lines = []

        # Which sheet / file structure.
        source_sheet = config.get('source_sheet', 'Sheet1')
        if source_sheet.endswith('*'):
            prefix = source_sheet[:-1]
            lines.append(
                f"• Data sheet: any sheet name starting with "
                f"'{prefix}' (wildcard match)."
            )
        else:
            lines.append(f"• Data sheet: '{source_sheet}' (exact).")

        header_row = config.get('header_row', 0)
        if header_row:
            lines.append(
                f"• Column headers are on row {header_row + 1} (the "
                f"file's title/merged-cell occupies row 1)."
            )

        # Pre-processor (Reliance).
        if config.get('pre_process') == 'reliance_po_sheet':
            lines.append(
                "• PO number and location are parsed from the title "
                "cell on row 1 ('<PO>  <Location>') and injected onto "
                "every data row."
            )

        # Case-insensitivity.
        if config.get('case_insensitive_cols'):
            lines.append(
                "• Column headers are matched case-insensitively and "
                "whitespace-tolerantly. 'HSN', 'Hsn', 'hsn' all match; "
                "'PO Number' and 'Po  Number' (double space) also "
                "match. Only applies to this marketplace."
            )

        # Compare basis explainer.
        compare_basis = config.get('compare_basis', 'cost')
        compare_label = config.get('compare_label', 'Cost')
        if compare_basis == 'landing':
            lines.append(
                f"• Price check basis: 'landing' — marketplace's "
                f"'{compare_label}' column is compared against "
                f"MRP × margin% (pre-GST)."
            )
        else:
            lines.append(
                f"• Price check basis: 'cost' — marketplace's "
                f"'{compare_label}' column is compared against "
                f"MRP × margin% ÷ GST divisor (post-GST)."
            )

        # HSN check.
        if config.get('hsn_col'):
            lines.append(
                f"• HSN cross-check: ENABLED. The '{config['hsn_col']}' "
                f"column is compared against 'HSN/SAC Code' in "
                f"Items_March master. Mismatches surface on the "
                f"Validation sheet and in Warnings."
            )

        # Amount formula.
        amount_col = config.get('amount_col')
        if amount_col:
            if isinstance(amount_col, dict) and amount_col.get('multiply'):
                factors = amount_col['multiply']
                formula = ' × '.join(factors)
                if amount_col.get('apply_margin'):
                    formula += ' × margin%'
                lines.append(
                    f"• Amount per row: computed as {formula}."
                )
            elif isinstance(amount_col, str):
                lines.append(
                    f"• Amount per row: read directly from "
                    f"'{amount_col}' column."
                )

        # v1.9.1: D365 Unit Price override (BlinkMP).
        if config.get('override_unit_price'):
            lines.append(
                "• D365 Sales Line Unit Price (col H): OVERRIDDEN — "
                "populated with our computed post-GST Cost Price "
                "per row. Forces the ERP to record the correct "
                "cost (e.g. BlinkMP's 75% margin) instead of the "
                "vendor master's default (BCPL's 70% for Blinkit)."
            )

        # Render the lines.
        for i, line in enumerate(lines):
            cell = ref_ws.cell(row=4 + i, column=1, value=line)
            cell.font = BODY_FONT
            cell.alignment = _Align(wrap_text=True, vertical='top')
            ref_ws.merge_cells(start_row=4 + i, start_column=1,
                                end_row=4 + i, end_column=4)

        # ── Per-column reference table ──────────────────────────────────
        table_start = 4 + len(lines) + 2
        ref_ws.cell(row=table_start, column=1,
                     value='COLUMN REFERENCE').font = SECTION_FONT

        # Table header row
        hdr_row = table_start + 1
        for col_idx, label in enumerate(
            ['Column header', 'Required?', 'Match mode',
             'What the engine does with it'], start=1,
        ):
            c = ref_ws.cell(row=hdr_row, column=col_idx, value=label)
            c.fill = HEADER_FILL
            c.font = WHITE_FONT
            c.alignment = _Align(horizontal='center', vertical='center')

        # Build the table rows from the config.
        case_insensitive = bool(config.get('case_insensitive_cols'))
        match_mode_scalar = (
            'Case-insensitive' if case_insensitive else 'Exact'
        )
        rows = []

        def _add_row(col_val, required, fill, explanation):
            """Helper: append one ref-table row for a config column."""
            if col_val is None:
                return
            if isinstance(col_val, list):
                display = ' / '.join(col_val)
                match_mode = f'Alias list ({match_mode_scalar})'
            elif col_val.startswith('__') and col_val.endswith('__'):
                display = col_val
                match_mode = 'Synthetic (set by pre-processor)'
            else:
                display = col_val
                match_mode = match_mode_scalar
            rows.append((display, required, match_mode, explanation, fill))

        # Required columns.
        _add_row(config.get('po_col'), 'Required', BLUE,
                  'PO number — grouped and emitted in the SO output.')
        _add_row(config.get('loc_col'), 'Required', BLUE,
                  'Delivery location — looked up in Ship-To B2B to '
                  'get Cust No + Ship-to code.')
        _add_row(config.get('qty_col'), 'Required', BLUE,
                  'Quantity per line.')

        item_resolution = config.get('item_resolution', 'from_column')
        if item_resolution == 'from_ean':
            _add_row(config.get('ean_col'), 'Required', BLUE,
                      'EAN/GTIN — looked up in Items_March master '
                      'to resolve the ERP Item No.')
        else:
            _add_row(config.get('item_col'), 'Required', BLUE,
                      'ERP Item No — used directly (no master '
                      'resolution needed).')

        # Validation columns.
        _add_row(config.get('fob_col'), 'Validation', GREEN,
                  f"Marketplace's stated {config.get('compare_label', 'price')} "
                  f"per unit — compared against our calculated price "
                  f"to produce the Validation sheet's diff.")

        ref_fob = config.get('ref_fob_col')
        if ref_fob:
            _add_row(ref_fob, 'Optional', GREEN,
                      'Reference-only comparison column. Surfaces a '
                      'parallel diff on the Raw Data sheet but does '
                      'NOT affect OK/MISMATCH status.')

        if config.get('hsn_col'):
            _add_row(config.get('hsn_col'), 'Validation', GREEN,
                      "Marketplace's HSN code — compared against "
                      "Items_March's 'HSN/SAC Code'. Mismatches are "
                      "flagged but don't abort the run.")

        # Amount column (when scalar).
        if isinstance(amount_col, str):
            _add_row(amount_col, 'Validation', GREEN,
                      'Per-row amount — summed into the email '
                      "report's Amount stat and the Summary sheet's "
                      'Total Amount column.')

        # Write the table body.
        for r_offset, (disp, req, mode, explain, fill) in enumerate(rows, start=1):
            r = hdr_row + r_offset
            # Column 1 — header cell with required-color fill
            hdr_cell_fill = PatternFill('solid',
                                          fgColor=fill.fgColor.rgb)
            hc = ref_ws.cell(row=r, column=1, value=disp)
            hc.fill = hdr_cell_fill
            hc.font = Font(color='FFFFFFFF', bold=True,
                           name='Aptos Display', size=10)
            hc.alignment = _Align(vertical='center')

            # Column 2 — required / optional / not read
            rc = ref_ws.cell(row=r, column=2, value=req)
            rc.font = BODY_FONT

            # Column 3 — match mode
            mc = ref_ws.cell(row=r, column=3, value=mode)
            mc.font = BODY_FONT

            # Column 4 — what the engine does (wraps)
            ec = ref_ws.cell(row=r, column=4, value=explain)
            ec.font = BODY_FONT
            ec.alignment = _Align(wrap_text=True, vertical='center')

        # ── Footer ──────────────────────────────────────────────────────
        footer_row = hdr_row + len(rows) + 2
        ref_ws.cell(
            row=footer_row, column=1,
            value=(f'Generated by Online PO Processor for {marketplace}. '
                   f'This description is derived from the live engine '
                   f'config — if behavior changes in a future version, '
                   f'downloading a fresh template will reflect it.'),
        ).font = Font(italic=True, color='FF6A6A6A',
                       name='Aptos Display', size=9)
        ref_ws.merge_cells(start_row=footer_row, start_column=1,
                            end_row=footer_row, end_column=4)

        # ── Column widths ──────────────────────────────────────────────
        ref_ws.column_dimensions['A'].width = 32
        ref_ws.column_dimensions['B'].width = 13
        ref_ws.column_dimensions['C'].width = 22
        ref_ws.column_dimensions['D'].width = 60

    # ── Run the app ────────────────────────────────────────────────────

    def run(self) -> None:
        """Start the Tkinter main loop. Blocks until the window closes."""
        self.root.mainloop()