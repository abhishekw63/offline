"""
exporter.sheets.raw_data_sheet
==============================

Writes the **Raw Data** sheet — verbatim copy of the marketplace's punch
file on the left, our calculated validation columns appended on the
right. Gives a single-workbook view of "what they sent us + what we
computed" so the user can audit without flipping between files.

Column layout
-------------
Left side: **all** columns from the input punch file, in original order,
with a slate-grey header fill. Values are copied with minimal
transformation — only pandas Timestamps are formatted as
``dd-mm-yyyy`` and NaN cells become empty strings.

Right side: **calculated** columns with a green header fill::

    Item No (Master) | MRP | Landing (m%) | GST Code | Cost Price
    [ Diffn with <ref_fob_col> ]?      ← OPTIONAL reference Diffn
    Diffn with <compare_label>         ← ACTIVE  (validation-driving)

The reference Diffn column appears only when ``ref_fob_col`` is set in
the marketplace config AND present in the punch file. Myntra uses this
to surface the legacy "Diffn with List price(FOB+Transport-Excise)"
alongside the active "Diffn with Landing Rate". The reference Diffn
uses a muted slate header and pale-grey row tint so the user can tell
at a glance it's informational, not validation-driving.

Row matching
------------
For each raw row we need to find the corresponding SORow that was
emitted by the engine. We build ``validation_lookup`` indexed under
BOTH ``(po, item_no)`` and ``(po, ean)``, so row matching works
whether the punch file has Item No (RK-style) or only EAN (Myntra-
style). The per-row code tries ``item_col`` first, then falls back to
``ean_col``.

Row styling
-----------
* All calc cells on a row share one background tint:
    - ``RAW_MISMATCH_BG`` (light red) when ``validation_status ==
      'MISMATCH'``
    - ``CALC_BG`` (light green) otherwise
* Reference Diffn cells additionally get ``REF_BG`` (pale grey) on
  non-mismatched rows — distinguishes them from the active Diffn
  without competing for attention on actually-problematic rows.
* Active Diffn gets bold red font on mismatch rows for emphasis.
"""

from __future__ import annotations
from typing import Dict, List, Tuple

import pandas as pd

from online_po_processor.config.marketplaces import MARKETPLACE_CONFIGS
from online_po_processor.data.models import ProcessingResult, SORow
from online_po_processor.exporter._styles import (
    CALC_BG, CALC_FILL, HEADER_FONT, MISMATCH_TEXT_FONT,
    RAW_HDR_FILL, RAW_MISMATCH_BG, REF_BG, REF_FILL, BORDER,
    auto_width, data_cell,
)

from openpyxl.styles import Alignment


def write(wb, result: ProcessingResult) -> None:
    """
    Append the 'Raw Data' sheet to ``wb``.

    No-op if ``result.raw_df`` is missing or empty (e.g. the engine
    couldn't read the input file).
    """
    if result.raw_df is None or result.raw_df.empty:
        return

    ws = wb.create_sheet('Raw Data')
    df = result.raw_df

    # Dual-keyed validation lookup — SORow by (po, item_no) AND (po, ean)
    validation_lookup = _build_validation_lookup(result)

    # ── Find the marketplace's config (for col names + ref_fob_col) ─────
    # v1.5.6: prefer the alias-resolved config stashed on the result
    # by the engine — it has list-valued columns (e.g. Myntra's
    # ``po_col = ['PO', 'PO Number']``) already collapsed to the
    # single name that actually appeared in the uploaded file.
    # Falling back to MARKETPLACE_CONFIGS here would re-introduce
    # the list and crash ``col in df.columns`` later.
    marketplace_cfg = result.resolved_config or next(
        (cfg for cfg in MARKETPLACE_CONFIGS.values()
         if cfg['party_name'] == result.marketplace),
        None,
    )
    ref_fob_col_name = (marketplace_cfg or {}).get('ref_fob_col')
    has_ref_diff = bool(ref_fob_col_name) and ref_fob_col_name in df.columns

    # ── v1.7.0: Source column + synthetic column handling ───────────────
    # Reliance's engine pre-processor injects '__po__' and '__loc__'
    # columns onto every data row (because the source file carries
    # those values out-of-band in a merged title cell, not in per-row
    # columns). For multi-file batches the merger also adds
    # '__source_file__' so we can trace which upload produced each
    # row.
    #
    # These are INTERNAL markers — we don't want them to appear in
    # the user-facing Raw Data sheet with the underscore names. What
    # we DO want is a single readable "Source" column that surfaces
    # the PO + location together ("5000466441 BHIWANDI (Reliance)") so
    # users scanning a 125-row combined batch can tell which rows
    # came from which PO at a glance.
    #
    # Implementation:
    #   * If any synthetic column exists, build a combined Source
    #     value per row and prepend it as a proper first column.
    #   * Filter the synthetic columns out of the displayed DataFrame
    #     so they don't appear twice.
    SYNTHETIC_COLS = {'__po__', '__loc__', '__source_file__'}
    has_source = (
        '__po__' in df.columns and '__loc__' in df.columns
    )

    # Build per-row source values BEFORE filtering the columns out.
    source_values: List[str] = []
    if has_source:
        for _, r in df.iterrows():
            po_val = r.get('__po__', '')
            loc_val = r.get('__loc__', '')
            po_str = str(po_val).strip() if pd.notna(po_val) else ''
            loc_str = str(loc_val).strip() if pd.notna(loc_val) else ''
            if po_str and loc_str:
                source_values.append(f"{po_str} {loc_str}")
            elif po_str:
                source_values.append(po_str)
            else:
                source_values.append('')

    # Display DataFrame: everything except synthetic marker columns.
    display_cols = [c for c in df.columns if c not in SYNTHETIC_COLS]
    source_offset = 1 if has_source else 0  # Source lands in col 1 if present

    # ── Header row ──────────────────────────────────────────────────────
    # If we're showing a Source column, it occupies column 1 with a
    # distinctive fill so it reads as meta (not marketplace data).
    if has_source:
        src_cell = ws.cell(row=1, column=1, value='Source')
        src_cell.font = HEADER_FONT
        src_cell.fill = CALC_FILL  # green — "calculated/derived", not raw
        src_cell.alignment = Alignment(horizontal='center', vertical='center')
        src_cell.border = BORDER

    # Original headers (shifted right by source_offset).
    for i, col_name in enumerate(display_cols):
        col_idx = source_offset + i + 1
        cell = ws.cell(row=1, column=col_idx, value=str(col_name))
        cell.font = HEADER_FONT
        cell.fill = RAW_HDR_FILL
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = BORDER

    orig_col_count = source_offset + len(display_cols)

    # ── Calc (right) header labels ──────────────────────────────────────
    diff_label = (f"Diffn with {result.compare_label}"
                  if result.compare_label else "Diffn")
    calc_headers = [
        'Item No (Master)',
        'MRP',
        f'Landing ({int(result.margin_pct * 100)}%)',
        'GST Code',
        'Cost Price',
    ]
    if has_ref_diff:
        # Reference Diffn goes BEFORE the active Diffn so the active
        # column is rightmost — that's the one the user glances at last
        # to confirm the status, so it gets visual precedence.
        calc_headers.append(f'Diffn with {ref_fob_col_name}')
    calc_headers.append(diff_label)

    # 0-based indices of reference and active Diffn columns within
    # calc_headers. ref_idx is None when no reference diff is shown.
    ref_idx = (len(calc_headers) - 2) if has_ref_diff else None
    active_idx = len(calc_headers) - 1

    # ── Calc header row ─────────────────────────────────────────────────
    for i, header_text in enumerate(calc_headers):
        col_idx = orig_col_count + i + 1
        cell = ws.cell(row=1, column=col_idx, value=header_text)
        cell.font = HEADER_FONT
        # Reference header gets the muted slate fill, everything else green.
        cell.fill = REF_FILL if (has_ref_diff and i == ref_idx) else CALC_FILL
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = BORDER

    # ── Data rows ───────────────────────────────────────────────────────
    config_item_col = (marketplace_cfg or {}).get('item_col')
    config_ean_col = (marketplace_cfg or {}).get('ean_col')
    config_po_col = (marketplace_cfg or {}).get('po_col')

    n_calc = len(calc_headers)
    base_c = orig_col_count + 1  # first calc column (1-based)

    for r, (_idx, raw_row) in enumerate(df.iterrows(), start=2):
        # v1.7.0: write the Source cell (col 1) if it's present.
        # Uses the precomputed values from source_values rather than
        # re-deriving per row.
        if has_source:
            data_cell(ws, r, 1, source_values[r - 2])

        # Write the marketplace's own raw columns, skipping the
        # synthetic __po__/__loc__/__source_file__ markers that we
        # already distilled into the Source column.
        _write_raw_row(
            ws, r, raw_row, display_cols, start_col=source_offset + 1,
        )

        # Find this row's validation match. The PO-column rename for
        # Reliance (po_col='__po__') is fine here because the raw_row
        # still carries __po__ even though we hid it from display.
        po_val, lookup_val = _derive_row_key(
            raw_row, df, config_po_col, config_item_col, config_ean_col,
        )
        vrow = validation_lookup.get((po_val, lookup_val))

        if vrow is not None:
            _write_calc_cells(
                ws, r, base_c, vrow, result,
                has_ref_diff, ref_idx, active_idx, n_calc,
            )
        else:
            # Likely a zero-qty row that the engine skipped — leave the
            # calc cells blank so the row aligns visually.
            for i in range(n_calc):
                data_cell(ws, r, base_c + i, '')

    auto_width(ws)
    ws.freeze_panes = 'A2'


# ── Helpers ────────────────────────────────────────────────────────────

def _build_validation_lookup(
    result: ProcessingResult,
) -> Dict[Tuple[str, str], SORow]:
    """
    Build the dual-keyed lookup table: SORow indexed by BOTH
    ``(po, item_no)`` AND ``(po, ean)`` so row matching works whether
    the raw file has Item No or only EAN.
    """
    lookup: Dict[Tuple[str, str], SORow] = {}
    for so_row in result.rows:
        lookup[(so_row.po_number, str(so_row.item_no))] = so_row
        if so_row.ean:
            lookup[(so_row.po_number, str(so_row.ean))] = so_row
    return lookup


def _write_raw_row(
    ws, r: int, raw_row: pd.Series,
    display_cols: List[str], start_col: int = 1,
) -> None:
    """
    Copy ``raw_row`` into sheet row ``r``, writing only the columns
    named in ``display_cols`` and starting at spreadsheet column
    ``start_col``.

    v1.7.0 reworked to take an explicit column list rather than
    reading ``df.columns`` so callers can exclude synthetic columns
    (``__po__``/``__loc__``/``__source_file__``) from the display
    while keeping them available on ``raw_row`` for lookups.

    Timestamps are formatted ``dd-mm-yyyy``. NaNs become empty strings.
    Everything else is passed through to openpyxl as-is.

    Args:
        ws:            Target worksheet.
        r:             Spreadsheet row number (1-based).
        raw_row:       Series for this row — may contain columns not
                       in ``display_cols``; those are ignored.
        display_cols:  Ordered list of column names to write.
        start_col:     Spreadsheet column (1-based) to start writing
                       at. Used to leave room for a leading Source
                       column when Reliance's pre-process ran.
    """
    for col_offset, col_name in enumerate(display_cols):
        val = raw_row[col_name]
        if isinstance(val, pd.Timestamp):
            val = val.strftime('%d-%m-%Y')
        elif pd.isna(val):
            val = ''
        data_cell(ws, r, start_col + col_offset, val)


def _derive_row_key(raw_row: pd.Series, df: pd.DataFrame,
                     po_col: str | None, item_col: str | None,
                     ean_col: str | None) -> Tuple[str, str]:
    """
    Build the ``(po, item-or-ean)`` key for this raw row.

    Strategy: read PO; try item_col first; if missing/empty, fall back to
    ean_col with the same float→int cleanup the engine uses
    (``8906121642599.0`` → ``'8906121642599'``).
    """
    po_val = (str(raw_row[po_col]).strip()
              if po_col and po_col in df.columns else '')

    lookup_val = ''

    # Try item column
    if item_col and item_col in df.columns:
        iv = raw_row[item_col]
        if pd.notna(iv):
            try:
                lookup_val = str(int(iv))
            except (ValueError, TypeError):
                lookup_val = str(iv).strip()

    # Fall back to EAN (with float→int cleanup)
    if not lookup_val and ean_col and ean_col in df.columns:
        ev = raw_row[ean_col]
        if pd.notna(ev):
            if isinstance(ev, (int, float)):
                try:
                    lookup_val = str(int(ev))
                except (ValueError, OverflowError):
                    lookup_val = str(ev).strip()
            else:
                lookup_val = str(ev).strip()

    return po_val, lookup_val


def _write_calc_cells(
    ws, r: int, base_c: int, vrow: SORow, result: ProcessingResult,
    has_ref_diff: bool, ref_idx: int | None, active_idx: int,
    n_calc: int,
) -> None:
    """
    Write the 6 or 7 appended calc cells for a single row and apply the
    row's background tint.

    Column order (base_c + offset)::

        +0  Item No (Master)
        +1  MRP
        +2  Landing (m%)
        +3  GST Code
        +4  Cost Price
        +ref_idx     (optional)  Reference Diffn
        +active_idx  (always)    Active Diffn
    """
    landing = (float(vrow.mrp) * result.margin_pct
               if vrow.mrp and not pd.isna(vrow.mrp) else None)

    data_cell(ws, r, base_c + 0, vrow.item_no)
    data_cell(ws, r, base_c + 1, vrow.mrp, '#,##0.00')
    data_cell(ws, r, base_c + 2,
               round(landing, 2) if landing else '', '#,##0.00')
    data_cell(ws, r, base_c + 3, vrow.gst_code)
    data_cell(ws, r, base_c + 4,
               round(vrow.cost_price_ref, 2) if vrow.cost_price_ref else '',
               '#,##0.00')

    if has_ref_diff and ref_idx is not None:
        data_cell(ws, r, base_c + ref_idx,
                   round(vrow.ref_diffn, 2)
                   if vrow.ref_diffn is not None else '',
                   '#,##0.00')

    data_cell(ws, r, base_c + active_idx,
               round(vrow.diffn, 2) if vrow.diffn is not None else '',
               '#,##0.00')

    # Apply row tint across the whole calc block
    is_mismatch = vrow.validation_status == 'MISMATCH'
    fill = RAW_MISMATCH_BG if is_mismatch else CALC_BG
    for i in range(n_calc):
        ws.cell(row=r, column=base_c + i).fill = fill

    # Re-tint reference Diffn grey on clean rows so it's visibly
    # distinct from the active column. Skipped on mismatch rows so the
    # red fill shows across everything (mismatch = user attention
    # priority; visual hierarchy beats column differentiation).
    if has_ref_diff and not is_mismatch and ref_idx is not None:
        ws.cell(row=r, column=base_c + ref_idx).fill = REF_BG

    # Bold red font on the active Diffn when mismatched
    if is_mismatch:
        ws.cell(row=r, column=base_c + active_idx).font = MISMATCH_TEXT_FONT