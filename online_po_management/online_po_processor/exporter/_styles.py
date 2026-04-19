"""
exporter._styles
================

Shared formatting constants and small helper functions used by every
sheet writer.

Single source of truth for colours, fonts, and the standard
``_hdr_cell`` / ``_data_cell`` / ``_auto_width`` patterns. Per-sheet
modules import from here rather than re-defining their own ``Font``
objects so the workbook has a consistent visual language.

Color palette
-------------
::

    Hex        Role                                         Used by
    ─────────  ───────────────────────────────────────────  ─────────────────
    1A237E     Standard header (deep blue)                  most sheets
    E65100     Warning header (orange)                      Warnings sheet
    37474F     Raw passthrough header (slate)               Raw Data (left)
    1B5E20     Calculated/validation header (green)         Validation, Raw
    455A64     Reference column header (muted slate)        Raw Data ref Diffn
    00C853     OK status pill (green)                       Summary, Validation
    FF5252     UNMAPPED pill (red)                          Summary
    E8F5E9     Calc cells row tint (light green)            Raw Data
    FFEBEE     Validation MISMATCH row tint (pink)          Validation
    FFCDD2     Raw Data MISMATCH row tint (light red)       Raw Data
    FFF3E0     NOT_IN_MASTER row tint (light orange)        Validation
    ECEFF1     Reference Diffn row tint (light grey)        Raw Data
    FFF59D     Location mismatch tint (pale yellow)         Summary
    D32F2F     Mismatch text emphasis (red)                 Validation, Raw
    9E9E9E     Template "not read" header                   PO template only
"""

from __future__ import annotations
from typing import Optional

from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


# ── Fills ──────────────────────────────────────────────────────────────────

HEADER_FILL = PatternFill('solid', fgColor='1A237E')   # deep blue
WARN_FILL = PatternFill('solid', fgColor='E65100')     # orange — Warnings sheet
RAW_HDR_FILL = PatternFill('solid', fgColor='37474F')  # slate — Raw Data passthrough headers
CALC_FILL = PatternFill('solid', fgColor='1B5E20')     # green — calc/validation headers
REF_FILL = PatternFill('solid', fgColor='455A64')      # muted slate — reference Diffn header

# Row-tint fills (applied to data cells, not headers)
OK_FILL = PatternFill('solid', fgColor='E8F5E9')
MISMATCH_FILL = PatternFill('solid', fgColor='FFEBEE')
NO_MASTER_FILL = PatternFill('solid', fgColor='FFF3E0')
CALC_BG = PatternFill('solid', fgColor='E8F5E9')
REF_BG = PatternFill('solid', fgColor='ECEFF1')
RAW_MISMATCH_BG = PatternFill('solid', fgColor='FFCDD2')
LOC_MISMATCH_FILL = PatternFill('solid', fgColor='FFF59D')

# Status pills (cell fill + bold font, applied per-cell)
STATUS_OK_FILL = PatternFill('solid', fgColor='00C853')
STATUS_BAD_FILL = PatternFill('solid', fgColor='FF5252')


# ── Fonts ──────────────────────────────────────────────────────────────────

HEADER_FONT = Font(bold=True, color='FFFFFF', name='Aptos Display', size=11)
DATA_FONT = Font(name='Aptos Display', size=11)

STATUS_OK_FONT = Font(name='Aptos Display', size=11, bold=True, color='000000')
STATUS_BAD_FONT = Font(name='Aptos Display', size=11, bold=True, color='FFFFFF')
MISMATCH_TEXT_FONT = Font(name='Aptos Display', size=11, bold=True,
                           color='D32F2F')
NOT_IN_MASTER_TEXT_FONT = Font(name='Aptos Display', size=11, bold=True,
                                color='E65100')

INFO_ITALIC_FONT = Font(name='Aptos Display', size=10, italic=True,
                         color='666666')
LEGEND_ITALIC_FONT = Font(name='Aptos Display', size=10, italic=True,
                           color='B7950B')

BOLD_DATA_FONT = Font(name='Aptos Display', size=11, bold=True)


# ── Borders ────────────────────────────────────────────────────────────────

THIN_SIDE = Side(style='thin', color='CCCCCC')
BORDER = Border(left=THIN_SIDE, right=THIN_SIDE,
                 top=THIN_SIDE, bottom=THIN_SIDE)


# ── Cell-writing helpers ───────────────────────────────────────────────────
#
# Every sheet writer ends up doing the same three things over and over:
# write a styled header cell, write a styled data cell, and auto-fit the
# column widths. Centralising them here means a future style tweak is
# one-line wide instead of seven-files wide.

def hdr_cell(ws, row: int, col: int, value,
              fill: Optional[PatternFill] = None,
              font: Optional[Font] = None):
    """
    Write and style a header cell.

    Args:
        ws:    Target worksheet.
        row:   1-based row number.
        col:   1-based column number.
        value: Header text.
        fill:  Override the default ``HEADER_FILL`` (e.g. ``WARN_FILL``).
        font:  Override the default ``HEADER_FONT``.

    Returns:
        The created cell (for further per-call tweaks).
    """
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = font or HEADER_FONT
    cell.fill = fill or HEADER_FILL
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = BORDER
    return cell


def data_cell(ws, row: int, col: int, value,
               number_format: Optional[str] = None):
    """
    Write and style a data cell.

    Args:
        ws:            Target worksheet.
        row:           1-based row number.
        col:           1-based column number.
        value:         Cell value (any openpyxl-acceptable type).
        number_format: Optional Excel number format string
                       (e.g. ``'#,##0.00'``).

    Returns:
        The created cell.
    """
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = DATA_FONT
    cell.border = BORDER
    if number_format:
        cell.number_format = number_format
    return cell


def auto_width(ws, max_width: int = 50) -> None:
    """
    Auto-fit each column's width based on its longest value.

    Args:
        ws:        Target worksheet.
        max_width: Cap on column width (Excel character units). Prevents
                   one extreme cell from blowing out the layout.
    """
    for col in ws.columns:
        letter = col[0].column_letter
        widest = max((len(str(c.value or '')) for c in col), default=8)
        ws.column_dimensions[letter].width = min(widest + 3, max_width)
