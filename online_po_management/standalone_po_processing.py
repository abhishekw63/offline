"""
╔═══════════════════════════════════════════════════════════════════════════════╗
║               ONLINE PO PROCESSOR — Marketplace SO Generator                 ║
║               Tkinter GUI Desktop Application                                ║
╠═══════════════════════════════════════════════════════════════════════════════╣
║  Author  : Agami AI / Vishal                                                ║
║  Version : 1.3.7                                                             ║
║  Purpose : Reads marketplace punch/PO files (Myntra, Bigbasket, Blink etc), ║
║            maps locations to Ship-to codes via a mapping registry,           ║
║            generates ERP-importable Headers (SO) + Lines (SO) sheets.        ║
║  Stack   : Python 3.13, Tkinter, pandas, openpyxl                           ║
╠═══════════════════════════════════════════════════════════════════════════════╣
║  CHANGELOG                                                                   ║
║    v1.3.7 — Bug fix: 'Download PO Template' crashed with KeyError 'item_col'║
║             for Myntra and RK after they switched to item_resolution=        ║
║             'from_ean' in v1.3.5/v1.3.6 (config no longer has item_col).    ║
║             Template generator now branches on item_resolution:              ║
║                from_column → item_col is required (BLUE)                    ║
║                from_ean    → ean_col is required (BLUE), promoted from      ║
║                              validation (GREEN)                              ║
║             Color legend in the generated template is unchanged in meaning. ║
║    v1.3.6 — RK config: switched to item_resolution='from_ean'.              ║
║             Real RK punch files don't have an 'Item no' column either —     ║
║             only 'External ID' (the EAN). Item No is resolved via master    ║
║             lookup against Items_March, same mechanism Myntra uses.         ║
║             RK template_headers also corrected to match RK's real columns: ║
║                PO | Vendor code | Order date | Product name | External ID  ║
║                | Accepted quantity | Ship-to location | Cost                ║
║                | Total accepted cost                                        ║
║             (Removed the non-existent 'Status' and 'Model number' that     ║
║             were in earlier versions.)                                      ║
║    v1.3.5 — Item No resolution mode per marketplace.                         ║
║             Real Myntra punch files do NOT include an 'Item no' column —    ║
║             only EAN/GTIN. Previous versions required item_col, which broke ║
║             on real Myntra files.                                            ║
║                                                                              ║
║             New config key 'item_resolution' with two modes:                 ║
║                'from_column' → take Item No from item_col   (RK)            ║
║                'from_ean'    → look up Item No in master via ean_col (Myntra)║
║                Default 'from_column' if omitted (backwards-compatible).     ║
║                                                                              ║
║             Other related changes:                                           ║
║              • Required-column validator only checks the column relevant     ║
║                to the chosen resolution mode.                                ║
║              • EAN extraction now handles float64 GTINs (e.g. Myntra exports║
║                GTIN as 8906121642599.0) by converting through int() → str.  ║
║              • Raw Data sheet's row-matching falls back to EAN when         ║
║                item_col is absent. validation_lookup is indexed under both  ║
║                (po, item_no) and (po, ean) so either path works.            ║
║              • If item_resolution='from_ean' and the row's EAN isn't in     ║
║                Items_March, the row is still emitted with status            ║
║                NOT_IN_MASTER and the EAN as the visible item value.         ║
║              • Both Myntra and RK configs explicitly tagged with their      ║
║                item_resolution value for clarity.                            ║
║                                                                              ║
║    v1.3.4 — Two changes:                                                     ║
║                                                                              ║
║         1) Reference Diffn column in Raw Data sheet.                         ║
║            New optional config key 'ref_fob_col'. When set, the Raw Data    ║
║            sheet shows BOTH:                                                 ║
║                Diffn with <ref_fob_col>     ← reference only, muted grey   ║
║                Diffn with <Label>           ← validation-driving, green    ║
║            Currently set on Myntra to surface the legacy "Diffn with        ║
║            List price(FOB+Transport-Excise)" alongside the active "Diffn   ║
║            with Landing Rate" — for visibility only, has zero effect on    ║
║            the OK/MISMATCH status. Reference column gets a slate-grey      ║
║            header and pale-grey row tint so it's clearly distinguishable   ║
║            from the validation column.                                      ║
║                                                                              ║
║         2) Output saved next to the input punch file.                        ║
║            Was: <script-dir>/output_online/<marketplace>_so_<ts>.xlsx       ║
║            Now: <punch-file-dir>/output/<marketplace>_so_<ts>.xlsx          ║
║            Means each month's output lives next to the inputs it was        ║
║            generated from — no more digging through a global folder. The   ║
║            'output/' subfolder is auto-created. Falls back to the old       ║
║            behaviour if the input path isn't recorded (defensive only).    ║
║                                                                              ║
║    v1.3.3 — Summary sheet: Location split into two columns for visual       ║
║             match-verification:                                              ║
║                  Location (Raw)    — value from the marketplace's punch     ║
║                                       file (what they sent us)              ║
║                  Location (Mapped) — canonical key from our Ship-To         ║
║                                       registry that we matched it to        ║
║             • When raw and mapped differ (case-insensitive), BOTH cells     ║
║               get a pale yellow highlight so loose / fuzzy matches are      ║
║               obvious at a glance — typically caused by mapping fuzzy       ║
║               match (e.g. "Bilaspur Warehouse - Gurgaon" → "Bilaspur").    ║
║             • A small legend row appears at the bottom of the sheet only   ║
║               when at least one mismatch exists.                             ║
║             • MappingLoader.lookup() now also returns 'matched_key' so the  ║
║               engine can record which canonical name was used.              ║
║    v1.3.2 — Per-file last-updated timestamp shown in the GUI:               ║
║             • A small grey sub-line appears under the Master and Mapping    ║
║               picker rows, e.g.                                              ║
║                   Items Master:  ✓ Items March.xlsx  [Browse]               ║
║                                  Updated: 19-Apr-2026 18:41                 ║
║             • Source: in-app history of when the user clicked "Update       ║
║               Bundled Files" — NOT filesystem mtime, so unrelated saves     ║
║               (e.g. opening the file in Excel) do not move the timestamp.   ║
║             • Stored as Calculation Data/.update_history.json (a tiny       ║
║               sidecar maintained by the script). Survives app restarts.    ║
║             • Sub-line is empty until the first explicit update — it does   ║
║               not lie about a timestamp it doesn't have.                     ║
║             • Manual file overrides (Browse → pick a different file)        ║
║               clear the sub-line, since user-picks aren't tracked.          ║
║             • Window height bumped 580 → 620 to fit the two extra lines.    ║
║    v1.3.1 — Validation sheet: 'Description' column added immediately after  ║
║             'EAN'. Pulled from Items_March master so the user can read      ║
║             what each EAN actually is at a glance, instead of cross-        ║
║             referencing back to the master file. All subsequent columns     ║
║             shift right by 1.                                                ║
║    v1.3 — Bundled master/mapping auto-loader:                                ║
║             • New 'Calculation Data/' folder beside the script holds two    ║
║               long-lived files that rarely change:                           ║
║                   Calculation Data/Items March.xlsx   (Items Master)        ║
║                   Calculation Data/Ship to B2B.xlsx   (Ship-To Mapping)    ║
║             • On startup the GUI auto-detects these. If found, the picker   ║
║               rows show "✓ Items March.xlsx (auto-loaded)" instead of       ║
║               "Not selected" — no more re-picking the same files every run. ║
║             • Pickers stay visible so user can override per-run with a      ║
║               different file if needed (e.g. testing).                       ║
║             • New "📁 Update Bundled Files" button copies a fresh master    ║
║               or mapping into Calculation Data/ in-place; subsequent runs   ║
║               then auto-load the updated version.                            ║
║             • Falls back gracefully: if the folder or files are missing,    ║
║               the user gets a clear log message and uses the picker as     ║
║               in v1.2.x.                                                     ║
║    v1.2.1 — Diffn display rounded to 2 decimals (0.00) instead of 4         ║
║             (0.0000). Sub-paisa precision was just floating-point dust       ║
║             after Myntra's clean integer landing math; 2 decimals matches   ║
║             rupee-paisa convention.                                          ║
║    v1.2 — Per-marketplace comparison basis (landing vs cost):                ║
║             • New config key `compare_basis`:                                ║
║                 'landing' → compare against MRP × margin% (pre-GST)         ║
║                             — used by MYNTRA (Myntra's "Landing Price"      ║
║                               column matches our pre-GST landing exactly,   ║
║                               no rounding noise)                            ║
║                 'cost'    → compare against MRP × margin% ÷ GST (post-GST) ║
║                             — used by RK (default, original v1.0 behavior) ║
║             • New config key `compare_label`: friendly label shown in the   ║
║               Validation sheet (e.g. "Landing Rate", "Cost").               ║
║             • Validation sheet:                                              ║
║                 - Marketplace-price column renamed dynamically to           ║
║                   "Marketplace <Label>" (e.g. "Marketplace Landing Rate").  ║
║                 - Diffn column renamed to "Difference with <Label>".        ║
║                 - "Our Cost Price" + "GST Code" still shown for reference   ║
║                   even when not used for comparison (the "naked CP").       ║
║             • Myntra fob_col switched: 'List price(FOB+Transport-Excise)'   ║
║               → 'Landing Price'.                                            ║
║    v1.1 — Color-coded PO template download:                                  ║
║             • BLUE  → Required columns (script fails without them)          ║
║             • GREEN → Validation columns (used for price check / lookup)    ║
║             • GREY  → Not read by script (kept only to mirror the           ║
║                       marketplace's native file format; safe to leave       ║
║                       blank)                                                ║
║    v1.0 — Initial release with Myntra + RK support, price validation        ║
║           against Items_March, raw data sheet, PO template download.        ║
╚═══════════════════════════════════════════════════════════════════════════════╝

═══════════════════════════════════════════════════════════════════════════════
  ARCHITECTURE
═══════════════════════════════════════════════════════════════════════════════

  ┌─────────────────────────────────────────────────────────────────┐
  │                    OnlinePOApp (GUI)                            │
  │  Select Marketplace → Load Files → Generate → Open Output      │
  │                                                                 │
  │  3 file inputs:                                                 │
  │    1. Items_March.xlsx    (Item Master — shared)                │
  │    2. Online_Mapping.xlsx (Ship-To B2B — location registry)    │
  │    3. Marketplace PO file (e.g., Myntra_Punch.xlsx)            │
  └────────────────────┬────────────────────────────────────────────┘
                       │
                       ▼
  ┌─────────────────────────────────────────────────────────────────┐
  │              MarketplaceEngine                                   │
  │  1. Load mapping (filter by selected marketplace)               │
  │  2. Parse PO file (marketplace-specific column config)          │
  │  3. Match Location → Cust No + Ship-to from mapping            │
  │  4. Generate output rows                                        │
  └────────────────────┬────────────────────────────────────────────┘
                       │
                       ▼
  ┌─────────────────────────────────────────────────────────────────┐
  │              SOExporter (Output)                                 │
  │  Sheet 1: Headers (SO)  → ERP SO header import                 │
  │  Sheet 2: Lines (SO)    → ERP SO line import (10K increments)  │
  │  Sheet 3: Summary       → Per-PO breakdown with mapping info   │
  │  Sheet 4: Warnings      → Unmapped locations, missing data     │
  └─────────────────────────────────────────────────────────────────┘

═══════════════════════════════════════════════════════════════════════════════
  INPUT FILES
═══════════════════════════════════════════════════════════════════════════════

  1. ITEMS MASTER (Items_March.xlsx)
     ───────────────────────────────
     Same master used by EKA Script and GT Mass Dump.
     Required columns: No., GTIN, Description, GST Group Code, Mrp
     Used here only for validation — Item No comes from the PO file directly.

  2. MAPPING FILE (Online_Mapping.xlsx, sheet: 'Ship-To B2B')
     ─────────────────────────────────────────────────────────
     Registry of all marketplace delivery locations.
     Required columns:
         Party          → Marketplace name (Myntra, Bigbasket, Blink, etc.)
         Del Location   → Delivery location name as used by the marketplace
         Cust No        → ERP Sell-to Customer Number
         Ship to        → ERP Ship-to Code (e.g., '20011_4')

  3. MARKETPLACE PO FILE (e.g., Myntra_Punch_09-04-2026.xlsx)
     ─────────────────────────────────────────────────────────
     Purchase order / punch file from the marketplace.
     Column mapping varies per marketplace (see MARKETPLACE_CONFIGS).

═══════════════════════════════════════════════════════════════════════════════
  MARKETPLACE CONFIGURATIONS
═══════════════════════════════════════════════════════════════════════════════

  Each marketplace has a different PO file format. The config defines:
      party_name       → Name to filter in mapping file (e.g., 'Myntra')
      po_col           → Column containing PO/SO number       [REQUIRED]
      loc_col          → Column containing delivery location  [REQUIRED]
      qty_col          → Column containing order quantity     [REQUIRED]

      item_resolution  → How to determine the canonical Item No:
                           'from_column' → take it directly from item_col
                                           (use when the marketplace provides
                                           a pre-resolved Item No, e.g. RK).
                                           Requires item_col.
                           'from_ean'    → look up the EAN in Items_March,
                                           Item No is master_info['item_no']
                                           (use when the marketplace only
                                           provides an EAN/GTIN, e.g. real
                                           Myntra files).
                                           Requires ean_col.
                         Default: 'from_column' if omitted.
      item_col         → Column containing Item No (only when
                         item_resolution = 'from_column')
      ean_col          → Column with EAN/GTIN for master lookup. REQUIRED
                         when item_resolution = 'from_ean'; otherwise just
                         used for price validation.
      price_col        → Column containing unit price (None = leave empty)
      fob_col          → Column with marketplace price        [VALIDATION]
      ref_fob_col      → Optional second marketplace price column shown
                         only as a reference Diffn in Raw Data (v1.3.4).
      default_margin   → Default margin % for landing cost
      compare_basis    → 'landing' (= MRP × margin%, pre-GST)
                         'cost'    (= MRP × margin% ÷ GST, post-GST)
                         Default 'cost' if omitted. Decides what we compare
                         the marketplace's fob_col against.
      compare_label    → Friendly label for Validation sheet (e.g. 'Landing
                         Rate', 'Cost'). Used in column headers like
                         "Marketplace <Label>" and "Difference with <Label>".
                         Default 'Price' if omitted.
      template_headers → Full column list for PO template download
                         (everything not in REQUIRED/VALIDATION is
                          rendered grey and left empty by the user)

  To add a new marketplace:
      1. Add entry to MARKETPLACE_CONFIGS dict
      2. Add name to MARKETPLACE_NAMES list
      3. The mapping sheet must have the party's locations

═══════════════════════════════════════════════════════════════════════════════
  OUTPUT — 4 EXCEL SHEETS
═══════════════════════════════════════════════════════════════════════════════

  Sheet 1: 'Headers (SO)' — One row per unique PO number
      Document Type | No. | Sell-to Customer No. | Ship-to Code |
      5 × date fields | External Document No. | Location Code = PICK |
      Supply Type = B2B | ...dimension columns (empty)

  Sheet 2: 'Lines (SO)' — One row per ordered item
      Document Type | Document No. | Line No. (10K increments) | Type = Item |
      No. | Location Code = PICK | Quantity | Unit Price (empty)

  Sheet 3: 'Summary' — Per-PO grouped info
      PO | Location | Cust No | Ship-to | Items | Total Qty | Status

  Sheet 4: 'Warnings' — Only if issues found
      PO | Location | Warning

═══════════════════════════════════════════════════════════════════════════════
  PO TEMPLATE — COLOR LEGEND (v1.1)
═══════════════════════════════════════════════════════════════════════════════

  When user clicks "Download PO Template", the generated Excel highlights
  each column header by its role in the script:

      ┌──────────┬──────────────────┬─────────────────────────────────┐
      │  COLOR   │  ROLE            │  WHAT IT MEANS FOR THE USER     │
      ├──────────┼──────────────────┼─────────────────────────────────┤
      │  BLUE    │  Required        │  MUST be filled — script fails  │
      │  (#1A237E)│ (po/loc/item/qty)│  without these columns          │
      ├──────────┼──────────────────┼─────────────────────────────────┤
      │  GREEN   │  Validation      │  Used for price check & master  │
      │  (#1B5E20)│ (fob_col, ean_col)│ lookup — strongly recommended  │
      ├──────────┼──────────────────┼─────────────────────────────────┤
      │  GREY    │  Not read        │  Kept only to mirror the        │
      │  (#9E9E9E)│ (everything else)│  marketplace's native file     │
      │           │                   │  format — safe to leave blank  │
      └──────────┴──────────────────┴─────────────────────────────────┘

═══════════════════════════════════════════════════════════════════════════════
  HOW TO ADD A NEW MARKETPLACE
═══════════════════════════════════════════════════════════════════════════════

  1. Get a sample PO file from the marketplace
  2. Identify which columns have: PO number, Location, Item No, Quantity
  3. Add config to MARKETPLACE_CONFIGS:
         'NewMarket': {
             'party_name': 'NewMarket',     # Must match 'Party' in mapping
             'po_col': 'PO Number',          # Column name for PO/SO number
             'loc_col': 'Warehouse',          # Column name for location
             'item_col': 'Item Code',         # Column name for item number
             'qty_col': 'Qty',                # Column name for quantity
             'price_col': None,               # None = leave empty in output
         }
  4. Add 'NewMarket' to MARKETPLACE_NAMES list
  5. Add locations to the mapping sheet (Ship-To B2B)

Requirements:
    pip install pandas openpyxl

Run:
    python online_po_processor.py
"""

# ═══════════════════════════════════════════════════════════════════════════════
#  IMPORTS
# ═══════════════════════════════════════════════════════════════════════════════

from __future__ import annotations
import os
import sys
import json
import platform
import shutil
import time
import logging
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
from dataclasses import dataclass, field
from pathlib import Path
from typing import List, Optional, Tuple, Dict, Any
from datetime import datetime

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ═══════════════════════════════════════════════════════════════════════════════
#  LOGGING
# ═══════════════════════════════════════════════════════════════════════════════

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s | %(levelname)s | %(message)s"
)


# ═══════════════════════════════════════════════════════════════════════════════
#  EXPIRY CHECK
# ═══════════════════════════════════════════════════════════════════════════════

EXPIRY_DATE = "30-06-2026"


def check_expiry():
    """Check if the application has expired. Shows popup and exits if so."""
    expiry = datetime.strptime(EXPIRY_DATE, "%d-%m-%Y").date()
    today = datetime.now().date()

    if today > expiry:
        root = tk.Tk()
        root.withdraw()
        messagebox.showerror(
            "Application Expired",
            f"This application expired on {EXPIRY_DATE}.\n\n"
            f"Please contact the administrator for an updated version."
        )
        root.destroy()
        sys.exit(0)

    days_remaining = (expiry - today).days
    if days_remaining <= 7:
        root = tk.Tk()
        root.withdraw()
        messagebox.showwarning(
            "Expiration Warning",
            f"⚠️ This application will expire in {days_remaining} day(s).\n\n"
            f"Expiry Date: {EXPIRY_DATE}\n"
            f"Please contact the administrator for renewal."
        )
        root.destroy()


# ═══════════════════════════════════════════════════════════════════════════════
#  BUNDLED FILES (v1.3) — Auto-loaded master & mapping
# ═══════════════════════════════════════════════════════════════════════════════
#
# The Items Master and Ship-To Mapping change rarely (master ~monthly, mapping
# ~weekly when new locations are added). Forcing the user to pick them on every
# run was friction. So the script now looks for them in a 'Calculation Data'
# folder sitting next to the script, with the exact filenames matching what the
# user already has on disk:
#
#   <script-folder>/
#       online_po_processor.py
#       Calculation Data/
#           Items March.xlsx     ← auto-loaded as Items Master
#           Ship to B2B.xlsx     ← auto-loaded as Ship-To Mapping
#
# If the folder or files are missing the GUI falls back to the manual picker
# behaviour from v1.2.x — fully backwards compatible.
#
# The file pickers stay visible after auto-load so the user can override with
# a different file for one-off testing without touching the bundled copy.
# An "Update Bundled Files" button in the GUI copies a fresh master/mapping
# into Calculation Data/ in-place; subsequent runs then auto-load the update.

BUNDLED_DATA_FOLDER = "Calculation Data"
BUNDLED_MASTER_NAME = "Items March.xlsx"
BUNDLED_MAPPING_NAME = "Ship to B2B.xlsx"


def _script_dir() -> Path:
    """
    Return the directory the script (or PyInstaller .exe) is running from.

    Handles three cases:
      • Frozen PyInstaller exe → directory of sys.executable
      • Normal .py run         → directory of __file__
      • Interactive / fallback → current working directory
    """
    if getattr(sys, 'frozen', False):
        return Path(sys.executable).parent
    try:
        return Path(__file__).parent
    except NameError:
        return Path.cwd()


def get_bundled_master_path() -> Optional[Path]:
    """
    Returns the absolute path to the bundled Items Master file if it exists,
    otherwise None. Called by the GUI on startup to pre-populate the master
    file path.
    """
    p = _script_dir() / BUNDLED_DATA_FOLDER / BUNDLED_MASTER_NAME
    return p if p.exists() else None


def get_bundled_mapping_path() -> Optional[Path]:
    """
    Returns the absolute path to the bundled Ship-To Mapping file if it
    exists, otherwise None.
    """
    p = _script_dir() / BUNDLED_DATA_FOLDER / BUNDLED_MAPPING_NAME
    return p if p.exists() else None


def get_bundled_data_folder(create: bool = False) -> Path:
    """
    Returns the absolute path to the Calculation Data folder.

    Args:
        create: If True, create the folder if it doesn't exist (used by the
                "Update Bundled Files" flow so the user can drop the first
                copy in even before the folder exists).
    """
    folder = _script_dir() / BUNDLED_DATA_FOLDER
    if create:
        folder.mkdir(parents=True, exist_ok=True)
    return folder


# ─────────────────────────────────────────────────────────────────────────────
#  v1.3.2 — In-app update history (JSON sidecar)
# ─────────────────────────────────────────────────────────────────────────────
#
# We track WHEN the user last clicked "Update Bundled Files" for each of the
# two bundled files. This is shown under each file row in the GUI as a small
# secondary line:
#
#     Items Master:    ✓ Items March.xlsx
#                       Updated: 19-Apr-2026 18:41
#
# Why a sidecar JSON instead of file mtime?
#   • Survives across runs without inferring intent from arbitrary filesystem
#     mtimes (which can change for non-update reasons — e.g. someone opened
#     the file in Excel without modifying it, or copied it from a backup).
#   • Records ONLY explicit user updates via the GUI — that's what the user
#     asked for ("when 'Update Bundled Files' was last clicked").
#
# Format (Calculation Data/.update_history.json):
#   {
#     "Items March.xlsx":  "2026-04-19T18:41:32",
#     "Ship to B2B.xlsx":  "2026-04-17T11:22:05"
#   }

UPDATE_HISTORY_FILE = ".update_history.json"


def _history_path() -> Path:
    """Path to the JSON sidecar inside Calculation Data/."""
    return _script_dir() / BUNDLED_DATA_FOLDER / UPDATE_HISTORY_FILE


def load_update_history() -> Dict[str, str]:
    """
    Load the in-app update history. Returns {filename: ISO-timestamp} or
    {} if the sidecar doesn't exist or is corrupt.
    """
    p = _history_path()
    if not p.exists():
        return {}
    try:
        with open(p, 'r', encoding='utf-8') as f:
            data = json.load(f)
        return data if isinstance(data, dict) else {}
    except (json.JSONDecodeError, OSError) as e:
        logging.warning(f"Could not read update history at {p}: {e}")
        return {}


def record_update(filename: str) -> None:
    """
    Stamp `filename` (e.g. 'Items March.xlsx') as just-updated NOW.

    Creates the Calculation Data/ folder if needed, then merges the new
    timestamp into the existing JSON. Failures are logged but do not abort
    the update flow — losing a timestamp shouldn't block a real file copy.
    """
    folder = get_bundled_data_folder(create=True)
    history = load_update_history()
    history[filename] = datetime.now().isoformat(timespec='seconds')
    try:
        with open(folder / UPDATE_HISTORY_FILE, 'w', encoding='utf-8') as f:
            json.dump(history, f, indent=2)
    except OSError as e:
        logging.warning(f"Could not write update history: {e}")


def get_update_timestamp(filename: str) -> Optional[str]:
    """
    Returns the recorded update timestamp for `filename` formatted for
    display (e.g. '19-Apr-2026 18:41'), or None if no record exists.
    """
    history = load_update_history()
    iso = history.get(filename)
    if not iso:
        return None
    try:
        dt = datetime.fromisoformat(iso)
        return dt.strftime('%d-%b-%Y %H:%M')
    except ValueError:
        return None


# ═══════════════════════════════════════════════════════════════════════════════
#  MARKETPLACE CONFIGURATIONS
# ═══════════════════════════════════════════════════════════════════════════════
# Each marketplace has a different PO file column layout.
# party_name must match the 'Party' column in the mapping sheet exactly.
# price_col = None means Unit Price is left empty in output.
#
# COLOR ROLE in PO template (v1.1):
#   • BLUE  (Required)   = po_col, loc_col, item_col, qty_col
#   • GREEN (Validation) = fob_col, ean_col
#   • GREY  (Not read)   = every other entry in template_headers

MARKETPLACE_CONFIGS: Dict[str, Dict[str, Any]] = {
    'Myntra': {
        'party_name': 'Myntra',            # Must match 'Party' in mapping sheet
        'po_col': 'PO',                    # [REQUIRED] PO/SO number
        'loc_col': 'Location',              # [REQUIRED] delivery location
        # v1.3.5: real Myntra punch files do NOT have an 'Item no' column —
        # only EAN/GTIN. So item resolution flips to 'from_ean': the engine
        # uses ean_col to look up the canonical Item No from Items_March.
        # 'item_col' is omitted; if Myntra ever adds an Item No column in
        # future, just add 'item_col': '<col-name>' AND set
        # item_resolution: 'from_column'.
        'item_resolution': 'from_ean',     # 'from_ean' or 'from_column'
        'qty_col': 'Quantity',              # [REQUIRED] order quantity
        'price_col': None,                  # None = leave Unit Price empty (WMS handles)
        'fob_col': 'Landing Price',         # [VALIDATION] marketplace price.
                                            #   v1.2: switched from
                                            #   'List price(FOB+Transport-Excise)'
                                            #   (post-GST, has rounding noise)
                                            #   → 'Landing Price' (pre-GST, exact
                                            #   match to MRP × margin%, clean 0).
        'ref_fob_col': 'List price(FOB+Transport-Excise)',  # v1.3.4: REFERENCE
                                            # ONLY — not used for the OK/MISMATCH
                                            # status. We capture this column from
                                            # the punch file and compute a second
                                            # diff (vs our naked Cost Price) for
                                            # visibility in the Raw Data sheet.
                                            # Useful when the user wants to see
                                            # "what would the diff have been if
                                            # we'd compared against List price?"
                                            # without it driving validation.
        'ean_col': 'Vendor Article Number', # [REQUIRED for from_ean] EAN/GTIN
                                            #   for master lookup. Item No will
                                            #   be resolved from this.
        'default_margin': 70,              # Default margin % for landing cost
        'compare_basis': 'landing',         # v1.2: compare against MRP × margin%
                                            #       (pre-GST). Myntra's "Landing
                                            #       Price" column is already this.
        'compare_label': 'Landing Rate',    # v1.2: shown in Validation sheet as
                                            #       "Marketplace Landing Rate"
                                            #       and "Difference with Landing Rate"
        'template_headers': ['PO', 'Location', 'SKU Id', 'Style Id', 'SKU Code',
                             'HSN Code', 'Brand', 'GTIN', 'Vendor Article Number',
                             'Vendor Article Name', 'Size', 'Colour', 'Mrp',
                             'Credit Period', 'Margin Type', 'Agreed Margin',
                             'Gross Margin', 'Quantity', 'FOB Amount',
                             'List price(FOB+Transport-Excise)', 'Landing Price',
                             'Estimated Delivery Date'],
    },
    'RK': {
        'party_name': 'RK',                # Must match 'Party' in mapping sheet
        'po_col': 'PO',                    # [REQUIRED] PO/SO number (alphanumeric like '2EH63D1K')
        'loc_col': 'Ship-to location',      # [REQUIRED] delivery location (codes like 'ISK3')
        # v1.3.6: real RK punch files do NOT have an 'Item no' column either
        # (just like Myntra). External ID is the EAN; we look up Item No
        # from Items_March via that.
        'item_resolution': 'from_ean',     # 'from_ean' or 'from_column'
        'qty_col': 'Accepted quantity',      # [REQUIRED] accepted order quantity
        'price_col': None,                  # None = leave Unit Price empty (WMS handles)
        'fob_col': 'Cost',                  # [VALIDATION] marketplace price (post-GST)
        'ean_col': 'External ID',           # [REQUIRED for from_ean] EAN/GTIN
                                            #   for master lookup. Item No will
                                            #   be resolved from this.
        'default_margin': 70,              # Default margin % for landing cost
        'compare_basis': 'cost',            # v1.2: compare against MRP × margin% ÷ GST
                                            #       (post-GST). RK's "Cost" column
                                            #       is the GST-divided value
                                            #       (verified: matches MRP×70%÷1.18
                                            #       to 2 decimals exactly).
        'compare_label': 'Cost',            # v1.2: shown in Validation sheet as
                                            #       "Marketplace Cost" and
                                            #       "Difference with Cost"
        'template_headers': ['PO', 'Vendor code', 'Order date', 'Product name',
                             'External ID', 'Accepted quantity', 'Ship-to location',
                             'Cost', 'Total accepted cost'],
    },
    # ┌─────────────────────────────────────────────────────────────────────┐
    # │ ADD NEW MARKETPLACES HERE                                           │
    # │                                                                     │
    # │ 'Bigbasket': {                                                      │
    # │     'party_name': 'Bigbasket',                                      │
    # │     'po_col': 'PO Number',                                          │
    # │     'loc_col': 'Delivery Location',                                 │
    # │     'item_col': 'Item Code',                                        │
    # │     'qty_col': 'Qty',                                               │
    # │     'fob_col': 'Unit Price',                                        │
    # │     'ean_col': 'EAN',                                               │
    # │     'price_col': None,                                              │
    # │     'default_margin': 60,                                           │
    # │ },                                                                  │
    # └─────────────────────────────────────────────────────────────────────┘
}

# List of marketplace names for the dropdown (order matters for display)
MARKETPLACE_NAMES: List[str] = list(MARKETPLACE_CONFIGS.keys())


# ═══════════════════════════════════════════════════════════════════════════════
#  DATA MODEL
# ═══════════════════════════════════════════════════════════════════════════════

@dataclass
class SORow:
    """Single line item for SO generation."""
    po_number: str           # PO/SO number from marketplace (e.g., 'MYNJ-RNEE230326-1')
    location: str            # Delivery location from file (e.g., 'Bilaspur')
    item_no: Any             # Item No (e.g., 200074)
    qty: int                 # Order quantity
    unit_price: Optional[float] = None  # Unit price (None = leave empty)
    cust_no: str = ''        # Sell-to Customer No. from mapping
    ship_to: str = ''        # Ship-to Code from mapping
    mapped: bool = False     # Whether location was found in mapping
    mapped_location: str = ''  # v1.3.3: The canonical mapping key that was
                               #         matched to (could differ from raw
                               #         `location` due to fuzzy matching).
                               #         Shown in Summary as "Location (Mapped)"
                               #         for visual verification of matches.
    ean: str = ''            # EAN/GTIN from punch file (for master lookup)
    description: str = ''    # v1.3.1: Item description from Items_March
                             #         (shown next to EAN in Validation sheet)
    fob_price: Optional[float] = None   # Marketplace FOB/Cost/Landing price from file
    ref_fob_price: Optional[float] = None  # v1.3.4: Optional REFERENCE marketplace
                                            #   price (e.g. Myntra's List price). NOT
                                            #   used for OK/MISMATCH status — purely
                                            #   for visibility in the Raw Data sheet.
    calc_price: Optional[float] = None  # Our calculated value used for the diff.
                                        #   compare_basis = 'cost'    → MRP × m% ÷ GST
                                        #   compare_basis = 'landing' → MRP × m%
    cost_price_ref: Optional[float] = None  # v1.2: Always the post-GST cost price
                                            #   (MRP × m% ÷ GST). Shown in the
                                            #   Validation sheet as "Our Cost Price"
                                            #   for reference (the "naked CP"),
                                            #   regardless of compare_basis.
    diffn: Optional[float] = None       # fob_price - calc_price (0 = OK, non-zero = flag)
    ref_diffn: Optional[float] = None   # v1.3.4: ref_fob_price - cost_price_ref
                                        #   (reference-only diff, see ref_fob_price).
                                        #   Always computed against the post-GST
                                        #   cost price regardless of compare_basis.
    mrp: Optional[float] = None         # MRP from Items_March
    gst_code: str = ''       # GST Code from Items_March
    validation_status: str = ''  # 'OK', 'MISMATCH', 'NOT_IN_MASTER'


@dataclass
class ProcessingResult:
    """Result from processing a marketplace PO file."""
    rows: List[SORow] = field(default_factory=list)
    warnings: List[Tuple[str, str, str]] = field(default_factory=list)  # (po, location, message)
    marketplace: str = ''
    input_file: str = ''           # basename of the punch file (for display)
    input_file_path: str = ''      # v1.3.4: FULL path to the punch file. Used so
                                   #         the output can be saved next to it
                                   #         in an 'output/' subfolder.
    margin_pct: float = 0.70  # Margin % as decimal (0.70 = 70%)
    raw_df: Any = None        # Original marketplace DataFrame for reference sheet
    compare_basis: str = 'cost'        # v1.2: 'landing' or 'cost' (from config)
    compare_label: str = 'Price'       # v1.2: friendly label (from config)


# ═══════════════════════════════════════════════════════════════════════════════
#  MAPPING LOADER
# ═══════════════════════════════════════════════════════════════════════════════

class MappingLoader:
    """
    Loads the Ship-To B2B mapping from the mapping Excel file.

    The mapping file has a sheet named 'Ship-To B2B' with columns:
        Party | Del Location | Cust No | Ship to

    Filtering by party_name gives us only the relevant marketplace locations.
    """

    def __init__(self):
        self.mappings: Dict[str, Dict[str, str]] = {}  # location → {cust_no, ship_to}
        self.party_name: str = ''
        self.total_loaded: int = 0

    def load(self, filepath: str, party_name: str, logs: List[Tuple[str, str, str]]) -> int:
        """
        Load mapping for a specific marketplace.

        Args:
            filepath: Path to mapping Excel file
            party_name: Marketplace name to filter by (e.g., 'Myntra')
            logs: Warning accumulator

        Returns: number of locations loaded for this marketplace
        """
        self.party_name = party_name
        self.mappings = {}

        try:
            df = pd.read_excel(filepath, sheet_name='Ship-To B2B', header=0)
        except ValueError:
            # Sheet not found — try first sheet
            logging.warning("Sheet 'Ship-To B2B' not found, trying first sheet")
            df = pd.read_excel(filepath, header=0)
        except Exception as e:
            logs.append(('', '', f"Cannot read mapping file: {e}"))
            return 0

        # ── Detect required columns ──
        col_map = {}
        for col in df.columns:
            cl = str(col).strip().lower()
            if cl == 'party':
                col_map['party'] = col
            elif cl in ('del location', 'delivery location', 'location'):
                col_map['location'] = col
            elif cl in ('cust no', 'cust no.', 'customer no', 'sell-to'):
                col_map['cust_no'] = col
            elif cl in ('ship to', 'ship-to', 'ship to code'):
                col_map['ship_to'] = col

        missing = [k for k in ('party', 'location', 'cust_no', 'ship_to') if k not in col_map]
        if missing:
            logs.append(('', '', f"Mapping file missing columns: {', '.join(missing)}. "
                                f"Available: {list(df.columns)}"))
            return 0

        # ── Filter by party and build lookup ──
        for _, row in df.iterrows():
            party = str(row[col_map['party']]).strip()
            if party.lower() != party_name.lower():
                continue

            location = str(row[col_map['location']]).strip()
            cust_no = str(row[col_map['cust_no']]).strip() if pd.notna(row[col_map['cust_no']]) else ''
            ship_to = str(row[col_map['ship_to']]).strip() if pd.notna(row[col_map['ship_to']]) else ''

            # Clean up cust_no — remove '.0' from float conversion
            if cust_no.endswith('.0'):
                cust_no = cust_no[:-2]

            if location and location.lower() != 'nan':
                self.mappings[location] = {
                    'cust_no': cust_no,
                    'ship_to': ship_to,
                }

        self.total_loaded = len(self.mappings)
        logging.info(f"Mapping: Loaded {self.total_loaded} locations for '{party_name}'")
        return self.total_loaded

    def lookup(self, location: str) -> Optional[Dict[str, str]]:
        """
        Look up a location in the mapping.

        First tries exact match, then case-insensitive match,
        then checks if any mapping key contains the location or vice versa.

        Returns: {cust_no, ship_to, matched_key} or None if not found.
                 v1.3.3: 'matched_key' is the actual key from the mapping
                 sheet that was matched (e.g. raw "Bilaspur Warehouse -
                 Gurgaon" might fuzzy-match to canonical "Bilaspur"). Used
                 by the Summary sheet's "Location (Mapped)" column so the
                 user can visually verify the match was correct.
        """
        if not location:
            return None

        loc_clean = location.strip()

        # 1. Exact match
        if loc_clean in self.mappings:
            return {**self.mappings[loc_clean], 'matched_key': loc_clean}

        # 2. Case-insensitive match
        loc_lower = loc_clean.lower()
        for key, val in self.mappings.items():
            if key.lower() == loc_lower:
                return {**val, 'matched_key': key}

        # 3. Contains match (location in key or key in location)
        for key, val in self.mappings.items():
            key_lower = key.lower()
            if loc_lower in key_lower or key_lower in loc_lower:
                logging.info(f"Mapping: Fuzzy match '{loc_clean}' → '{key}'")
                return {**val, 'matched_key': key}

        return None


# ═══════════════════════════════════════════════════════════════════════════════
#  MASTER LOADER
# ═══════════════════════════════════════════════════════════════════════════════

class MasterLoader:
    """
    Loads Items_March.xlsx for price validation.
    Indexes by both GTIN (EAN) and No. (item code) for flexible lookup.
    """

    def __init__(self):
        self.master: Dict[str, Dict] = {}  # EAN/ItemNo → {mrp, gst_code, description}

    def load(self, filepath: str) -> int:
        """Load Items_March and build lookup. Returns item count."""
        df = pd.read_excel(filepath, header=0)
        df['GTIN_str'] = df['GTIN'].astype(str).str.strip()
        self.master = {}

        for _, r in df.iterrows():
            desc = str(r.get('Description', '')) if pd.notna(r.get('Description')) else ''
            gst = str(r['GST Group Code']) if pd.notna(r.get('GST Group Code')) else ''
            mrp = r.get('Mrp')
            item_no = str(r['No.']).strip()

            entry = {'item_no': item_no, 'mrp': mrp, 'gst_code': gst, 'description': desc}

            # Index by GTIN
            self.master[r['GTIN_str']] = entry
            # Also index by item code
            if item_no not in self.master:
                self.master[item_no] = entry

        return len(df)

    @staticmethod
    def calc_cost_price(mrp, gst_code: str, margin_pct: float) -> Optional[float]:
        """
        Calculate cost price (post-GST): MRP × margin% ÷ GST divisor.

        Args:
            mrp: Maximum Retail Price
            gst_code: Tax code from Items_March
            margin_pct: Margin as decimal (e.g., 0.70 for 70%)

        GST codes in Items_March and their divisors:
            0-G      (9 items)    → 0% GST  → ÷ 1.00
            G-3      (1 item)     → 3% GST  → ÷ 1.03
            G-5      (1084 items) → 5% GST  → ÷ 1.05
            G-5-S    (108 items)  → 5% GST  → ÷ 1.05
            G-12     (67 items)   → 12% GST → ÷ 1.12
            G-18     (2022 items) → 18% GST → ÷ 1.18
            G-18-S   (1364 items) → 18% GST → ÷ 1.18

        Returns: calculated cost price, or None if MRP is missing
        """
        if mrp is None or pd.isna(mrp):
            return None
        landing = float(mrp) * margin_pct
        gst = str(gst_code).strip().upper()

        # 0% GST (0-G, G-0, empty)
        if gst in ('0-G', 'G-0', 'G-0-S', '0', '') or gst == 'NAN':
            return landing
        # 3% GST (G-3)
        if gst in ('G-3', 'G-3-S'):
            return landing / 1.03
        # 5% GST (G-5, G-5-S)
        if '5' in gst and '18' not in gst and '12' not in gst:
            return landing / 1.05
        # 12% GST (G-12, G-12-S)
        if '12' in gst:
            return landing / 1.12
        # 18% GST (G-18, G-18-S)
        if '18' in gst:
            return landing / 1.18
        # Unknown — default to 18%
        return landing / 1.18

    @staticmethod
    def calc_landing_price(mrp, margin_pct: float) -> Optional[float]:
        """
        v1.2: Calculate the pre-GST landing rate: MRP × margin%.

        Used when a marketplace's price column is itself pre-GST (e.g.
        Myntra's "Landing Price"). No GST divisor — keeps the math
        rounding-noise-free, so a clean 0 difference is achievable.

        Args:
            mrp: Maximum Retail Price
            margin_pct: Margin as decimal (e.g., 0.70 for 70%)

        Returns: MRP × margin%, or None if MRP is missing
        """
        if mrp is None or pd.isna(mrp):
            return None
        return float(mrp) * margin_pct


    def lookup(self, key: str) -> Optional[Dict]:
        """Look up by EAN or Item No. Returns {mrp, gst_code, ...} or None."""
        key_clean = str(key).strip()
        # Try direct
        if key_clean in self.master:
            return self.master[key_clean]
        # Try stripping leading zeros (EAN sometimes has leading 0)
        stripped = key_clean.lstrip('0')
        if stripped in self.master:
            return self.master[stripped]
        return None


# ═══════════════════════════════════════════════════════════════════════════════
#  MARKETPLACE ENGINE
# ═══════════════════════════════════════════════════════════════════════════════

class MarketplaceEngine:
    """
    Processes a marketplace PO file using the appropriate column config,
    mapping data, and Items_March for price validation.
    """

    # Threshold for flagging price mismatches (in rupees)
    DIFFN_THRESHOLD = 1.0

    def __init__(self, mapping: MappingLoader, master: Optional[MasterLoader] = None):
        self.mapping = mapping
        self.master = master

    def process(self, filepath: str, config: Dict[str, Any],
                margin_pct: float = 0.70) -> ProcessingResult:
        """
        Process a marketplace PO file.

        Args:
            filepath: Path to the PO/punch Excel file
            config: Marketplace column configuration from MARKETPLACE_CONFIGS

        Returns: ProcessingResult with rows and warnings
        """
        result = ProcessingResult(
            marketplace=config['party_name'],
            input_file=os.path.basename(filepath),
            input_file_path=str(filepath),  # v1.3.4: full path for output location
            compare_basis=config.get('compare_basis', 'cost'),
            compare_label=config.get('compare_label', 'Price'),
        )

        # ── Read file ──
        try:
            df = pd.read_excel(filepath, header=0)
        except Exception as e:
            result.warnings.append(('', '', f"Cannot read file: {e}"))
            return result

        logging.info(f"Read {len(df)} rows from {os.path.basename(filepath)}")

        # Store raw data for reference sheet in output
        result.raw_df = df

        # ── Validate required columns ──
        # v1.3.5: validation depends on item_resolution.
        #   from_column → item_col is required
        #   from_ean    → ean_col is required (item_col may be absent)
        item_resolution = config.get('item_resolution', 'from_column')

        required_cols: Dict[str, str] = {
            'po': config['po_col'],
            'loc': config['loc_col'],
            'qty': config['qty_col'],
        }
        if item_resolution == 'from_ean':
            ean_required = config.get('ean_col')
            if not ean_required:
                result.warnings.append(('', '',
                    "Config error: item_resolution='from_ean' requires ean_col."))
                return result
            required_cols['ean'] = ean_required
        else:  # 'from_column' (default)
            item_required = config.get('item_col')
            if not item_required:
                result.warnings.append(('', '',
                    "Config error: item_resolution='from_column' requires item_col."))
                return result
            required_cols['item'] = item_required

        for key, col_name in required_cols.items():
            if col_name not in df.columns:
                result.warnings.append(('', '', f"Required column '{col_name}' not found. "
                                               f"Available: {list(df.columns)[:15]}..."))
                return result

        price_col = config.get('price_col')
        if price_col and price_col not in df.columns:
            logging.warning(f"Price column '{price_col}' not found — will leave empty")
            price_col = None

        # ── Validate comparison column for price validation ──
        fob_col_name = config.get('fob_col')
        fob_col_valid = fob_col_name and fob_col_name in df.columns
        if fob_col_name and not fob_col_valid:
            result.warnings.append(('', '',
                f"Validation column '{fob_col_name}' not found in file — "
                f"price validation will be skipped. Available: {list(df.columns)[:10]}..."))
            logging.warning(f"FOB/Cost column '{fob_col_name}' not found — validation disabled")

        ean_col_name = config.get('ean_col')
        ean_col_valid = ean_col_name and ean_col_name in df.columns
        if ean_col_name and not ean_col_valid:
            result.warnings.append(('', '',
                f"EAN column '{ean_col_name}' not found in file — "
                f"master lookup will use Item No only. Available: {list(df.columns)[:10]}..."))
            logging.warning(f"EAN column '{ean_col_name}' not found — using Item No for lookup")

        # ── Process each row ──
        warned_locations = set()  # Track (po, location) to avoid duplicate warnings
        for _, row in df.iterrows():
            po = str(row[config['po_col']]).strip()
            location = str(row[config['loc_col']]).strip() if pd.notna(row[config['loc_col']]) else ''
            qty_raw = row[config['qty_col']]

            # Skip rows with no PO
            if po.lower() == 'nan':
                continue

            # Parse quantity early — skip zero-qty rows before doing any
            # expensive lookups.
            try:
                qty = int(float(qty_raw)) if pd.notna(qty_raw) else 0
            except (ValueError, TypeError):
                qty = 0

            if qty <= 0:
                continue

            # ── Extract EAN for master lookup ──
            # v1.3.5: extracted BEFORE item resolution because Myntra-style
            # configs (item_resolution='from_ean') need it to derive Item No.
            ean = ''
            ean_col = config.get('ean_col')
            if ean_col and ean_col in df.columns:
                ean_raw = row[ean_col]
                if pd.notna(ean_raw):
                    # GTINs may arrive as float64 (e.g. 8906121642599.0) —
                    # strip the trailing .0 by going through int().
                    if isinstance(ean_raw, (int, float)):
                        try:
                            ean = str(int(ean_raw))
                        except (ValueError, OverflowError):
                            ean = str(ean_raw).strip()
                    else:
                        ean = str(ean_raw).strip()

            # ── Resolve Item No (per item_resolution mode) ──
            # v1.3.5: two paths.
            #   from_column → item_col is read from row (e.g. RK)
            #   from_ean    → master lookup by EAN, take master_info['item_no']
            #                 (e.g. real Myntra files which lack Item No)
            item_no = None
            if item_resolution == 'from_ean':
                if not ean:
                    # No EAN to look up — can't resolve. Skip this row.
                    warn_key = ('NO_EAN', po)
                    if warn_key not in warned_locations:
                        warned_locations.add(warn_key)
                        result.warnings.append((
                            po, '',
                            f"Row skipped: ean_col '{ean_col}' is empty for PO {po}. "
                            f"item_resolution='from_ean' requires a non-empty EAN."
                        ))
                    continue

                if not self.master:
                    # Can't resolve from EAN without a master loaded. Skip.
                    warn_key = ('NO_MASTER', 'global')
                    if warn_key not in warned_locations:
                        warned_locations.add(warn_key)
                        result.warnings.append((
                            '', '',
                            "Cannot resolve Item No: item_resolution='from_ean' "
                            "requires the Items_March master to be loaded."
                        ))
                    continue

                master_info_for_item = self.master.lookup(ean)
                if not master_info_for_item:
                    # EAN not in master — record as unresolvable. We use the
                    # EAN itself as the visible item value so the row still
                    # appears in output (in NOT_IN_MASTER state).
                    item_no = ean
                else:
                    # Use the canonical Item No from master.
                    resolved = master_info_for_item.get('item_no', '')
                    try:
                        item_no = int(resolved)
                    except (ValueError, TypeError):
                        item_no = str(resolved).strip()

            else:  # 'from_column' — original v1.2 behavior
                item_raw = row[config['item_col']]
                if pd.isna(item_raw):
                    continue  # No item value → skip
                try:
                    item_no = int(item_raw)
                except (ValueError, TypeError):
                    item_no = str(item_raw).strip()

            # Parse unit price (if configured)
            unit_price = None
            if price_col:
                try:
                    p = row[price_col]
                    unit_price = float(p) if pd.notna(p) else None
                except (ValueError, TypeError):
                    unit_price = None

            # ── Extract marketplace FOB/Cost price ──
            fob_price = None
            fob_col = config.get('fob_col')
            if fob_col and fob_col in df.columns:
                try:
                    fob_raw = row[fob_col]
                    fob_price = float(fob_raw) if pd.notna(fob_raw) else None
                except (ValueError, TypeError):
                    fob_price = None

            # v1.3.4: extract OPTIONAL reference price (for visibility only —
            # not used for OK/MISMATCH logic). Currently used by Myntra to
            # surface the legacy 'List price(FOB+Transport-Excise)' diff
            # alongside the active 'Landing Rate' diff in the Raw Data sheet.
            ref_fob_price = None
            ref_fob_col = config.get('ref_fob_col')
            if ref_fob_col and ref_fob_col in df.columns:
                try:
                    rfob_raw = row[ref_fob_col]
                    ref_fob_price = float(rfob_raw) if pd.notna(rfob_raw) else None
                except (ValueError, TypeError):
                    ref_fob_price = None

            # ── Validate price against Items_March ──
            calc_price = None        # value used for the diff (depends on compare_basis)
            cost_price_ref = None    # always the post-GST cost price (naked CP, for display)
            mrp = None
            gst_code = ''
            description = ''         # v1.3.1: filled when master lookup hits
            diffn = None
            ref_diffn = None         # v1.3.4: reference diff (vs cost_price_ref)
            validation_status = ''

            # v1.2: pick comparison basis from config
            #   'landing' → MRP × margin%        (pre-GST, e.g. Myntra Landing Price)
            #   'cost'    → MRP × margin% ÷ GST  (post-GST, e.g. RK Cost) — default
            compare_basis = config.get('compare_basis', 'cost')
            compare_label = config.get('compare_label', 'Price')

            if self.master:
                # Try lookup by EAN first, then by Item No
                master_info = self.master.lookup(ean) if ean else None
                if not master_info:
                    master_info = self.master.lookup(str(item_no))

                if master_info:
                    mrp = master_info['mrp']
                    gst_code = master_info['gst_code']
                    # v1.3.1: capture description for the Validation sheet
                    description = master_info.get('description', '')

                    # Warn if GST code is not in known set
                    known_gst = {'0-G', 'G-3', 'G-3-S', 'G-5', 'G-5-S', 'G-12', 'G-12-S', 'G-18', 'G-18-S', ''}
                    gst_upper = str(gst_code).strip().upper()
                    if gst_upper not in known_gst and gst_upper != 'NAN':
                        warn_key = ('GST', gst_upper)
                        if warn_key not in warned_locations:
                            warned_locations.add(warn_key)
                            result.warnings.append((
                                po, str(item_no),
                                f"Unknown GST code '{gst_code}' for Item {item_no} — "
                                f"defaulting to 18%. Please verify in Items_March."
                            ))
                            logging.warning(f"Unknown GST code '{gst_code}' for Item {item_no}")

                    # Always compute the post-GST cost price for the "naked CP"
                    # reference column shown in the Validation sheet.
                    cost_price_ref = MasterLoader.calc_cost_price(mrp, gst_code, margin_pct)

                    # v1.3.4: reference diff (vs naked CP) — used purely for
                    # display in the Raw Data sheet. Always computed against
                    # cost_price_ref (post-GST) regardless of compare_basis,
                    # because the reference column is the post-GST one
                    # (e.g. Myntra's List price).
                    if cost_price_ref is not None and ref_fob_price is not None:
                        ref_diffn = ref_fob_price - cost_price_ref

                    # Pick what we ACTUALLY compare against, based on basis.
                    if compare_basis == 'landing':
                        calc_price = MasterLoader.calc_landing_price(mrp, margin_pct)
                    else:  # 'cost' (default)
                        calc_price = cost_price_ref

                    if calc_price is not None and fob_price is not None:
                        diffn = fob_price - calc_price
                        if abs(diffn) <= self.DIFFN_THRESHOLD:
                            validation_status = 'OK'
                        else:
                            validation_status = 'MISMATCH'
                            warn_key = ('VALIDATION', str(item_no))
                            if warn_key not in warned_locations:
                                warned_locations.add(warn_key)
                                result.warnings.append((
                                    po, str(item_no),
                                    f"{compare_label} mismatch: Item {item_no}, "
                                    f"Marketplace={fob_price:.2f}, "
                                    f"Calculated={calc_price:.2f}, "
                                    f"Diff={diffn:.2f}"
                                ))
                    else:
                        validation_status = 'NO_PRICE'
                else:
                    validation_status = 'NOT_IN_MASTER'

            # ── Lookup location in mapping ──
            mapping_result = self.mapping.lookup(location)

            if mapping_result:
                cust_no = mapping_result['cust_no']
                ship_to = mapping_result['ship_to']
                # v1.3.3: capture the canonical mapping key we matched to
                # (could differ from raw `location` when fuzzy match was used)
                mapped_location = mapping_result.get('matched_key', location)
                mapped = True
            else:
                cust_no = ''
                ship_to = ''
                mapped_location = ''
                mapped = False
                warn_key = (po, location)
                if warn_key not in warned_locations:
                    warned_locations.add(warn_key)
                    result.warnings.append((
                        po, location,
                        f"Location '{location}' not found in mapping for {config['party_name']}. "
                        f"Cust No and Ship-to left empty."
                    ))

            result.rows.append(SORow(
                po_number=po,
                location=location,
                item_no=item_no,
                qty=qty,
                unit_price=unit_price,
                cust_no=cust_no,
                ship_to=ship_to,
                mapped=mapped,
                mapped_location=mapped_location,
                ean=ean,
                description=description,
                fob_price=fob_price,
                ref_fob_price=ref_fob_price,
                calc_price=calc_price,
                cost_price_ref=cost_price_ref,
                diffn=diffn,
                ref_diffn=ref_diffn,
                mrp=mrp,
                gst_code=gst_code,
                validation_status=validation_status,
            ))

        logging.info(f"Processed {len(result.rows)} items across "
                     f"{len(set(r.po_number for r in result.rows))} PO(s)")
        return result


# ═══════════════════════════════════════════════════════════════════════════════
#  SO EXPORTER
# ═══════════════════════════════════════════════════════════════════════════════

class SOExporter:
    """Writes the output Excel with Headers (SO), Lines (SO), Summary, Warnings."""

    # ── Formatting constants ──
    HEADER_FILL = PatternFill('solid', fgColor='1A237E')
    HEADER_FONT = Font(bold=True, color='FFFFFF', name='Aptos Display', size=11)
    DATA_FONT = Font(name='Aptos Display', size=11)
    THIN_SIDE = Side(style='thin', color='CCCCCC')
    BORDER = Border(left=THIN_SIDE, right=THIN_SIDE, top=THIN_SIDE, bottom=THIN_SIDE)

    # Warning header — orange
    WARN_FILL = PatternFill('solid', fgColor='E65100')

    def _hdr_cell(self, ws, row, col, value, fill=None):
        """Create a formatted header cell."""
        cell = ws.cell(row=row, column=col, value=value)
        cell.font = self.HEADER_FONT
        cell.fill = fill or self.HEADER_FILL
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = self.BORDER
        return cell

    def _data_cell(self, ws, row, col, value, fmt=None):
        """Create a formatted data cell."""
        cell = ws.cell(row=row, column=col, value=value)
        cell.font = self.DATA_FONT
        cell.border = self.BORDER
        if fmt:
            cell.number_format = fmt
        return cell

    def _auto_width(self, ws, max_w=50):
        """Auto-fit column widths based on content."""
        for col in ws.columns:
            letter = col[0].column_letter
            w = max((len(str(c.value or '')) for c in col), default=8)
            ws.column_dimensions[letter].width = min(w + 3, max_w)

    def export(self, result: ProcessingResult) -> Optional[Path]:
        """
        Write output Excel.
        Returns output path on success, None if no data.

        v1.3.4: Output is saved next to the input punch file in an 'output/'
        subfolder, e.g.:

            D:\\PO\\Myntra\\April\\Myntra_Punch_17-04-2026.xlsx       ← input
            D:\\PO\\Myntra\\April\\output\\myntra_so_19-04-2026_*.xlsx ← output

        Falls back to the script's working directory + 'output_online/' if
        the input path isn't recorded for some reason (defensive — should
        not normally happen since the engine populates input_file_path).
        """
        if not result.rows:
            messagebox.showwarning("No Data", "No valid rows found.\nNothing to export.")
            return None

        # ── Prepare output path ──
        # v1.3.4: save next to the input file
        if result.input_file_path:
            input_dir = Path(result.input_file_path).parent
            output_folder = input_dir / 'output'
        else:
            # Defensive fallback (shouldn't trigger in normal flow)
            output_folder = Path("output_online")

        output_folder.mkdir(parents=True, exist_ok=True)
        today = datetime.now().strftime("%d-%m-%Y_%H%M%S")
        marketplace = result.marketplace.lower().replace(' ', '_')
        file_path = output_folder / f"{marketplace}_so_{today}.xlsx"

        # ── Create workbook ──
        wb = Workbook()
        wb.remove(wb.active)

        self._write_headers_so(wb, result)
        self._write_lines_so(wb, result)
        self._write_summary(wb, result)
        self._write_validation(wb, result)
        self._write_warnings(wb, result)
        self._write_raw_data(wb, result)

        wb.save(str(file_path))
        logging.info(f"Output saved: {file_path}")
        return file_path

    def _write_headers_so(self, wb, result: ProcessingResult):
        """
        Sheet 1: 'Headers (SO)' — One row per unique PO number.
        """
        ws = wb.create_sheet('Headers (SO)')
        headers = [
            'Document Type', 'No.', 'Sell-to Customer No.', 'Ship-to Code',
            'Posting Date', 'Order Date', 'Document Date',
            'Invoice From Date', 'Invoice To Date',
            'External Document No.', 'Location Code', 'Dimension Set ID',
            'Supply Type', 'Voucher Narration',
            'Brand Code (Dimension)', 'Channel Code (Dimension)',
            'Catagory (Dimension)', 'Geography Code (Dimension)'
        ]
        for c, h in enumerate(headers, 1):
            self._hdr_cell(ws, 1, c, h)

        today_str = datetime.now().strftime("%d-%m-%Y")

        # Collect unique POs preserving order
        seen = set()
        unique_pos = []
        for row in result.rows:
            if row.po_number not in seen:
                seen.add(row.po_number)
                unique_pos.append(row)

        r = 2
        for row in unique_pos:
            self._data_cell(ws, r, 1, 'Order')              # Document Type
            self._data_cell(ws, r, 2, row.po_number)         # No.
            self._data_cell(ws, r, 3, row.cust_no)           # Sell-to Customer No.
            self._data_cell(ws, r, 4, row.ship_to)           # Ship-to Code
            self._data_cell(ws, r, 5, today_str)             # Posting Date
            self._data_cell(ws, r, 6, today_str)             # Order Date
            self._data_cell(ws, r, 7, today_str)             # Document Date
            self._data_cell(ws, r, 8, today_str)             # Invoice From Date
            self._data_cell(ws, r, 9, today_str)             # Invoice To Date
            self._data_cell(ws, r, 10, row.po_number)        # External Document No.
            self._data_cell(ws, r, 11, 'PICK')               # Location Code (always PICK)
            self._data_cell(ws, r, 12, '')                   # Dimension Set ID
            self._data_cell(ws, r, 13, 'B2B')                # Supply Type
            # Columns 14-18: empty dimension columns
            r += 1

        self._auto_width(ws)

    def _write_lines_so(self, wb, result: ProcessingResult):
        """
        Sheet 2: 'Lines (SO)' — One row per ordered item.
        Line No. increments by 10000, resets per new PO.
        """
        ws = wb.create_sheet('Lines (SO)')
        headers = ['Document Type', 'Document No.', 'Line No.', 'Type',
                   'No.', 'Location Code', 'Quantity', 'Unit Price']
        for c, h in enumerate(headers, 1):
            self._hdr_cell(ws, 1, c, h)

        r = 2
        current_po = None
        line_no = 0

        for row in result.rows:
            # Reset line counter on new PO
            if row.po_number != current_po:
                current_po = row.po_number
                line_no = 0

            line_no += 10000

            self._data_cell(ws, r, 1, 'Order')              # Document Type
            self._data_cell(ws, r, 2, row.po_number)         # Document No.
            self._data_cell(ws, r, 3, line_no)               # Line No.
            self._data_cell(ws, r, 4, 'Item')                # Type
            self._data_cell(ws, r, 5, row.item_no)           # No. (Item No)
            self._data_cell(ws, r, 6, 'PICK')                # Location Code
            self._data_cell(ws, r, 7, row.qty)               # Quantity
            self._data_cell(ws, r, 8, '')                    # Unit Price (empty)
            r += 1

        self._auto_width(ws)

    def _write_summary(self, wb, result: ProcessingResult):
        """
        Sheet 3: 'Summary' — Per-PO grouped info for verification.

        v1.3.3: Location column split into two for visual verification:
            • Location (Raw)    — what the marketplace's punch file says
            • Location (Mapped) — the canonical key from our Ship-To registry
                                   that we matched the raw value to

        These should usually match exactly. When they don't (e.g. fuzzy
        match: "Bilaspur Warehouse - Gurgaon" → "Bilaspur"), both cells
        get a pale yellow fill so the user can eyeball the matches and
        spot any wrong/loose mappings at a glance.
        """
        ws = wb.create_sheet('Summary')
        # v1.3.3: 2 location columns instead of 1; everything else shifts +1
        headers = ['PO', 'Location (Raw)', 'Location (Mapped)',
                   'Cust No', 'Ship-to', 'Items', 'Total Qty', 'Status']
        for c, h in enumerate(headers, 1):
            self._hdr_cell(ws, 1, c, h)

        # Status column index (always last). Used for color-coding below.
        status_col = len(headers)  # = 8 in v1.3.3

        # Highlight for raw vs mapped mismatches (pale yellow)
        # — applied to BOTH location cells so the diff is obvious.
        loc_mismatch_fill = PatternFill('solid', fgColor='FFF59D')

        # Group by PO. v1.3.3: also capture mapped_location.
        po_groups: Dict[str, dict] = {}
        for row in result.rows:
            if row.po_number not in po_groups:
                po_groups[row.po_number] = {
                    'location': row.location,
                    'mapped_location': row.mapped_location,
                    'cust_no': row.cust_no,
                    'ship_to': row.ship_to,
                    'items': 0,
                    'qty': 0,
                    'mapped': row.mapped,
                }
            po_groups[row.po_number]['items'] += 1
            po_groups[row.po_number]['qty'] += row.qty

        r = 2
        for po, info in po_groups.items():
            status = 'OK' if info['mapped'] else 'UNMAPPED'

            self._data_cell(ws, r, 1, po)
            self._data_cell(ws, r, 2, info['location'])           # Raw
            self._data_cell(ws, r, 3, info['mapped_location'])    # Mapped
            self._data_cell(ws, r, 4, info['cust_no'])
            self._data_cell(ws, r, 5, info['ship_to'])
            self._data_cell(ws, r, 6, info['items'])
            self._data_cell(ws, r, 7, info['qty'])
            self._data_cell(ws, r, status_col, status)

            # v1.3.3: highlight when raw ≠ mapped (case-insensitive). Means
            # we used a fuzzy match and the user should eyeball the result.
            raw = (info['location'] or '').strip().lower()
            mapped_loc = (info['mapped_location'] or '').strip().lower()
            if info['mapped'] and raw and mapped_loc and raw != mapped_loc:
                ws.cell(row=r, column=2).fill = loc_mismatch_fill
                ws.cell(row=r, column=3).fill = loc_mismatch_fill

            # Color status cell
            status_cell = ws.cell(row=r, column=status_col)
            if status == 'OK':
                status_cell.fill = PatternFill('solid', fgColor='00C853')
                status_cell.font = Font(name='Aptos Display', size=11, bold=True, color='000000')
            else:
                status_cell.fill = PatternFill('solid', fgColor='FF5252')
                status_cell.font = Font(name='Aptos Display', size=11, bold=True, color='FFFFFF')

            r += 1

        # Totals row (column indices shifted +1 from v1.3.2)
        total_items = sum(g['items'] for g in po_groups.values())
        total_qty = sum(g['qty'] for g in po_groups.values())
        self._data_cell(ws, r, 1, 'TOTAL')
        ws.cell(row=r, column=1).font = Font(name='Aptos Display', size=11, bold=True)
        self._data_cell(ws, r, 6, total_items)
        ws.cell(row=r, column=6).font = Font(name='Aptos Display', size=11, bold=True)
        self._data_cell(ws, r, 7, total_qty)
        ws.cell(row=r, column=7).font = Font(name='Aptos Display', size=11, bold=True)

        # Info row — marketplace and margin
        r += 2
        info_font = Font(name='Aptos Display', size=10, italic=True, color='666666')
        margin_str = f"{int(result.margin_pct * 100)}%"
        ws.cell(row=r, column=1, value=f"Marketplace: {result.marketplace}  |  "
                                        f"Margin: {margin_str}  |  "
                                        f"File: {result.input_file}  |  "
                                        f"Generated: {datetime.now().strftime('%d-%m-%Y %H:%M')}").font = info_font

        # v1.3.3: legend for the yellow highlight, only printed if there
        # was at least one mismatch — otherwise it's noise.
        any_loc_mismatch = any(
            (g['mapped'] and (g['location'] or '').strip().lower()
             != (g['mapped_location'] or '').strip().lower()
             and g['location'] and g['mapped_location'])
            for g in po_groups.values()
        )
        if any_loc_mismatch:
            r += 1
            ws.cell(row=r, column=1,
                    value="🟨 Yellow = raw and mapped location differ "
                          "(fuzzy match) — please verify."
                    ).font = Font(name='Aptos Display', size=10,
                                   italic=True, color='B7950B')

        self._auto_width(ws)

    def _write_validation(self, wb, result: ProcessingResult):
        """
        Sheet: 'Validation' — Price check.

        v1.2: column layout adapts to the marketplace's compare_basis:

            Common columns (always shown):
                PO | Item No | EAN | MRP | Landing (m%) | GST Code |
                Our Cost Price                         ← naked CP, always shown
                                                         for reference even when
                                                         not used for the diff
                Marketplace <Label>                    ← e.g. "Marketplace
                                                         Landing Rate" for
                                                         Myntra, "Marketplace
                                                         Cost" for RK
                Difference with <Label>                ← clear, marketplace-aware
                                                         column name
                Status                                 ← OK / MISMATCH /
                                                         NOT_IN_MASTER

            For compare_basis = 'landing' (Myntra):
                The "Marketplace Landing Rate" is compared against
                "Landing (m%)" (= MRP × m%, pre-GST). Diff is therefore
                clean (no GST rounding noise).

            For compare_basis = 'cost' (RK, default):
                The "Marketplace Cost" is compared against
                "Our Cost Price" (= MRP × m% ÷ GST, post-GST).
                Diff may have small rounding noise (≤ 1 rupee threshold).
        """
        ws = wb.create_sheet('Validation')

        label = result.compare_label or 'Price'
        margin_pct_int = int(result.margin_pct * 100)

        # v1.3.1: 'Description' column added immediately after 'EAN' so the
        # user can read what each EAN actually is at a glance, instead of
        # cross-referencing back to the master. All subsequent columns
        # shifted right by 1.
        headers = ['PO', 'Item No', 'EAN', 'Description', 'MRP',
                   f'Landing ({margin_pct_int}%)', 'GST Code',
                   'Our Cost Price',
                   f'Marketplace {label}',
                   f'Difference with {label}',
                   'Status']

        # Green header for our calculated columns (MRP, Landing, GST, Cost Price),
        # standard blue header for everything else.
        # v1.3.1: indices shifted +1 from v1.2 because of the new Description col.
        calc_fill = PatternFill('solid', fgColor='1B5E20')
        calc_col_indices = {5, 6, 7, 8}  # MRP, Landing, GST Code, Our Cost Price
        for c, h in enumerate(headers, 1):
            fill = calc_fill if c in calc_col_indices else self.HEADER_FILL
            self._hdr_cell(ws, 1, c, h, fill=fill)

        # Color fills for status
        ok_fill = PatternFill('solid', fgColor='E8F5E9')
        mismatch_fill = PatternFill('solid', fgColor='FFEBEE')
        no_master_fill = PatternFill('solid', fgColor='FFF3E0')

        # Number of columns (used for full-row highlighting)
        n_cols = len(headers)
        status_col = n_cols  # Status is always the last column

        r = 2
        mismatches = 0
        for row in result.rows:
            self._data_cell(ws, r, 1, row.po_number)
            self._data_cell(ws, r, 2, row.item_no)
            self._data_cell(ws, r, 3, row.ean)
            self._data_cell(ws, r, 4, row.description)         # v1.3.1: new col
            self._data_cell(ws, r, 5, row.mrp, '#,##0.00' if row.mrp else None)

            # Landing cost (= MRP × margin%)
            landing = float(row.mrp) * result.margin_pct if row.mrp and not pd.isna(row.mrp) else None
            self._data_cell(ws, r, 6, round(landing, 2) if landing else '', '#,##0.00')

            self._data_cell(ws, r, 7, row.gst_code)

            # Our Cost Price (naked CP) — always shown, regardless of compare_basis
            self._data_cell(ws, r, 8,
                            round(row.cost_price_ref, 2) if row.cost_price_ref else '',
                            '#,##0.00')

            # Marketplace value (whatever fob_col extracted)
            self._data_cell(ws, r, 9,
                            round(row.fob_price, 2) if row.fob_price else '',
                            '#,##0.00')

            # Difference (against calc_price, which is landing OR cost depending on basis)
            # v1.2.1: display rounded to 2 decimals (0.00) — finer precision is
            # just floating-point dust and adds visual noise.
            self._data_cell(ws, r, 10,
                            round(row.diffn, 2) if row.diffn is not None else '',
                            '#,##0.00')

            self._data_cell(ws, r, status_col, row.validation_status)

            # Row highlighting based on status
            if row.validation_status == 'MISMATCH':
                mismatches += 1
                for c in range(1, n_cols + 1):
                    ws.cell(row=r, column=c).fill = mismatch_fill
                ws.cell(row=r, column=status_col).font = Font(
                    name='Aptos Display', size=11, bold=True, color='D32F2F')
            elif row.validation_status == 'OK':
                ws.cell(row=r, column=status_col).fill = PatternFill(
                    'solid', fgColor='00C853')
                ws.cell(row=r, column=status_col).font = Font(
                    name='Aptos Display', size=11, bold=True, color='000000')
            elif row.validation_status == 'NOT_IN_MASTER':
                for c in range(1, n_cols + 1):
                    ws.cell(row=r, column=c).fill = no_master_fill
                ws.cell(row=r, column=status_col).font = Font(
                    name='Aptos Display', size=11, bold=True, color='E65100')

            r += 1

        # Summary row — also notes the comparison basis used
        r += 1
        total = len(result.rows)
        ok_count = sum(1 for row in result.rows if row.validation_status == 'OK')
        basis_note = (f"basis={result.compare_basis} "
                      f"(compared against '{label}')")
        ws.cell(row=r, column=1,
                value=f"Total: {total} items | OK: {ok_count} | "
                      f"Mismatches: {mismatches} | "
                      f"Margin: {margin_pct_int}% | {basis_note}").font = \
            Font(name='Aptos Display', size=10, italic=True, color='666666')

        self._auto_width(ws)
        ws.freeze_panes = 'A2'

    def _write_warnings(self, wb, result: ProcessingResult):
        """
        Sheet 4: 'Warnings' — Only created if warnings exist.
        """
        if not result.warnings:
            return

        ws = wb.create_sheet('Warnings')
        headers = ['PO', 'Location', 'Warning']
        for c, h in enumerate(headers, 1):
            self._hdr_cell(ws, 1, c, h, fill=self.WARN_FILL)

        for r, (po, loc, msg) in enumerate(result.warnings, 2):
            self._data_cell(ws, r, 1, po)
            self._data_cell(ws, r, 2, loc)
            self._data_cell(ws, r, 3, msg)

        self._auto_width(ws)

    def _write_raw_data(self, wb, result: ProcessingResult):
        """
        Sheet: 'Raw Data' — Original marketplace data + calculated validation columns.

        Copies ALL original columns from the input file, then appends calculated
        columns from the Items_March master lookup:

            Item No (Master) | MRP | Landing (m%) | GST Code | Cost Price |
            [Diffn with <ref_label>]?  |  Diffn with <Label>

        The optional reference Diffn column appears when the marketplace's
        config defines `ref_fob_col` (v1.3.4+). For Myntra this surfaces the
        old "Diffn with List price(FOB+Transport-Excise)" alongside the
        active "Diffn with Landing Rate" — purely for visibility, the
        reference diff has zero effect on validation status.

        This gives a single-workbook reference: original marketplace data +
        our price validation, without needing to open the source file separately.
        """
        if result.raw_df is None or result.raw_df.empty:
            return

        ws = wb.create_sheet('Raw Data')
        df = result.raw_df

        # ── Calculate columns to append ──
        # Build a lookup from result.rows. v1.3.5: index by BOTH (po, item_no)
        # and (po, ean), so the row-matching loop below can find the validation
        # row whether the punch file has Item No (RK) or only EAN (Myntra).
        validation_lookup: Dict[tuple, SORow] = {}
        for row in result.rows:
            validation_lookup[(row.po_number, str(row.item_no))] = row
            if row.ean:
                validation_lookup[(row.po_number, str(row.ean))] = row

        # ── Original column headers (dark grey) ──
        raw_hdr_fill = PatternFill('solid', fgColor='37474F')
        orig_col_count = len(df.columns)
        for c, col_name in enumerate(df.columns, 1):
            cell = ws.cell(row=1, column=c, value=str(col_name))
            cell.font = self.HEADER_FONT
            cell.fill = raw_hdr_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = self.BORDER

        # ── Look up the marketplace's ref_fob_col (if any) so we can label
        # ── the reference Diffn column with its name (v1.3.4).
        marketplace_cfg = None
        for cfg in MARKETPLACE_CONFIGS.values():
            if cfg['party_name'] == result.marketplace:
                marketplace_cfg = cfg
                break
        ref_fob_col_name = (marketplace_cfg or {}).get('ref_fob_col')
        # If the punch file doesn't actually have the ref column, hide
        # the reference Diffn entirely — no point showing an empty column.
        has_ref_diff = bool(ref_fob_col_name) and (ref_fob_col_name in df.columns)

        # ── Calculated column headers (green) ──
        # v1.2: Diffn column name follows the compare_label so it's clear
        # which marketplace value is being compared against.
        diff_label = f"Diffn with {result.compare_label}" if result.compare_label \
            else "Diffn"
        calc_headers = ['Item No (Master)', 'MRP',
                        f'Landing ({int(result.margin_pct*100)}%)',
                        'GST Code', 'Cost Price']
        if has_ref_diff:
            # v1.3.4: insert reference Diffn BEFORE the active Diffn so the
            # active (validation-driving) column ends up rightmost — matches
            # the user's reading flow (look at our calc → compare → final diff).
            calc_headers.append(f'Diffn with {ref_fob_col_name}')
        calc_headers.append(diff_label)

        calc_fill = PatternFill('solid', fgColor='1B5E20')   # active (validation)
        ref_fill = PatternFill('solid', fgColor='455A64')    # reference (muted)

        # Index in calc_headers of the reference diff column (if present)
        # and the active diff column. Used for per-cell tinting later.
        ref_idx = (len(calc_headers) - 2) if has_ref_diff else None  # 0-based
        active_idx = len(calc_headers) - 1                            # 0-based

        for i, h in enumerate(calc_headers):
            c = orig_col_count + i + 1
            cell = ws.cell(row=1, column=c, value=h)
            cell.font = self.HEADER_FONT
            # Reference Diffn header gets the muted fill so the user can tell
            # at a glance it's not the validation-driving column.
            cell.fill = ref_fill if (has_ref_diff and i == ref_idx) else calc_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = self.BORDER

        # ── Data rows: original + calculated ──
        # v1.3.5: derive the per-row "lookup value" from whichever column
        # the marketplace actually has. RK gives us item_col; Myntra only
        # gives us ean_col.
        config_item_col = (marketplace_cfg or {}).get('item_col')
        config_ean_col = (marketplace_cfg or {}).get('ean_col')
        config_po_col = (marketplace_cfg or {}).get('po_col')

        calc_bg = PatternFill('solid', fgColor='E8F5E9')      # Light green for calc cells
        ref_bg = PatternFill('solid', fgColor='ECEFF1')       # Light grey for ref cells
        mismatch_bg = PatternFill('solid', fgColor='FFCDD2')  # Light red for mismatch

        n_calc = len(calc_headers)

        for r, (_, row) in enumerate(df.iterrows(), 2):
            # ── Write original columns ──
            for c, col_name in enumerate(df.columns, 1):
                val = row[col_name]
                if isinstance(val, pd.Timestamp):
                    val = val.strftime('%d-%m-%Y')
                elif pd.isna(val):
                    val = ''
                self._data_cell(ws, r, c, val)

            # ── Write calculated columns ──
            po_val = str(row[config_po_col]).strip() if config_po_col and config_po_col in df.columns else ''

            # Build the lookup key — try item_col first, then ean_col.
            # validation_lookup is indexed under both, so either path works.
            lookup_val = ''
            if config_item_col and config_item_col in df.columns:
                iv = row[config_item_col]
                try:
                    lookup_val = str(int(iv)) if pd.notna(iv) else ''
                except (ValueError, TypeError):
                    lookup_val = str(iv).strip() if pd.notna(iv) else ''

            if not lookup_val and config_ean_col and config_ean_col in df.columns:
                # Item col missing or empty — fall back to EAN. Same float→int
                # treatment as the engine uses (8906121642599.0 → '8906121642599').
                ev = row[config_ean_col]
                if pd.notna(ev):
                    if isinstance(ev, (int, float)):
                        try:
                            lookup_val = str(int(ev))
                        except (ValueError, OverflowError):
                            lookup_val = str(ev).strip()
                    else:
                        lookup_val = str(ev).strip()

            # Find matching validation row
            vrow = validation_lookup.get((po_val, lookup_val))

            base_c = orig_col_count + 1
            if vrow:
                landing = float(vrow.mrp) * result.margin_pct if vrow.mrp and not pd.isna(vrow.mrp) else None

                self._data_cell(ws, r, base_c, vrow.item_no)                     # Item No (Master)
                self._data_cell(ws, r, base_c + 1, vrow.mrp, '#,##0.00')         # MRP
                self._data_cell(ws, r, base_c + 2,
                                round(landing, 2) if landing else '', '#,##0.00')  # Landing
                self._data_cell(ws, r, base_c + 3, vrow.gst_code)                # GST Code
                # v1.2: always show the naked CP (post-GST cost price) here,
                # even when compare_basis = 'landing'. This is the reference
                # value the user wants visible regardless of what's compared.
                self._data_cell(ws, r, base_c + 4,
                                round(vrow.cost_price_ref, 2) if vrow.cost_price_ref else '',
                                '#,##0.00')  # Cost Price (naked CP)

                # v1.3.4: reference Diffn column (if configured)
                if has_ref_diff:
                    self._data_cell(ws, r, base_c + ref_idx,
                                    round(vrow.ref_diffn, 2) if vrow.ref_diffn is not None else '',
                                    '#,##0.00')

                # Active Diffn (validation-driving)
                self._data_cell(ws, r, base_c + active_idx,
                                round(vrow.diffn, 2) if vrow.diffn is not None else '',
                                '#,##0.00')

                # Apply background color across all calc cells
                is_mismatch = vrow.validation_status == 'MISMATCH'
                fill = mismatch_bg if is_mismatch else calc_bg
                for i in range(n_calc):
                    ws.cell(row=r, column=base_c + i).fill = fill

                # v1.3.4: re-tint the reference Diffn column with a muted
                # grey background so it's clearly distinguished from the
                # validation-driving Diffn. Only apply when NOT mismatched
                # (mismatch overrides everything for visibility).
                if has_ref_diff and not is_mismatch:
                    ws.cell(row=r, column=base_c + ref_idx).fill = ref_bg

                # Bold red Diffn if mismatch (active column only)
                if is_mismatch:
                    ws.cell(row=r, column=base_c + active_idx).font = Font(
                        name='Aptos Display', size=11, bold=True, color='D32F2F')
            else:
                # No matching validation row (qty=0 items in original file)
                for i in range(n_calc):
                    self._data_cell(ws, r, base_c + i, '')

        self._auto_width(ws)
        ws.freeze_panes = 'A2'


# ═══════════════════════════════════════════════════════════════════════════════
#  FILE OPENER (cross-platform)
# ═══════════════════════════════════════════════════════════════════════════════

def open_file(file_path: Path):
    """Opens the file using the OS default application."""
    try:
        system = platform.system()
        if system == "Windows":
            os.startfile(str(file_path))
        elif system == "Darwin":
            import subprocess as sp
            sp.Popen(["open", str(file_path)])
        else:
            import subprocess as sp
            sp.Popen(["xdg-open", str(file_path)])
    except Exception as e:
        messagebox.showerror("Open File Error", f"Could not open file:\n{e}")


# ═══════════════════════════════════════════════════════════════════════════════
#  TKINTER GUI
# ═══════════════════════════════════════════════════════════════════════════════

class OnlinePOApp:
    """GUI for Online Marketplace PO → SO generation."""

    def __init__(self):
        self.root = tk.Tk()
        self.root.title("Online PO Processor — Marketplace SO Generator")
        # v1.3.2: bumped height again (+40px) to fit the per-file timestamp
        # sub-lines under the Master/Mapping rows ("Updated: 19-Apr-2026 18:41").
        self.root.geometry("520x620")
        self.root.resizable(False, False)

        # ── State ──
        self.master_path: Optional[str] = None
        self.mapping_path: Optional[str] = None
        self.po_path: Optional[str] = None
        self.last_output: Optional[Path] = None

        # v1.3: track whether master/mapping came from the bundled folder
        # (vs user-picked). Used so the GUI can show "(auto-loaded)" hints
        # and so the "Update Bundled Files" flow knows what's currently in use.
        self.master_is_bundled: bool = False
        self.mapping_is_bundled: bool = False

        # ── Engine ──
        self.mapping_loader = MappingLoader()
        self.exporter = SOExporter()

        self._build_ui()

        # v1.3: auto-load bundled master/mapping AFTER UI is built so we can
        # log to the panel and update the picker labels in one go.
        self._auto_load_bundled_files()

    def _build_ui(self):
        """Build the GUI layout."""

        # ── Title ──
        tk.Label(
            self.root, text="Online PO Processor",
            font=("Arial", 14, "bold")
        ).pack(pady=(12, 2))

        tk.Label(
            self.root, text="Marketplace PO → ERP Sales Order Import",
            font=("Arial", 9), fg="gray"
        ).pack(pady=(0, 10))

        # ── Marketplace selector + Margin input ──
        mkt_frame = tk.Frame(self.root)
        mkt_frame.pack(fill='x', padx=20, pady=(0, 8))

        tk.Label(mkt_frame, text="Marketplace:", font=("Arial", 10, "bold")).pack(side='left')
        self.marketplace_var = tk.StringVar(value=MARKETPLACE_NAMES[0] if MARKETPLACE_NAMES else '')
        self.marketplace_dropdown = ttk.Combobox(
            mkt_frame, textvariable=self.marketplace_var,
            values=MARKETPLACE_NAMES, state='readonly', width=20
        )
        self.marketplace_dropdown.pack(side='left', padx=8)
        self.marketplace_dropdown.bind('<<ComboboxSelected>>', self._on_marketplace_change)

        # Margin % input — user can override per run
        tk.Label(mkt_frame, text="Margin:", font=("Arial", 10, "bold")).pack(side='left', padx=(12, 0))
        self.margin_var = tk.StringVar(value=str(self._get_default_margin()))
        self.margin_entry = tk.Entry(mkt_frame, textvariable=self.margin_var, width=5,
                                      font=("Arial", 10), justify='center')
        self.margin_entry.pack(side='left', padx=4)
        tk.Label(mkt_frame, text="%", font=("Arial", 10)).pack(side='left')
        tk.Label(mkt_frame, text="(Landing Cost)", font=("Arial", 8), fg="gray").pack(side='left', padx=4)

        # ── File selectors ──
        files_frame = tk.LabelFrame(self.root, text="Input Files", font=("Arial", 10, "bold"),
                                     padx=10, pady=8)
        files_frame.pack(fill='x', padx=20, pady=(0, 8))

        # Items Master
        # v1.3.2: ts_var holds the "Updated: <timestamp>" sub-line. Empty
        # string means the sub-line is invisible (renders as a blank label)
        # which is fine for the case where no update has been recorded yet.
        self.master_var = tk.StringVar(value="Not selected")
        self.master_ts_var = tk.StringVar(value="")
        self._file_row(files_frame, "Items Master:", self.master_var,
                        self._select_master, ts_var=self.master_ts_var)

        # Mapping file
        self.mapping_var = tk.StringVar(value="Not selected")
        self.mapping_ts_var = tk.StringVar(value="")
        self._file_row(files_frame, "Ship-To Mapping:", self.mapping_var,
                        self._select_mapping, ts_var=self.mapping_ts_var)

        # PO file — no timestamp sub-line (it's a per-run input, not bundled)
        self.po_var = tk.StringVar(value="Not selected")
        self._file_row(files_frame, "Marketplace PO:", self.po_var, self._select_po)

        # ── Buttons ──
        btn_frame = tk.Frame(self.root)
        btn_frame.pack(pady=8)

        tk.Button(btn_frame, text="▶  Generate SO", width=20, font=("Arial", 10, "bold"),
                  bg="#00C853", fg="white", command=self.generate).pack(pady=4)

        self.open_btn = tk.Button(btn_frame, text="📂  Open Last Output", width=20,
                                   state=tk.DISABLED, command=self.open_last)
        self.open_btn.pack(pady=4)

        tk.Button(btn_frame, text="📋  Download PO Template", width=20,
                  command=self._download_template).pack(pady=4)

        # v1.3: Update bundled master/mapping in-place — copies a fresh file
        # into 'Calculation Data/' so subsequent runs auto-load the update.
        tk.Button(btn_frame, text="📁  Update Bundled Files", width=20,
                  command=self._update_bundled_files).pack(pady=4)

        # ── Status ──
        self.status_var = tk.StringVar(value="Status: Waiting — select files and generate")
        self.status_label = tk.Label(
            self.root, textvariable=self.status_var,
            font=("Arial", 10), fg="gray", wraplength=460
        )
        self.status_label.pack(pady=6)

        # ── Log ──
        log_frame = tk.LabelFrame(self.root, text="Log", font=("Arial", 9))
        log_frame.pack(fill='both', expand=True, padx=20, pady=(0, 12))

        scroll = ttk.Scrollbar(log_frame, orient='vertical')
        scroll.pack(side='right', fill='y')

        self.log_text = tk.Text(log_frame, height=6, font=("Consolas", 9),
                                 state='disabled', wrap='word',
                                 yscrollcommand=scroll.set)
        self.log_text.pack(fill='both', expand=True)
        scroll.config(command=self.log_text.yview)

    def _file_row(self, parent, label, var, command, ts_var=None):
        """
        Create a file selector row: label + filename + browse button.

        v1.3.2: If `ts_var` is provided, also renders a small grey sub-line
        below showing the in-app last-updated timestamp, e.g.:

            Items Master:    ✓ Items March.xlsx              [Browse]
                              Updated: 19-Apr-2026 18:41

        The sub-line is bound to a StringVar so refreshes (after auto-load
        or "Update Bundled Files") update the visible text immediately.
        For files where update tracking doesn't apply (the marketplace PO),
        ts_var is omitted and no sub-line is drawn.
        """
        # Container holds both the main row and the optional sub-line, so
        # they stay vertically grouped.
        container = tk.Frame(parent)
        container.pack(fill='x', pady=3)

        row = tk.Frame(container)
        row.pack(fill='x')

        tk.Label(row, text=label, font=("Arial", 9), width=16, anchor='w').pack(side='left')
        tk.Label(row, textvariable=var, font=("Arial", 9), fg="blue",
                 width=28, anchor='w').pack(side='left', padx=4)
        tk.Button(row, text="Browse", width=8, command=command).pack(side='right')

        # Sub-line: aligned under the filename column (16 + padx ≈ 20 chars
        # of leading whitespace via a left-padded sub-frame so it visually
        # hangs under the file name).
        if ts_var is not None:
            sub_row = tk.Frame(container)
            sub_row.pack(fill='x', anchor='w')
            # Indent matches the label width (16) so the timestamp sits
            # under the filename, not under the "Items Master:" label.
            tk.Label(sub_row, text="", width=16).pack(side='left')
            tk.Label(sub_row, textvariable=ts_var, font=("Arial", 8),
                      fg="#777777", anchor='w').pack(side='left', padx=4)

    def _log(self, msg: str):
        """Append message to the log panel."""
        self.log_text.config(state='normal')
        ts = time.strftime("%H:%M:%S")
        self.log_text.insert('end', f"[{ts}] {msg}\n")
        self.log_text.see('end')
        self.log_text.config(state='disabled')

    def _get_default_margin(self) -> int:
        """Get the default margin % for the currently selected marketplace."""
        mkt = self.marketplace_var.get() if hasattr(self, 'marketplace_var') else ''
        if mkt and mkt in MARKETPLACE_CONFIGS:
            return MARKETPLACE_CONFIGS[mkt].get('default_margin', 70)
        return 70

    def _on_marketplace_change(self, event=None):
        """Update margin input when marketplace selection changes."""
        margin = self._get_default_margin()
        self.margin_var.set(str(margin))
        self._log(f"Marketplace changed to {self.marketplace_var.get()}, margin set to {margin}%")

    def _get_margin(self) -> float:
        """
        Get the margin % from the input field.
        Returns as decimal (e.g., 70 → 0.70).
        Falls back to default if input is invalid.
        """
        try:
            val = float(self.margin_var.get().strip())
            if val <= 0 or val > 100:
                raise ValueError
            return val / 100.0
        except (ValueError, AttributeError):
            default = self._get_default_margin()
            self._log(f"Invalid margin input, using default {default}%")
            return default / 100.0

    # ── BUNDLED FILES (v1.3) ──────────────────────────────────────────────────
    #
    # The Items Master and Ship-To Mapping live in 'Calculation Data/' next to
    # the script. On startup we look for them and pre-populate the picker rows
    # so the user doesn't have to re-select the same files every run.
    #
    # When the user edits the master or mapping and wants future runs to use
    # the new version, they click "Update Bundled Files" — which copies their
    # picked file into Calculation Data/, replacing what was there.

    def _auto_load_bundled_files(self):
        """
        Look for Items Master + Ship-To Mapping in Calculation Data/ and
        pre-populate the picker labels if found.

        Called once from __init__ after the UI is built. Does not abort
        startup if files are missing — just logs a hint and leaves the
        pickers in their default "Not selected" state.
        """
        master_p = get_bundled_master_path()
        mapping_p = get_bundled_mapping_path()

        # Master
        if master_p:
            self.master_path = str(master_p)
            self.master_is_bundled = True
            self.master_var.set(f"✓ {master_p.name} (auto-loaded)")
            # v1.3.2: surface the in-app last-updated timestamp
            self._refresh_ts_label(self.master_ts_var, master_p.name)
            self._log(f"Auto-loaded master from {BUNDLED_DATA_FOLDER}/{master_p.name}")
        else:
            self._log(f"No bundled master at {BUNDLED_DATA_FOLDER}/{BUNDLED_MASTER_NAME} "
                      f"— pick one manually or use 'Update Bundled Files'")

        # Mapping
        if mapping_p:
            self.mapping_path = str(mapping_p)
            self.mapping_is_bundled = True
            self.mapping_var.set(f"✓ {mapping_p.name} (auto-loaded)")
            # v1.3.2: surface the in-app last-updated timestamp
            self._refresh_ts_label(self.mapping_ts_var, mapping_p.name)
            self._log(f"Auto-loaded mapping from {BUNDLED_DATA_FOLDER}/{mapping_p.name}")
        else:
            self._log(f"No bundled mapping at {BUNDLED_DATA_FOLDER}/{BUNDLED_MAPPING_NAME} "
                      f"— pick one manually or use 'Update Bundled Files'")

    def _update_bundled_files(self):
        """
        Replace the bundled master and/or mapping in 'Calculation Data/'.

        Workflow:
          1. Ask user which file to update (master / mapping / both).
          2. For each chosen type, open a file picker and copy the chosen
             file into Calculation Data/<canonical_name>.xlsx (overwriting).
          3. Refresh the in-memory paths and picker labels so the rest of
             the session uses the new file.
          4. Log what happened for audit.

        The folder is created if it doesn't exist (first-time setup).
        """
        # Step 1: which file to update?
        choice_dialog = tk.Toplevel(self.root)
        choice_dialog.title("Update Bundled Files")
        choice_dialog.geometry("380x200")
        choice_dialog.resizable(False, False)
        choice_dialog.transient(self.root)
        choice_dialog.grab_set()

        tk.Label(choice_dialog, text="Which file do you want to update?",
                 font=("Arial", 11, "bold")).pack(pady=(15, 8))
        tk.Label(choice_dialog, text=f"Files will be copied into:\n{get_bundled_data_folder()}",
                 font=("Arial", 9), fg="gray", justify='center').pack(pady=(0, 10))

        choice = {'value': None}

        def pick(kind):
            choice['value'] = kind
            choice_dialog.destroy()

        btn_row = tk.Frame(choice_dialog)
        btn_row.pack(pady=4)
        tk.Button(btn_row, text="Items Master", width=14,
                  command=lambda: pick('master')).pack(side='left', padx=4)
        tk.Button(btn_row, text="Ship-To Mapping", width=14,
                  command=lambda: pick('mapping')).pack(side='left', padx=4)

        tk.Button(choice_dialog, text="Both", width=30,
                  command=lambda: pick('both')).pack(pady=4)
        tk.Button(choice_dialog, text="Cancel", width=30,
                  command=choice_dialog.destroy).pack(pady=4)

        # Wait for the user to make a choice (or close the dialog)
        self.root.wait_window(choice_dialog)

        if not choice['value']:
            return  # user cancelled

        # Step 2-4: do the actual updates
        target_folder = get_bundled_data_folder(create=True)
        updated_any = False

        if choice['value'] in ('master', 'both'):
            updated_any |= self._do_update_one_bundled(
                kind_label='Items Master',
                source_title='Select new Items Master file to bundle',
                target_path=target_folder / BUNDLED_MASTER_NAME,
                on_success=self._refresh_master_after_update,
            )

        if choice['value'] in ('mapping', 'both'):
            updated_any |= self._do_update_one_bundled(
                kind_label='Ship-To Mapping',
                source_title='Select new Ship-To Mapping file to bundle',
                target_path=target_folder / BUNDLED_MAPPING_NAME,
                on_success=self._refresh_mapping_after_update,
            )

        if updated_any:
            messagebox.showinfo(
                "Bundled Files Updated",
                f"Bundled files updated in:\n{target_folder}\n\n"
                f"Future runs will auto-load the new version."
            )

    def _do_update_one_bundled(self, kind_label: str, source_title: str,
                                target_path: Path, on_success) -> bool:
        """
        Helper: prompt user to pick a source file and copy it to target_path.

        Args:
            kind_label:   Display label e.g. 'Items Master' (for log/dialog text)
            source_title: Title of the file-picker dialog
            target_path:  Where to copy to (e.g. Calculation Data/Items March.xlsx)
            on_success:   Callback to refresh the GUI after a successful copy

        Returns: True if a copy was performed, False if user cancelled.
        """
        src = filedialog.askopenfilename(
            title=source_title,
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]
        )
        if not src:
            self._log(f"Update cancelled for {kind_label}")
            return False

        try:
            shutil.copy2(src, str(target_path))
            # v1.3.2: stamp the in-app update history BEFORE refreshing the GUI
            # so the refresh callback picks up the new timestamp.
            record_update(target_path.name)
            self._log(f"Bundled {kind_label} updated → {target_path}")
            on_success()
            return True
        except Exception as e:
            self._log(f"ERROR copying {kind_label}: {e}")
            messagebox.showerror("Update Failed",
                                  f"Could not copy {kind_label}:\n{e}")
            return False

    def _refresh_ts_label(self, ts_var: tk.StringVar, filename: str) -> None:
        """
        v1.3.2: Update a timestamp StringVar based on the in-app update history.

        Reads the JSON sidecar for `filename` and sets ts_var to either
        "Updated: <date>" or empty string (which renders as nothing).
        """
        ts = get_update_timestamp(filename)
        ts_var.set(f"Updated: {ts}" if ts else "")

    def _refresh_master_after_update(self):
        """Re-point the in-memory master path to the freshly bundled file."""
        p = get_bundled_master_path()
        if p:
            self.master_path = str(p)
            self.master_is_bundled = True
            self.master_var.set(f"✓ {p.name} (auto-loaded)")
            # v1.3.2: refresh the timestamp sub-line so user sees their
            # update reflected immediately.
            self._refresh_ts_label(self.master_ts_var, p.name)

    def _refresh_mapping_after_update(self):
        """Re-point the in-memory mapping path to the freshly bundled file."""
        p = get_bundled_mapping_path()
        if p:
            self.mapping_path = str(p)
            self.mapping_is_bundled = True
            self.mapping_var.set(f"✓ {p.name} (auto-loaded)")
            # v1.3.2: refresh the timestamp sub-line
            self._refresh_ts_label(self.mapping_ts_var, p.name)

    # ── FILE SELECTION ─────────────────────────────────────────────────────────

    def _select_master(self):
        """
        Manually pick an Items Master file.

        v1.3: Marks the master as user-picked (not bundled) so the GUI shows
        the actual filename rather than "(auto-loaded)". The bundled file in
        Calculation Data/ is not touched — use "Update Bundled Files" for that.
        """
        path = filedialog.askopenfilename(
            title="Select Items Master file",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]
        )
        if path:
            self.master_path = path
            self.master_is_bundled = False
            self.master_var.set(os.path.basename(path))
            # v1.3.2: clear the bundled-file timestamp — manual picks aren't
            # in our update history.
            self.master_ts_var.set("")
            self._log(f"Master (manual override): {os.path.basename(path)}")

    def _select_mapping(self):
        """
        Manually pick a Ship-To B2B mapping file.

        v1.3: Marks the mapping as user-picked (not bundled). Bundled file
        in Calculation Data/ is not touched.
        """
        path = filedialog.askopenfilename(
            title="Select Mapping File (Ship-To B2B)",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]
        )
        if path:
            self.mapping_path = path
            self.mapping_is_bundled = False
            self.mapping_var.set(os.path.basename(path))
            # v1.3.2: clear the bundled-file timestamp
            self.mapping_ts_var.set("")
            self._log(f"Mapping (manual override): {os.path.basename(path)}")

    def _select_po(self):
        """Select the marketplace PO/punch file."""
        path = filedialog.askopenfilename(
            title="Select Marketplace PO File",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]
        )
        if path:
            self.po_path = path
            self.po_var.set(os.path.basename(path))
            self._log(f"PO file: {os.path.basename(path)}")

    # ── PROCESSING ─────────────────────────────────────────────────────────────

    def generate(self):
        """Main processing: load mapping → parse PO → generate output."""

        # ── Validate inputs ──
        marketplace = self.marketplace_var.get()
        if not marketplace or marketplace not in MARKETPLACE_CONFIGS:
            messagebox.showwarning("No Marketplace", "Please select a marketplace.")
            return

        if not self.mapping_path:
            messagebox.showwarning("No Mapping", "Please select the Ship-To mapping file.")
            return

        if not self.po_path:
            messagebox.showwarning("No PO File", "Please select the marketplace PO file.")
            return

        config = MARKETPLACE_CONFIGS[marketplace]
        margin_pct = self._get_margin()
        start_time = time.time()

        self.status_var.set("Processing...")
        self.status_label.config(fg="blue")
        self.root.update()

        self._log(f"Marketplace: {marketplace} | Margin: {int(margin_pct * 100)}%")

        # ── Load mapping ──
        self._log(f"Loading mapping for '{marketplace}'...")
        warnings: List[Tuple[str, str, str]] = []
        loc_count = self.mapping_loader.load(self.mapping_path, config['party_name'], warnings)

        if loc_count == 0:
            self._log("ERROR: No mapping locations found!")
            for _, _, msg in warnings:
                self._log(f"  {msg}")
            self.status_var.set("Failed — mapping load error")
            self.status_label.config(fg="red")
            return

        self._log(f"Loaded {loc_count} locations for {marketplace}")

        # ── Load Items_March (for price validation) ──
        master_loader = None
        if self.master_path:
            self._log(f"Loading Items_March for validation...")
            master_loader = MasterLoader()
            try:
                item_count = master_loader.load(self.master_path)
                self._log(f"Loaded {item_count:,} items from master")
            except Exception as e:
                self._log(f"WARNING: Master load failed: {e} — skipping validation")
                master_loader = None

        # ── Process PO file ──
        self._log(f"Processing {os.path.basename(self.po_path)}...")
        engine = MarketplaceEngine(self.mapping_loader, master=master_loader)
        result = engine.process(self.po_path, config, margin_pct=margin_pct)
        result.margin_pct = margin_pct

        if not result.rows:
            self._log("ERROR: No valid rows extracted!")
            for _, _, msg in result.warnings:
                self._log(f"  WARNING: {msg}")
            self.status_var.set("Failed — no data extracted")
            self.status_label.config(fg="red")
            return

        # ── Log results ──
        unique_pos = set(r.po_number for r in result.rows)
        total_qty = sum(r.qty for r in result.rows)
        unmapped = sum(1 for r in result.rows if not r.mapped)
        mapped_pos = len([po for po in unique_pos
                          if any(r.mapped for r in result.rows if r.po_number == po)])

        self._log(f"Extracted: {len(result.rows)} items, {len(unique_pos)} PO(s), {total_qty} total qty")
        if result.warnings:
            self._log(f"Warnings: {len(result.warnings)}")
            for po, loc, msg in result.warnings[:5]:
                self._log(f"  [{po}] {msg}")
            if len(result.warnings) > 5:
                self._log(f"  ... and {len(result.warnings) - 5} more (see Warnings sheet)")

        # ── Export ──
        self._log("Writing output...")
        output_path = self.exporter.export(result)

        elapsed = time.time() - start_time

        if output_path:
            self.last_output = output_path
            self.open_btn.config(state=tk.NORMAL)

            status_msg = (f"Done — {len(result.rows)} items, {len(unique_pos)} PO(s), "
                          f"{total_qty} qty | {elapsed:.2f}s")
            if result.warnings:
                status_msg += f" | {len(result.warnings)} warning(s)"
                self.status_label.config(fg="orange")
            else:
                self.status_label.config(fg="darkgreen")

            self.status_var.set(status_msg)
            self._log(f"Saved: {output_path}")

            answer = messagebox.askyesno(
                "SO Generated",
                f"Sales Order generated successfully!\n\n"
                f"Marketplace : {marketplace}\n"
                f"PO(s)       : {len(unique_pos)}\n"
                f"Items       : {len(result.rows)}\n"
                f"Total Qty   : {total_qty}\n"
                f"Warnings    : {len(result.warnings)}\n"
                f"Time        : {elapsed:.2f}s\n\n"
                f"Do you want to open the output file?"
            )
            if answer:
                open_file(output_path)
        else:
            self.status_var.set("Failed — no output generated")
            self.status_label.config(fg="red")

    def open_last(self):
        """Open the last generated output file."""
        if self.last_output and self.last_output.exists():
            open_file(self.last_output)
        else:
            messagebox.showwarning("Not Found", "Output file not found.")

    # ─────────────────────────────────────────────────────────────────────────
    #  PO TEMPLATE DOWNLOAD (v1.1)
    # ─────────────────────────────────────────────────────────────────────────

    def _download_template(self):
        """
        Generate a blank PO template for the currently selected marketplace.

        v1.1 change — color-coded headers so the user knows which columns to
        actually fill in:

            ┌──────────┬──────────────────────────────────────────────────┐
            │  COLOR   │  ROLE                                            │
            ├──────────┼──────────────────────────────────────────────────┤
            │  BLUE    │  REQUIRED — script fails without these columns  │
            │  (#1A237E)│ → po_col, loc_col, item_col, qty_col            │
            ├──────────┼──────────────────────────────────────────────────┤
            │  GREEN   │  VALIDATION — used for price check & master     │
            │  (#1B5E20)│  lookup. Strongly recommended but not fatal.   │
            │           │ → fob_col, ean_col                              │
            ├──────────┼──────────────────────────────────────────────────┤
            │  GREY    │  NOT READ — kept only to mirror the              │
            │  (#9E9E9E)│  marketplace's native file format. Safe to     │
            │           │  leave blank when filling the template.         │
            └──────────┴──────────────────────────────────────────────────┘

        The template includes a 3-row legend below the header explaining each
        color, plus a final orange instruction line.
        """
        marketplace = self.marketplace_var.get()
        if not marketplace or marketplace not in MARKETPLACE_CONFIGS:
            messagebox.showwarning("No Marketplace", "Please select a marketplace first.")
            return

        config = MARKETPLACE_CONFIGS[marketplace]

        # ── Ask user where to save ──
        save_path = filedialog.asksaveasfilename(
            title=f"Save {marketplace} PO Template",
            defaultextension=".xlsx",
            initialfile=f"{marketplace}_PO_Template.xlsx",
            filetypes=[("Excel files", "*.xlsx")]
        )
        if not save_path:
            return

        try:
            wb = Workbook()
            ws = wb.active
            ws.title = f'{marketplace} PO'

            # ── Categorise each column into one of three roles ──
            # v1.3.7: 'item_col' is now optional. For marketplaces using
            # item_resolution='from_ean' (Myntra, RK), there is no item_col;
            # the EAN column takes its place as the required identifier.
            item_resolution = config.get('item_resolution', 'from_column')

            required_cols = {config['po_col'], config['loc_col'],
                             config['qty_col']}
            if item_resolution == 'from_ean':
                # ean_col becomes REQUIRED in this mode (it's how Item No
                # gets resolved). Validation upgrades from green→blue.
                if config.get('ean_col'):
                    required_cols.add(config['ean_col'])
            else:  # 'from_column'
                if config.get('item_col'):
                    required_cols.add(config['item_col'])

            validation_cols = set()
            # ean_col is in validation_cols ONLY when not already required
            if config.get('ean_col') and config['ean_col'] not in required_cols:
                validation_cols.add(config['ean_col'])
            if config.get('fob_col'):
                validation_cols.add(config['fob_col'])

            # ── Build column list ──
            # Use template_headers if defined (full marketplace format),
            # otherwise build a minimal list of just the required + validation cols.
            headers = config.get('template_headers')
            if not headers:
                # Construct a minimal template from the required+validation set,
                # preserving stable order for predictability.
                headers = [config['po_col'], config['loc_col']]
                if item_resolution == 'from_column' and config.get('item_col'):
                    headers.append(config['item_col'])
                headers.append(config['qty_col'])
                if config.get('ean_col') and config['ean_col'] not in headers:
                    headers.append(config['ean_col'])
                if config.get('fob_col') and config['fob_col'] not in headers:
                    headers.append(config['fob_col'])

            # ── Fill / font styles per role ──
            required_fill   = PatternFill('solid', fgColor='1A237E')   # Dark blue
            validation_fill = PatternFill('solid', fgColor='1B5E20')   # Dark green
            unused_fill     = PatternFill('solid', fgColor='9E9E9E')   # Mid grey

            hdr_font_white = Font(bold=True, color='FFFFFF',
                                   name='Aptos Display', size=11)
            hdr_font_dim   = Font(bold=True, color='EEEEEE',
                                   name='Aptos Display', size=11, italic=True)

            # ── Write header row with color coding ──
            for c, h in enumerate(headers, 1):
                cell = ws.cell(row=1, column=c, value=h)

                if h in required_cols:
                    # BLUE — must be filled
                    cell.fill = required_fill
                    cell.font = hdr_font_white
                elif h in validation_cols:
                    # GREEN — used for validation
                    cell.fill = validation_fill
                    cell.font = hdr_font_white
                else:
                    # GREY — script ignores these
                    cell.fill = unused_fill
                    cell.font = hdr_font_dim

                cell.alignment = Alignment(horizontal='center', vertical='center')
                ws.column_dimensions[get_column_letter(c)].width = max(len(h) + 4, 12)

            # ── Legend rows (rows 3, 4, 5) ──
            # Each row: a colored tag in column A + a description merged across cols B–H
            legend_row = 3
            legend_items = [
                ('1A237E', 'FFFFFF', 'REQUIRED',
                 f'Script fails without these — fill them in: '
                 f'{", ".join(sorted(required_cols))}'),
                ('1B5E20', 'FFFFFF', 'VALIDATION',
                 f'Used for price check & master lookup: '
                 f'{", ".join(sorted(validation_cols)) or "(none)"}'),
                ('9E9E9E', 'FFFFFF', 'NOT READ',
                 'Optional — kept only to match marketplace file format; '
                 'can stay blank'),
            ]

            for fg, fc, label, desc in legend_items:
                # Coloured tag
                tag = ws.cell(row=legend_row, column=1, value=label)
                tag.fill = PatternFill('solid', fgColor=fg)
                tag.font = Font(bold=True, color=fc, name='Aptos Display', size=10)
                tag.alignment = Alignment(horizontal='center')

                # Description (italic, dim)
                desc_cell = ws.cell(row=legend_row, column=2, value=desc)
                desc_cell.font = Font(name='Aptos Display', size=10,
                                       color='333333', italic=True)
                # Merge across B..min(H, last_col) for nice readability
                ws.merge_cells(start_row=legend_row, start_column=2,
                                end_row=legend_row, end_column=min(8, len(headers)))
                legend_row += 1

            # ── Final instruction row (orange italic) ──
            ws.cell(row=legend_row + 1, column=1,
                    value=f'← {marketplace} PO template. Fill data rows below the header. '
                          f'Only the BLUE & GREEN columns are read by the script.').font = \
                Font(name='Aptos Display', size=10, color='FF6600', italic=True)

            ws.freeze_panes = 'A2'
            wb.save(save_path)

            self._log(f"{marketplace} template saved → {save_path}")
            messagebox.showinfo(
                "Template Saved",
                f"{marketplace} PO template saved to:\n{save_path}\n\n"
                f"Header colours:\n"
                f"  • Blue  = Required (must fill)\n"
                f"  • Green = Validation (recommended)\n"
                f"  • Grey  = Not read by script"
            )
        except Exception as e:
            self._log(f"Template save failed: {e}")
            messagebox.showerror("Error", f"Failed to save template:\n{e}")

    def run(self):
        """Start the Tkinter main loop."""
        self.root.mainloop()


# ═══════════════════════════════════════════════════════════════════════════════
#  ENTRY POINT
# ═══════════════════════════════════════════════════════════════════════════════

def main():
    check_expiry()
    app = OnlinePOApp()
    app.run()


if __name__ == "__main__":
    main()