"""
GT Mass Dump Generator — Django Web Engine (v2.4 parity)
=========================================================

Ported from the standalone v2.4 Tkinter app for Django web context.
No Tkinter dependency — all popups/dialogs replaced with return values.
Output is returned as BytesIO for HttpResponse attachment.

Features (matching standalone v2.4):
    - TemplateValidator (strict: BC Code + Order Qty + PO Number required)
    - MetadataExtractor (Distributor, City, State, Location, SO Number)
    - Location Code mapping (AHD→PICK, BLR→DS_BL_OFF1)
    - 7-sheet Excel output:
        1. Headers (SO)       — ERP Sales Order headers
        2. Lines (SO)         — ERP Sales Order lines
        3. Sales Lines        — Detailed flat reference
        4. Sales Header       — Grouped summary per SO
        5. SKU Summary        — Demand pivot per BC Code
        6. File → SO Mapping  — ALL files (success/fail/warn)
        7. Warnings           — Red-highlighted critical issues
    - ProcessResult tracks attempted_files for full traceability
"""

from __future__ import annotations

import io
import logging
import re
from dataclasses import dataclass, field
from pathlib import Path
from typing import List, Optional, Tuple, Dict, Any
from datetime import datetime

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

logger = logging.getLogger(__name__)


# ═══════════════════════════════════════════════════════════════════════════════
#  CONSTANTS
# ═══════════════════════════════════════════════════════════════════════════════

LOCATION_CODE_MAP: Dict[str, str] = {
    'AHD': 'PICK',
    'BLR': 'DS_BL_OFF1',
}

STATE_LIKE_VALUES = {
    "up", "mp", "ap", "hp", "uk", "jk", "wb", "tn", "kl", "ka",
    "gj", "rj", "hr", "pb", "br", "od", "as", "mh", "cg", "jh",
    "north", "south", "east", "west", "central",
    "uttar pradesh", "madhya pradesh", "rajasthan", "punjab",
    "maharashtra", "gujarat", "karnataka", "tamil nadu",
    "haryana", "delhi", "u.p", "u.p.", "m.p", "m.p.",
}


# ═══════════════════════════════════════════════════════════════════════════════
#  HELPERS
# ═══════════════════════════════════════════════════════════════════════════════

def safe_str_val(row_vals, idx: Optional[int], as_int_str: bool = False) -> str:
    """Safely extract a string from row_vals[idx]. Returns '' if None/NaN."""
    if idx is None:
        return ''
    val = row_vals[idx]
    if pd.isna(val):
        return ''
    if as_int_str and isinstance(val, (int, float)):
        return str(int(val))
    return str(val).strip()


# ═══════════════════════════════════════════════════════════════════════════════
#  DATA MODEL
# ═══════════════════════════════════════════════════════════════════════════════

@dataclass
class OrderRow:
    """Single ordered item extracted from a GT-Mass file."""
    so_number: str
    item_no: str
    ean: str
    category: str
    description: str
    qty: int
    tester_qty: int
    distributor: str
    city: str
    state: str
    location: str
    location_code: str
    source_file: str


@dataclass
class ProcessResult:
    """
    Aggregated result from processing all uploaded files.

    Fields:
        rows            : All OrderRow objects across all files
        failed_files    : [(filename, reason)] — files that couldn't be parsed
        warned_files    : [(filename, warning)] — non-fatal issues
        attempted_files : ALL filenames in upload order (for File → SO Mapping)
    """
    rows: List[OrderRow] = field(default_factory=list)
    failed_files: List[Tuple[str, str]] = field(default_factory=list)
    warned_files: List[Tuple[str, str]] = field(default_factory=list)
    attempted_files: List[str] = field(default_factory=list)


# ═══════════════════════════════════════════════════════════════════════════════
#  SO NUMBER FORMATTER
# ═══════════════════════════════════════════════════════════════════════════════

class SONumberFormatter:
    """Extracts SO number from filename digits (fallback)."""

    @staticmethod
    def from_filename(filename: str) -> Optional[str]:
        """
        Extract first digit sequence from filename.
        Example: "SOGTM6325.xlsx" → "SO/GTM/6325"
        """
        stem = Path(filename).stem
        match = re.search(r"\d+", stem)

        if not match:
            logger.warning(f"No digits in filename: {filename}")
            return None

        return f"SO/GTM/{match.group()}"


# ═══════════════════════════════════════════════════════════════════════════════
#  FILE READER — reads Django InMemoryUploadedFile into DataFrame
# ═══════════════════════════════════════════════════════════════════════════════

class FileReader:
    """
    Reads uploaded Excel files into raw DataFrames (no header).

    Handles Django's InMemoryUploadedFile by reading .read() bytes
    into a BytesIO wrapper for pandas.
    """

    @staticmethod
    def read(file_obj: Any, filename: str) -> pd.DataFrame:
        """
        Read an uploaded file into a raw DataFrame.

        Args:
            file_obj : Django InMemoryUploadedFile (has .read() and .name)
            filename : Display name for logging

        Returns:
            DataFrame with integer column indices and no header.

        Raises:
            RuntimeError: if file cannot be read.
        """
        try:
            file_obj.seek(0)
            data = file_obj.read()
            buf = io.BytesIO(data)

            ext = Path(filename).suffix.lower()

            if ext in (".xlsx", ".xlsm"):
                df = pd.read_excel(buf, header=None, engine="openpyxl")
            elif ext == ".xls":
                df = pd.read_excel(buf, header=None, engine="xlrd")
            else:
                raise RuntimeError(f"Unsupported format: '{ext}'")

            logger.info(f"{filename} — read ({len(df)} rows)")
            return df

        except RuntimeError:
            raise
        except Exception as e:
            raise RuntimeError(f"Cannot read '{filename}': {e}")


# ═══════════════════════════════════════════════════════════════════════════════
#  TEMPLATE VALIDATOR
# ═══════════════════════════════════════════════════════════════════════════════

class TemplateValidator:
    """
    Validates that an uploaded file's first sheet matches the GT-Mass template.

    Hard rejections (file skipped entirely):
        1. No header row with 'BC Code' AND 'Order Qty'
        2. No 'PO Number' label with a value

    Soft checks (file still processes, warning logged):
        - Missing Location → MetadataExtractor fires ❌ CRITICAL warning
    """

    @staticmethod
    def validate(file_obj: Any, filename: str) -> Tuple[bool, Optional[str]]:
        """
        Run template compliance checks.

        Args:
            file_obj : Django InMemoryUploadedFile
            filename : Display name

        Returns:
            (is_valid, reason) — reason is None if valid.
        """
        try:
            raw_df = FileReader.read(file_obj, filename)
        except RuntimeError as e:
            return False, str(e)

        # Rule 1: Header row with BC Code + Order Qty
        header_row = TemplateValidator._find_header_row(raw_df)

        if header_row is None:
            return False, (
                "Template violation: header row not found. "
                "File must have a row with BOTH 'BC Code' AND 'Order Qty'."
            )

        # Rule 2: PO Number label with value
        meta_df = raw_df.iloc[:header_row]
        po_found, po_has_value = TemplateValidator._check_po_number(meta_df)

        if not po_found:
            return False, (
                "Template violation: missing 'PO Number' label in meta rows."
            )

        if not po_has_value:
            return False, (
                "Template violation: 'PO Number' label exists but value is empty."
            )

        return True, None

    @staticmethod
    def _find_header_row(raw_df: pd.DataFrame) -> Optional[int]:
        """Scan for the row containing both 'BC Code' and 'Order Qty'."""
        for i, row_vals in enumerate(raw_df.values):
            vals = [str(v).lower() for v in row_vals]

            if "bc code" in vals and any("order qty" in v for v in vals):
                return i

        return None

    @staticmethod
    def _check_po_number(meta_df: pd.DataFrame) -> Tuple[bool, bool]:
        """
        Check if 'PO Number' label exists with a non-empty adjacent value.

        Returns:
            (label_found, has_value)
        """
        for _, row in meta_df.iterrows():
            for col_idx in range(min(len(row) - 1, 10)):
                if pd.isna(row.iloc[col_idx]):
                    continue

                if str(row.iloc[col_idx]).strip().lower() == "po number":
                    # Check next 1-2 cells for a value
                    for offset in range(1, 3):
                        check = col_idx + offset

                        if check >= len(row):
                            break

                        val = row.iloc[check]

                        if pd.notna(val) and str(val).strip() and str(val).strip().lower() != 'nan':
                            return True, True

                    return True, False

        return False, False


# ═══════════════════════════════════════════════════════════════════════════════
#  METADATA EXTRACTOR
# ═══════════════════════════════════════════════════════════════════════════════

class MetadataExtractor:
    """Extracts meta fields (SO#, Distributor, City, State, Location) from header rows."""

    @staticmethod
    def extract(raw_df: pd.DataFrame, header_row: int) -> Tuple[dict, List[str]]:
        """
        Scan rows 0..header_row-1 for meta field labels and values.

        Returns:
            (meta_dict, warnings_list)
        """
        distributor = ""
        city = ""
        location = ""
        so_number = ""
        state_values: List[str] = []
        warnings: List[str] = []

        meta_df = raw_df.iloc[:header_row]

        for _, row in meta_df.iterrows():
            # LEFT SIDE: Col A (label) + Col B (value)
            label = ""
            if pd.notna(row.iloc[0]):
                label = str(row.iloc[0]).strip().lower()

            value = ""
            if pd.notna(row.iloc[1]):
                value = str(row.iloc[1]).strip()
                if value.lower() == "nan":
                    value = ""

            if label == "distributor name" and not distributor:
                distributor = value

            elif label == "city" and not city:
                city = value

            elif label == "state":
                state_values.append(value)

            # RIGHT SIDE: scan cols 0-9 for "PO Number" / "Location"
            for col_idx in range(min(len(row) - 1, 10)):
                if pd.isna(row.iloc[col_idx]):
                    continue

                cell_text = str(row.iloc[col_idx]).strip().lower()

                if cell_text == "location":
                    for vi in range(col_idx + 1, min(col_idx + 3, len(row))):
                        lv = row.iloc[vi]
                        if pd.notna(lv) and str(lv).strip() and str(lv).strip().lower() != 'nan':
                            location = str(lv).strip()
                            break

                elif cell_text == "po number" and not so_number:
                    for vi in range(col_idx + 1, min(col_idx + 3, len(row))):
                        pv = row.iloc[vi]
                        if pd.notna(pv) and str(pv).strip() and str(pv).strip().lower() != 'nan':
                            so_number = str(pv).strip()
                            break

        # Resolve state
        state = next((s for s in reversed(state_values) if s), "")

        # Map Location → ERP Location Code
        location_code = ""
        if location:
            location_code = LOCATION_CODE_MAP.get(location.upper().strip(), location)

        # Warnings
        if not distributor:
            warnings.append("Distributor Name is blank.")

        if not city:
            warnings.append("City is blank.")

        if not state:
            warnings.append("State is blank.")

        if not location_code:
            warnings.append(
                "❌ CRITICAL: Location Code is EMPTY — "
                "ERP import will fail without Location Code."
            )

        if distributor and distributor.strip().lower() in STATE_LIKE_VALUES:
            warnings.append(
                f"Distributor '{distributor}' looks like a state — verify."
            )

        return {
            "distributor": distributor,
            "city": city,
            "state": state,
            "location": location,
            "location_code": location_code,
            "so_number": so_number,
        }, warnings


# ═══════════════════════════════════════════════════════════════════════════════
#  EXCEL PARSER
# ═══════════════════════════════════════════════════════════════════════════════

class ExcelParser:
    """Parses a single GT-Mass file into OrderRow objects."""

    BC_COLUMN = "bc code"
    QTY_COLUMN = "order qty"
    TESTER_COLUMN = "tester qty"

    def parse(self, file_obj: Any, filename: str) -> Tuple[List[OrderRow], List[str]]:
        """
        Parse one uploaded file end-to-end.

        Args:
            file_obj : Django InMemoryUploadedFile
            filename : Display name

        Returns:
            (list of OrderRow, list of warnings)

        Raises:
            RuntimeError: if file has broken structure.
        """
        logger.info(f"Parsing: {filename}")

        warnings: List[str] = []

        # Read raw
        raw_df = FileReader.read(file_obj, filename)

        # Find header row
        header_row = self._find_header_row(raw_df)

        # Extract meta
        meta, meta_warnings = MetadataExtractor.extract(raw_df, header_row)
        warnings.extend(meta_warnings)

        # Resolve SO number
        so_number, so_warnings = self._resolve_so_number(meta, filename)
        warnings.extend(so_warnings)

        # Build data table
        df = raw_df.iloc[header_row + 1:].copy()
        df.columns = raw_df.iloc[header_row].values
        df = df.reset_index(drop=True)

        # Extract rows
        rows, extract_warnings = self._extract_rows(df, so_number, meta, filename)
        warnings.extend(extract_warnings)

        return rows, warnings

    def _find_header_row(self, raw_df: pd.DataFrame) -> int:
        """Find the row containing 'BC Code' + 'Order Qty'."""
        for i, row_vals in enumerate(raw_df.values):
            vals = [str(v).lower() for v in row_vals]
            if "bc code" in vals and any("order qty" in v for v in vals):
                return i

        raise RuntimeError("Header row not found — no 'BC Code' + 'Order Qty'.")

    def _resolve_so_number(self, meta: dict, filename: str) -> Tuple[str, List[str]]:
        """Resolve SO: file PO Number → filename digits → UNKNOWN."""
        warnings: List[str] = []
        so = meta.get("so_number", "")

        if so:
            return so, warnings

        so = SONumberFormatter.from_filename(filename)
        if so:
            warnings.append(f"SO from filename: '{so}'. Fill PO Number field.")
            return so, warnings

        warnings.append("SO not found — using 'SO/GTM/UNKNOWN'.")
        return "SO/GTM/UNKNOWN", warnings

    def _extract_rows(
        self, df: pd.DataFrame, so_number: str,
        meta: dict, filename: str
    ) -> Tuple[List[OrderRow], List[str]]:
        """Build OrderRow list from the data table."""
        warnings: List[str] = []

        bc_col, qty_col, tester_col, ean_col, cat_col, desc_col = self._detect_columns(df)

        if bc_col is None:
            raise RuntimeError("'BC Code' column not found.")
        if qty_col is None:
            raise RuntimeError("'Order Qty' column not found.")
        if tester_col is None:
            warnings.append("'Tester Qty' not found — defaulting to 0.")

        bc_idx = df.columns.get_loc(bc_col)
        qty_idx = df.columns.get_loc(qty_col)
        tester_idx = df.columns.get_loc(tester_col) if tester_col else None
        ean_idx = df.columns.get_loc(ean_col) if ean_col else None
        cat_idx = df.columns.get_loc(cat_col) if cat_col else None
        desc_idx = df.columns.get_loc(desc_col) if desc_col else None

        rows: List[OrderRow] = []

        for rv in df.values:
            bc = rv[bc_idx]
            if pd.isna(bc):
                continue
            try:
                bc = int(bc)
            except (ValueError, TypeError):
                continue

            qty = self._clean_qty(rv[qty_idx])
            tqty = self._clean_qty(rv[tester_idx]) if tester_idx is not None else 0

            if qty <= 0 and tqty <= 0:
                continue

            rows.append(OrderRow(
                so_number=so_number,
                item_no=str(bc),
                ean=safe_str_val(rv, ean_idx, as_int_str=True),
                category=safe_str_val(rv, cat_idx),
                description=safe_str_val(rv, desc_idx),
                qty=qty,
                tester_qty=tqty,
                distributor=meta["distributor"],
                city=meta["city"],
                state=meta["state"],
                location=meta["location"],
                location_code=meta["location_code"],
                source_file=filename,
            ))

        if not rows:
            warnings.append("No ordered rows — all quantities are 0.")

        return rows, warnings

    def _detect_columns(self, df) -> Tuple[Optional[str], ...]:
        """Find BC Code, Order Qty, Tester Qty, EAN, Category, Description."""
        bc = qty = tester = ean = cat = desc = None

        for col in df.columns:
            name = str(col).strip().lower()
            if name == self.BC_COLUMN:
                bc = col
            if self.QTY_COLUMN in name:
                qty = col
            if self.TESTER_COLUMN in name:
                tester = col
            if name == 'ean' and not ean:
                ean = col
            if name == 'category' and not cat:
                cat = col
            if 'article description' in name:
                desc = col
            elif name == 'description' and not desc:
                desc = col

        return bc, qty, tester, ean, cat, desc

    @staticmethod
    def _clean_qty(value) -> int:
        """Clean quantity cell: NaN → 0, commas stripped, float→int."""
        if pd.isna(value):
            return 0
        value = str(value).strip()
        if value in ("", "-"):
            return 0
        value = value.replace(",", "")
        try:
            return int(float(value))
        except (ValueError, TypeError):
            return 0


# ═══════════════════════════════════════════════════════════════════════════════
#  DUMP EXPORTER — writes 7-sheet Excel to BytesIO
# ═══════════════════════════════════════════════════════════════════════════════

class DumpExporter:
    """Writes the 7-sheet output Excel to memory (BytesIO) for HTTP response."""

    HEADER_FILL = PatternFill('solid', fgColor='1A237E')
    HEADER_FONT = Font(bold=True, color='FFFFFF', name='Aptos Display', size=11)
    DATA_FONT = Font(name='Aptos Display', size=11)
    THIN_SIDE = Side(style='thin', color='CCCCCC')
    BORDER = Border(left=THIN_SIDE, right=THIN_SIDE, top=THIN_SIDE, bottom=THIN_SIDE)

    def _hdr_cell(self, ws, row, col, value):
        """Create a formatted header cell."""
        cell = ws.cell(row=row, column=col, value=value)
        cell.font = self.HEADER_FONT
        cell.fill = self.HEADER_FILL
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = self.BORDER
        return cell

    def _data_cell(self, ws, row, col, value, fmt=None):
        """Create a formatted data cell."""
        cell = ws.cell(row=row, column=col, value=value)
        cell.font = self.DATA_FONT
        cell.border = self.BORDER
        if fmt:
            cell.number_format = fmt
        return cell

    def _auto_width(self, ws, max_w=50):
        """Auto-fit column widths."""
        for col in ws.columns:
            letter = col[0].column_letter
            w = max((len(str(c.value or '')) for c in col), default=8)
            ws.column_dimensions[letter].width = min(w + 3, max_w)

    def export_to_memory(self, result: ProcessResult) -> Optional[io.BytesIO]:
        """
        Write all 7 sheets to an in-memory Excel file.

        Args:
            result: ProcessResult from the engine

        Returns:
            BytesIO buffer with the Excel file, or None if no data.
        """
        if not result.rows and not result.attempted_files:
            return None

        wb = Workbook()
        wb.remove(wb.active)

        if result.rows:
            self._write_headers_so(wb, result)
            self._write_lines_so(wb, result)
            self._write_sales_lines(wb, result)
            self._write_sales_header(wb, result)
            self._write_sku_summary(wb, result)

        self._write_file_so_mapping(wb, result)
        self._write_warnings(wb, result)

        if not wb.sheetnames:
            wb.create_sheet('Empty')

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        return output

    # ─────────────────────────────────────────────────────────────────
    #  Sheet writers (same logic as standalone v2.4)
    # ─────────────────────────────────────────────────────────────────

    def _write_headers_so(self, wb, result: ProcessResult):
        """Sheet 1: Headers (SO) — one row per unique SO."""
        ws = wb.create_sheet('Headers (SO)')

        headers = [
            'Document Type', 'No.', 'Sell-to Customer No.', 'Ship-to Code',
            'Posting Date', 'Order Date', 'Document Date',
            'Invoice From Date', 'Invoice To Date',
            'External Document No.', 'Location Code', 'Dimension Set ID',
            'Supply Type', 'Voucher Narration',
            'Brand Code (Dimension)', 'Channel Code (Dimension)',
            'Catagory (Dimension)', 'Geography Code (Dimension)',
        ]

        for c, h in enumerate(headers, 1):
            self._hdr_cell(ws, 1, c, h)

        today_str = datetime.now().strftime("%d-%m-%Y")

        seen = set()
        unique_sos = []
        for row in result.rows:
            if row.so_number not in seen:
                seen.add(row.so_number)
                unique_sos.append(row)

        r = 2
        for row in unique_sos:
            self._data_cell(ws, r, 1, 'Order')
            self._data_cell(ws, r, 2, row.so_number)
            self._data_cell(ws, r, 3, '')
            self._data_cell(ws, r, 4, '')

            for c in range(5, 10):
                self._data_cell(ws, r, c, today_str)

            self._data_cell(ws, r, 10, row.so_number)
            self._data_cell(ws, r, 11, row.location_code)
            self._data_cell(ws, r, 12, '')
            self._data_cell(ws, r, 13, 'B2B')
            r += 1

        self._auto_width(ws)

    def _write_lines_so(self, wb, result: ProcessResult):
        """Sheet 2: Lines (SO) — one row per item, 10K line increments."""
        ws = wb.create_sheet('Lines (SO)')

        headers = [
            'Document Type', 'Document No.', 'Line No.', 'Type',
            'No.', 'Location Code', 'Quantity', 'Unit Price',
        ]

        for c, h in enumerate(headers, 1):
            self._hdr_cell(ws, 1, c, h)

        r = 2
        current_so = None
        line_no = 0

        for row in result.rows:
            if row.so_number != current_so:
                current_so = row.so_number
                line_no = 0
            line_no += 10000

            self._data_cell(ws, r, 1, 'Order')
            self._data_cell(ws, r, 2, row.so_number)
            self._data_cell(ws, r, 3, line_no)
            self._data_cell(ws, r, 4, 'Item')
            self._data_cell(ws, r, 5, row.item_no)
            self._data_cell(ws, r, 6, row.location_code)
            self._data_cell(ws, r, 7, row.qty)
            self._data_cell(ws, r, 8, '')
            r += 1

        self._auto_width(ws)

    def _write_sales_lines(self, wb, result: ProcessResult):
        """Sheet 3: Sales Lines — detailed flat list."""
        ws = wb.create_sheet('Sales Lines')

        headers = [
            'SO Number', 'EAN', 'BC Code', 'Category',
            'Article Description', 'Order Qty', 'Tester Qty',
        ]

        for c, h in enumerate(headers, 1):
            self._hdr_cell(ws, 1, c, h)

        for r, row in enumerate(result.rows, 2):
            self._data_cell(ws, r, 1, row.so_number)
            self._data_cell(ws, r, 2, row.ean)
            self._data_cell(ws, r, 3, row.item_no)
            self._data_cell(ws, r, 4, row.category)
            self._data_cell(ws, r, 5, row.description)
            self._data_cell(ws, r, 6, row.qty)
            self._data_cell(ws, r, 7, row.tester_qty)

        self._auto_width(ws)

    def _write_sales_header(self, wb, result: ProcessResult):
        """Sheet 4: Sales Header — grouped summary per SO."""
        ws = wb.create_sheet('Sales Header')

        headers = [
            'SO Number', 'Order Qty', 'Tester Qty', 'Total Qty',
            'Distributor', 'City', 'State', 'Location',
        ]

        for c, h in enumerate(headers, 1):
            self._hdr_cell(ws, 1, c, h)

        so_groups: Dict[str, dict] = {}
        for row in result.rows:
            if row.so_number not in so_groups:
                so_groups[row.so_number] = {
                    'oq': 0, 'tq': 0,
                    'd': row.distributor, 'c': row.city,
                    's': row.state, 'l': row.location,
                }
            so_groups[row.so_number]['oq'] += row.qty
            so_groups[row.so_number]['tq'] += row.tester_qty

        r = 2
        for so, info in so_groups.items():
            self._data_cell(ws, r, 1, so)
            self._data_cell(ws, r, 2, info['oq'])
            self._data_cell(ws, r, 3, info['tq'])
            self._data_cell(ws, r, 4, info['oq'] + info['tq'])
            self._data_cell(ws, r, 5, info['d'])
            self._data_cell(ws, r, 6, info['c'])
            self._data_cell(ws, r, 7, info['s'])
            self._data_cell(ws, r, 8, info['l'])
            r += 1

        self._auto_width(ws)

    def _write_sku_summary(self, wb, result: ProcessResult):
        """Sheet 5: SKU Summary — demand pivot per BC Code."""
        ws = wb.create_sheet('SKU Summary')

        headers = ['BC Code', 'Description', 'Category', 'Order Qty', 'Tester Qty', 'Total Qty']
        for c, h in enumerate(headers, 1):
            self._hdr_cell(ws, 1, c, h)

        sku: Dict[str, dict] = {}
        for row in result.rows:
            if row.item_no not in sku:
                sku[row.item_no] = {'d': row.description, 'c': row.category, 'oq': 0, 'tq': 0}
            sku[row.item_no]['oq'] += row.qty
            sku[row.item_no]['tq'] += row.tester_qty
            if not sku[row.item_no]['d'] and row.description:
                sku[row.item_no]['d'] = row.description
            if not sku[row.item_no]['c'] and row.category:
                sku[row.item_no]['c'] = row.category

        sorted_skus = sorted(sku.items(), key=lambda x: x[1]['oq'] + x[1]['tq'], reverse=True)

        r = 2
        go = gt = 0
        for item, info in sorted_skus:
            t = info['oq'] + info['tq']
            go += info['oq']
            gt += info['tq']
            self._data_cell(ws, r, 1, item)
            self._data_cell(ws, r, 2, info['d'])
            self._data_cell(ws, r, 3, info['c'])
            self._data_cell(ws, r, 4, info['oq'])
            self._data_cell(ws, r, 5, info['tq'])
            self._data_cell(ws, r, 6, t)
            r += 1

        bold = Font(name='Aptos Display', size=11, bold=True)
        ws.cell(row=r, column=1, value='GRAND TOTAL').font = bold
        ws.cell(row=r, column=2, value=f'{len(sorted_skus)} unique SKUs').font = bold
        ws.cell(row=r, column=4, value=go).font = bold
        ws.cell(row=r, column=5, value=gt).font = bold
        ws.cell(row=r, column=6, value=go + gt).font = bold
        for c in range(1, 7):
            ws.cell(row=r, column=c).border = self.BORDER

        self._auto_width(ws)

    def _write_file_so_mapping(self, wb, result: ProcessResult):
        """
        Sheet 6: File → SO Mapping — every uploaded file gets an entry.

        Status:
            ✅ OK       — parsed cleanly
            ⚠️ WARNING  — parsed with warnings
            ❌ FAILED   — rejected at validation or parse stage
        """
        ws = wb.create_sheet('File → SO Mapping')

        for c, h in enumerate(['Sr No', 'Filename', 'SO Number', 'Status'], 1):
            self._hdr_cell(ws, 1, c, h)

        file_to_so = {}
        for row in result.rows:
            if row.source_file not in file_to_so:
                file_to_so[row.source_file] = row.so_number

        failed_map = {f: r for f, r in result.failed_files}
        warned_set = {f for f, _ in result.warned_files}

        red_fill = PatternFill('solid', fgColor='FFCDD2')
        red_font = Font(name='Aptos Display', size=11, bold=True, color='D32F2F')
        yellow_fill = PatternFill('solid', fgColor='FFF9C4')

        sr = 1
        ok_count = warn_count = fail_count = 0

        for filename in result.attempted_files:
            r = sr + 1

            if filename in failed_map:
                self._data_cell(ws, r, 1, sr)
                self._data_cell(ws, r, 2, filename)
                self._data_cell(ws, r, 3, f"❌ FAILED: {failed_map[filename]}")
                self._data_cell(ws, r, 4, '❌ FAILED')
                for c in range(1, 5):
                    ws.cell(row=r, column=c).fill = red_fill
                    ws.cell(row=r, column=c).font = red_font
                fail_count += 1

            elif filename in file_to_so:
                so = file_to_so[filename]

                if filename in warned_set:
                    self._data_cell(ws, r, 1, sr)
                    self._data_cell(ws, r, 2, filename)
                    self._data_cell(ws, r, 3, so)
                    self._data_cell(ws, r, 4, '⚠️ WARNING')
                    for c in range(1, 5):
                        ws.cell(row=r, column=c).fill = yellow_fill
                    warn_count += 1
                else:
                    self._data_cell(ws, r, 1, sr)
                    self._data_cell(ws, r, 2, filename)
                    self._data_cell(ws, r, 3, so)
                    self._data_cell(ws, r, 4, '✅ OK')
                    ok_count += 1
            else:
                self._data_cell(ws, r, 1, sr)
                self._data_cell(ws, r, 2, filename)
                self._data_cell(ws, r, 3, '(no data)')
                self._data_cell(ws, r, 4, '❌ FAILED')
                for c in range(1, 5):
                    ws.cell(row=r, column=c).fill = red_fill
                    ws.cell(row=r, column=c).font = red_font
                fail_count += 1

            sr += 1

        # Summary row
        summary_r = sr + 1
        bold = Font(name='Aptos Display', size=11, bold=True)
        ws.cell(row=summary_r, column=1, value='TOTAL').font = bold
        ws.cell(row=summary_r, column=2,
                value=f'{len(result.attempted_files)} file(s) attempted').font = bold
        ws.cell(row=summary_r, column=3,
                value=f'✅ {ok_count} OK | ⚠️ {warn_count} warn | ❌ {fail_count} failed').font = bold
        for c in range(1, 5):
            ws.cell(row=summary_r, column=c).border = self.BORDER

        self._auto_width(ws)

    def _write_warnings(self, wb, result: ProcessResult):
        """Sheet 7: Warnings — failures first (red), then warnings."""
        if not result.warned_files and not result.failed_files:
            return

        ws = wb.create_sheet('Warnings')
        for c, h in enumerate(['File', 'Type', 'Message'], 1):
            self._hdr_cell(ws, 1, c, h)

        red_fill = PatternFill('solid', fgColor='FFCDD2')
        red_font = Font(name='Aptos Display', size=11, bold=True, color='D32F2F')

        r = 2

        for fname, reason in result.failed_files:
            self._data_cell(ws, r, 1, fname)
            self._data_cell(ws, r, 2, '❌ FAILED')
            self._data_cell(ws, r, 3, reason)
            for c in range(1, 4):
                ws.cell(row=r, column=c).fill = red_fill
                ws.cell(row=r, column=c).font = red_font
            r += 1

        for fname, warning in result.warned_files:
            is_critical = '❌ CRITICAL' in warning
            self._data_cell(ws, r, 1, fname)
            self._data_cell(ws, r, 2, '❌ CRITICAL' if is_critical else '⚠️ WARNING')
            self._data_cell(ws, r, 3, warning)
            if is_critical:
                for c in range(1, 4):
                    ws.cell(row=r, column=c).fill = red_fill
                    ws.cell(row=r, column=c).font = red_font
            r += 1

        self._auto_width(ws)


# ═══════════════════════════════════════════════════════════════════════════════
#  MAIN AUTOMATION ENGINE
# ═══════════════════════════════════════════════════════════════════════════════

class GTMassAutomation:
    """
    Orchestrates validation, parsing, and export for Django web context.

    Replaces the simple process_files → List[OrderRow] with a full
    ProcessResult that tracks attempted files, failures, and warnings.
    """

    def __init__(self):
        self.validator = TemplateValidator()
        self.parser = ExcelParser()
        self.exporter = DumpExporter()

    def process_files(self, file_objects: List[Any]) -> ProcessResult:
        """
        Process all uploaded files.

        Args:
            file_objects: List of Django InMemoryUploadedFile objects

        Returns:
            ProcessResult with rows, warnings, failures, attempted_files
        """
        result = ProcessResult()

        for file_obj in file_objects:
            fname = file_obj.name
            result.attempted_files.append(fname)

            # Validate template compliance
            is_valid, reason = self.validator.validate(file_obj, fname)

            if not is_valid:
                result.failed_files.append((fname, reason))
                logger.error(f"{fname} REJECTED: {reason}")
                continue

            # Reset file pointer after validation read it
            file_obj.seek(0)

            # Parse
            try:
                rows, warnings = self.parser.parse(file_obj, fname)
                result.rows.extend(rows)

                for w in warnings:
                    result.warned_files.append((fname, w))
                    logger.warning(f"{fname}: {w}")

            except RuntimeError as e:
                result.failed_files.append((fname, str(e)))
                logger.error(f"{fname} FAILED: {e}")

            except (ValueError, KeyError, TypeError) as e:
                result.failed_files.append((fname, f"Data error: {e}"))
                logger.error(f"{fname} DATA: {e}")

            except Exception as e:
                result.failed_files.append((fname, f"Unexpected: {e}"))
                logger.error(f"{fname} UNEXPECTED: {e}")

        logger.info(
            f"Done — {len(result.attempted_files)} attempted | "
            f"{len(result.rows)} rows | "
            f"{len({r.so_number for r in result.rows})} SOs | "
            f"{len(result.failed_files)} failed | "
            f"{len(result.warned_files)} warnings"
        )

        return result