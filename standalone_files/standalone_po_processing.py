"""
╔═══════════════════════════════════════════════════════════════════════════════╗
║               ONLINE PO PROCESSOR — Marketplace SO Generator                 ║
║               Tkinter GUI Desktop Application                                ║
╠═══════════════════════════════════════════════════════════════════════════════╣
║  Author  : Agami AI / Vishal                                                ║
║  Version : 1.0                                                               ║
║  Purpose : Reads marketplace punch/PO files (Myntra, Bigbasket, Blink etc), ║
║            maps locations to Ship-to codes via a mapping registry,           ║
║            generates ERP-importable Headers (SO) + Lines (SO) sheets.        ║
║  Stack   : Python 3.13, Tkinter, pandas, openpyxl                           ║
╚═══════════════════════════════════════════════════════════════════════════════╝

═══════════════════════════════════════════════════════════════════════════════
  ARCHITECTURE
═══════════════════════════════════════════════════════════════════════════════

  ┌─────────────────────────────────────────────────────────────────┐
  │                    OnlinePOApp (GUI)                            │
  │  Select Marketplace → Load Files → Generate → Open Output      │
  │                                                                 │
  │  3 file inputs:                                                 │
  │    1. Items_March.xlsx    (Item Master — shared)                │
  │    2. Online_Mapping.xlsx (Ship-To B2B — location registry)    │
  │    3. Marketplace PO file (e.g., Myntra_Punch.xlsx)            │
  └────────────────────┬────────────────────────────────────────────┘
                       │
                       ▼
  ┌─────────────────────────────────────────────────────────────────┐
  │              MarketplaceEngine                                   │
  │  1. Load mapping (filter by selected marketplace)               │
  │  2. Parse PO file (marketplace-specific column config)          │
  │  3. Match Location → Cust No + Ship-to from mapping            │
  │  4. Generate output rows                                        │
  └────────────────────┬────────────────────────────────────────────┘
                       │
                       ▼
  ┌─────────────────────────────────────────────────────────────────┐
  │              SOExporter (Output)                                 │
  │  Sheet 1: Headers (SO)  → ERP SO header import                 │
  │  Sheet 2: Lines (SO)    → ERP SO line import (10K increments)  │
  │  Sheet 3: Summary       → Per-PO breakdown with mapping info   │
  │  Sheet 4: Warnings      → Unmapped locations, missing data     │
  └─────────────────────────────────────────────────────────────────┘

═══════════════════════════════════════════════════════════════════════════════
  INPUT FILES
═══════════════════════════════════════════════════════════════════════════════

  1. ITEMS MASTER (Items_March.xlsx)
     ───────────────────────────────
     Same master used by EKA Script and GT Mass Dump.
     Required columns: No., GTIN, Description, GST Group Code, Mrp
     Used here only for validation — Item No comes from the PO file directly.

  2. MAPPING FILE (Online_Mapping.xlsx, sheet: 'Ship-To B2B')
     ─────────────────────────────────────────────────────────
     Registry of all marketplace delivery locations.
     Required columns:
         Party          → Marketplace name (Myntra, Bigbasket, Blink, etc.)
         Del Location   → Delivery location name as used by the marketplace
         Cust No        → ERP Sell-to Customer Number
         Ship to        → ERP Ship-to Code (e.g., '20011_4')

  3. MARKETPLACE PO FILE (e.g., Myntra_Punch_09-04-2026.xlsx)
     ─────────────────────────────────────────────────────────
     Purchase order / punch file from the marketplace.
     Column mapping varies per marketplace (see MARKETPLACE_CONFIGS).

═══════════════════════════════════════════════════════════════════════════════
  MARKETPLACE CONFIGURATIONS
═══════════════════════════════════════════════════════════════════════════════

  Each marketplace has a different PO file format. The config defines:
      party_name  → Name to filter in mapping file (e.g., 'Myntra')
      po_col      → Column containing PO/SO number
      loc_col     → Column containing delivery location
      item_col    → Column containing Item No (BC Code / Item no)
      qty_col     → Column containing order quantity
      price_col   → Column containing unit price (None = leave empty)

  To add a new marketplace:
      1. Add entry to MARKETPLACE_CONFIGS dict
      2. Add name to MARKETPLACE_NAMES list
      3. The mapping sheet must have the party's locations

═══════════════════════════════════════════════════════════════════════════════
  OUTPUT — 4 EXCEL SHEETS
═══════════════════════════════════════════════════════════════════════════════

  Sheet 1: 'Headers (SO)' — One row per unique PO number
      Document Type | No. | Sell-to Customer No. | Ship-to Code |
      5 × date fields | External Document No. | Location Code = PICK |
      Supply Type = B2B | ...dimension columns (empty)

  Sheet 2: 'Lines (SO)' — One row per ordered item
      Document Type | Document No. | Line No. (10K increments) | Type = Item |
      No. | Location Code = PICK | Quantity | Unit Price (empty)

  Sheet 3: 'Summary' — Per-PO grouped info
      PO | Location | Cust No | Ship-to | Items | Total Qty | Status

  Sheet 4: 'Warnings' — Only if issues found
      PO | Location | Warning

═══════════════════════════════════════════════════════════════════════════════
  HOW TO ADD A NEW MARKETPLACE
═══════════════════════════════════════════════════════════════════════════════

  1. Get a sample PO file from the marketplace
  2. Identify which columns have: PO number, Location, Item No, Quantity
  3. Add config to MARKETPLACE_CONFIGS:
         'NewMarket': {
             'party_name': 'NewMarket',     # Must match 'Party' in mapping
             'po_col': 'PO Number',          # Column name for PO/SO number
             'loc_col': 'Warehouse',          # Column name for location
             'item_col': 'Item Code',         # Column name for item number
             'qty_col': 'Qty',                # Column name for quantity
             'price_col': None,               # None = leave empty in output
         }
  4. Add 'NewMarket' to MARKETPLACE_NAMES list
  5. Add locations to the mapping sheet (Ship-To B2B)

Requirements:
    pip install pandas openpyxl

Run:
    python online_po_processor.py
"""

# ═══════════════════════════════════════════════════════════════════════════════
#  IMPORTS
# ═══════════════════════════════════════════════════════════════════════════════

from __future__ import annotations
import os
import sys
import platform
import time
import logging
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
from dataclasses import dataclass, field
from pathlib import Path
from typing import List, Optional, Tuple, Dict, Any
from datetime import datetime

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ═══════════════════════════════════════════════════════════════════════════════
#  LOGGING
# ═══════════════════════════════════════════════════════════════════════════════

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s | %(levelname)s | %(message)s"
)


# ═══════════════════════════════════════════════════════════════════════════════
#  EXPIRY CHECK
# ═══════════════════════════════════════════════════════════════════════════════

EXPIRY_DATE = "30-06-2026"


def check_expiry():
    """Check if the application has expired. Shows popup and exits if so."""
    expiry = datetime.strptime(EXPIRY_DATE, "%d-%m-%Y").date()
    today = datetime.now().date()

    if today > expiry:
        root = tk.Tk()
        root.withdraw()
        messagebox.showerror(
            "Application Expired",
            f"This application expired on {EXPIRY_DATE}.\n\n"
            f"Please contact the administrator for an updated version."
        )
        root.destroy()
        sys.exit(0)

    days_remaining = (expiry - today).days
    if days_remaining <= 7:
        root = tk.Tk()
        root.withdraw()
        messagebox.showwarning(
            "Expiration Warning",
            f"⚠️ This application will expire in {days_remaining} day(s).\n\n"
            f"Expiry Date: {EXPIRY_DATE}\n"
            f"Please contact the administrator for renewal."
        )
        root.destroy()


# ═══════════════════════════════════════════════════════════════════════════════
#  MARKETPLACE CONFIGURATIONS
# ═══════════════════════════════════════════════════════════════════════════════
# Each marketplace has a different PO file column layout.
# party_name must match the 'Party' column in the mapping sheet exactly.
# price_col = None means Unit Price is left empty in output.

MARKETPLACE_CONFIGS: Dict[str, Dict[str, Any]] = {
    'Myntra': {
        'party_name': 'Myntra',            # Must match 'Party' in mapping sheet
        'po_col': 'PO',                    # Column: PO/SO number
        'loc_col': 'Location',              # Column: delivery location
        'item_col': 'Item no',              # Column: Item No (already resolved)
        'qty_col': 'Quantity',              # Column: order quantity
        'price_col': None,                  # None = leave Unit Price empty (WMS handles)
        'fob_col': 'List price(FOB+Transport-Excise)',  # Marketplace price to validate against
        'ean_col': 'Vendor Article Number', # Column with EAN/GTIN for master lookup
        'default_margin': 70,              # Default margin % for landing cost calculation
        'template_headers': ['PO', 'Location', 'SKU Id', 'Style Id', 'SKU Code',
                             'HSN Code', 'Brand', 'GTIN', 'Vendor Article Number',
                             'Vendor Article Name', 'Size', 'Colour', 'Mrp',
                             'Credit Period', 'Margin Type', 'Agreed Margin',
                             'Gross Margin', 'Quantity', 'FOB Amount',
                             'List price(FOB+Transport-Excise)', 'Landing Price',
                             'Estimated Delivery Date'],
    },
    'RK': {
        'party_name': 'RK',               # Must match 'Party' in mapping sheet
        'po_col': 'PO',                    # Column: PO/SO number (alphanumeric like '2EH63D1K')
        'loc_col': 'Ship-to location',      # Column: delivery location (codes like 'ISK3', 'BLR4')
        'item_col': 'Item no',              # Column: Item No (already resolved)
        'qty_col': 'Accepted quantity',      # Column: accepted order quantity
        'price_col': None,                  # None = leave Unit Price empty (WMS handles)
        'fob_col': 'Cost',                  # Marketplace price to validate against
        'ean_col': 'External ID',           # Column with EAN/GTIN for master lookup
        'default_margin': 70,              # Default margin % for landing cost calculation
        'template_headers': ['PO', 'Order date', 'Status', 'Product name', 'External ID',
                             'Model number', 'Accepted quantity', 'Ship-to location',
                             'Cost', 'Total accepted cost'],
    },
    # ┌─────────────────────────────────────────────────────────────────────┐
    # │ ADD NEW MARKETPLACES HERE                                           │
    # │                                                                     │
    # │ 'Bigbasket': {                                                      │
    # │     'party_name': 'Bigbasket',                                      │
    # │     'po_col': 'PO Number',                                          │
    # │     'loc_col': 'Delivery Location',                                 │
    # │     'item_col': 'Item Code',                                        │
    # │     'qty_col': 'Qty',                                               │
    # │     'fob_col': 'Unit Price',                                        │
    # │     'ean_col': 'EAN',                                               │
    # │     'price_col': None,                                              │
    # │     'default_margin': 60,                                           │
    # │ },                                                                  │
    # └─────────────────────────────────────────────────────────────────────┘
}

# List of marketplace names for the dropdown (order matters for display)
MARKETPLACE_NAMES: List[str] = list(MARKETPLACE_CONFIGS.keys())


# ═══════════════════════════════════════════════════════════════════════════════
#  DATA MODEL
# ═══════════════════════════════════════════════════════════════════════════════

@dataclass
class SORow:
    """Single line item for SO generation."""
    po_number: str           # PO/SO number from marketplace (e.g., 'MYNJ-RNEE230326-1')
    location: str            # Delivery location from file (e.g., 'Bilaspur')
    item_no: Any             # Item No (e.g., 200074)
    qty: int                 # Order quantity
    unit_price: Optional[float] = None  # Unit price (None = leave empty)
    cust_no: str = ''        # Sell-to Customer No. from mapping
    ship_to: str = ''        # Ship-to Code from mapping
    mapped: bool = False     # Whether location was found in mapping
    ean: str = ''            # EAN/GTIN from punch file (for master lookup)
    fob_price: Optional[float] = None   # Marketplace FOB/Cost price from file
    calc_price: Optional[float] = None  # Our calculated cost price (MRP × margin% ÷ GST)
    diffn: Optional[float] = None       # fob_price - calc_price (0 = OK, non-zero = flag)
    mrp: Optional[float] = None         # MRP from Items_March
    gst_code: str = ''       # GST Code from Items_March
    validation_status: str = ''  # 'OK', 'MISMATCH', 'NOT_IN_MASTER'


@dataclass
class ProcessingResult:
    """Result from processing a marketplace PO file."""
    rows: List[SORow] = field(default_factory=list)
    warnings: List[Tuple[str, str, str]] = field(default_factory=list)  # (po, location, message)
    marketplace: str = ''
    input_file: str = ''
    margin_pct: float = 0.70  # Margin % as decimal (0.70 = 70%)
    raw_df: Any = None        # Original marketplace DataFrame for reference sheet


# ═══════════════════════════════════════════════════════════════════════════════
#  MAPPING LOADER
# ═══════════════════════════════════════════════════════════════════════════════

class MappingLoader:
    """
    Loads the Ship-To B2B mapping from the mapping Excel file.

    The mapping file has a sheet named 'Ship-To B2B' with columns:
        Party | Del Location | Cust No | Ship to

    Filtering by party_name gives us only the relevant marketplace locations.
    """

    def __init__(self):
        self.mappings: Dict[str, Dict[str, str]] = {}  # location → {cust_no, ship_to}
        self.party_name: str = ''
        self.total_loaded: int = 0

    def load(self, filepath: str, party_name: str, logs: List[Tuple[str, str, str]]) -> int:
        """
        Load mapping for a specific marketplace.

        Args:
            filepath: Path to mapping Excel file
            party_name: Marketplace name to filter by (e.g., 'Myntra')
            logs: Warning accumulator

        Returns: number of locations loaded for this marketplace
        """
        self.party_name = party_name
        self.mappings = {}

        try:
            df = pd.read_excel(filepath, sheet_name='Ship-To B2B', header=0)
        except ValueError:
            # Sheet not found — try first sheet
            logging.warning("Sheet 'Ship-To B2B' not found, trying first sheet")
            df = pd.read_excel(filepath, header=0)
        except Exception as e:
            logs.append(('', '', f"Cannot read mapping file: {e}"))
            return 0

        # ── Detect required columns ──
        col_map = {}
        for col in df.columns:
            cl = str(col).strip().lower()
            if cl == 'party':
                col_map['party'] = col
            elif cl in ('del location', 'delivery location', 'location'):
                col_map['location'] = col
            elif cl in ('cust no', 'cust no.', 'customer no', 'sell-to'):
                col_map['cust_no'] = col
            elif cl in ('ship to', 'ship-to', 'ship to code'):
                col_map['ship_to'] = col

        missing = [k for k in ('party', 'location', 'cust_no', 'ship_to') if k not in col_map]
        if missing:
            logs.append(('', '', f"Mapping file missing columns: {', '.join(missing)}. "
                                f"Available: {list(df.columns)}"))
            return 0

        # ── Filter by party and build lookup ──
        for _, row in df.iterrows():
            party = str(row[col_map['party']]).strip()
            if party.lower() != party_name.lower():
                continue

            location = str(row[col_map['location']]).strip()
            cust_no = str(row[col_map['cust_no']]).strip() if pd.notna(row[col_map['cust_no']]) else ''
            ship_to = str(row[col_map['ship_to']]).strip() if pd.notna(row[col_map['ship_to']]) else ''

            # Clean up cust_no — remove '.0' from float conversion
            if cust_no.endswith('.0'):
                cust_no = cust_no[:-2]

            if location and location.lower() != 'nan':
                self.mappings[location] = {
                    'cust_no': cust_no,
                    'ship_to': ship_to,
                }

        self.total_loaded = len(self.mappings)
        logging.info(f"Mapping: Loaded {self.total_loaded} locations for '{party_name}'")
        return self.total_loaded

    def lookup(self, location: str) -> Optional[Dict[str, str]]:
        """
        Look up a location in the mapping.

        First tries exact match, then case-insensitive match,
        then checks if any mapping key contains the location or vice versa.

        Returns: {cust_no, ship_to} or None if not found
        """
        if not location:
            return None

        loc_clean = location.strip()

        # 1. Exact match
        if loc_clean in self.mappings:
            return self.mappings[loc_clean]

        # 2. Case-insensitive match
        loc_lower = loc_clean.lower()
        for key, val in self.mappings.items():
            if key.lower() == loc_lower:
                return val

        # 3. Contains match (location in key or key in location)
        for key, val in self.mappings.items():
            key_lower = key.lower()
            if loc_lower in key_lower or key_lower in loc_lower:
                logging.info(f"Mapping: Fuzzy match '{loc_clean}' → '{key}'")
                return val

        return None


# ═══════════════════════════════════════════════════════════════════════════════
#  MARKETPLACE ENGINE
# ═══════════════════════════════════════════════════════════════════════════════

# ═══════════════════════════════════════════════════════════════════════════════
#  MASTER LOADER
# ═══════════════════════════════════════════════════════════════════════════════

class MasterLoader:
    """
    Loads Items_March.xlsx for price validation.
    Indexes by both GTIN (EAN) and No. (item code) for flexible lookup.
    """

    def __init__(self):
        self.master: Dict[str, Dict] = {}  # EAN/ItemNo → {mrp, gst_code, description}

    def load(self, filepath: str) -> int:
        """Load Items_March and build lookup. Returns item count."""
        df = pd.read_excel(filepath, header=0)
        df['GTIN_str'] = df['GTIN'].astype(str).str.strip()
        self.master = {}

        for _, r in df.iterrows():
            desc = str(r.get('Description', '')) if pd.notna(r.get('Description')) else ''
            gst = str(r['GST Group Code']) if pd.notna(r.get('GST Group Code')) else ''
            mrp = r.get('Mrp')
            item_no = str(r['No.']).strip()

            entry = {'item_no': item_no, 'mrp': mrp, 'gst_code': gst, 'description': desc}

            # Index by GTIN
            self.master[r['GTIN_str']] = entry
            # Also index by item code
            if item_no not in self.master:
                self.master[item_no] = entry

        return len(df)

    @staticmethod
    def calc_cost_price(mrp, gst_code: str, margin_pct: float) -> Optional[float]:
        """
        Calculate cost price: MRP × margin% ÷ GST divisor.

        Args:
            mrp: Maximum Retail Price
            gst_code: Tax code from Items_March
            margin_pct: Margin as decimal (e.g., 0.70 for 70%)

        GST codes in Items_March and their divisors:
            0-G      (9 items)    → 0% GST  → ÷ 1.00
            G-3      (1 item)     → 3% GST  → ÷ 1.03
            G-5      (1084 items) → 5% GST  → ÷ 1.05
            G-5-S    (108 items)  → 5% GST  → ÷ 1.05
            G-12     (67 items)   → 12% GST → ÷ 1.12
            G-18     (2022 items) → 18% GST → ÷ 1.18
            G-18-S   (1364 items) → 18% GST → ÷ 1.18

        Returns: calculated cost price, or None if MRP is missing
        """
        if mrp is None or pd.isna(mrp):
            return None
        landing = float(mrp) * margin_pct
        gst = str(gst_code).strip().upper()

        # 0% GST (0-G, G-0, empty)
        if gst in ('0-G', 'G-0', 'G-0-S', '0', '') or gst == 'NAN':
            return landing
        # 3% GST (G-3)
        if gst in ('G-3', 'G-3-S'):
            return landing / 1.03
        # 5% GST (G-5, G-5-S)
        if '5' in gst and '18' not in gst and '12' not in gst:
            return landing / 1.05
        # 12% GST (G-12, G-12-S)
        if '12' in gst:
            return landing / 1.12
        # 18% GST (G-18, G-18-S)
        if '18' in gst:
            return landing / 1.18
        # Unknown — default to 18%
        return landing / 1.18

    def lookup(self, key: str) -> Optional[Dict]:
        """Look up by EAN or Item No. Returns {mrp, gst_code, ...} or None."""
        key_clean = str(key).strip()
        # Try direct
        if key_clean in self.master:
            return self.master[key_clean]
        # Try stripping leading zeros (EAN sometimes has leading 0)
        stripped = key_clean.lstrip('0')
        if stripped in self.master:
            return self.master[stripped]
        return None


# ═══════════════════════════════════════════════════════════════════════════════
#  MARKETPLACE ENGINE
# ═══════════════════════════════════════════════════════════════════════════════

class MarketplaceEngine:
    """
    Processes a marketplace PO file using the appropriate column config,
    mapping data, and Items_March for price validation.
    """

    # Threshold for flagging price mismatches (in rupees)
    DIFFN_THRESHOLD = 1.0

    def __init__(self, mapping: MappingLoader, master: Optional[MasterLoader] = None):
        self.mapping = mapping
        self.master = master

    def process(self, filepath: str, config: Dict[str, Any],
                margin_pct: float = 0.70) -> ProcessingResult:
        """
        Process a marketplace PO file.

        Args:
            filepath: Path to the PO/punch Excel file
            config: Marketplace column configuration from MARKETPLACE_CONFIGS

        Returns: ProcessingResult with rows and warnings
        """
        result = ProcessingResult(
            marketplace=config['party_name'],
            input_file=os.path.basename(filepath),
        )

        # ── Read file ──
        try:
            df = pd.read_excel(filepath, header=0)
        except Exception as e:
            result.warnings.append(('', '', f"Cannot read file: {e}"))
            return result

        logging.info(f"Read {len(df)} rows from {os.path.basename(filepath)}")

        # Store raw data for reference sheet in output
        result.raw_df = df

        # ── Validate required columns ──
        required_cols = {
            'po': config['po_col'],
            'loc': config['loc_col'],
            'item': config['item_col'],
            'qty': config['qty_col'],
        }

        for key, col_name in required_cols.items():
            if col_name not in df.columns:
                result.warnings.append(('', '', f"Required column '{col_name}' not found. "
                                               f"Available: {list(df.columns)[:15]}..."))
                return result

        price_col = config.get('price_col')
        if price_col and price_col not in df.columns:
            logging.warning(f"Price column '{price_col}' not found — will leave empty")
            price_col = None

        # ── Validate comparison column for price validation ──
        fob_col_name = config.get('fob_col')
        fob_col_valid = fob_col_name and fob_col_name in df.columns
        if fob_col_name and not fob_col_valid:
            result.warnings.append(('', '',
                f"Validation column '{fob_col_name}' not found in file — "
                f"price validation will be skipped. Available: {list(df.columns)[:10]}..."))
            logging.warning(f"FOB/Cost column '{fob_col_name}' not found — validation disabled")

        ean_col_name = config.get('ean_col')
        ean_col_valid = ean_col_name and ean_col_name in df.columns
        if ean_col_name and not ean_col_valid:
            result.warnings.append(('', '',
                f"EAN column '{ean_col_name}' not found in file — "
                f"master lookup will use Item No only. Available: {list(df.columns)[:10]}..."))
            logging.warning(f"EAN column '{ean_col_name}' not found — using Item No for lookup")

        # ── Process each row ──
        warned_locations = set()  # Track (po, location) to avoid duplicate warnings
        for _, row in df.iterrows():
            po = str(row[config['po_col']]).strip()
            location = str(row[config['loc_col']]).strip() if pd.notna(row[config['loc_col']]) else ''
            item_raw = row[config['item_col']]
            qty_raw = row[config['qty_col']]

            # Skip empty rows
            if pd.isna(item_raw) or po.lower() == 'nan':
                continue

            # Parse item number
            try:
                item_no = int(item_raw)
            except (ValueError, TypeError):
                item_no = str(item_raw).strip()

            # Parse quantity
            try:
                qty = int(float(qty_raw)) if pd.notna(qty_raw) else 0
            except (ValueError, TypeError):
                qty = 0

            if qty <= 0:
                continue

            # Parse unit price (if configured)
            unit_price = None
            if price_col:
                try:
                    p = row[price_col]
                    unit_price = float(p) if pd.notna(p) else None
                except (ValueError, TypeError):
                    unit_price = None

            # ── Extract EAN for master lookup ──
            ean = ''
            ean_col = config.get('ean_col')
            if ean_col and ean_col in df.columns:
                ean_raw = row[ean_col]
                if pd.notna(ean_raw):
                    ean = str(int(ean_raw)) if isinstance(ean_raw, (int, float)) else str(ean_raw).strip()

            # ── Extract marketplace FOB/Cost price ──
            fob_price = None
            fob_col = config.get('fob_col')
            if fob_col and fob_col in df.columns:
                try:
                    fob_raw = row[fob_col]
                    fob_price = float(fob_raw) if pd.notna(fob_raw) else None
                except (ValueError, TypeError):
                    fob_price = None

            # ── Validate price against Items_March ──
            calc_price = None
            mrp = None
            gst_code = ''
            diffn = None
            validation_status = ''

            if self.master:
                # Try lookup by EAN first, then by Item No
                master_info = self.master.lookup(ean) if ean else None
                if not master_info:
                    master_info = self.master.lookup(str(item_no))

                if master_info:
                    mrp = master_info['mrp']
                    gst_code = master_info['gst_code']

                    # Warn if GST code is not in known set
                    known_gst = {'0-G', 'G-3', 'G-3-S', 'G-5', 'G-5-S', 'G-12', 'G-12-S', 'G-18', 'G-18-S', ''}
                    gst_upper = str(gst_code).strip().upper()
                    if gst_upper not in known_gst and gst_upper != 'NAN':
                        warn_key = ('GST', gst_upper)
                        if warn_key not in warned_locations:
                            warned_locations.add(warn_key)
                            result.warnings.append((
                                po, str(item_no),
                                f"Unknown GST code '{gst_code}' for Item {item_no} — "
                                f"defaulting to 18%. Please verify in Items_March."
                            ))
                            logging.warning(f"Unknown GST code '{gst_code}' for Item {item_no}")

                    calc_price = MasterLoader.calc_cost_price(mrp, gst_code, margin_pct)

                    if calc_price is not None and fob_price is not None:
                        diffn = fob_price - calc_price
                        if abs(diffn) <= self.DIFFN_THRESHOLD:
                            validation_status = 'OK'
                        else:
                            validation_status = 'MISMATCH'
                            warn_key = ('VALIDATION', str(item_no))
                            if warn_key not in warned_locations:
                                warned_locations.add(warn_key)
                                result.warnings.append((
                                    po, str(item_no),
                                    f"Price mismatch: Item {item_no}, "
                                    f"Marketplace={fob_price:.2f}, "
                                    f"Calculated={calc_price:.2f}, "
                                    f"Diff={diffn:.2f}"
                                ))
                    else:
                        validation_status = 'NO_PRICE'
                else:
                    validation_status = 'NOT_IN_MASTER'

            # ── Lookup location in mapping ──
            mapping_result = self.mapping.lookup(location)

            if mapping_result:
                cust_no = mapping_result['cust_no']
                ship_to = mapping_result['ship_to']
                mapped = True
            else:
                cust_no = ''
                ship_to = ''
                mapped = False
                warn_key = (po, location)
                if warn_key not in warned_locations:
                    warned_locations.add(warn_key)
                    result.warnings.append((
                        po, location,
                        f"Location '{location}' not found in mapping for {config['party_name']}. "
                        f"Cust No and Ship-to left empty."
                    ))

            result.rows.append(SORow(
                po_number=po,
                location=location,
                item_no=item_no,
                qty=qty,
                unit_price=unit_price,
                cust_no=cust_no,
                ship_to=ship_to,
                mapped=mapped,
                ean=ean,
                fob_price=fob_price,
                calc_price=calc_price,
                diffn=diffn,
                mrp=mrp,
                gst_code=gst_code,
                validation_status=validation_status,
            ))

        logging.info(f"Processed {len(result.rows)} items across "
                     f"{len(set(r.po_number for r in result.rows))} PO(s)")
        return result


# ═══════════════════════════════════════════════════════════════════════════════
#  SO EXPORTER
# ═══════════════════════════════════════════════════════════════════════════════

class SOExporter:
    """Writes the output Excel with Headers (SO), Lines (SO), Summary, Warnings."""

    # ── Formatting constants ──
    HEADER_FILL = PatternFill('solid', fgColor='1A237E')
    HEADER_FONT = Font(bold=True, color='FFFFFF', name='Aptos Display', size=11)
    DATA_FONT = Font(name='Aptos Display', size=11)
    THIN_SIDE = Side(style='thin', color='CCCCCC')
    BORDER = Border(left=THIN_SIDE, right=THIN_SIDE, top=THIN_SIDE, bottom=THIN_SIDE)

    # Warning header — orange
    WARN_FILL = PatternFill('solid', fgColor='E65100')

    def _hdr_cell(self, ws, row, col, value, fill=None):
        """Create a formatted header cell."""
        cell = ws.cell(row=row, column=col, value=value)
        cell.font = self.HEADER_FONT
        cell.fill = fill or self.HEADER_FILL
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = self.BORDER
        return cell

    def _data_cell(self, ws, row, col, value, fmt=None):
        """Create a formatted data cell."""
        cell = ws.cell(row=row, column=col, value=value)
        cell.font = self.DATA_FONT
        cell.border = self.BORDER
        if fmt:
            cell.number_format = fmt
        return cell

    def _auto_width(self, ws, max_w=50):
        """Auto-fit column widths based on content."""
        for col in ws.columns:
            letter = col[0].column_letter
            w = max((len(str(c.value or '')) for c in col), default=8)
            ws.column_dimensions[letter].width = min(w + 3, max_w)

    def export(self, result: ProcessingResult) -> Optional[Path]:
        """
        Write output Excel.
        Returns output path on success, None if no data.
        """
        if not result.rows:
            messagebox.showwarning("No Data", "No valid rows found.\nNothing to export.")
            return None

        # ── Prepare output path ──
        output_folder = Path("output_online")
        output_folder.mkdir(exist_ok=True)
        today = datetime.now().strftime("%d-%m-%Y_%H%M%S")
        marketplace = result.marketplace.lower().replace(' ', '_')
        file_path = output_folder / f"{marketplace}_so_{today}.xlsx"

        # ── Create workbook ──
        wb = Workbook()
        wb.remove(wb.active)

        self._write_headers_so(wb, result)
        self._write_lines_so(wb, result)
        self._write_summary(wb, result)
        self._write_validation(wb, result)
        self._write_warnings(wb, result)
        self._write_raw_data(wb, result)

        wb.save(str(file_path))
        logging.info(f"Output saved: {file_path}")
        return file_path

    def _write_headers_so(self, wb, result: ProcessingResult):
        """
        Sheet 1: 'Headers (SO)' — One row per unique PO number.
        """
        ws = wb.create_sheet('Headers (SO)')
        headers = [
            'Document Type', 'No.', 'Sell-to Customer No.', 'Ship-to Code',
            'Posting Date', 'Order Date', 'Document Date',
            'Invoice From Date', 'Invoice To Date',
            'External Document No.', 'Location Code', 'Dimension Set ID',
            'Supply Type', 'Voucher Narration',
            'Brand Code (Dimension)', 'Channel Code (Dimension)',
            'Catagory (Dimension)', 'Geography Code (Dimension)'
        ]
        for c, h in enumerate(headers, 1):
            self._hdr_cell(ws, 1, c, h)

        today_str = datetime.now().strftime("%d-%m-%Y")

        # Collect unique POs preserving order
        seen = set()
        unique_pos = []
        for row in result.rows:
            if row.po_number not in seen:
                seen.add(row.po_number)
                unique_pos.append(row)

        r = 2
        for row in unique_pos:
            self._data_cell(ws, r, 1, 'Order')              # Document Type
            self._data_cell(ws, r, 2, row.po_number)         # No.
            self._data_cell(ws, r, 3, row.cust_no)           # Sell-to Customer No.
            self._data_cell(ws, r, 4, row.ship_to)           # Ship-to Code
            self._data_cell(ws, r, 5, today_str)             # Posting Date
            self._data_cell(ws, r, 6, today_str)             # Order Date
            self._data_cell(ws, r, 7, today_str)             # Document Date
            self._data_cell(ws, r, 8, today_str)             # Invoice From Date
            self._data_cell(ws, r, 9, today_str)             # Invoice To Date
            self._data_cell(ws, r, 10, row.po_number)        # External Document No.
            self._data_cell(ws, r, 11, 'PICK')               # Location Code (always PICK)
            self._data_cell(ws, r, 12, '')                   # Dimension Set ID
            self._data_cell(ws, r, 13, 'B2B')                # Supply Type
            # Columns 14-18: empty dimension columns
            r += 1

        self._auto_width(ws)

    def _write_lines_so(self, wb, result: ProcessingResult):
        """
        Sheet 2: 'Lines (SO)' — One row per ordered item.
        Line No. increments by 10000, resets per new PO.
        """
        ws = wb.create_sheet('Lines (SO)')
        headers = ['Document Type', 'Document No.', 'Line No.', 'Type',
                   'No.', 'Location Code', 'Quantity', 'Unit Price']
        for c, h in enumerate(headers, 1):
            self._hdr_cell(ws, 1, c, h)

        r = 2
        current_po = None
        line_no = 0

        for row in result.rows:
            # Reset line counter on new PO
            if row.po_number != current_po:
                current_po = row.po_number
                line_no = 0

            line_no += 10000

            self._data_cell(ws, r, 1, 'Order')              # Document Type
            self._data_cell(ws, r, 2, row.po_number)         # Document No.
            self._data_cell(ws, r, 3, line_no)               # Line No.
            self._data_cell(ws, r, 4, 'Item')                # Type
            self._data_cell(ws, r, 5, row.item_no)           # No. (Item No)
            self._data_cell(ws, r, 6, 'PICK')                # Location Code
            self._data_cell(ws, r, 7, row.qty)               # Quantity
            self._data_cell(ws, r, 8, '')                    # Unit Price (empty)
            r += 1

        self._auto_width(ws)

    def _write_summary(self, wb, result: ProcessingResult):
        """
        Sheet 3: 'Summary' — Per-PO grouped info for verification.
        """
        ws = wb.create_sheet('Summary')
        headers = ['PO', 'Location', 'Cust No', 'Ship-to', 'Items', 'Total Qty', 'Status']
        for c, h in enumerate(headers, 1):
            self._hdr_cell(ws, 1, c, h)

        # Group by PO
        po_groups: Dict[str, dict] = {}
        for row in result.rows:
            if row.po_number not in po_groups:
                po_groups[row.po_number] = {
                    'location': row.location,
                    'cust_no': row.cust_no,
                    'ship_to': row.ship_to,
                    'items': 0,
                    'qty': 0,
                    'mapped': row.mapped,
                }
            po_groups[row.po_number]['items'] += 1
            po_groups[row.po_number]['qty'] += row.qty

        r = 2
        for po, info in po_groups.items():
            status = 'OK' if info['mapped'] else 'UNMAPPED'

            self._data_cell(ws, r, 1, po)
            self._data_cell(ws, r, 2, info['location'])
            self._data_cell(ws, r, 3, info['cust_no'])
            self._data_cell(ws, r, 4, info['ship_to'])
            self._data_cell(ws, r, 5, info['items'])
            self._data_cell(ws, r, 6, info['qty'])
            self._data_cell(ws, r, 7, status)

            # Color status cell
            status_cell = ws.cell(row=r, column=7)
            if status == 'OK':
                status_cell.fill = PatternFill('solid', fgColor='00C853')
                status_cell.font = Font(name='Aptos Display', size=11, bold=True, color='000000')
            else:
                status_cell.fill = PatternFill('solid', fgColor='FF5252')
                status_cell.font = Font(name='Aptos Display', size=11, bold=True, color='FFFFFF')

            r += 1

        # Totals row
        total_items = sum(g['items'] for g in po_groups.values())
        total_qty = sum(g['qty'] for g in po_groups.values())
        self._data_cell(ws, r, 1, 'TOTAL')
        ws.cell(row=r, column=1).font = Font(name='Aptos Display', size=11, bold=True)
        self._data_cell(ws, r, 5, total_items)
        ws.cell(row=r, column=5).font = Font(name='Aptos Display', size=11, bold=True)
        self._data_cell(ws, r, 6, total_qty)
        ws.cell(row=r, column=6).font = Font(name='Aptos Display', size=11, bold=True)

        # Info row — marketplace and margin
        r += 2
        info_font = Font(name='Aptos Display', size=10, italic=True, color='666666')
        margin_str = f"{int(result.margin_pct * 100)}%"
        ws.cell(row=r, column=1, value=f"Marketplace: {result.marketplace}  |  "
                                        f"Margin: {margin_str}  |  "
                                        f"File: {result.input_file}  |  "
                                        f"Generated: {datetime.now().strftime('%d-%m-%Y %H:%M')}").font = info_font

        self._auto_width(ws)

    def _write_validation(self, wb, result: ProcessingResult):
        """
        Sheet: 'Validation' — Price check: marketplace price vs calculated cost price.

        Shows per-item: Item No, EAN, MRP, Landing Cost, GST Code, Cost Price,
        Marketplace Price, Difference, Status (OK / MISMATCH / NOT_IN_MASTER)

        Items with MISMATCH are highlighted in red for easy identification.
        """
        ws = wb.create_sheet('Validation')

        headers = ['PO', 'Item No', 'EAN', 'MRP', f'Landing ({int(result.margin_pct*100)}%)',
                   'GST Code', 'Our Cost Price', 'Marketplace Price', 'Diffn', 'Status']

        # Green header for calc columns, normal for others
        calc_fill = PatternFill('solid', fgColor='1B5E20')
        for c, h in enumerate(headers, 1):
            fill = calc_fill if c in (4, 5, 6, 7) else self.HEADER_FILL
            self._hdr_cell(ws, 1, c, h, fill=fill)

        # Color fills for status
        ok_fill = PatternFill('solid', fgColor='E8F5E9')
        mismatch_fill = PatternFill('solid', fgColor='FFEBEE')
        no_master_fill = PatternFill('solid', fgColor='FFF3E0')

        r = 2
        mismatches = 0
        for row in result.rows:
            self._data_cell(ws, r, 1, row.po_number)
            self._data_cell(ws, r, 2, row.item_no)
            self._data_cell(ws, r, 3, row.ean)
            self._data_cell(ws, r, 4, row.mrp, '#,##0.00' if row.mrp else None)

            # Landing cost
            landing = float(row.mrp) * result.margin_pct if row.mrp and not pd.isna(row.mrp) else None
            self._data_cell(ws, r, 5, round(landing, 2) if landing else '', '#,##0.00')

            self._data_cell(ws, r, 6, row.gst_code)
            self._data_cell(ws, r, 7, round(row.calc_price, 2) if row.calc_price else '', '#,##0.00')
            self._data_cell(ws, r, 8, round(row.fob_price, 2) if row.fob_price else '', '#,##0.00')
            self._data_cell(ws, r, 9, round(row.diffn, 4) if row.diffn is not None else '', '#,##0.0000')
            self._data_cell(ws, r, 10, row.validation_status)

            # Row highlighting based on status
            if row.validation_status == 'MISMATCH':
                mismatches += 1
                for c in range(1, 11):
                    ws.cell(row=r, column=c).fill = mismatch_fill
                ws.cell(row=r, column=10).font = Font(name='Aptos Display', size=11,
                                                       bold=True, color='D32F2F')
            elif row.validation_status == 'OK':
                ws.cell(row=r, column=10).fill = PatternFill('solid', fgColor='00C853')
                ws.cell(row=r, column=10).font = Font(name='Aptos Display', size=11,
                                                       bold=True, color='000000')
            elif row.validation_status == 'NOT_IN_MASTER':
                for c in range(1, 11):
                    ws.cell(row=r, column=c).fill = no_master_fill
                ws.cell(row=r, column=10).font = Font(name='Aptos Display', size=11,
                                                       bold=True, color='E65100')

            r += 1

        # Summary row
        r += 1
        total = len(result.rows)
        ok_count = sum(1 for row in result.rows if row.validation_status == 'OK')
        ws.cell(row=r, column=1, value=f"Total: {total} items | OK: {ok_count} | "
                                        f"Mismatches: {mismatches} | "
                                        f"Margin: {int(result.margin_pct*100)}%").font = \
            Font(name='Aptos Display', size=10, italic=True, color='666666')

        self._auto_width(ws)
        ws.freeze_panes = 'A2'

    def _write_warnings(self, wb, result: ProcessingResult):
        """
        Sheet 4: 'Warnings' — Only created if warnings exist.
        """
        if not result.warnings:
            return

        ws = wb.create_sheet('Warnings')
        headers = ['PO', 'Location', 'Warning']
        for c, h in enumerate(headers, 1):
            self._hdr_cell(ws, 1, c, h, fill=self.WARN_FILL)

        for r, (po, loc, msg) in enumerate(result.warnings, 2):
            self._data_cell(ws, r, 1, po)
            self._data_cell(ws, r, 2, loc)
            self._data_cell(ws, r, 3, msg)

        self._auto_width(ws)

    def _write_raw_data(self, wb, result: ProcessingResult):
        """
        Sheet: 'Raw Data' — Original marketplace data + calculated validation columns.

        Copies ALL original columns from the input file, then appends calculated
        columns (Item No, MRP, Landing Cost, GST Code, Cost Price, Diffn) from
        the Items_March master lookup — same as the manual verification columns.

        This gives a single-workbook reference: original marketplace data +
        our price validation, without needing to open the source file separately.
        """
        if result.raw_df is None or result.raw_df.empty:
            return

        ws = wb.create_sheet('Raw Data')
        df = result.raw_df

        # ── Calculate columns to append ──
        # Build a lookup from result.rows by (po_number, item_no) → validation data
        validation_lookup: Dict[tuple, SORow] = {}
        for row in result.rows:
            key = (row.po_number, str(row.item_no))
            validation_lookup[key] = row

        # ── Original column headers (dark grey) ──
        raw_hdr_fill = PatternFill('solid', fgColor='37474F')
        orig_col_count = len(df.columns)
        for c, col_name in enumerate(df.columns, 1):
            cell = ws.cell(row=1, column=c, value=str(col_name))
            cell.font = self.HEADER_FONT
            cell.fill = raw_hdr_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = self.BORDER

        # ── Calculated column headers (green) ──
        calc_headers = ['Item No (Master)', 'MRP', f'Landing ({int(result.margin_pct*100)}%)',
                        'GST Code', 'Cost Price', 'Diffn']
        calc_fill = PatternFill('solid', fgColor='1B5E20')
        for i, h in enumerate(calc_headers):
            c = orig_col_count + i + 1
            cell = ws.cell(row=1, column=c, value=h)
            cell.font = self.HEADER_FONT
            cell.fill = calc_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = self.BORDER

        # ── Data rows: original + calculated ──
        config_item_col = None
        config_po_col = None
        for cfg in MARKETPLACE_CONFIGS.values():
            if cfg['party_name'] == result.marketplace:
                config_item_col = cfg.get('item_col')
                config_po_col = cfg.get('po_col')
                break

        calc_bg = PatternFill('solid', fgColor='E8F5E9')  # Light green for calc cells
        mismatch_bg = PatternFill('solid', fgColor='FFCDD2')  # Light red for mismatch

        for r, (_, row) in enumerate(df.iterrows(), 2):
            # ── Write original columns ──
            for c, col_name in enumerate(df.columns, 1):
                val = row[col_name]
                if isinstance(val, pd.Timestamp):
                    val = val.strftime('%d-%m-%Y')
                elif pd.isna(val):
                    val = ''
                self._data_cell(ws, r, c, val)

            # ── Write calculated columns ──
            po_val = str(row[config_po_col]).strip() if config_po_col and config_po_col in df.columns else ''
            item_val = ''
            if config_item_col and config_item_col in df.columns:
                iv = row[config_item_col]
                try:
                    item_val = str(int(iv)) if pd.notna(iv) else ''
                except (ValueError, TypeError):
                    item_val = str(iv).strip() if pd.notna(iv) else ''

            # Find matching validation row
            vrow = validation_lookup.get((po_val, item_val))

            base_c = orig_col_count + 1
            if vrow:
                landing = float(vrow.mrp) * result.margin_pct if vrow.mrp and not pd.isna(vrow.mrp) else None

                self._data_cell(ws, r, base_c, vrow.item_no)                     # Item No (Master)
                self._data_cell(ws, r, base_c + 1, vrow.mrp, '#,##0.00')         # MRP
                self._data_cell(ws, r, base_c + 2,
                                round(landing, 2) if landing else '', '#,##0.00')  # Landing
                self._data_cell(ws, r, base_c + 3, vrow.gst_code)                # GST Code
                self._data_cell(ws, r, base_c + 4,
                                round(vrow.calc_price, 2) if vrow.calc_price else '', '#,##0.00')  # Cost Price
                self._data_cell(ws, r, base_c + 5,
                                round(vrow.diffn, 4) if vrow.diffn is not None else '', '#,##0.0000')  # Diffn

                # Apply background color
                is_mismatch = vrow.validation_status == 'MISMATCH'
                fill = mismatch_bg if is_mismatch else calc_bg
                for i in range(6):
                    ws.cell(row=r, column=base_c + i).fill = fill

                # Bold red Diffn if mismatch
                if is_mismatch:
                    ws.cell(row=r, column=base_c + 5).font = Font(
                        name='Aptos Display', size=11, bold=True, color='D32F2F')
            else:
                # No matching validation row (qty=0 items in original file)
                for i in range(6):
                    self._data_cell(ws, r, base_c + i, '')

        self._auto_width(ws)
        ws.freeze_panes = 'A2'


# ═══════════════════════════════════════════════════════════════════════════════
#  FILE OPENER (cross-platform)
# ═══════════════════════════════════════════════════════════════════════════════

def open_file(file_path: Path):
    """Opens the file using the OS default application."""
    try:
        system = platform.system()
        if system == "Windows":
            os.startfile(str(file_path))
        elif system == "Darwin":
            import subprocess as sp
            sp.Popen(["open", str(file_path)])
        else:
            import subprocess as sp
            sp.Popen(["xdg-open", str(file_path)])
    except Exception as e:
        messagebox.showerror("Open File Error", f"Could not open file:\n{e}")


# ═══════════════════════════════════════════════════════════════════════════════
#  TKINTER GUI
# ═══════════════════════════════════════════════════════════════════════════════

class OnlinePOApp:
    """GUI for Online Marketplace PO → SO generation."""

    def __init__(self):
        self.root = tk.Tk()
        self.root.title("Online PO Processor — Marketplace SO Generator")
        self.root.geometry("520x520")
        self.root.resizable(False, False)

        # ── State ──
        self.master_path: Optional[str] = None
        self.mapping_path: Optional[str] = None
        self.po_path: Optional[str] = None
        self.last_output: Optional[Path] = None

        # ── Engine ──
        self.mapping_loader = MappingLoader()
        self.exporter = SOExporter()

        self._build_ui()

    def _build_ui(self):
        """Build the GUI layout."""

        # ── Title ──
        tk.Label(
            self.root, text="Online PO Processor",
            font=("Arial", 14, "bold")
        ).pack(pady=(12, 2))

        tk.Label(
            self.root, text="Marketplace PO → ERP Sales Order Import",
            font=("Arial", 9), fg="gray"
        ).pack(pady=(0, 10))

        # ── Marketplace selector + Margin input ──
        mkt_frame = tk.Frame(self.root)
        mkt_frame.pack(fill='x', padx=20, pady=(0, 8))

        tk.Label(mkt_frame, text="Marketplace:", font=("Arial", 10, "bold")).pack(side='left')
        self.marketplace_var = tk.StringVar(value=MARKETPLACE_NAMES[0] if MARKETPLACE_NAMES else '')
        self.marketplace_dropdown = ttk.Combobox(
            mkt_frame, textvariable=self.marketplace_var,
            values=MARKETPLACE_NAMES, state='readonly', width=20
        )
        self.marketplace_dropdown.pack(side='left', padx=8)
        self.marketplace_dropdown.bind('<<ComboboxSelected>>', self._on_marketplace_change)

        # Margin % input — user can override per run
        tk.Label(mkt_frame, text="Margin:", font=("Arial", 10, "bold")).pack(side='left', padx=(12, 0))
        self.margin_var = tk.StringVar(value=str(self._get_default_margin()))
        self.margin_entry = tk.Entry(mkt_frame, textvariable=self.margin_var, width=5,
                                      font=("Arial", 10), justify='center')
        self.margin_entry.pack(side='left', padx=4)
        tk.Label(mkt_frame, text="%", font=("Arial", 10)).pack(side='left')
        tk.Label(mkt_frame, text="(Landing Cost)", font=("Arial", 8), fg="gray").pack(side='left', padx=4)

        # ── File selectors ──
        files_frame = tk.LabelFrame(self.root, text="Input Files", font=("Arial", 10, "bold"),
                                     padx=10, pady=8)
        files_frame.pack(fill='x', padx=20, pady=(0, 8))

        # Items Master
        self.master_var = tk.StringVar(value="Not selected")
        self._file_row(files_frame, "Items Master:", self.master_var, self._select_master)

        # Mapping file
        self.mapping_var = tk.StringVar(value="Not selected")
        self._file_row(files_frame, "Ship-To Mapping:", self.mapping_var, self._select_mapping)

        # PO file
        self.po_var = tk.StringVar(value="Not selected")
        self._file_row(files_frame, "Marketplace PO:", self.po_var, self._select_po)

        # ── Buttons ──
        btn_frame = tk.Frame(self.root)
        btn_frame.pack(pady=8)

        tk.Button(btn_frame, text="▶  Generate SO", width=20, font=("Arial", 10, "bold"),
                  bg="#00C853", fg="white", command=self.generate).pack(pady=4)

        self.open_btn = tk.Button(btn_frame, text="📂  Open Last Output", width=20,
                                   state=tk.DISABLED, command=self.open_last)
        self.open_btn.pack(pady=4)

        tk.Button(btn_frame, text="📋  Download PO Template", width=20,
                  command=self._download_template).pack(pady=4)

        # ── Status ──
        self.status_var = tk.StringVar(value="Status: Waiting — select files and generate")
        self.status_label = tk.Label(
            self.root, textvariable=self.status_var,
            font=("Arial", 10), fg="gray", wraplength=460
        )
        self.status_label.pack(pady=6)

        # ── Log ──
        log_frame = tk.LabelFrame(self.root, text="Log", font=("Arial", 9))
        log_frame.pack(fill='both', expand=True, padx=20, pady=(0, 12))

        scroll = ttk.Scrollbar(log_frame, orient='vertical')
        scroll.pack(side='right', fill='y')

        self.log_text = tk.Text(log_frame, height=6, font=("Consolas", 9),
                                 state='disabled', wrap='word',
                                 yscrollcommand=scroll.set)
        self.log_text.pack(fill='both', expand=True)
        scroll.config(command=self.log_text.yview)

    def _file_row(self, parent, label, var, command):
        """Create a file selector row: label + filename + browse button."""
        row = tk.Frame(parent)
        row.pack(fill='x', pady=3)

        tk.Label(row, text=label, font=("Arial", 9), width=16, anchor='w').pack(side='left')
        tk.Label(row, textvariable=var, font=("Arial", 9), fg="blue",
                 width=28, anchor='w').pack(side='left', padx=4)
        tk.Button(row, text="Browse", width=8, command=command).pack(side='right')

    def _log(self, msg: str):
        """Append message to the log panel."""
        self.log_text.config(state='normal')
        ts = time.strftime("%H:%M:%S")
        self.log_text.insert('end', f"[{ts}] {msg}\n")
        self.log_text.see('end')
        self.log_text.config(state='disabled')

    def _get_default_margin(self) -> int:
        """Get the default margin % for the currently selected marketplace."""
        mkt = self.marketplace_var.get() if hasattr(self, 'marketplace_var') else ''
        if mkt and mkt in MARKETPLACE_CONFIGS:
            return MARKETPLACE_CONFIGS[mkt].get('default_margin', 70)
        return 70

    def _on_marketplace_change(self, event=None):
        """Update margin input when marketplace selection changes."""
        margin = self._get_default_margin()
        self.margin_var.set(str(margin))
        self._log(f"Marketplace changed to {self.marketplace_var.get()}, margin set to {margin}%")

    def _get_margin(self) -> float:
        """
        Get the margin % from the input field.
        Returns as decimal (e.g., 70 → 0.70).
        Falls back to default if input is invalid.
        """
        try:
            val = float(self.margin_var.get().strip())
            if val <= 0 or val > 100:
                raise ValueError
            return val / 100.0
        except (ValueError, AttributeError):
            default = self._get_default_margin()
            self._log(f"Invalid margin input, using default {default}%")
            return default / 100.0

    # ── FILE SELECTION ─────────────────────────────────────────────────────────

    def _select_master(self):
        """Select Items_March master file."""
        path = filedialog.askopenfilename(
            title="Select Items_March.xlsx",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]
        )
        if path:
            self.master_path = path
            self.master_var.set(os.path.basename(path))
            self._log(f"Master: {os.path.basename(path)}")

    def _select_mapping(self):
        """Select the Ship-To B2B mapping file."""
        path = filedialog.askopenfilename(
            title="Select Mapping File (Ship-To B2B)",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]
        )
        if path:
            self.mapping_path = path
            self.mapping_var.set(os.path.basename(path))
            self._log(f"Mapping: {os.path.basename(path)}")

    def _select_po(self):
        """Select the marketplace PO/punch file."""
        path = filedialog.askopenfilename(
            title="Select Marketplace PO File",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]
        )
        if path:
            self.po_path = path
            self.po_var.set(os.path.basename(path))
            self._log(f"PO file: {os.path.basename(path)}")

    # ── PROCESSING ─────────────────────────────────────────────────────────────

    def generate(self):
        """Main processing: load mapping → parse PO → generate output."""

        # ── Validate inputs ──
        marketplace = self.marketplace_var.get()
        if not marketplace or marketplace not in MARKETPLACE_CONFIGS:
            messagebox.showwarning("No Marketplace", "Please select a marketplace.")
            return

        if not self.mapping_path:
            messagebox.showwarning("No Mapping", "Please select the Ship-To mapping file.")
            return

        if not self.po_path:
            messagebox.showwarning("No PO File", "Please select the marketplace PO file.")
            return

        config = MARKETPLACE_CONFIGS[marketplace]
        margin_pct = self._get_margin()
        start_time = time.time()

        self.status_var.set("Processing...")
        self.status_label.config(fg="blue")
        self.root.update()

        self._log(f"Marketplace: {marketplace} | Margin: {int(margin_pct * 100)}%")

        # ── Load mapping ──
        self._log(f"Loading mapping for '{marketplace}'...")
        warnings: List[Tuple[str, str, str]] = []
        loc_count = self.mapping_loader.load(self.mapping_path, config['party_name'], warnings)

        if loc_count == 0:
            self._log("ERROR: No mapping locations found!")
            for _, _, msg in warnings:
                self._log(f"  {msg}")
            self.status_var.set("Failed — mapping load error")
            self.status_label.config(fg="red")
            return

        self._log(f"Loaded {loc_count} locations for {marketplace}")

        # ── Load Items_March (for price validation) ──
        master_loader = None
        if self.master_path:
            self._log(f"Loading Items_March for validation...")
            master_loader = MasterLoader()
            try:
                item_count = master_loader.load(self.master_path)
                self._log(f"Loaded {item_count:,} items from master")
            except Exception as e:
                self._log(f"WARNING: Master load failed: {e} — skipping validation")
                master_loader = None

        # ── Process PO file ──
        self._log(f"Processing {os.path.basename(self.po_path)}...")
        engine = MarketplaceEngine(self.mapping_loader, master=master_loader)
        result = engine.process(self.po_path, config, margin_pct=margin_pct)
        result.margin_pct = margin_pct

        if not result.rows:
            self._log("ERROR: No valid rows extracted!")
            for _, _, msg in result.warnings:
                self._log(f"  WARNING: {msg}")
            self.status_var.set("Failed — no data extracted")
            self.status_label.config(fg="red")
            return

        # ── Log results ──
        unique_pos = set(r.po_number for r in result.rows)
        total_qty = sum(r.qty for r in result.rows)
        unmapped = sum(1 for r in result.rows if not r.mapped)
        mapped_pos = len([po for po in unique_pos
                          if any(r.mapped for r in result.rows if r.po_number == po)])

        self._log(f"Extracted: {len(result.rows)} items, {len(unique_pos)} PO(s), {total_qty} total qty")
        if result.warnings:
            self._log(f"Warnings: {len(result.warnings)}")
            for po, loc, msg in result.warnings[:5]:
                self._log(f"  [{po}] {msg}")
            if len(result.warnings) > 5:
                self._log(f"  ... and {len(result.warnings) - 5} more (see Warnings sheet)")

        # ── Export ──
        self._log("Writing output...")
        output_path = self.exporter.export(result)

        elapsed = time.time() - start_time

        if output_path:
            self.last_output = output_path
            self.open_btn.config(state=tk.NORMAL)

            status_msg = (f"Done — {len(result.rows)} items, {len(unique_pos)} PO(s), "
                          f"{total_qty} qty | {elapsed:.2f}s")
            if result.warnings:
                status_msg += f" | {len(result.warnings)} warning(s)"
                self.status_label.config(fg="orange")
            else:
                self.status_label.config(fg="darkgreen")

            self.status_var.set(status_msg)
            self._log(f"Saved: {output_path}")

            answer = messagebox.askyesno(
                "SO Generated",
                f"Sales Order generated successfully!\n\n"
                f"Marketplace : {marketplace}\n"
                f"PO(s)       : {len(unique_pos)}\n"
                f"Items       : {len(result.rows)}\n"
                f"Total Qty   : {total_qty}\n"
                f"Warnings    : {len(result.warnings)}\n"
                f"Time        : {elapsed:.2f}s\n\n"
                f"Do you want to open the output file?"
            )
            if answer:
                open_file(output_path)
        else:
            self.status_var.set("Failed — no output generated")
            self.status_label.config(fg="red")

    def open_last(self):
        """Open the last generated output file."""
        if self.last_output and self.last_output.exists():
            open_file(self.last_output)
        else:
            messagebox.showwarning("Not Found", "Output file not found.")

    def _download_template(self):
        """
        Generate a blank PO template for the currently selected marketplace.
        Uses the column headers from the marketplace config (template_headers)
        or falls back to the required columns if template_headers not defined.
        """
        marketplace = self.marketplace_var.get()
        if not marketplace or marketplace not in MARKETPLACE_CONFIGS:
            messagebox.showwarning("No Marketplace", "Please select a marketplace first.")
            return

        config = MARKETPLACE_CONFIGS[marketplace]

        save_path = filedialog.asksaveasfilename(
            title=f"Save {marketplace} PO Template",
            defaultextension=".xlsx",
            initialfile=f"{marketplace}_PO_Template.xlsx",
            filetypes=[("Excel files", "*.xlsx")]
        )
        if not save_path:
            return

        try:
            wb = Workbook()
            ws = wb.active
            ws.title = f'{marketplace} PO'

            # Use template_headers if defined, else build from required columns
            headers = config.get('template_headers')
            if not headers:
                headers = [config['po_col'], config['loc_col'], config['item_col'], config['qty_col']]
                if config.get('ean_col'):
                    headers.append(config['ean_col'])
                if config.get('fob_col'):
                    headers.append(config['fob_col'])

            hdr_fill = PatternFill('solid', fgColor='1A237E')
            hdr_font = Font(bold=True, color='FFFFFF', name='Aptos Display', size=11)

            for c, h in enumerate(headers, 1):
                cell = ws.cell(row=1, column=c, value=h)
                cell.font = hdr_font
                cell.fill = hdr_fill
                cell.alignment = Alignment(horizontal='center')
                ws.column_dimensions[get_column_letter(c)].width = max(len(h) + 4, 12)

            # Note row
            ws.cell(row=3, column=1,
                    value=f'← {marketplace} PO template. Fill in data rows below the header. '
                          f'Required: {config["po_col"]}, {config["loc_col"]}, '
                          f'{config["item_col"]}, {config["qty_col"]}').font = \
                Font(name='Aptos Display', size=11, color='FF6600', italic=True)

            ws.freeze_panes = 'A2'
            wb.save(save_path)
            self._log(f"{marketplace} template saved → {save_path}")
            messagebox.showinfo("Template Saved",
                                f"{marketplace} PO template saved to:\n{save_path}")
        except Exception as e:
            self._log(f"Template save failed: {e}")
            messagebox.showerror("Error", f"Failed to save template:\n{e}")

    def run(self):
        """Start the Tkinter main loop."""
        self.root.mainloop()


# ═══════════════════════════════════════════════════════════════════════════════
#  ENTRY POINT
# ═══════════════════════════════════════════════════════════════════════════════

def main():
    check_expiry()
    app = OnlinePOApp()
    app.run()


if __name__ == "__main__":
    main()