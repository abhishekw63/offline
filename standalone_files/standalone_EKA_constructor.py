"""
╔═══════════════════════════════════════════════════════════════════════════════╗
║                     RENEE PO PROCESSOR — EKA Script                         ║
║                     Tkinter GUI Desktop Application                          ║
╠═══════════════════════════════════════════════════════════════════════════════╣
║  Author  : Agami AI / Vishal                                                ║
║  Version : 1.2                                                               ║
║  Tested  : 23 PO files (9 Airport + 10 Kiosk + 4 EBO), 4660 master items   ║
║  Stack   : Python 3.13, Tkinter, pandas, openpyxl                           ║
╚═══════════════════════════════════════════════════════════════════════════════╝

═══════════════════════════════════════════════════════════════════════════════
  CHANGELOG
═══════════════════════════════════════════════════════════════════════════════

  v1.2 — Auto-fill + D365 TO Export + Tester TO Uniqueness
  ─────────────────────────────────────────────────────────
    ┌─────────────────────────────────────────────────────────────────────┐
    │  CHANGE                     │  IMPACT                               │
    ├─────────────────────────────┼───────────────────────────────────────┤
    │  1. Tester TO date_code +1  │  TO/AHDEB/TT/18427 instead of 18426  │
    │     (both segments)         │  → unique last digits for Excel find  │
    ├─────────────────────────────┼───────────────────────────────────────┤
    │  2. Standalone auto-fill    │  Rename file to Location code         │
    │     from EKA_DATA           │  (e.g., EBO_AMD01.xlsx) → auto-fills  │
    │                             │  TO, Transfer-to, Posting Group       │
    ├─────────────────────────────┼───────────────────────────────────────┤
    │  3. D365 TO Package Export  │  New D365TOExporter class. Fills      │
    │     (NEW class)             │  Transfer Header + Transfer Line      │
    │                             │  template via ZIP/XML regex.          │
    ├─────────────────────────────┼───────────────────────────────────────┤
    │  4. EKA_DATA shared across  │  Moved to shared section. Available   │
    │     both modes              │  in standalone (filename lookup) and  │
    │                             │  special order (broadcast).           │
    └─────────────────────────────┴───────────────────────────────────────┘

  v1.1 — Special Order Engine + Full Documentation
  v1.0 — Initial release (23 PO files tested)

═══════════════════════════════════════════════════════════════════════════════
  ARCHITECTURE OVERVIEW
═══════════════════════════════════════════════════════════════════════════════

  ┌──────────────────────────────────────────────────────────────────────┐
  │                        ReneePOApp (GUI)                             │
  │                                                                      │
  │  SHARED (v1.2):                                                      │
  │    ┌────────────┐  ┌───────────────┐                                │
  │    │ Master File │  │ EKA_DATA      │  ← shared across both modes   │
  │    │ Selector    │  │ (Location DB) │                                │
  │    └──────┬─────┘  └──────┬────────┘                                │
  │           │               │                                          │
  │  ┌────────┴───────────────┴──────────────────────────────┐          │
  │  │                    MODE SELECTOR                       │          │
  │  │         [STANDALONE]          [SPECIAL ORDER]          │          │
  │  └──────────┬────────────────────────┬───────────────────┘          │
  │             │                        │                               │
  │  ┌──────────┴──────────┐  ┌──────────┴──────────┐                  │
  │  │  Segment 1:          │  │  Segment 2:          │                  │
  │  │  Per-location PO     │  │  Product broadcast   │                  │
  │  │  files renamed to    │  │  across all EKA       │                  │
  │  │  Location code       │  │  locations             │                  │
  │  │  (EBO_AMD01.xlsx)    │  │  (Special_Order.xlsx) │                  │
  │  │                      │  │                        │                  │
  │  │  v1.2: auto-fill     │  │  Same as v1.1 but     │                  │
  │  │  TO/Transfer/Posting │  │  tester date_code+1   │                  │
  │  │  from EKA_DATA       │  │                        │                  │
  │  └──────────┬──────────┘  └──────────┬──────────┘                  │
  │             │                        │                               │
  │             ▼                        ▼                               │
  │  ┌─────────────────────────────────────────────────┐                │
  │  │           ExcelWriter.write()                    │                │
  │  │  Headers/Lines (TO/SO) | Final Data | Summary   │                │
  │  │  | Unmatched | Tester Master | SO Reference     │                │
  │  └─────────────────────────────────────────────────┘                │
  │             │                                                        │
  │             ▼                                                        │
  │  ┌─────────────────────────────────────────────────┐                │
  │  │     D365TOExporter (NEW v1.2)                    │                │
  │  │  Fills D365 Transfer Order template              │                │
  │  │  (Transfer Header + Transfer Line sheets)        │                │
  │  └─────────────────────────────────────────────────┘                │
  └──────────────────────────────────────────────────────────────────────┘

═══════════════════════════════════════════════════════════════════════════════
  TO NUMBER FORMAT (v1.2 — both segments)
═══════════════════════════════════════════════════════════════════════════════

  Pattern:
      Regular: {Prefix}/{ShortCode}/{MM}/{DDMYY}
      Tester:  {Prefix}/{ShortCode}/TT/{DDMYY + 1}

  Where:
      Prefix    = 'TO' or 'SO' (from EKA_DATA)
      ShortCode = e.g., 'AHDEB', 'CHNAP' (from EKA_DATA)
      MM        = month number, zero-padded ('01'-'12')
      TT        = literal 'TT' for tester orders
      DDMYY     = DD (day) + M (month, not zero-padded) + YY (year)

  Example for April 18, 2026:
      Regular: TO/AHDEB/04/18426    ← date_code = 18426
      Tester:  TO/AHDEB/TT/18427   ← date_code = 18426 + 1 = 18427

  ⚠ The +1 on tester ensures UNIQUE last digits for Excel search.
    Before v1.2, both had the same date_code → searching by last 4 digits
    returned both regular and tester rows. Now they're always different.

═══════════════════════════════════════════════════════════════════════════════
  FILENAME CONVENTION (v1.2 — for standalone auto-fill)
═══════════════════════════════════════════════════════════════════════════════

  For standalone mode to auto-fill TO/Transfer/Posting, rename PO files
  to the ERP Location code from EKA_DATA's 'Location' column:

      EBO_AMD01.xlsx       → matches Location 'EBO_AMD01' → Ahmedabad EBO
      AP_PUNE01.xlsx       → matches Location 'AP_PUNE01' → Pune Airport
      FK_AMD_01.xlsx       → matches Location 'FK_AMD_01' → FK Ahmedabad

  The script:
      1. Strips .xlsx from filename
      2. Matches against 'Location' column in loaded EKA_DATA
      3. If matched → auto-fills TO number, Transfer-to Code, Posting Group
      4. If not matched → WARNING, fields left empty (manual fill needed)

  ⚠ EKA_DATA must be loaded BEFORE processing for auto-fill to work.
    If EKA_DATA not loaded, standalone works as before (fields empty).

═══════════════════════════════════════════════════════════════════════════════
  INPUT FILES — WHAT MUST BE STANDARDIZED
═══════════════════════════════════════════════════════════════════════════════

1. MASTER FILE (Items_March.xlsx)
   ─────────────────────────────
   - This is the product catalog / item master from the ERP system.
   - Must have these columns (exact names):
       No.              → Item number (used in output)
       GTIN             → EAN / barcode (used as lookup key)
       Description      → Product name (for reference)
       GST Group Code   → Tax slab: 'G-18-S', 'G-5-S', 'G-0' etc.
       HSN/SAC Code     → HSN code (not used by script, but in master)
       Mrp              → Maximum Retail Price (used for cost calculation)
   - The script indexes by GTIN and also by No. (item code) for non-stock items.

2. PO FILES (per airport/location)
   ────────────────────────────────
   Each PO file MUST have exactly 5 sheets with these EXACT names:
       'PO'         → Main product order sheet
       'PWP'        → Purchase With Purchase (promotional items)
       'GWP'        → Gift With Purchase (gifting items)
       'Non Stock'  → Non-stock operational items
       'Summary'    → Order summary (not read by script, for reference)

   PO SHEET — Required columns (EXACT names, case-sensitive):
   ┌─────────────┬──────────────────────────────────────────────────────────┐
   │ Column Name │ Description                                            │
   ├─────────────┼──────────────────────────────────────────────────────────┤
   │ EAN         │ Product barcode. MUST match GTIN in Items_March.       │
   │ Order Qty   │ Regular order quantity. Blank/0 = no order.            │
   │ Tester Qty  │ Tester quantity. Blank/0 = no tester.                  │
   └─────────────┴──────────────────────────────────────────────────────────┘
   ⚠ These 3 column names are FIXED. The script will ERROR if not found.
   ⚠ Other columns (Rank, Category, SKU Code, Product Name, Brand, MRP,
     Available, etc.) can be in any order — they are NOT read by the script.
     All product info is looked up from Items_March via EAN.

   PWP SHEET — Fixed structure (DO NOT change):
       Col A: Sr. No.
       Col B: Product Name    → Must be one of: 'Stay With Me - Mini',
                                'Perfume', 'Crème Mini'
       Col C: Avail.Qty
       Col D: Req.Qty         → Demand quantity

   GWP SHEET — Fixed structure (DO NOT change):
       Col A: Sr. No.
       Col B: EAN             → Must match GTIN in Items_March
       Col C: Product Name
       Col D: Avail.Qty
       Col E: Req.Qty         → Demand quantity

   NON STOCK SHEET — Fixed structure (DO NOT change):
       Col A: Sr. No.
       Col B: Product Name    → Must match hard-coded names (see NON_STOCK_EAN_MAP)
       Col C: QTY             → Demand quantity

3. EKA_DATA.xlsx (v1.2: shared across both modes)
   ────────────────────────────────────────────────
   Required columns:
       Short Name              → Display name (e.g., 'Ahmedabad EBO')
       Prefix                  → 'TO' or 'SO' (Transfer Order vs Sales Order)
       Short Code              → Code for TO number (e.g., 'AHDEB', 'CHNAP')
       Transfer Code           → Transfer-to Code for TO, or Ship-to for SO
       Type                    → 'EBO', 'Airport', or 'Kiosk'
       Gen. Biz. Posting Group → e.g., 'OFF-EBO', 'OFF-AIRPORT'
   Optional columns:
       Location                → ERP Location Code (e.g., 'EBO_AMD01')
                                  v1.2: used for standalone filename matching
       Bill to                 → Sell-to Customer No. (for SO locations)
       Ship to                 → Ship-to Code (e.g., '20329_1')
       Status                  → 'Active' or 'Inactive'

═══════════════════════════════════════════════════════════════════════════════
  CALCULATION LOGIC
═══════════════════════════════════════════════════════════════════════════════

  For regular PO orders:
      1. EAN → lookup Items_March → get Item No, MRP, GST Code
      2. Landing Cost = MRP × 0.60  (60% of MRP)
      3. Cost Price (Unit Price):
          - If GST Code contains 'G-18' → Landing Cost / 1.18
          - If GST Code contains 'G-12' → Landing Cost / 1.12
          - If GST Code contains 'G-5'  → Landing Cost / 1.05
          - If GST Code contains 'G-3'  → Landing Cost / 1.03
          - If GST Code is 'G-0' or ''  → Landing Cost as-is
          - Default (unknown GST)       → Landing Cost / 1.18

  For testers (product testers, PWP, GWP, Non-Stock):
      - Unit Price = ₹0.54 (flat rate for all tester/promotional items)

═══════════════════════════════════════════════════════════════════════════════
  SPECIAL RULES — PWP / GWP / NON-STOCK
═══════════════════════════════════════════════════════════════════════════════

  PWP (Purchase With Purchase):
      - 'Stay With Me - Mini' → IGNORED (not shipped)
      - 'Crème Mini'          → IGNORED (not shipped)
      - 'Perfume'             → SPLIT equally across 4 perfume EANs:
            8906121642674 (RENEE BLOOM 8ML NFS)
            8906121647495 (RENEE FLIRT 8ML NFS)
            8906121647501 (RENEE MADAME 8ML NFS)
            8906121645743 (RENEE RED NOIR 8ML NFS)
          Example: demand=10 → 3+3+2+2 (remainder goes to first EANs)

  GWP (Gift With Purchase):
      - All items have EANs → looked up from Items_March directly.
      - Goes into tester TO at ₹0.54.

  Non-Stock:
      - Each item name is mapped to a hard-coded EAN/code (see NON_STOCK_EAN_MAP).
      - That code is then looked up in Items_March to get Item No.
      - If name not in map → output name directly (manual fix needed).
      - Goes into tester TO at ₹0.54.

═══════════════════════════════════════════════════════════════════════════════
  OUTPUT — FINAL DATA SHEET
═══════════════════════════════════════════════════════════════════════════════

  Columns:
      TO                      → Transfer Order number
                                 v1.2: auto-filled from EKA_DATA
      Item                    → Item No from Items_March
      Qty                     → Order/Tester quantity
      Unit Price              → Calculated cost price (PO) or ₹0.54 (testers)
      Transfer-to Code        → Location code
                                 v1.2: auto-filled from EKA_DATA
      Gen. Bus. Posting Group → Posting group
                                 v1.2: auto-filled from EKA_DATA
      Source                  → PO / TESTER / PWP / GWP / NON_STOCK
      Location                → Source filename
      EAN                     → Original EAN
      Product Name            → Product description
      Lookup Status           → OK / NOT_FOUND / UNKNOWN / NO_MAP

  Row ordering per location:
      1. Regular PO orders  (Source = PO)
      2. PWP orders         (Source = PWP)
      3. Product testers    (Source = TESTER)
      4. GWP items          (Source = GWP)
      5. Non-Stock items    (Source = NON_STOCK)

═══════════════════════════════════════════════════════════════════════════════
  WHAT TO DO WHEN THINGS GO WRONG
═══════════════════════════════════════════════════════════════════════════════

  Check the LOG panel in the GUI. Every issue is logged:
      🔴 ERROR  → Missing sheet or column. File cannot be processed.
      🟡 WARN   → EAN not found, unknown item, missing mapping. Row output
                   with name/EAN for manual fix.
      🔵 INFO   → Normal operation details (counts, PWP splits, skips).

  Common issues:
      - "PO sheet: 'Order Qty' column not found"
          → Column header in Excel is not exactly 'Order Qty'. Fix the header.
      - "PO row 45: EAN 890612164XXXX not found in master"
          → New product not yet in Items_March. Add it to master file.
      - "Non-Stock: 'NewItem' not in hard-coded map"
          → New non-stock item. Add to NON_STOCK_EAN_MAP in this script.
      - "PWP: Unknown item 'SomeName'"
          → New PWP item. Add handling in process_pwp() in this script.
      - "EKA: 'filename' not found in Location column"
          → v1.2: File not renamed to Location code. Rename to match EKA_DATA.

═══════════════════════════════════════════════════════════════════════════════
  HOW TO ADD NEW ITEMS
═══════════════════════════════════════════════════════════════════════════════

  New Non-Stock item:
      1. Find its EAN or internal code from Items_March
      2. Add to NON_STOCK_EAN_MAP: 'Exact Name': 'EAN_OR_CODE',
      3. Add same name to blank template Non Stock sheet

  New PWP item:
      1. If should be ignored: add name to PWP_IGNORE
      2. If should be processed: add elif branch in process_pwp()

  New GWP item:
      1. Ensure EAN exists in Items_March (auto-resolved)

  New Perfume variant:
      1. Add EAN to PERFUME_EANS list

  New EKA Location:
      1. Add row to EKA_DATA.xlsx with all required columns
      2. Set Status = 'Active'

═══════════════════════════════════════════════════════════════════════════════
  D365 TO EXPORT (NEW in v1.2)
═══════════════════════════════════════════════════════════════════════════════

  Template format (from EKA_Sample_Package.xlsx):
      Sheet 1 'Transfer Header':
          Row 1: metadata (ignored by script)
          Row 3: column headers
          Row 4+: data rows
          Columns: No. | Transfer-from Code | Transfer-to Code |
                   Posting Date | In-Transit Code | Direct Transfer |
                   Gen. Bus. Posting Group | ...dimensions...

      Sheet 2 'Transfer Line':
          Row 1: metadata (ignored)
          Row 3: column headers
          Row 4+: data rows
          Columns: Document No. | Line No. | Item No. | Quantity |
                   Unit of Measure | Qty. to Ship | Qty. to Receive |
                   Dimension Set ID | Transfer Price

  The D365TOExporter fills this template using the same ZIP/XML regex
  approach as the GT Mass D365 SO exporter. If data exceeds template
  row capacity, new <row> elements are injected before filling.

Requirements:
    pip install pandas openpyxl

Run:
    python renee_po_processor.py
"""

# ═══════════════════════════════════════════════════════════════════════════════
#  IMPORTS
# ═══════════════════════════════════════════════════════════════════════════════

import tkinter as tk                    # GUI framework (standard library)
from tkinter import ttk, filedialog, messagebox  # Themed widgets, dialogs, popups
import threading                        # Background thread (keeps UI responsive)
import os                               # File/path operations
import time                             # Timestamps for logs and output filenames
import re                               # Regex (reserved for future use)
import math                             # Math functions (sin/cos for toggle switch rays)
import shutil                           # File copy for D365 template
import zipfile                          # ZIP manipulation for D365 XML editing
from dataclasses import dataclass, field  # Structured data containers
from typing import List, Dict, Optional, Any, Tuple  # Type hints
from pathlib import Path                # Cross-platform path handling

import pandas as pd                     # Excel reading for master file
from openpyxl import load_workbook, Workbook  # Excel R/W for PO files
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ═══════════════════════════════════════════════════════════════════════════════
#  THEME SYSTEM (Dark / Light)
# ═══════════════════════════════════════════════════════════════════════════════

class Theme:
    """
    Switchable dark/light theme palette.

    All GUI colors are read from Theme.palette so that toggling the theme
    updates every widget. Each color is accessed via a class method shortcut:
        Theme.bg()       → background
        Theme.surface()  → card/panel background
        Theme.accent()   → primary accent (cyan in dark, blue in light)
    """

    DARK = {
        'BG':       "#0F1117",    # Main background
        'SURFACE':  "#1A1D27",    # Card/panel background
        'SURFACE2': "#22263A",    # Secondary surface (drop zones, alt rows)
        'ACCENT':   "#00D4FF",    # Primary accent (cyan)
        'ACCENT2':  "#7B61FF",    # Secondary accent (purple)
        'GREEN':    "#00E676",    # Success / OK status
        'RED':      "#FF5252",    # Error / NOT_FOUND
        'AMBER':    "#FFB300",    # Warning / in-progress
        'PINK':     "#FF4081",    # Branding accent
        'TEXT':     "#E8EAF6",    # Primary text
        'TEXT_DIM': "#6B7280",    # Secondary/muted text
        'BORDER_C': "#2D3250",    # Border color
        'LIST_SEL': "#22263A",    # Listbox selection background
    }

    LIGHT = {
        'BG':       "#F0F2F5",
        'SURFACE':  "#FFFFFF",
        'SURFACE2': "#E8EAF0",
        'ACCENT':   "#0077B6",
        'ACCENT2':  "#5C3D99",
        'GREEN':    "#00A651",
        'RED':      "#D32F2F",
        'AMBER':    "#E65100",
        'PINK':     "#C2185B",
        'TEXT':     "#1A1A2E",
        'TEXT_DIM': "#5F6368",
        'BORDER_C': "#C4C7D0",
        'LIST_SEL': "#D6E4FF",
    }

    _current = 'dark'
    _palette = DARK.copy()

    @classmethod
    def is_dark(cls) -> bool:
        return cls._current == 'dark'

    @classmethod
    def toggle(cls):
        cls._current = 'light' if cls._current == 'dark' else 'dark'
        cls._palette = cls.DARK.copy() if cls._current == 'dark' else cls.LIGHT.copy()

    @classmethod
    def get(cls, key: str) -> str:
        return cls._palette[key]

    # ── Shortcut accessors ──
    @classmethod
    def bg(cls): return cls._palette['BG']
    @classmethod
    def surface(cls): return cls._palette['SURFACE']
    @classmethod
    def surface2(cls): return cls._palette['SURFACE2']
    @classmethod
    def accent(cls): return cls._palette['ACCENT']
    @classmethod
    def accent2(cls): return cls._palette['ACCENT2']
    @classmethod
    def green(cls): return cls._palette['GREEN']
    @classmethod
    def red(cls): return cls._palette['RED']
    @classmethod
    def amber(cls): return cls._palette['AMBER']
    @classmethod
    def pink(cls): return cls._palette['PINK']
    @classmethod
    def text(cls): return cls._palette['TEXT']
    @classmethod
    def text_dim(cls): return cls._palette['TEXT_DIM']
    @classmethod
    def border(cls): return cls._palette['BORDER_C']
    @classmethod
    def list_sel(cls): return cls._palette['LIST_SEL']


# ── Static accent colors (don't change with theme toggle) ──
GREEN    = "#00E676"
RED      = "#FF5252"
AMBER    = "#FFB300"
PINK     = "#FF4081"
ACCENT   = "#00D4FF"
ACCENT2  = "#7B61FF"

# ── Font constants ──
FONT_TITLE = ("Aptos Display", 18, "bold")
FONT_SUB   = ("Aptos Display", 11)
FONT_LABEL = ("Aptos Display", 11, "bold")
FONT_MONO  = ("Aptos Display", 11)
FONT_BTN   = ("Aptos Display", 11, "bold")


# ═══════════════════════════════════════════════════════════════════════════════
#  TOGGLE SWITCH WIDGET
# ═══════════════════════════════════════════════════════════════════════════════

class ToggleSwitch(tk.Canvas):
    """
    Animated toggle switch with sun/moon icons for light/dark mode.

    Visual:
        Dark mode  (is_on=False): dark track, grey knob with crescent moon
        Light mode (is_on=True):  yellow track, white knob with sun + rays

    Animation: knob slides from left→right or right→left in 12ms steps.
    """

    def __init__(self, parent, command=None, width=56, height=28, **kw):
        super().__init__(
            parent, width=width, height=height,
            highlightthickness=0, bd=0, cursor='hand2', **kw
        )
        self.w = width
        self.h = height
        self.pad = 3
        self.knob_r = (height - 2 * self.pad) // 2
        self.is_on = False    # False = dark (moon), True = light (sun)
        self.anim_pos = self.pad + self.knob_r  # current knob X position
        self.command = command
        self._draw()
        self.bind('<Button-1>', self._on_click)

    def _draw(self):
        """Redraw the entire toggle switch (track + knob + icon)."""
        self.delete('all')
        r = self.h // 2

        # Track colors
        if self.is_on:
            track_bg = '#FFD54F'    # warm yellow track (light mode)
            knob_fill = '#FFFFFF'
        else:
            track_bg = '#37474F'    # dark track (dark mode)
            knob_fill = '#B0BEC5'

        # ── Rounded track (pill shape) ──
        self.create_oval(0, 0, self.h, self.h, fill=track_bg, outline='')
        self.create_oval(self.w - self.h, 0, self.w, self.h, fill=track_bg, outline='')
        self.create_rectangle(r, 0, self.w - r, self.h, fill=track_bg, outline='')

        # ── Knob (circle) ──
        kx = self.anim_pos
        ky = self.h // 2
        kr = self.knob_r
        self.create_oval(kx - kr, ky - kr, kx + kr, ky + kr,
                         fill=knob_fill, outline='#CCCCCC')

        # ── Icon on knob ──
        if self.is_on:
            # Sun: small circle + 8 rays
            self.create_oval(kx - 4, ky - 4, kx + 4, ky + 4,
                             fill='#FF8F00', outline='')
            for angle in range(0, 360, 45):
                rad = math.radians(angle)
                x1 = kx + 6 * math.cos(rad)
                y1 = ky + 6 * math.sin(rad)
                x2 = kx + 8 * math.cos(rad)
                y2 = ky + 8 * math.sin(rad)
                self.create_line(x1, y1, x2, y2, fill='#FF8F00', width=1.5)
        else:
            # Moon: crescent via overlapping circles
            self.create_oval(kx - 5, ky - 5, kx + 5, ky + 5,
                             fill='#78909C', outline='')
            self.create_oval(kx - 2, ky - 6, kx + 6, ky + 4,
                             fill=knob_fill, outline='')

    def _on_click(self, event=None):
        """Handle click: toggle state, animate, call callback."""
        self.is_on = not self.is_on
        self._animate()
        if self.command:
            self.command()

    def _animate(self):
        """Animate knob sliding to new position in 12ms steps."""
        target = (self.w - self.pad - self.knob_r) if self.is_on else (self.pad + self.knob_r)
        step = 3 if self.is_on else -3

        def _step():
            if (step > 0 and self.anim_pos < target) or (step < 0 and self.anim_pos > target):
                self.anim_pos += step
                self._draw()
                self.after(12, _step)
            else:
                self.anim_pos = target
                self._draw()

        _step()


# ═══════════════════════════════════════════════════════════════════════════════
#  CORE DATA STRUCTURES
# ═══════════════════════════════════════════════════════════════════════════════

@dataclass
class OutputRow:
    """
    Single row in the Final Data output.

    Fields:
        to             : Transfer Order / Sales Order number
                          v1.2: auto-filled from EKA_DATA in standalone mode
        item_no        : Item No from Items_March, or product name if unresolved
        qty            : Quantity (order or tester)
        unit_price     : Calculated cost price (PO) or ₹0.54 (testers)
        transfer_to    : Transfer-to Code / Location Code
                          v1.2: auto-filled from EKA_DATA
        posting_group  : Gen. Bus. Posting Group
                          v1.2: auto-filled from EKA_DATA
        source         : Origin type: 'PO', 'TESTER', 'PWP', 'GWP', 'NON_STOCK'
        ean            : Original EAN/barcode
        product_name   : Product description (for reference)
        lookup_status  : 'OK', 'NOT_FOUND', 'UNKNOWN', 'NO_MAP'
    """
    to: str = ''
    item_no: Any = ''
    qty: int = 0
    unit_price: float = 0.0
    transfer_to: str = ''
    posting_group: str = ''
    source: str = ''
    ean: str = ''
    product_name: str = ''
    lookup_status: str = ''


@dataclass
class LocationResult:
    """
    Complete processing result for one PO file (one location).

    Contains separate lists for each order type, plus unmatched EANs
    and processing logs for the GUI log panel.
    """
    filename: str
    regular_orders: List[OutputRow] = field(default_factory=list)   # PO orders (calculated price)
    tester_orders: List[OutputRow] = field(default_factory=list)    # Product testers (₹0.54)
    pwp_orders: List[OutputRow] = field(default_factory=list)       # PWP items (₹0.54)
    gwp_orders: List[OutputRow] = field(default_factory=list)       # GWP items (₹0.54)
    nonstock_orders: List[OutputRow] = field(default_factory=list)  # Non-stock items (₹0.54)
    unmatched: List[Dict] = field(default_factory=list)             # EANs not found in master
    logs: List[tuple] = field(default_factory=list)                 # (level, message) for GUI log


# ═══════════════════════════════════════════════════════════════════════════════
#  PROCESSING ENGINE
# ═══════════════════════════════════════════════════════════════════════════════

class POEngine:
    """
    Core processing logic — no GUI dependency.

    Can be used standalone:
        engine = POEngine()
        engine.load_master('Items_March.xlsx')
        result = engine.process_file('EBO_AMD01.xlsx')

    Responsibilities:
        - Load and index Items_March master file
        - Validate PO file structure (sheets, columns)
        - Process PO/PWP/GWP/Non-Stock sheets
        - Calculate cost prices with GST adjustment
    """

    # ┌─────────────────────────────────────────────────────────────────────────┐
    # │ HARD-CODED MAPPINGS — UPDATE HERE WHEN NEW ITEMS ARE ADDED             │
    # │                                                                         │
    # │ Non-Stock product names → EAN/GTIN or internal item code               │
    # │ ⚠ Names MUST match EXACTLY what's in the Excel Non Stock sheet         │
    # │ ⚠ EAN/codes MUST exist in Items_March (GTIN or No. column)            │
    # └─────────────────────────────────────────────────────────────────────────┘
    NON_STOCK_EAN_MAP = {
        'Cotton Rolls':      'OPM-RSK-CR500-RE',        # → Item 400039
        'Mirrors':           'OPM-RSK-PU-LMS-RE',       # → Item 400037
        'Carry Bag (Small)': '8904473106011',            # → Item 300077
        'Carry Bag (Big)':   '8904473106004',            # → Item 300076
        'Cleansers':         '8906121643572',            # → Item 200101
        'Calculator':        'OPM-CAL-SK-RE',            # → Item 400111
        'Blotters':          'RCPL_PB',                  # → Item 400060
        'Swabs':             'OPM-NMS-OT-P100-SWB',     # → Item 400057
        'Bill Roll':         'OPM-TPR-VL-TSC-100-150',  # → Item 400088
        'Renee Notebook':    'RCPL_NOTEPAD',             # → Item 400059
        'Pen':               'RCPL_PEN',                 # → Item 400061
    }

    # ┌─────────────────────────────────────────────────────────────────────────┐
    # │ PERFUME PWP — 4 EANs to split demand equally                           │
    # │ When PWP has 'Perfume' qty N: N÷4 per EAN, remainder to first EANs    │
    # │ Example: N=10 → 3, 3, 2, 2                                            │
    # └─────────────────────────────────────────────────────────────────────────┘
    PERFUME_EANS = [
        '8906121642674',  # RENEE BLOOM 8ML NFS
        '8906121647495',  # RENEE FLIRT 8ML NFS
        '8906121647501',  # RENEE MADAME 8ML NFS
        '8906121645743',  # RENEE RED NOIR 8ML NFS
    ]

    # ┌─────────────────────────────────────────────────────────────────────────┐
    # │ PWP ITEMS TO IGNORE — not shipped                                      │
    # └─────────────────────────────────────────────────────────────────────────┘
    PWP_IGNORE = {'Stay With Me - Mini', 'Crème Mini'}

    def __init__(self):
        self.master: Dict[str, Dict] = {}  # Indexed by GTIN and item code

    def load_master(self, path: str) -> int:
        """
        Load Items_March.xlsx and build lookup dictionary.

        Indexed by BOTH GTIN (EAN) and No. (item code) so that:
            - PO items can be looked up by EAN
            - Non-stock items can be looked up by internal codes

        Args:
            path: Path to Items_March.xlsx

        Returns:
            Number of rows loaded from master file.
        """
        df = pd.read_excel(path, header=0)
        df['GTIN_str'] = df['GTIN'].astype(str).str.strip()
        self.master = {}

        # Get column indices for fast array access
        gtin_idx = df.columns.get_loc('GTIN_str')
        desc_idx = df.columns.get_loc('Description') if 'Description' in df.columns else None
        no_idx = df.columns.get_loc('No.')
        mrp_idx = df.columns.get_loc('Mrp')
        gst_idx = df.columns.get_loc('GST Group Code') if 'GST Group Code' in df.columns else None

        for r_vals in df.values:
            desc = str(r_vals[desc_idx]) if desc_idx is not None and pd.notna(r_vals[desc_idx]) else ''
            gst_val = str(r_vals[gst_idx]) if gst_idx is not None and pd.notna(r_vals[gst_idx]) else ''

            # Primary index: by GTIN (EAN barcode)
            self.master[r_vals[gtin_idx]] = {
                'item_no': r_vals[no_idx],
                'mrp': r_vals[mrp_idx],
                'gst_code': gst_val,
                'description': desc,
            }

            # Secondary index: by No. (item code) — for non-stock
            item_code = str(r_vals[no_idx]).strip()
            if item_code not in self.master:
                self.master[item_code] = {
                    'item_no': r_vals[no_idx],
                    'mrp': r_vals[mrp_idx],
                    'gst_code': gst_val,
                    'description': desc,
                }

        return len(df)

    @staticmethod
    def calc_cost_price(mrp, gst_code: str) -> Optional[float]:
        """
        Calculate unit price for regular PO orders.

        Formula:
            Landing Cost = MRP × 60%
            Cost Price   = Landing Cost ÷ (1 + GST rate)

        GST codes in Items_March and their divisors:
            0-G      (9 items)    → 0% GST  → ÷ 1.00
            G-3      (1 item)     → 3% GST  → ÷ 1.03
            G-5      (1084 items) → 5% GST  → ÷ 1.05
            G-5-S    (108 items)  → 5% GST  → ÷ 1.05
            G-12     (67 items)   → 12% GST → ÷ 1.12
            G-18     (2022 items) → 18% GST → ÷ 1.18
            G-18-S   (1364 items) → 18% GST → ÷ 1.18

        ⚠ This is ONLY for regular PO orders. Testers use flat ₹0.54.

        Args:
            mrp      : Maximum Retail Price from Items_March
            gst_code : GST Group Code string

        Returns:
            Calculated cost price, or None if MRP is invalid.
        """
        if mrp is None or pd.isna(mrp):
            return None

        landing = float(mrp) * 0.60
        gst = str(gst_code).strip().upper()

        # 0% GST
        if gst in ('0-G', 'G-0', 'G-0-S', '0', '') or gst == 'NAN':
            return landing
        # 3% GST
        if gst in ('G-3', 'G-3-S'):
            return landing / 1.03
        # 5% GST
        if '5' in gst and '18' not in gst and '12' not in gst:
            return landing / 1.05
        # 12% GST
        if '12' in gst:
            return landing / 1.12
        # 18% GST
        if '18' in gst:
            return landing / 1.18
        # Unknown — default to 18%
        return landing / 1.18

    def _detect_po_columns(self, ws, logs: Optional[List] = None) -> Dict[str, int]:
        """
        Scan header row of PO sheet for 3 required columns.

        Standard names (exact match):
            'EAN'        → ean
            'Order Qty'  → order_qty
            'Tester Qty' → tester_qty

        Fallbacks with alert:
            'Tester'    → tester_qty (alert: rename to 'Tester Qty')
            'order Qty' → order_qty  (alert: rename to 'Order Qty')

        Args:
            ws   : openpyxl worksheet
            logs : optional list to append alerts to

        Returns:
            Dict mapping internal key → column index (0-based).
        """
        hmap = {}
        all_headers = {}

        for cell in list(ws.iter_rows(min_row=1, max_row=1))[0]:
            val = str(cell.value or '').strip()
            idx = cell.column - 1
            if val:
                all_headers[val] = idx

            if val == 'EAN':
                hmap['ean'] = idx
            elif val == 'Order Qty':
                hmap['order_qty'] = idx
            elif val == 'Tester Qty':
                hmap['tester_qty'] = idx
            elif val == 'Tester' and 'tester_qty' not in hmap:
                all_headers['__tester_fallback'] = idx

        # Fallback: accept 'Tester' if 'Tester Qty' not found
        if 'tester_qty' not in hmap:
            if '__tester_fallback' in all_headers:
                hmap['tester_qty'] = all_headers['__tester_fallback']
                if logs is not None:
                    logs.append(('alert',
                        "Auto-fixed: 'Tester' → 'Tester Qty'. "
                        "Please rename column to 'Tester Qty' in this file."))
            elif 'Tester' in all_headers:
                hmap['tester_qty'] = all_headers['Tester']
                if logs is not None:
                    logs.append(('alert',
                        "Auto-fixed: 'Tester' → 'Tester Qty'. "
                        "Please rename column."))

        # Fallback: accept 'order Qty' (lowercase o)
        if 'order_qty' not in hmap and 'order Qty' in all_headers:
            hmap['order_qty'] = all_headers['order Qty']
            if logs is not None:
                logs.append(('alert',
                    "Auto-fixed: 'order Qty' → 'Order Qty'. "
                    "Please rename column."))

        return hmap

    def _safe_int(self, val) -> int:
        """Safely convert cell value to int. Returns 0 for None/empty/errors."""
        try:
            if val is None or str(val).strip() in ('', '#N/A', 'None'):
                return 0
            return int(float(val))
        except (ValueError, TypeError):
            return 0

    def _ean_str(self, raw) -> str:
        """Convert raw EAN cell value to clean string. Handles float→int conversion."""
        if raw is None:
            return ''
        return str(int(raw)) if isinstance(raw, (int, float)) else str(raw).strip()

    def process_po_sheet(self, ws, col_map: Dict, logs: List) -> Tuple[List[OutputRow], List[OutputRow], List[Dict]]:
        """
        Process PO sheet → separate into regular orders and tester orders.

        For each row with EAN:
            - If Order Qty > 0 → regular order at calculated cost price
            - If Tester Qty > 0 → tester at ₹0.54
            - If EAN not in master → warning, output with '?EAN:...'

        Skips rows where EAN is empty or row contains 'TOTAL'.

        Args:
            ws      : openpyxl worksheet for PO sheet
            col_map : column index mapping from _detect_po_columns
            logs    : list to append processing messages

        Returns:
            Tuple of (regular_orders, tester_orders, unmatched_list)
        """
        regular, testers, unmatched = [], [], []

        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=False):
            ean_raw = row[col_map['ean']].value
            if ean_raw is None:
                continue
            if any(c.value and str(c.value).upper() == 'TOTAL' for c in row):
                continue

            row_num = row[0].row
            ean = self._ean_str(ean_raw)
            order_qty = self._safe_int(row[col_map['order_qty']].value) if 'order_qty' in col_map else 0
            tester_qty = self._safe_int(row[col_map['tester_qty']].value) if 'tester_qty' in col_map else 0

            # Master lookup
            info = self.master.get(ean) or self.master.get(ean.lstrip('0'))

            if info:
                item_no = info['item_no']
                gst_code = info['gst_code']

                # Warn on unknown GST codes
                known_gst = {'0-G', 'G-3', 'G-3-S', 'G-5', 'G-5-S',
                             'G-12', 'G-12-S', 'G-18', 'G-18-S', ''}
                gst_upper = str(gst_code).strip().upper()
                if gst_upper not in known_gst and gst_upper != 'NAN':
                    logs.append(('warn',
                        f"PO row {row_num}: Unknown GST '{gst_code}' "
                        f"for Item {item_no} — defaulting to 18%"))

                cost = self.calc_cost_price(info['mrp'], gst_code)
                status = 'OK'
            else:
                item_no = f'?EAN:{ean}'
                cost = None
                status = 'NOT_FOUND'
                unmatched.append({
                    'ean': ean, 'product_name': '',
                    'order_qty': order_qty, 'tester_qty': tester_qty,
                })
                logs.append(('warn', f"PO row {row_num}: EAN {ean} not found in master"))

            if order_qty > 0:
                regular.append(OutputRow(
                    item_no=item_no, qty=order_qty,
                    unit_price=cost or 0, source='PO',
                    ean=ean, lookup_status=status,
                ))

            if tester_qty > 0:
                testers.append(OutputRow(
                    item_no=item_no, qty=tester_qty,
                    unit_price=0.54, source='TESTER',
                    ean=ean, lookup_status=status,
                ))

        return regular, testers, unmatched

    def process_pwp(self, ws, logs: List) -> List[OutputRow]:
        """
        Process PWP (Purchase With Purchase) sheet.

        Rules:
            'Stay With Me - Mini' → SKIP (in PWP_IGNORE)
            'Crème Mini'          → SKIP (in PWP_IGNORE)
            'Perfume'             → SPLIT into 4 EANs equally

        All at ₹0.54. Sheet: Col A=Sr.No, B=Name, C=Avail, D=Req.Qty
        """
        rows = []

        for row in ws.iter_rows(min_row=2, max_row=20, values_only=False):
            a, b = row[0].value, row[1].value
            d = row[3].value
            if a is None or str(a).strip().upper() == 'TOTAL':
                continue

            qty = self._safe_int(d)
            name = str(b or '').strip()

            # Ignore list
            if name in self.PWP_IGNORE:
                if qty > 0:
                    logs.append(('info', f"PWP: '{name}' qty={qty} → skipped (ignore list)"))
                continue

            if qty <= 0:
                continue

            # Perfume → split across 4 EANs
            if 'perfume' in name.lower():
                base_qty = qty // 4
                remainder = qty % 4
                logs.append(('info',
                    f"PWP: Perfume qty={qty} → split 4 EANs "
                    f"({base_qty}+{base_qty}+{base_qty}+{base_qty}, "
                    f"remainder={remainder})"))

                for i, ean in enumerate(self.PERFUME_EANS):
                    eq = base_qty + (1 if i < remainder else 0)
                    if eq <= 0:
                        continue

                    info = self.master.get(ean)
                    if info:
                        item_no = info['item_no']
                    else:
                        item_no = f'?EAN:{ean}'
                        logs.append(('warn', f"PWP: Perfume EAN {ean} not in master"))

                    rows.append(OutputRow(
                        item_no=item_no, qty=eq, unit_price=0.54,
                        source='PWP', ean=ean,
                        product_name=f'Perfume ({ean})',
                        lookup_status='OK' if info else 'NOT_FOUND',
                    ))
                continue

            # Unknown PWP item
            logs.append(('warn', f"PWP: Unknown '{name}' qty={qty} → outputting name"))
            rows.append(OutputRow(
                item_no=name, qty=qty, unit_price=0.54,
                source='PWP', product_name=name, lookup_status='UNKNOWN',
            ))

        return rows

    def process_gwp(self, ws, logs: List) -> List[OutputRow]:
        """
        Process GWP (Gift With Purchase) sheet.

        Each GWP item has an EAN → looked up from Items_March.
        All priced at ₹0.54.

        Sheet: Col A=Sr.No, B=EAN, C=Name, D=Avail, E=Req.Qty
        """
        rows = []

        for row in ws.iter_rows(min_row=2, max_row=20, values_only=False):
            a, ean_raw, name, _, req_raw = [row[i].value for i in range(5)]
            if a is None or str(a).strip().upper() == 'TOTAL':
                continue

            qty = self._safe_int(req_raw)
            if qty > 0:
                ean = self._ean_str(ean_raw)
                name_str = str(name or '')
                info = self.master.get(ean)

                if info:
                    item_no = info['item_no']
                else:
                    item_no = name_str
                    logs.append(('warn',
                        f"GWP: EAN {ean} ({name_str}) not found → using name"))

                rows.append(OutputRow(
                    item_no=item_no, qty=qty, unit_price=0.54,
                    source='GWP', ean=ean, product_name=name_str,
                    lookup_status='OK' if info else 'NOT_FOUND',
                ))

        return rows

    def process_non_stock(self, ws, logs: List) -> List[OutputRow]:
        """
        Process Non Stock sheet.

        Lookup chain:
            1. Product name → NON_STOCK_EAN_MAP → EAN/code
            2. EAN/code → Items_March → Item No
            3. Not in map → WARNING, output name for manual fix

        All at ₹0.54. Sheet: Col A=Sr.No, B=Name, C=QTY
        """
        rows = []

        for row in ws.iter_rows(min_row=2, max_row=20, values_only=False):
            a, b, c = row[0].value, row[1].value, row[2].value
            if a is None or str(a).strip().upper() == 'TOTAL':
                continue

            qty = self._safe_int(c)
            if qty <= 0:
                continue

            name = str(b or '').strip()

            # Look up EAN from hard-coded map
            ean = self.NON_STOCK_EAN_MAP.get(name, '')
            if not ean:
                logs.append(('warn',
                    f"Non-Stock: '{name}' qty={qty} → not in map, "
                    f"outputting name directly"))
                rows.append(OutputRow(
                    item_no=name, qty=qty, unit_price=0.54,
                    source='NON_STOCK', ean='', product_name=name,
                    lookup_status='NO_MAP',
                ))
                continue

            info = self.master.get(ean)
            if info:
                item_no = info['item_no']
                status = 'OK'
            else:
                item_no = name
                status = 'NOT_FOUND'
                logs.append(('warn',
                    f"Non-Stock: '{name}' code={ean} → not in master"))

            rows.append(OutputRow(
                item_no=item_no, qty=qty, unit_price=0.54,
                source='NON_STOCK', ean=ean, product_name=name,
                lookup_status=status,
            ))

        return rows

    def validate_file(self, filepath: str) -> List[tuple]:
        """
        Pre-processing validation pass — checks structure without extracting.

        Validates:
            1. Required sheets: PO, PWP, GWP, Non Stock
            2. PO columns: EAN, Order Qty, Tester Qty
            3. PWP item names
            4. Non-Stock item names
            5. GWP EANs against master

        Returns:
            List of (level, message) tuples:
                'error' → blocking, 'warn' → non-blocking, 'info' → normal
        """
        logs = []
        has_blocking = False

        try:
            wb = load_workbook(filepath, data_only=True)
        except Exception as e:
            return [('error', f"Cannot open: {e}")]

        # ── 1. Sheet validation ──
        required = ['PO', 'PWP', 'GWP', 'Non Stock']
        for sheet in required:
            if sheet not in wb.sheetnames:
                logs.append(('error', f"Missing sheet: '{sheet}'"))
                has_blocking = True

        found = [s for s in required if s in wb.sheetnames]
        logs.append(('info', f"Sheets: {', '.join(found)} of {len(required)}"))

        # ── 2. PO column validation ──
        if 'PO' in wb.sheetnames:
            ws_po = wb['PO']
            col_map = self._detect_po_columns(ws_po, logs)

            if 'ean' not in col_map:
                logs.append(('error', "PO: 'EAN' not found"))
                has_blocking = True
            if 'order_qty' not in col_map:
                logs.append(('error', "PO: 'Order Qty' not found"))
                has_blocking = True
            if 'tester_qty' not in col_map:
                logs.append(('error', "PO: 'Tester Qty' not found"))
                has_blocking = True

            if not has_blocking:
                data_rows = 0
                for row in ws_po.iter_rows(min_row=2, max_row=ws_po.max_row, values_only=False):
                    ean_raw = row[col_map['ean']].value
                    if ean_raw and not any(
                        c.value and str(c.value).upper() == 'TOTAL' for c in row
                    ):
                        data_rows += 1
                logs.append(('info', f"PO: {data_rows} data rows"))

        # ── 3. PWP item names ──
        if 'PWP' in wb.sheetnames:
            known_pwp = self.PWP_IGNORE | {'Perfume', 'perfume'}
            for row in wb['PWP'].iter_rows(min_row=2, max_row=20, values_only=False):
                a, b = row[0].value, row[1].value
                d = row[3].value
                if a is None or str(a).strip().upper() == 'TOTAL':
                    continue
                name = str(b or '').strip()
                qty = self._safe_int(d)
                if qty > 0 and name.lower() not in {n.lower() for n in known_pwp}:
                    logs.append(('warn', f"PWP: Unknown '{name}' qty={qty}"))

        # ── 4. Non-Stock names ──
        if 'Non Stock' in wb.sheetnames:
            for row in wb['Non Stock'].iter_rows(min_row=2, max_row=20, values_only=False):
                a, b, c = row[0].value, row[1].value, row[2].value
                if a is None or str(a).strip().upper() == 'TOTAL':
                    continue
                name = str(b or '').strip()
                qty = self._safe_int(c)
                if qty > 0 and name not in self.NON_STOCK_EAN_MAP:
                    logs.append(('warn', f"Non-Stock: '{name}' qty={qty} — not in map"))

        # ── 5. GWP EANs ──
        if 'GWP' in wb.sheetnames and self.master:
            for row in wb['GWP'].iter_rows(min_row=2, max_row=20, values_only=False):
                a, ean_raw, name, _, req_raw = [row[i].value for i in range(5)]
                if a is None or str(a).strip().upper() == 'TOTAL':
                    continue
                qty = self._safe_int(req_raw)
                if qty > 0:
                    ean = self._ean_str(ean_raw)
                    if not self.master.get(ean):
                        logs.append(('warn',
                            f"GWP: EAN {ean} ({name}) qty={qty} — not in master"))

        return logs

    def process_file(self, filepath: str) -> LocationResult:
        """
        Process a single PO file → LocationResult with all order types.

        Opens the workbook, validates sheets exist, then processes each
        sheet (PO, PWP, GWP, Non Stock) independently.

        Args:
            filepath: Path to the PO Excel file

        Returns:
            LocationResult with all orders, unmatched EANs, and logs.
        """
        wb = load_workbook(filepath, data_only=True)
        res = LocationResult(filename=Path(filepath).name)

        # Check required sheets
        for sheet in ['PO', 'PWP', 'GWP', 'Non Stock']:
            if sheet not in wb.sheetnames:
                res.logs.append(('error', f"Sheet '{sheet}' not found"))

        # ── PO Sheet ──
        if 'PO' in wb.sheetnames:
            ws_po = wb['PO']
            col_map = self._detect_po_columns(ws_po, res.logs)

            if 'ean' not in col_map:
                res.logs.append(('error', "PO: 'EAN' not found"))
            if 'order_qty' not in col_map:
                res.logs.append(('error', "PO: 'Order Qty' not found"))
            if 'tester_qty' not in col_map:
                res.logs.append(('error', "PO: 'Tester Qty' not found"))

            if 'ean' in col_map:
                res.regular_orders, res.tester_orders, res.unmatched = \
                    self.process_po_sheet(ws_po, col_map, res.logs)
                res.logs.append(('info',
                    f"PO: {len(res.regular_orders)} orders, "
                    f"{len(res.tester_orders)} testers"))
            else:
                res.logs.append(('error', "PO: Skipping — no EAN column"))

        # ── PWP ──
        if 'PWP' in wb.sheetnames:
            res.pwp_orders = self.process_pwp(wb['PWP'], res.logs)

        # ── GWP ──
        if 'GWP' in wb.sheetnames:
            res.gwp_orders = self.process_gwp(wb['GWP'], res.logs)

        # ── Non Stock ──
        if 'Non Stock' in wb.sheetnames:
            res.nonstock_orders = self.process_non_stock(wb['Non Stock'], res.logs)

        return res


# ═══════════════════════════════════════════════════════════════════════════════
#  EXCEL WRITER
# ═══════════════════════════════════════════════════════════════════════════════
# Takes a list of LocationResult objects and writes a formatted Excel workbook
# with 8+ sheets: Headers/Lines (TO/SO), Final Data, Summary, Unmatched EANs,
# Tester Items Master, and optionally SO Reference.
#
# All formatting uses 'Aptos Display' size 11, with color-coded Source and
# Status columns for quick visual scanning.
#
# Sheet 1 'Lines (TO)':       Transfer Order line items (Item, Qty, Price)
# Sheet 2 'Lines (SO)':       Sales Order line items
# Sheet 3 'Headers (SO)':     SO headers (one row per SO number)
# Sheet 4 'Headers (TO)':     TO headers (one row per TO number)
# Sheet 5 'Final Data':       All rows (PO + PWP + TESTER + GWP + NON_STOCK)
# Sheet 6 'Summary':          Per-location qty breakdown with totals
# Sheet 7 'Unmatched EANs':   Failed EAN lookups for manual fix
# Sheet 8 'Tester Items Master': PWP/GWP/Non-Stock resolution audit
# Sheet 9 'SO Reference':     Special Order source data (optional)

class ExcelWriter:
    """
    Writes formatted Excel output with color-coded Source and Status.

    Sheets produced:
        1. Lines (TO)          — Transfer Order line items
        2. Lines (SO)          — Sales Order line items
        3. Headers (SO)        — SO headers (one per SO number)
        4. Headers (TO)        — TO headers (one per TO number)
        5. Final Data          — All rows with metadata
        6. Summary             — Per-location qty breakdown
        7. Unmatched EANs      — Failed lookups
        8. Tester Items Master — PWP/GWP/Non-Stock audit
        9. SO Reference        — (Special Order mode only)
    """

    HEADER_FILL = PatternFill('solid', fgColor='1A237E')
    HEADER_FONT = Font(bold=True, color='FFFFFF', name='Aptos Display', size=11)
    THIN_SIDE = Side(style='thin', color='CCCCCC')
    BORDER = Border(left=THIN_SIDE, right=THIN_SIDE, top=THIN_SIDE, bottom=THIN_SIDE)

    SOURCE_COLORS = {
        'PO': ('000000', 'FFFFFF'), 'TESTER': ('7B61FF', 'FFFFFF'),
        'PWP': ('FF6600', 'FFFFFF'), 'GWP': ('00BCD4', 'FFFFFF'),
        'NON_STOCK': ('795548', 'FFFFFF'),
    }
    STATUS_COLORS = {
        'OK': ('00C853', '000000'), 'NOT_FOUND': ('FF5252', 'FFFFFF'),
        'NEEDS_EAN': ('FFB300', '000000'), 'NEEDS_ITEM_NO': ('FFB300', '000000'),
    }

    @classmethod
    def write(cls, results, output_path, eka_locations=None, master=None, so_products=None):
        """Write output Excel with all sheets."""
        wb = Workbook()
        wb.remove(wb.active)

        loc_lookup = {}
        if eka_locations:
            for loc in eka_locations:
                loc_lookup[loc['short_name']] = loc

        cls._write_lines_to(wb, results, loc_lookup)
        cls._write_lines_so(wb, results, loc_lookup)
        cls._write_headers_so(wb, results, loc_lookup)
        cls._write_headers_to(wb, results, loc_lookup)
        cls._write_final_data(wb, results)
        cls._write_summary(wb, results)
        cls._write_unmatched(wb, results)
        cls._write_tester_master(wb, results)

        if so_products and master:
            cls._write_so_reference(wb, so_products, master)

        wb.save(output_path)

    @classmethod
    def _hdr_cell(cls, ws, row, col, value):
        """Create a formatted header cell (navy bg, white bold text)."""
        cell = ws.cell(row=row, column=col, value=value)
        cell.font = cls.HEADER_FONT
        cell.fill = cls.HEADER_FILL
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = cls.BORDER
        return cell

    @classmethod
    def _data_cell(cls, ws, row, col, value, fmt=None):
        """Create a formatted data cell."""
        cell = ws.cell(row=row, column=col, value=value)
        cell.font = Font(name='Aptos Display', size=11)
        cell.border = cls.BORDER
        if fmt:
            cell.number_format = fmt
        return cell

    @classmethod
    def _auto_width(cls, ws, max_w=50):
        """Auto-fit column widths based on content."""
        for col in ws.columns:
            letter = col[0].column_letter
            w = max((len(str(c.value or '')) for c in col), default=8)
            ws.column_dimensions[letter].width = min(w + 3, max_w)

    # ── Headers (TO) ──────────────────────────────────────────────────────────

    @classmethod
    def _write_headers_to(cls, wb, results, loc_lookup):
        """Sheet: Headers (TO) — one row per unique TO number."""
        ws = wb.create_sheet('Headers (TO)')
        headers = [
            'No.', 'Transfer-from Code', 'Transfer-to Code', 'Posting Date',
            'In-Transit Code', 'Direct Transfer', 'Gen. Bus. Posting Group',
            'Brand Code (Dimension)', 'Channel Code (Dimension)',
            'Catagory (Dimension)', 'Department Code (Dimension)',
            'Geography Code (Dimension)',
        ]
        for c, h in enumerate(headers, 1):
            cls._hdr_cell(ws, 1, c, h)

        r = 2
        today_str = time.strftime("%d-%m-%Y")
        seen_to = set()

        for res in results:
            all_rows = (res.regular_orders + res.tester_orders +
                       res.pwp_orders + res.gwp_orders + res.nonstock_orders)

            for item in all_rows:
                to_num = item.to
                if to_num and to_num.startswith('TO/') and to_num not in seen_to:
                    seen_to.add(to_num)
                    loc = loc_lookup.get(res.filename, {})

                    cls._data_cell(ws, r, 1, to_num)
                    cls._data_cell(ws, r, 2, 'PICK')
                    cls._data_cell(ws, r, 3, loc.get('transfer_code', item.transfer_to))
                    cls._data_cell(ws, r, 4, today_str)
                    cls._data_cell(ws, r, 5, 'IN TRANSIT')
                    cls._data_cell(ws, r, 6, 'false')
                    cls._data_cell(ws, r, 7, loc.get('posting_group', item.posting_group))
                    r += 1

        cls._auto_width(ws)

    # ── Headers (SO) ──────────────────────────────────────────────────────────

    @classmethod
    def _write_headers_so(cls, wb, results, loc_lookup):
        """Sheet: Headers (SO) — one row per unique SO number."""
        ws = wb.create_sheet('Headers (SO)')
        headers = [
            'Document Type', 'No.', 'Sell-to Customer No.', 'Ship-to Code',
            'Posting Date', 'Order Date', 'Document Date',
            'Invoice From Date', 'Invoice To Date',
            'External Document No.', 'Location Code', 'Dimension Set ID',
            'Supply Type', 'Voucher Narration',
            'Brand Code (Dimension)', 'Channel Code (Dimension)',
            'Catagory (Dimension)', 'Geography Code (Dimension)',
        ]
        for c, h in enumerate(headers, 1):
            cls._hdr_cell(ws, 1, c, h)

        r = 2
        today_str = time.strftime("%d-%m-%Y")
        seen_so = set()

        for res in results:
            all_rows = (res.regular_orders + res.tester_orders +
                       res.pwp_orders + res.gwp_orders + res.nonstock_orders)

            for item in all_rows:
                so_num = item.to
                if so_num and so_num.startswith('SO/') and so_num not in seen_so:
                    seen_so.add(so_num)
                    loc = loc_lookup.get(res.filename, {})
                    bill_to = loc.get('bill_to', '')
                    ship_to = loc.get('ship_to', '')

                    cls._data_cell(ws, r, 1, 'Order')
                    cls._data_cell(ws, r, 2, so_num)
                    cls._data_cell(ws, r, 3, bill_to)
                    cls._data_cell(ws, r, 4, ship_to)
                    for c in range(5, 10):
                        cls._data_cell(ws, r, c, today_str)
                    cls._data_cell(ws, r, 10, so_num)
                    cls._data_cell(ws, r, 11, 'PICK')
                    cls._data_cell(ws, r, 12, '')
                    cls._data_cell(ws, r, 13, 'B2B')
                    r += 1

        cls._auto_width(ws)

    # ── Lines (TO) ────────────────────────────────────────────────────────────

    @classmethod
    def _write_lines_to(cls, wb, results, loc_lookup):
        """Sheet: Lines (TO) — one row per item, line no increments by 10000."""
        ws = wb.create_sheet('Lines (TO)')
        headers = [
            'Document No.', 'Line No.', 'Item No.', 'Quantity',
            'Unit of Measure', 'Qty. to Ship', 'Qty. to Receive',
            'Dimension Set ID', 'Transfer Price',
        ]
        for c, h in enumerate(headers, 1):
            cls._hdr_cell(ws, 1, c, h)

        r = 2
        current_to = None
        line_no = 0

        for res in results:
            all_rows = (res.regular_orders + res.pwp_orders +
                       res.tester_orders + res.gwp_orders + res.nonstock_orders)

            for item in all_rows:
                if not item.to or not item.to.startswith('TO/'):
                    continue

                if item.to != current_to:
                    current_to = item.to
                    line_no = 0
                line_no += 10000

                cls._data_cell(ws, r, 1, item.to)
                cls._data_cell(ws, r, 2, line_no)
                cls._data_cell(ws, r, 3, item.item_no)
                cls._data_cell(ws, r, 4, item.qty)
                cls._data_cell(ws, r, 5, 'Piece-1')
                cls._data_cell(ws, r, 6, '')
                cls._data_cell(ws, r, 7, '')
                cls._data_cell(ws, r, 8, '')
                cls._data_cell(ws, r, 9,
                    round(item.unit_price, 10) if item.unit_price else 0,
                    '#,##0.0000000000')
                r += 1

        cls._auto_width(ws)

    # ── Lines (SO) ────────────────────────────────────────────────────────────

    @classmethod
    def _write_lines_so(cls, wb, results, loc_lookup):
        """Sheet: Lines (SO) — one row per item within each SO number."""
        ws = wb.create_sheet('Lines (SO)')
        headers = [
            'Document Type', 'Document No.', 'Line No.', 'Type',
            'No.', 'Location Code', 'Quantity', 'Unit Price',
        ]
        for c, h in enumerate(headers, 1):
            cls._hdr_cell(ws, 1, c, h)

        r = 2
        current_so = None
        line_no = 0

        for res in results:
            all_rows = (res.regular_orders + res.pwp_orders +
                       res.tester_orders + res.gwp_orders + res.nonstock_orders)

            for item in all_rows:
                if not item.to or not item.to.startswith('SO/'):
                    continue

                if item.to != current_so:
                    current_so = item.to
                    line_no = 0
                line_no += 10000

                cls._data_cell(ws, r, 1, 'Order')
                cls._data_cell(ws, r, 2, item.to)
                cls._data_cell(ws, r, 3, line_no)
                cls._data_cell(ws, r, 4, 'Item')
                cls._data_cell(ws, r, 5, item.item_no)
                cls._data_cell(ws, r, 6, 'PICK')
                cls._data_cell(ws, r, 7, item.qty)
                cls._data_cell(ws, r, 8,
                    round(item.unit_price, 10) if item.unit_price else 0,
                    '#,##0.0000000000')
                r += 1

        cls._auto_width(ws)

    # ── Final Data ────────────────────────────────────────────────────────────

    @classmethod
    def _write_final_data(cls, wb, results):
        """Sheet: Final Data — all order rows with color-coded Source/Status."""
        ws = wb.create_sheet('Final Data')
        headers = [
            'TO', 'Item', 'Qty', 'Unit Price', 'Transfer-to Code',
            'Gen. Bus. Posting Group', 'Source', 'Location',
            'EAN', 'Product Name', 'Lookup Status',
        ]
        for c, h in enumerate(headers, 1):
            cls._hdr_cell(ws, 1, c, h)

        r = 2
        for res in results:
            loc = res.filename.replace('.xlsx', '').replace('_NEW_PO', '').replace('_New_PO', '')

            def write_row(item, row_num):
                cls._data_cell(ws, row_num, 1, item.to)
                cls._data_cell(ws, row_num, 2, item.item_no)
                cls._data_cell(ws, row_num, 3, item.qty)
                cls._data_cell(ws, row_num, 4,
                    round(item.unit_price, 10) if item.unit_price else 0,
                    '#,##0.0000000000')
                cls._data_cell(ws, row_num, 5, item.transfer_to)
                cls._data_cell(ws, row_num, 6, item.posting_group)

                # Source with color
                src_cell = cls._data_cell(ws, row_num, 7, item.source)
                sc = cls.SOURCE_COLORS.get(item.source, ('333333', 'FFFFFF'))
                src_cell.fill = PatternFill('solid', fgColor=sc[0])
                src_cell.font = Font(name='Aptos Display', size=11, bold=True, color=sc[1])
                src_cell.alignment = Alignment(horizontal='center')

                cls._data_cell(ws, row_num, 8, loc)
                cls._data_cell(ws, row_num, 9, item.ean)
                cls._data_cell(ws, row_num, 10, item.product_name)

                # Status with color
                st_cell = cls._data_cell(ws, row_num, 11, item.lookup_status)
                stc = cls.STATUS_COLORS.get(item.lookup_status, ('666666', 'FFFFFF'))
                st_cell.fill = PatternFill('solid', fgColor=stc[0])
                st_cell.font = Font(name='Aptos Display', size=11, bold=True, color=stc[1])
                st_cell.alignment = Alignment(horizontal='center')

                return row_num + 1

            for item in res.regular_orders: r = write_row(item, r)
            for item in res.pwp_orders: r = write_row(item, r)
            for item in res.tester_orders: r = write_row(item, r)
            for item in res.gwp_orders: r = write_row(item, r)
            for item in res.nonstock_orders: r = write_row(item, r)
            r += 1  # Separator between locations

        cls._auto_width(ws)
        ws.freeze_panes = 'A2'
        ws.auto_filter.ref = ws.dimensions

    # ── Summary ───────────────────────────────────────────────────────────────

    @classmethod
    def _write_summary(cls, wb, results):
        """Sheet: Summary — per-location qty breakdown with totals row."""
        ws = wb.create_sheet('Summary')
        headers = [
            'Location', 'PO Qty', 'PO Items', 'Tester Qty', 'Tester Items',
            'PWP Qty', 'GWP Qty', 'Non-Stock Qty', 'Total Qty', 'Unmatched EANs',
        ]
        for c, h in enumerate(headers, 1):
            cls._hdr_cell(ws, 1, c, h)

        for i, res in enumerate(results, 2):
            loc = res.filename.replace('.xlsx', '').replace('_NEW_PO', '').replace('_New_PO', '')
            po_q = sum(r.qty for r in res.regular_orders)
            tt_q = sum(r.qty for r in res.tester_orders)
            pw_q = sum(r.qty for r in res.pwp_orders)
            gw_q = sum(r.qty for r in res.gwp_orders)
            ns_q = sum(r.qty for r in res.nonstock_orders)
            total = po_q + tt_q + pw_q + gw_q + ns_q

            cls._data_cell(ws, i, 1, loc)
            cls._data_cell(ws, i, 2, po_q)
            cls._data_cell(ws, i, 3, len(res.regular_orders))
            cls._data_cell(ws, i, 4, tt_q)
            cls._data_cell(ws, i, 5, len(res.tester_orders))
            cls._data_cell(ws, i, 6, pw_q)
            cls._data_cell(ws, i, 7, gw_q)
            cls._data_cell(ws, i, 8, ns_q)
            cls._data_cell(ws, i, 9, total)
            cls._data_cell(ws, i, 10, len(res.unmatched))

            if res.unmatched:
                ws.cell(row=i, column=10).fill = PatternFill('solid', fgColor='FF5252')
                ws.cell(row=i, column=10).font = Font(
                    name='Aptos Display', size=11, bold=True, color='FFFFFF')

        # Totals row
        tr = len(results) + 2
        cls._data_cell(ws, tr, 1, 'TOTAL')
        ws.cell(row=tr, column=1).font = Font(name='Aptos Display', size=11, bold=True)
        for c in range(2, 11):
            total = sum(ws.cell(row=r, column=c).value or 0 for r in range(2, tr))
            cls._data_cell(ws, tr, c, total)
            ws.cell(row=tr, column=c).font = Font(name='Aptos Display', size=11, bold=True)

        cls._auto_width(ws)
        ws.freeze_panes = 'A2'

    # ── Unmatched EANs ────────────────────────────────────────────────────────

    @classmethod
    def _write_unmatched(cls, wb, results):
        """Sheet: Unmatched EANs — failed EAN lookups for manual fix."""
        ws = wb.create_sheet('Unmatched EANs')
        for c, h in enumerate(['Location', 'EAN', 'Product Name', 'Order Qty', 'Tester Qty'], 1):
            cls._hdr_cell(ws, 1, c, h)

        r = 2
        for res in results:
            loc = res.filename.replace('.xlsx', '')
            for u in res.unmatched:
                cls._data_cell(ws, r, 1, loc)
                cls._data_cell(ws, r, 2, u['ean'])
                cls._data_cell(ws, r, 3, u['product_name'])
                cls._data_cell(ws, r, 4, u['order_qty'])
                cls._data_cell(ws, r, 5, u['tester_qty'])
                r += 1

        if r == 2:
            cls._data_cell(ws, 2, 1, 'No unmatched EANs — all lookups resolved! ✓')
            ws.cell(row=2, column=1).font = Font(name='Aptos Display', size=11, color='00C853')
            ws.merge_cells('A2:E2')

        cls._auto_width(ws)

    # ── Tester Items Master ───────────────────────────────────────────────────

    @classmethod
    def _write_tester_master(cls, wb, results):
        """Sheet: Tester Items Master — PWP/GWP/Non-Stock resolution audit."""
        ws = wb.create_sheet('Tester Items Master')
        for c, h in enumerate(['Type', 'Product Name', 'EAN', 'Item No', 'Status', 'Used In Locations'], 1):
            cls._hdr_cell(ws, 1, c, h)

        # Collect unique items across all locations
        items_map = {}
        for res in results:
            loc = res.filename.replace('.xlsx', '').replace('_NEW_PO', '').replace('_New_PO', '')
            for item_list in [res.pwp_orders, res.gwp_orders, res.nonstock_orders]:
                for item in item_list:
                    key = (item.source, item.ean or item.product_name)
                    if key not in items_map:
                        items_map[key] = {
                            'source': item.source, 'name': item.product_name,
                            'ean': item.ean, 'item_no': item.item_no,
                            'status': item.lookup_status, 'locations': set(),
                        }
                    items_map[key]['locations'].add(loc)
                    if item.lookup_status == 'OK':
                        items_map[key]['status'] = 'OK'
                        items_map[key]['item_no'] = item.item_no

        TYPE_COLORS = {'PWP': 'FF6600', 'GWP': '00BCD4', 'NON_STOCK': '795548'}

        r = 2
        for key in sorted(items_map.keys(), key=lambda k: (k[0], k[1])):
            entry = items_map[key]
            is_ok = entry['status'] == 'OK'

            cls._data_cell(ws, r, 1, entry['source'])
            tc = TYPE_COLORS.get(entry['source'], '333333')
            ws.cell(row=r, column=1).fill = PatternFill('solid', fgColor=tc)
            ws.cell(row=r, column=1).font = Font(
                name='Aptos Display', size=11, bold=True, color='FFFFFF')

            cls._data_cell(ws, r, 2, entry['name'])
            cls._data_cell(ws, r, 3, entry['ean'])
            cls._data_cell(ws, r, 4, entry['item_no'] if is_ok else '')

            status_text = 'OK' if is_ok else entry['status']
            cls._data_cell(ws, r, 5, status_text)
            if is_ok:
                ws.cell(row=r, column=5).fill = PatternFill('solid', fgColor='00C853')
                ws.cell(row=r, column=5).font = Font(
                    name='Aptos Display', size=11, bold=True, color='000000')
            else:
                ws.cell(row=r, column=5).fill = PatternFill('solid', fgColor='FF5252')
                ws.cell(row=r, column=5).font = Font(
                    name='Aptos Display', size=11, bold=True, color='FFFFFF')

            cls._data_cell(ws, r, 6, ', '.join(sorted(entry['locations'])))
            r += 1

        if r == 2:
            cls._data_cell(ws, 2, 1, 'No PWP/GWP/Non-Stock items found')
            ws.merge_cells('A2:F2')

        cls._auto_width(ws)

    # ── SO Reference ──────────────────────────────────────────────────────────

    @classmethod
    def _write_so_reference(cls, wb, so_products, master):
        """Sheet: SO Reference — Special Order source data with calculated prices."""
        ws = wb.create_sheet('SO Reference')
        headers = [
            'Description', 'EAN', 'EBO Qty', 'Airport Qty', 'Kiosk Qty',
            'Tester Qty', 'Item No', 'MRP', 'GST Code',
            'Landing (×0.6)', 'Cost Price',
        ]
        calc_hdr_fill = PatternFill('solid', fgColor='1B5E20')
        for c, h in enumerate(headers, 1):
            cell = cls._hdr_cell(ws, 1, c, h)
            if c >= 7:
                cell.fill = calc_hdr_fill

        r = 2
        for prod in so_products:
            ean = prod['ean']
            info = master.get(ean) or master.get(ean.lstrip('0'))

            if info:
                item_no = info['item_no']
                mrp = info['mrp']
                gst_code = info.get('gst_code', '')
                description = info.get('description', '')
                landing = float(mrp) * 0.60 if mrp and not pd.isna(mrp) else 0
                cost_price = POEngine.calc_cost_price(mrp, gst_code) or 0
            else:
                item_no = f'?EAN:{ean}'
                mrp = ''
                gst_code = ''
                description = ''
                landing = cost_price = 0

            cls._data_cell(ws, r, 1, description)
            cls._data_cell(ws, r, 2, ean)
            cls._data_cell(ws, r, 3, prod.get('ebo_qty', 0))
            cls._data_cell(ws, r, 4, prod.get('airport_qty', 0))
            cls._data_cell(ws, r, 5, prod.get('kiosk_qty', 0))
            cls._data_cell(ws, r, 6, prod.get('tester_qty', 0))

            calc_fill = PatternFill('solid', fgColor='E8F5E9')
            for ci in range(7, 12):
                ws.cell(row=r, column=ci).fill = calc_fill

            cls._data_cell(ws, r, 7, item_no)
            cls._data_cell(ws, r, 8, mrp, '#,##0.00' if mrp else None)
            cls._data_cell(ws, r, 9, gst_code)
            cls._data_cell(ws, r, 10, round(landing, 2) if landing else 0, '#,##0.00')
            cls._data_cell(ws, r, 11,
                round(cost_price, 10) if cost_price else 0, '#,##0.0000000000')
            r += 1

        cls._auto_width(ws)
        ws.freeze_panes = 'A2'


# ═══════════════════════════════════════════════════════════════════════════════
#  SPECIAL ORDER ENGINE
# ═══════════════════════════════════════════════════════════════════════════════
#
# Unlike the standalone PO engine which reads individual files per location,
# the Special Order engine reads:
#   1. EKA_DATA.xlsx — locations with Prefix, Short Code, Transfer Code, Type
#   2. Special_Order.xlsx — products with per-type quantities
#   3. Items_March.xlsx — shared master (loaded by POEngine)
#
# Output: Same Final Data format as standalone, but with TO number,
# Transfer-to Code, and Gen. Biz. Posting Group PRE-FILLED from EKA_DATA.
#
# TO Number Pattern (v1.2 — same for both segments):
#     Regular: {Prefix}/{ShortCode}/{MM}/{DDMYY}
#     Tester:  {Prefix}/{ShortCode}/TT/{DDMYY + 1}
#
#   Example (April 18, 2026):
#     Regular: TO/CHNAP/04/18426
#     Tester:  TO/CHNAP/TT/18427   ← date_code + 1 for unique Excel search
#
# v1.2 changes:
#   - generate_to_number: uses MM (month) for regular, TT for tester
#   - tester date_code = regular date_code + 1
#   - load_eka_data: stores 'location' field for standalone filename matching

class SpecialOrderEngine:
    """
    Processes Special Order broadcasts across all EKA locations.

    Flow:
        1. load_eka_data()      → parse locations with metadata
        2. load_special_order() → parse products with per-type quantities
        3. validate()           → check structure, columns, EANs
        4. process()            → generate LocationResult per location
    """

    # ┌─────────────────────────────────────────────────────────────────────────┐
    # │ REQUIRED COLUMNS in EKA_DATA                                            │
    # │ These column names must exist (exact match) in the EKA_DATA file.       │
    # │                                                                         │
    # │ 'Gen. Biz. Posting Group' may have extra spaces in source file —       │
    # │ the loader uses fuzzy matching for this column.                          │
    # └─────────────────────────────────────────────────────────────────────────┘
    EKA_REQUIRED_COLS = [
        'Short Name',
        'Prefix',
        'Short Code',
        'Transfer Code',
        'Type',
        'Gen. Biz. Posting Group',
    ]

    # ┌─────────────────────────────────────────────────────────────────────────┐
    # │ REQUIRED COLUMNS in Special Order                                       │
    # │ Standard names with fallbacks (same alert pattern as standalone PO).    │
    # │                                                                         │
    # │ Standard          Fallback       Alert if fallback used                 │
    # │ ─────────────     ──────────     ──────────────────────                 │
    # │ EAN               (none)         ERROR if missing                       │
    # │ EBO Qty           EBO            Alert: rename to 'EBO Qty'            │
    # │ Airport Qty       Airport        Alert: rename to 'Airport Qty'        │
    # │ Kiosk Qty         Kiosk          Alert: rename to 'Kiosk Qty'          │
    # │ Tester Qty        Tester         Alert: rename to 'Tester Qty'         │
    # └─────────────────────────────────────────────────────────────────────────┘
    SO_STANDARD_COLS = {
        'ean':         ('EAN', []),
        'ebo_qty':     ('EBO Qty', ['EBO']),
        'airport_qty': ('Airport Qty', ['Airport']),
        'kiosk_qty':   ('Kiosk Qty', ['Kiosk']),
        'tester_qty':  ('Tester Qty', ['Tester', 'Tester']),
    }

    # ┌─────────────────────────────────────────────────────────────────────────┐
    # │ Type → qty column mapping                                               │
    # │ Determines which quantity column to use for each location type.         │
    # │                                                                         │
    # │ EBO     locations → ebo_qty column                                      │
    # │ Airport locations → airport_qty column                                  │
    # │ Kiosk   locations → kiosk_qty column                                    │
    # └─────────────────────────────────────────────────────────────────────────┘
    TYPE_QTY_MAP = {
        'EBO':     'ebo_qty',
        'Airport': 'airport_qty',
        'Kiosk':   'kiosk_qty',
    }

    def __init__(self, master: Dict[str, Dict]):
        """
        Args:
            master: The Items_March lookup dict (shared with POEngine).
                    Indexed by GTIN and by No.
        """
        self.master = master
        self.locations: List[Dict] = []       # Parsed from EKA_DATA
        self.products: List[Dict] = []        # Parsed from Special Order
        self.so_col_map: Dict[str, int] = {}  # Column name → index

    def _safe_int(self, val) -> int:
        """Safely convert cell value to int. Returns 0 for None/empty/errors."""
        try:
            if val is None or str(val).strip() in ('', '#N/A', 'None'):
                return 0
            return int(float(val))
        except (ValueError, TypeError):
            return 0

    def _ean_str(self, raw) -> str:
        """Convert raw EAN cell value to clean string."""
        if raw is None:
            return ''
        return str(int(raw)) if isinstance(raw, (int, float)) else str(raw).strip()

    # ┌─────────────────────────────────────────────────────────────────────────┐
    # │ TO NUMBER GENERATION (v1.2)                                             │
    # │                                                                         │
    # │ Pattern: {Prefix}/{ShortCode}/{Segment}/{DateCode}                      │
    # │                                                                         │
    # │ Segment:                                                                │
    # │   Regular → month number, zero-padded: '01', '02', ..., '12'           │
    # │   Tester  → literal 'TT'                                               │
    # │                                                                         │
    # │ DateCode: DD + M + YY                                                   │
    # │   DD = day (zero-padded)                                                │
    # │   M  = month (NOT zero-padded: 1-9 for Jan-Sep, 10-12 for Oct-Dec)     │
    # │   YY = last 2 digits of year                                            │
    # │                                                                         │
    # │ v1.2 CHANGE: Tester date_code = regular date_code + 1                  │
    # │   This ensures the last digits are unique when searching in Excel.      │
    # │                                                                         │
    # │ Examples (April 18, 2026):                                              │
    # │   Regular: TO/AHDEB/04/18426    ← date_code = 18426                    │
    # │   Tester:  TO/AHDEB/TT/18427   ← date_code = 18426 + 1 = 18427       │
    # │                                                                         │
    # │ This method is used by BOTH segments:                                   │
    # │   - Standalone: called when EKA_DATA filename match is found            │
    # │   - Special Order: called for every location during process()           │
    # └─────────────────────────────────────────────────────────────────────────┘

    @staticmethod
    def generate_to_number(prefix: str, short_code: str, is_tester: bool) -> str:
        """
        Generate TO/SO number for any mode.

        Args:
            prefix     : 'TO' or 'SO' (from EKA_DATA Prefix column)
            short_code : Location short code (e.g., 'AHDEB', 'CHNAP')
            is_tester  : True for tester, False for regular

        Returns:
            Full TO/SO number string.

        Examples:
            generate_to_number('TO', 'AHDEB', False)  → 'TO/AHDEB/04/18426'
            generate_to_number('TO', 'AHDEB', True)   → 'TO/AHDEB/TT/18427'
            generate_to_number('SO', 'PUNEB', False)  → 'SO/PUNEB/04/18426'
            generate_to_number('SO', 'PUNEB', True)   → 'SO/PUNEB/TT/18427'
        """
        from datetime import date
        today = date.today()

        dd = today.strftime('%d')           # '18' (zero-padded day)
        m = str(today.month)                 # '4'  (NOT zero-padded month)
        yy = today.strftime('%y')            # '26' (2-digit year)
        date_code = int(f"{dd}{m}{yy}")      # 18426

        if is_tester:
            # v1.2: increment by 1 for unique Excel search
            date_code += 1                   # 18426 → 18427
            segment = 'TT'
        else:
            segment = f"{today.month:02d}"   # '04' (zero-padded for TO number)

        return f"{prefix}/{short_code}/{segment}/{date_code}"

    def load_eka_data(self, filepath: str, logs: List[tuple]) -> int:
        """
        Load EKA_DATA.xlsx → parse locations with metadata.

        Required columns:
            Short Name, Prefix, Short Code, Transfer Code,
            Type, Gen. Biz. Posting Group

        Optional columns:
            Location (v1.2: for standalone filename matching)
            Bill to, Ship to (for SO headers)
            Status ('Active' or 'Inactive')

        Args:
            filepath: Path to EKA_DATA.xlsx
            logs:     List to append messages to

        Returns:
            Number of active locations loaded.
        """
        wb = load_workbook(filepath, data_only=True)
        ws = wb[wb.sheetnames[0]]

        # ── Build header map ──
        header_map = {}
        for cell in list(ws.iter_rows(min_row=1, max_row=1))[0]:
            val = str(cell.value or '').strip()
            if val:
                header_map[val] = cell.column - 1

        # ── Find required columns (with fuzzy match for Gen. Biz.) ──
        col_idx = {}
        for req in self.EKA_REQUIRED_COLS:
            if req in header_map:
                col_idx[req] = header_map[req]
            else:
                # Fuzzy match for columns with extra spaces
                for h, idx in header_map.items():
                    if (req.lower().replace(' ', '').replace('.', '') in
                            h.lower().replace(' ', '').replace('.', '')):
                        col_idx[req] = idx
                        break
                else:
                    logs.append(('error',
                        f"EKA_DATA: Column '{req}' not found. "
                        f"Available: {list(header_map.keys())}"))

        if len(col_idx) < len(self.EKA_REQUIRED_COLS):
            return 0

        # ── v1.2: Find optional 'Location' column for filename matching ──
        location_col_idx = None
        for h, idx in header_map.items():
            if h.strip().lower() == 'location':
                location_col_idx = idx
                break

        # ── Parse rows ──
        self.locations = []
        skipped_inactive = 0

        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=False):
            short_name = row[col_idx['Short Name']].value
            if not short_name or str(short_name).strip() == '':
                continue

            prefix = str(row[col_idx['Prefix']].value or 'TO').strip()
            short_code = str(row[col_idx['Short Code']].value or '').strip()
            transfer_code = str(row[col_idx['Transfer Code']].value or '').strip()
            loc_type = str(row[col_idx['Type']].value or '').strip()
            posting = str(row[col_idx['Gen. Biz. Posting Group']].value or '').strip()

            # v1.2: Read Location column
            location_code = ''
            if location_col_idx is not None:
                location_code = str(row[location_col_idx].value or '').strip()

            # Read optional columns (Bill to, Ship to, Status)
            bill_to = ''
            ship_to = ''
            status = 'Active'
            for h, idx in header_map.items():
                h_lower = h.lower().strip()
                if 'bill' in h_lower and 'to' in h_lower:
                    bill_to = str(row[idx].value or '').strip()
                elif 'ship' in h_lower and 'to' in h_lower:
                    ship_to = str(row[idx].value or '').strip()
                elif h_lower == 'status':
                    status = str(row[idx].value or 'Active').strip()

            # Skip Inactive locations
            if status.lower() == 'inactive':
                skipped_inactive += 1
                logs.append(('info',
                    f"EKA_DATA: '{str(short_name).strip()}' Inactive — skipping"))
                continue

            if not short_code:
                logs.append(('warn',
                    f"EKA_DATA: '{short_name}' has no Short Code — skipping"))
                continue

            self.locations.append({
                'short_name': str(short_name).strip(),
                'prefix': prefix,
                'short_code': short_code,
                'transfer_code': transfer_code,
                'location': location_code,      # v1.2: for filename matching
                'type': loc_type,
                'posting_group': posting,
                'bill_to': bill_to,
                'ship_to': ship_to,
            })

        if skipped_inactive:
            logs.append(('info',
                f"EKA_DATA: Skipped {skipped_inactive} Inactive location(s)"))
        logs.append(('info',
            f"EKA_DATA: Loaded {len(self.locations)} Active locations"))

        return len(self.locations)

    def _detect_so_columns(self, ws, logs: List[tuple]) -> Dict[str, int]:
        """
        Detect Special Order columns with fallback + alert.

        Same pattern as PO sheet column detection:
            'EAN'         → (no fallback, error if missing)
            'EBO Qty'     → fallback: 'EBO'
            'Airport Qty' → fallback: 'Airport'
            'Kiosk Qty'   → fallback: 'Kiosk'
            'Tester Qty'  → fallback: 'Tester'
        """
        hmap = {}
        all_headers = {}

        for cell in list(ws.iter_rows(min_row=1, max_row=1))[0]:
            val = str(cell.value or '').strip()
            idx = cell.column - 1
            if val:
                all_headers[val] = idx

        for key, (standard, fallbacks) in self.SO_STANDARD_COLS.items():
            if standard in all_headers:
                hmap[key] = all_headers[standard]
            else:
                for fb in fallbacks:
                    if fb in all_headers:
                        hmap[key] = all_headers[fb]
                        logs.append(('alert',
                            f"Auto-fixed: '{fb}' → '{standard}'. "
                            f"Please rename column to '{standard}'."))
                        break

        return hmap

    def load_special_order(self, filepath: str, logs: List[tuple]) -> int:
        """
        Load Special_Order.xlsx → parse products with per-type quantities.

        Args:
            filepath: Path to Special Order Excel file
            logs:     List to append messages to

        Returns:
            Number of products loaded.
        """
        wb = load_workbook(filepath, data_only=True)
        ws = wb[wb.sheetnames[0]]

        self.so_col_map = self._detect_so_columns(ws, logs)

        # Validate required columns
        if 'ean' not in self.so_col_map:
            logs.append(('error', "Special Order: 'EAN' not found — cannot process"))
            return 0

        missing_qty = []
        for key in ('ebo_qty', 'airport_qty', 'kiosk_qty', 'tester_qty'):
            if key not in self.so_col_map:
                std_name = self.SO_STANDARD_COLS[key][0]
                missing_qty.append(std_name)

        if missing_qty:
            logs.append(('error',
                f"Special Order: Missing columns: {', '.join(missing_qty)}"))
            return 0

        # Parse products
        self.products = []
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=False):
            ean_raw = row[self.so_col_map['ean']].value
            if ean_raw is None:
                continue

            ean = self._ean_str(ean_raw)
            ebo_qty = self._safe_int(row[self.so_col_map['ebo_qty']].value)
            airport_qty = self._safe_int(row[self.so_col_map['airport_qty']].value)
            kiosk_qty = self._safe_int(row[self.so_col_map['kiosk_qty']].value)
            tester_qty = self._safe_int(row[self.so_col_map['tester_qty']].value)

            self.products.append({
                'ean': ean,
                'ebo_qty': ebo_qty,
                'airport_qty': airport_qty,
                'kiosk_qty': kiosk_qty,
                'tester_qty': tester_qty,
            })

        logs.append(('info', f"Special Order: Loaded {len(self.products)} products"))
        return len(self.products)

    def validate(self, logs: List[tuple]) -> bool:
        """
        Validate EANs against master and location types.

        Returns True if no blocking errors.
        """
        if not self.locations:
            logs.append(('error', "No locations loaded from EKA_DATA"))
            return False
        if not self.products:
            logs.append(('error', "No products loaded from Special Order"))
            return False

        # Check EANs against master
        missing = 0
        for prod in self.products:
            ean = prod['ean']
            info = self.master.get(ean) or self.master.get(ean.lstrip('0'))
            if not info:
                logs.append(('warn', f"Special Order: EAN {ean} not found in master"))
                missing += 1

        if missing:
            logs.append(('warn',
                f"Special Order: {missing} EAN(s) not found — will output with ?EAN:"))
        else:
            logs.append(('info',
                f"Special Order: All {len(self.products)} EANs found in master ✓"))

        # Validate location types
        valid_types = set(self.TYPE_QTY_MAP.keys())
        for loc in self.locations:
            if loc['type'] not in valid_types:
                logs.append(('warn',
                    f"EKA_DATA: '{loc['short_name']}' has unknown Type "
                    f"'{loc['type']}' — will skip"))

        return True

    def process(self, logs: List[tuple]) -> List[LocationResult]:
        """
        Process Special Order → generate LocationResult per location.

        For each location × each product:
            - Regular row: qty from EBO/Airport/Kiosk column based on Type
            - Tester row: qty from Tester column at ₹0.54
            - TO number auto-generated (v1.2: tester date_code + 1)
            - Transfer-to Code and Gen. Biz. Posting Group pre-filled

        Returns:
            List of LocationResult, one per location.
        """
        results = []

        for loc in self.locations:
            loc_type = loc['type']
            qty_key = self.TYPE_QTY_MAP.get(loc_type)
            if not qty_key:
                logs.append(('warn',
                    f"Skipping '{loc['short_name']}' — unknown Type '{loc_type}'"))
                continue

            # Generate TO numbers (v1.2: tester gets date_code + 1)
            to_regular = self.generate_to_number(
                loc['prefix'], loc['short_code'], is_tester=False)
            to_tester = self.generate_to_number(
                loc['prefix'], loc['short_code'], is_tester=True)

            res = LocationResult(filename=loc['short_name'])
            unmatched = []

            for prod in self.products:
                ean = prod['ean']
                regular_qty = prod[qty_key]
                tester_qty = prod['tester_qty']

                # Master lookup
                info = self.master.get(ean) or self.master.get(ean.lstrip('0'))

                if info:
                    item_no = info['item_no']
                    cost = POEngine.calc_cost_price(info['mrp'], info['gst_code'])
                    product_name = str(info.get('description', ''))
                    status = 'OK'
                else:
                    item_no = f'?EAN:{ean}'
                    cost = None
                    product_name = ''
                    status = 'NOT_FOUND'
                    unmatched.append({
                        'ean': ean, 'product_name': '',
                        'order_qty': regular_qty, 'tester_qty': tester_qty,
                    })

                # Regular order row
                if regular_qty > 0:
                    res.regular_orders.append(OutputRow(
                        to=to_regular,
                        item_no=item_no,
                        qty=regular_qty,
                        unit_price=cost or 0,
                        transfer_to=loc['transfer_code'],
                        posting_group=loc['posting_group'],
                        source='PO',
                        ean=ean,
                        product_name=product_name,
                        lookup_status=status,
                    ))

                # Tester row
                if tester_qty > 0:
                    res.tester_orders.append(OutputRow(
                        to=to_tester,
                        item_no=item_no,
                        qty=tester_qty,
                        unit_price=0.54,
                        transfer_to=loc['transfer_code'],
                        posting_group=loc['posting_group'],
                        source='TESTER',
                        ean=ean,
                        product_name=product_name,
                        lookup_status=status,
                    ))

            res.unmatched = unmatched
            po_q = sum(r.qty for r in res.regular_orders)
            tt_q = sum(r.qty for r in res.tester_orders)
            res.logs.append(('info',
                f"PO: {len(res.regular_orders)} items ({po_q} qty), "
                f"Testers: {len(res.tester_orders)} items ({tt_q} qty)"))

            results.append(res)

        logs.append(('info',
            f"Special Order: Generated {len(results)} location results"))
        return results


# ═══════════════════════════════════════════════════════════════════════════════
#  D365 TRANSFER ORDER EXPORTER — NEW in v1.2
# ═══════════════════════════════════════════════════════════════════════════════
#
# Fills a D365 Transfer Order template with processed data.
# Uses the same ZIP/XML regex approach as the GT Mass D365 SO exporter.
#
# Template format (from EKA_Sample_Package.xlsx):
#     Sheet 1 'Transfer Header':
#         Row 1: metadata (e.g., 'TO - MILAN', 'Transfer Header', '5740')
#         Row 2: (empty)
#         Row 3: column headers
#         Row 4+: pre-formatted empty data rows
#
#     Sheet 2 'Transfer Line':
#         Row 1: metadata
#         Row 2: (empty)
#         Row 3: column headers
#         Row 4+: pre-formatted empty data rows
#
# If data exceeds template row capacity, new <row> elements are injected
# into the XML before filling (same inject_row technique as GT Mass).

class D365TOExporter:
    """
    Fills a D365 Transfer Order template via ZIP/XML manipulation.

    Template must have:
        - Sheet 1 named 'Transfer Header' with data starting at row 4
        - Sheet 2 named 'Transfer Line' with data starting at row 4
    """

    @staticmethod
    def export(results: List[LocationResult],
               template_path: str,
               output_path: str) -> str:
        """
        Fill the D365 TO template with processed data.

        Args:
            results       : List[LocationResult] with processed data
            template_path : Path to the D365 TO template .xlsx
            output_path   : Where to save the filled file

        Returns:
            output_path on success.
        """
        import re as re_mod

        shutil.copy2(template_path, output_path)
        today_str = time.strftime("%d-%m-%Y")

        # ── Collect unique TO/SO numbers with metadata ──
        unique_tos = []     # (to_number, transfer_to_code, posting_group)
        seen = set()

        for res in results:
            all_rows = (
                res.regular_orders + res.pwp_orders +
                res.tester_orders + res.gwp_orders + res.nonstock_orders
            )
            for item in all_rows:
                if item.to and item.to not in seen:
                    seen.add(item.to)
                    unique_tos.append((
                        item.to,
                        item.transfer_to,
                        item.posting_group,
                    ))

        # ── Collect all line items in order ──
        all_lines = []      # (doc_no, item_no, qty, unit_price)

        for res in results:
            all_rows = (
                res.regular_orders + res.pwp_orders +
                res.tester_orders + res.gwp_orders + res.nonstock_orders
            )
            for item in all_rows:
                if item.to:
                    all_lines.append((
                        item.to,
                        item.item_no,
                        item.qty,
                        item.unit_price,
                    ))

        # ── Read ZIP contents ──
        zip_contents = {}
        with zipfile.ZipFile(output_path, 'r') as z:
            for zi in z.namelist():
                zip_contents[zi] = z.read(zi)

        # ── Extend sharedStrings.xml ──
        ss_xml = zip_contents['xl/sharedStrings.xml'].decode('utf-8')
        existing = re_mod.findall(r'<t[^>]*>([^<]*)</t>', ss_xml)
        string_map = {s: i for i, s in enumerate(existing)}

        # Collect all new strings we'll need
        new_strings = {'PICK', 'IN TRANSIT', 'false', today_str, 'Piece-1'}
        for to_num, tc, pg in unique_tos:
            new_strings.add(to_num)
            if tc:
                new_strings.add(tc)
            if pg:
                new_strings.add(pg)

        # Assign indices to new strings
        next_idx = len(existing)
        for s in sorted(new_strings):
            if s not in string_map:
                string_map[s] = next_idx
                next_idx += 1

        # Rebuild sharedStrings XML
        total_count = next_idx
        si_items = [''] * total_count
        for s, idx in string_map.items():
            esc = (s.replace('&', '&amp;')
                    .replace('<', '&lt;')
                    .replace('>', '&gt;'))
            si_items[idx] = f'<si><t>{esc}</t></si>'

        zip_contents['xl/sharedStrings.xml'] = (
            '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>\r\n'
            f'<sst xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main" '
            f'count="{total_count}" uniqueCount="{total_count}">'
            + ''.join(si_items) + '</sst>'
        ).encode('utf-8')

        # ── Helper: fill a cell in XML ──
        def fill_cell(xml, col, row_num, value, is_string=True):
            """
            Replace a cell (empty or pre-filled) with new data.

            Handles two XML patterns:
                Empty cell:    <c r="A4" s="5"/>
                Pre-filled:   <c r="A4" s="5" t="s"><v>31</v></c>
                               <c r="A4" s="5"><v>123</v></c>
            """
            ref = f"{col}{row_num}"

            if is_string:
                idx = string_map.get(str(value), 0)
                new_cell = f'<c r="{ref}" s="\\1" t="s"><v>{idx}</v></c>'
            else:
                new_cell = f'<c r="{ref}" s="\\1"><v>{value}</v></c>'

            # Try 1: match pre-filled cell with value (t="s" or plain)
            pat_filled = f'<c r="{ref}" s="(\\d+)"[^/]*>.*?</c>'
            result = re_mod.sub(pat_filled, new_cell, xml, count=1, flags=re_mod.DOTALL)
            if result != xml:
                return result

            # Try 2: match empty self-closing cell
            pat_empty = f'<c r="{ref}" s="(\\d+)"\\s*/>'
            return re_mod.sub(pat_empty, new_cell, xml, count=1)

        # ── Helper: inject a new row when template has fewer rows than data ──
        def inject_row(xml, row_num, columns, style_id):
            """Add an empty row element before </sheetData>."""
            cells = ''.join(
                f'<c r="{c}{row_num}" s="{style_id}"/>' for c in columns
            )
            new_row = (
                f'<row r="{row_num}" spans="1:{len(columns)}" '
                f'x14ac:dyDescent="0.3">{cells}</row>'
            )
            return xml.replace('</sheetData>', new_row + '</sheetData>')

        # ── Sheet 1: Transfer Header ──
        s1 = zip_contents['xl/worksheets/sheet1.xml'].decode('utf-8')
        s1_existing = re_mod.findall(r'<row r="(\d+)"', s1)
        s1_max_row = max(int(r) for r in s1_existing) if s1_existing else 3
        hdr_cols = list('ABCDEFGHIJKL')  # 12 columns

        # Inject rows ONLY beyond the last existing template row
        last_needed_row = 3 + len(unique_tos)  # data starts at row 4
        if last_needed_row > s1_max_row:
            for extra in range(s1_max_row + 1, last_needed_row + 1):
                s1 = inject_row(s1, extra, hdr_cols, '11')

        # Fill header rows (data starts at row 4)
        for i, (to_num, tc, pg) in enumerate(unique_tos):
            r = i + 4
            s1 = fill_cell(s1, 'A', r, to_num)           # No.
            s1 = fill_cell(s1, 'B', r, 'PICK')            # Transfer-from Code
            s1 = fill_cell(s1, 'C', r, tc or '')           # Transfer-to Code
            s1 = fill_cell(s1, 'D', r, today_str)          # Posting Date
            s1 = fill_cell(s1, 'E', r, 'IN TRANSIT')       # In-Transit Code
            s1 = fill_cell(s1, 'F', r, 'false')            # Direct Transfer
            if pg:
                s1 = fill_cell(s1, 'G', r, pg)             # Gen. Bus. Posting Group

        zip_contents['xl/worksheets/sheet1.xml'] = s1.encode('utf-8')

        # ── Sheet 2: Transfer Line ──
        s2 = zip_contents['xl/worksheets/sheet2.xml'].decode('utf-8')
        s2_existing = re_mod.findall(r'<row r="(\d+)"', s2)
        s2_max_row = max(int(r) for r in s2_existing) if s2_existing else 3
        line_cols = list('ABCDEFGHI')  # 9 columns

        # Inject rows ONLY beyond the last existing template row
        last_needed_line = 3 + len(all_lines)
        if last_needed_line > s2_max_row:
            for extra in range(s2_max_row + 1, last_needed_line + 1):
                s2 = inject_row(s2, extra, line_cols, '8')

        # Fill line rows
        current_doc = None
        line_no = 0
        for i, (doc_no, item_no, qty, price) in enumerate(all_lines):
            # Reset line counter per document
            if doc_no != current_doc:
                current_doc = doc_no
                line_no = 0
            line_no += 10000

            r = i + 4
            s2 = fill_cell(s2, 'A', r, doc_no)                          # Document No.
            s2 = fill_cell(s2, 'B', r, line_no, is_string=False)        # Line No.

            # Item No. — try numeric first, fallback to string
            try:
                s2 = fill_cell(s2, 'C', r, int(str(item_no)), is_string=False)
            except (ValueError, TypeError):
                s2 = fill_cell(s2, 'C', r, str(item_no))

            s2 = fill_cell(s2, 'D', r, qty, is_string=False)            # Quantity
            s2 = fill_cell(s2, 'E', r, 'Piece-1')                        # Unit of Measure
            # F: Qty. to Ship (empty)
            # G: Qty. to Receive (empty)
            # H: Dimension Set ID (empty)
            s2 = fill_cell(s2, 'I', r,                                  # Transfer Price
                price if price else 0, is_string=False)

        zip_contents['xl/worksheets/sheet2.xml'] = s2.encode('utf-8')

        # ── Cleanup: remove unused rows, update dimensions ──
        last_hdr = 3 + len(unique_tos)
        last_line = 3 + len(all_lines)

        # Clean Transfer Header
        s1c = zip_contents['xl/worksheets/sheet1.xml'].decode('utf-8')
        for r in range(last_hdr + 1, 200):
            s1c = re_mod.sub(
                rf'<row r="{r}"[^>]*>.*?</row>', '', s1c, flags=re_mod.DOTALL)
        s1c = re_mod.sub(
            r'<dimension ref="[^"]*"/>',
            f'<dimension ref="A1:L{last_hdr}"/>', s1c)
        zip_contents['xl/worksheets/sheet1.xml'] = s1c.encode('utf-8')

        # Clean Transfer Line
        s2c = zip_contents['xl/worksheets/sheet2.xml'].decode('utf-8')
        for r in range(last_line + 1, 1000):
            s2c = re_mod.sub(
                rf'<row r="{r}"[^>]*>.*?</row>', '', s2c, flags=re_mod.DOTALL)
        s2c = re_mod.sub(
            r'<dimension ref="[^"]*"/>',
            f'<dimension ref="A1:I{last_line}"/>', s2c)
        zip_contents['xl/worksheets/sheet2.xml'] = s2c.encode('utf-8')

        # Update table refs if they exist
        for tbl in ['xl/tables/table1.xml', 'xl/tables/table2.xml']:
            if tbl in zip_contents:
                t = zip_contents[tbl].decode('utf-8')
                if 'table1' in tbl:
                    t = re_mod.sub(
                        r'ref="A3:[A-Z]+\d+"', f'ref="A3:L{last_hdr}"', t)
                else:
                    t = re_mod.sub(
                        r'ref="A3:[A-Z]+\d+"', f'ref="A3:I{last_line}"', t)
                zip_contents[tbl] = t.encode('utf-8')

        # ── Write final ZIP ──
        with zipfile.ZipFile(output_path, 'w', zipfile.ZIP_DEFLATED) as zo:
            for name, data in zip_contents.items():
                zo.writestr(name, data)

        return output_path


# ═══════════════════════════════════════════════════════════════════════════════
#  GUI APPLICATION — v1.2
# ═══════════════════════════════════════════════════════════════════════════════
#
# ReneePOApp is the Tkinter main window. It handles:
#   - Master file loading (Items_March.xlsx)
#   - EKA_DATA loading (v1.2: shared across both modes)
#   - PO file selection (individual files or entire folder)
#   - Blank template download for store teams
#   - Two-phase processing: VALIDATE → EXTRACT (in background thread)
#   - v1.2: Standalone auto-fill from EKA_DATA via filename
#   - v1.2: D365 TO package export button
#   - Dark/Light theme toggling with live widget updates
#   - Log panel showing all validation + extraction messages
#   - Stats panel showing last run quantity totals
#
# Thread safety: Processing runs in _process_worker() on a daemon thread.
# All GUI updates from the worker are dispatched via self.after(0, ...).
#
# Theme system: Every widget is registered via _reg_theme(widget, role).
# On toggle, _toggle_theme() iterates all registered widgets and applies
# new colors based on their role string.

class ReneePOApp(tk.Tk):
    """
    Main application window.

    v1.2 additions:
        - EKA_DATA in shared section (above mode toggle)
        - eka_locations stored for both modes
        - Standalone auto-fill: filename → EKA Location → TO/Transfer/Posting
        - D365 TO export button in bottom bar
        - last_results stored for D365 export
    """

    def __init__(self):
        super().__init__()
        self.title("RENEE PO Processor v1.2")
        self.geometry("1020x760")
        self.resizable(True, True)
        self.configure(bg=Theme.bg())
        self.minsize(860, 620)

        # ── Segment 1: Standalone PO state ──
        self.master_path: Optional[str] = None
        self.po_files: List[str] = []
        self.last_output: Optional[str] = None
        self.is_running = False
        self.engine = POEngine()

        # ── v1.2: Shared EKA state (used by both modes) ──
        self.eka_path: Optional[str] = None
        self.eka_locations: List[Dict] = []

        # ── Segment 2: Special Order state ──
        self.so_path: Optional[str] = None
        self.active_mode = 'standalone'

        # ── v1.2: Store last results for D365 export ──
        self.last_results: List[LocationResult] = []

        # Track themed widgets for live re-theming
        self._themed_widgets: List[tuple] = []

        self._build_ui()

    # ── THEME MANAGEMENT ───────────────────────────────────────────────────────

    def _reg_theme(self, widget, role: str):
        """Register widget for theme updates."""
        self._themed_widgets.append((widget, role))
        return widget

    def _toggle_theme(self):
        """Switch dark↔light and re-theme entire UI."""
        Theme.toggle()
        self.configure(bg=Theme.bg())

        role_map = {
            'bg':       lambda: {'bg': Theme.bg()},
            'surface':  lambda: {'bg': Theme.surface()},
            'surface2': lambda: {'bg': Theme.surface2()},
            'border':   lambda: {'highlightbackground': Theme.border()},
            'bg+text':       lambda: {'bg': Theme.bg(), 'fg': Theme.text()},
            'bg+text_dim':   lambda: {'bg': Theme.bg(), 'fg': Theme.text_dim()},
            'bg+accent':     lambda: {'bg': Theme.bg(), 'fg': Theme.accent()},
            'surface+text':  lambda: {'bg': Theme.surface(), 'fg': Theme.text()},
            'surface+text_dim': lambda: {'bg': Theme.surface(), 'fg': Theme.text_dim()},
            'surface+green': lambda: {'bg': Theme.surface(), 'fg': Theme.green()},
            'surface+accent2': lambda: {'bg': Theme.surface(), 'fg': Theme.accent2()},
            'surface+pink':  lambda: {'bg': Theme.surface(), 'fg': Theme.pink()},
            'surface2+text_dim': lambda: {'bg': Theme.surface2(), 'fg': Theme.text_dim()},
            'surface2+accent2': lambda: {'bg': Theme.surface2(), 'fg': Theme.accent2()},
            'listbox':  lambda: {'bg': Theme.surface(), 'fg': Theme.text(),
                                 'selectbackground': Theme.list_sel(),
                                 'selectforeground': Theme.accent()},
            'log':      lambda: {'bg': Theme.surface(), 'fg': Theme.text_dim()},
            'progress': lambda: {'bg': Theme.surface2()},
        }

        for widget, role in self._themed_widgets:
            try:
                fn = role_map.get(role)
                if fn:
                    widget.config(**fn())
            except tk.TclError:
                pass

        if hasattr(self, 'theme_toggle'):
            self.theme_toggle.config(bg=Theme.surface())
            self.theme_toggle._draw()
        apply_style()

    # ── UI CONSTRUCTION ────────────────────────────────────────────────────────

    def _build_ui(self):
        """Build the complete UI: header, left panel, right panel, bottom bar."""

        # ── Header bar ──
        hdr = tk.Frame(self, bg=Theme.surface(), height=60)
        hdr.pack(fill='x')
        hdr.pack_propagate(False)
        self._reg_theme(hdr, 'surface')

        title_lbl = tk.Label(hdr, text="▶ RENEE PO PROCESSOR v1.2",
                             font=FONT_TITLE, bg=Theme.surface(), fg=Theme.pink())
        title_lbl.pack(side='left', padx=20, pady=12)
        self._reg_theme(title_lbl, 'surface+pink')

        sub_lbl = tk.Label(hdr,
            text="EBO / Kiosk / Airport PO → Final Data  //  EKA Script",
            font=FONT_SUB, bg=Theme.surface(), fg=Theme.text_dim())
        sub_lbl.pack(side='left', padx=4)
        self._reg_theme(sub_lbl, 'surface+text_dim')

        # Theme toggle
        toggle_frame = tk.Frame(hdr, bg=Theme.surface())
        toggle_frame.pack(side='right', padx=16)
        self._reg_theme(toggle_frame, 'surface')
        tk.Label(toggle_frame, text="🌙", font=("Segoe UI Emoji", 11),
                 bg=Theme.surface()).pack(side='left', padx=(0, 4))
        self.theme_toggle = ToggleSwitch(toggle_frame, command=self._toggle_theme,
                                          width=52, height=26, bg=Theme.surface())
        self.theme_toggle.pack(side='left')
        tk.Label(toggle_frame, text="☀️", font=("Segoe UI Emoji", 11),
                 bg=Theme.surface()).pack(side='left', padx=(4, 0))

        # ── Body (left + right panels) ──
        body = tk.Frame(self, bg=Theme.bg())
        body.pack(fill='both', expand=True, padx=16, pady=12)
        body.columnconfigure(0, weight=3)
        body.columnconfigure(1, weight=2)
        body.rowconfigure(0, weight=1)
        self._reg_theme(body, 'bg')

        left = tk.Frame(body, bg=Theme.bg())
        right = tk.Frame(body, bg=Theme.bg())
        left.grid(row=0, column=0, sticky='nsew', padx=(0, 8))
        right.grid(row=0, column=1, sticky='nsew')
        self._reg_theme(left, 'bg')
        self._reg_theme(right, 'bg')

        self._build_left(left)
        self._build_right(right)
        self._build_bottom()

    def _build_left(self, parent):
        """Build left panel: Master, EKA_DATA, mode selector, standalone/special panels."""

        # ── SHARED: Master File ──
        self._section(parent, "00  //  MASTER FILE (Items_March)")

        master_frame = tk.Frame(parent, bg=Theme.surface(),
                                highlightthickness=1, highlightbackground=Theme.border())
        master_frame.pack(fill='x', pady=(0, 10))
        self._reg_theme(master_frame, 'surface')
        self._reg_theme(master_frame, 'border')

        master_inner = tk.Frame(master_frame, bg=Theme.surface())
        master_inner.pack(fill='x', padx=12, pady=10)
        self._reg_theme(master_inner, 'surface')

        self.master_var = tk.StringVar(value="No master file loaded")
        self.master_count_var = tk.StringVar(value="")

        top_row = tk.Frame(master_inner, bg=Theme.surface())
        top_row.pack(fill='x')
        self._reg_theme(top_row, 'surface')
        self._btn(top_row, "📂 SELECT MASTER XLSX", self._select_master, ACCENT).pack(side='left')

        mc_lbl = tk.Label(top_row, textvariable=self.master_count_var,
                          font=FONT_MONO, bg=Theme.surface(), fg=GREEN)
        mc_lbl.pack(side='right')
        self._reg_theme(mc_lbl, 'surface+green')

        mv_lbl = tk.Label(master_inner, textvariable=self.master_var,
                          font=FONT_MONO, bg=Theme.surface(), fg=Theme.text_dim(),
                          wraplength=400, anchor='w', justify='left')
        mv_lbl.pack(anchor='w', pady=(6, 0))
        self._reg_theme(mv_lbl, 'surface+text_dim')

        # ── SHARED: EKA_DATA (v1.2: moved here from Special Order panel) ──
        self._section(parent, "00b  //  EKA DATA (Location Registry — shared)")

        eka_frame = tk.Frame(parent, bg=Theme.surface(),
                             highlightthickness=1, highlightbackground=Theme.border())
        eka_frame.pack(fill='x', pady=(0, 10))
        self._reg_theme(eka_frame, 'surface')
        self._reg_theme(eka_frame, 'border')

        eka_inner = tk.Frame(eka_frame, bg=Theme.surface())
        eka_inner.pack(fill='x', padx=12, pady=10)
        self._reg_theme(eka_inner, 'surface')

        self.eka_var = tk.StringVar(
            value="No EKA_DATA loaded (optional for standalone, required for special)")
        self.eka_count_var = tk.StringVar(value="")

        eka_top = tk.Frame(eka_inner, bg=Theme.surface())
        eka_top.pack(fill='x')
        self._reg_theme(eka_top, 'surface')
        self._btn(eka_top, "📂 SELECT EKA_DATA", self._select_eka, ACCENT2).pack(side='left')

        eka_c = tk.Label(eka_top, textvariable=self.eka_count_var,
                         font=FONT_MONO, bg=Theme.surface(), fg=GREEN)
        eka_c.pack(side='right')
        self._reg_theme(eka_c, 'surface+green')

        eka_l = tk.Label(eka_inner, textvariable=self.eka_var,
                         font=FONT_MONO, bg=Theme.surface(), fg=Theme.text_dim(),
                         wraplength=400, anchor='w')
        eka_l.pack(anchor='w', pady=(6, 0))
        self._reg_theme(eka_l, 'surface+text_dim')

        # ── MODE SELECTOR ──
        mode_frame = tk.Frame(parent, bg=Theme.bg())
        mode_frame.pack(fill='x', pady=(6, 4))
        self._reg_theme(mode_frame, 'bg')

        self.btn_standalone = tk.Label(mode_frame, text="▸ STANDALONE PO FILES",
            font=FONT_LABEL, bg=Theme.accent(), fg=Theme.bg(),
            cursor='hand2', padx=12, pady=5,
            highlightthickness=1, highlightbackground=Theme.accent())
        self.btn_standalone.pack(side='left', padx=(0, 4))
        self.btn_standalone.bind('<Button-1>', lambda e: self._switch_mode('standalone'))

        self.btn_special = tk.Label(mode_frame, text="▸ SPECIAL ORDER",
            font=FONT_LABEL, bg=Theme.surface2(), fg=Theme.text_dim(),
            cursor='hand2', padx=12, pady=5,
            highlightthickness=1, highlightbackground=Theme.border())
        self.btn_special.pack(side='left', padx=(0, 4))
        self.btn_special.bind('<Button-1>', lambda e: self._switch_mode('special'))

        # ══════════════════════════════════════════════════════════════════════
        #  SEGMENT 1: STANDALONE PO FILES
        # ══════════════════════════════════════════════════════════════════════

        self.frame_standalone = tk.Frame(parent, bg=Theme.bg())
        self.frame_standalone.pack(fill='both', expand=True)
        self._reg_theme(self.frame_standalone, 'bg')

        # ── Drop Zone ──
        dz = tk.Frame(self.frame_standalone, bg=Theme.surface2(),
                      highlightthickness=1, highlightbackground=Theme.border())
        dz.pack(fill='x', pady=(0, 8))
        self._reg_theme(dz, 'surface2')
        self._reg_theme(dz, 'border')

        dz_inner = tk.Frame(dz, bg=Theme.surface2())
        dz_inner.pack(fill='x', padx=1, pady=1)
        self._reg_theme(dz_inner, 'surface2')

        dz_lbl1 = tk.Label(dz_inner, text="DROP ZONE",
            font=("Aptos Display", 11, "bold"), bg=Theme.surface2(), fg=ACCENT2)
        dz_lbl1.pack(pady=(10, 2))
        self._reg_theme(dz_lbl1, 'surface2+accent2')

        dz_lbl2 = tk.Label(dz_inner,
            text="Add PO files (rename to Location code, e.g. EBO_AMD01.xlsx)",
            font=FONT_MONO, bg=Theme.surface2(), fg=Theme.text_dim())
        dz_lbl2.pack(pady=(0, 10))
        self._reg_theme(dz_lbl2, 'surface2+text_dim')

        btn_row = tk.Frame(dz_inner, bg=Theme.surface2())
        btn_row.pack(pady=(0, 4))
        self._reg_theme(btn_row, 'surface2')
        self._btn(btn_row, "+ ADD FILES", self._add_files, ACCENT).pack(side='left', padx=4)
        self._btn(btn_row, "+ ADD FOLDER", self._add_folder, ACCENT2).pack(side='left', padx=4)
        self._btn(btn_row, "✕ CLEAR ALL", self._clear_files, RED).pack(side='left', padx=4)

        btn_row2 = tk.Frame(dz_inner, bg=Theme.surface2())
        btn_row2.pack(pady=(0, 10))
        self._reg_theme(btn_row2, 'surface2')
        self._btn(btn_row2, "📋 DOWNLOAD BLANK TEMPLATE",
                  self._download_template, GREEN).pack(side='left', padx=4)

        # ── File list ──
        list_frame = tk.Frame(self.frame_standalone, bg=Theme.surface(),
                              highlightthickness=1, highlightbackground=Theme.border())
        list_frame.pack(fill='both', expand=True, pady=(0, 8))
        self._reg_theme(list_frame, 'surface')
        self._reg_theme(list_frame, 'border')

        list_hdr = tk.Frame(list_frame, bg=Theme.surface2())
        list_hdr.pack(fill='x')
        self._reg_theme(list_hdr, 'surface2')

        lh1 = tk.Label(list_hdr, text="  FILE", font=FONT_LABEL,
                        bg=Theme.surface2(), fg=Theme.text_dim(), width=42, anchor='w')
        lh1.pack(side='left', padx=4, pady=4)
        self._reg_theme(lh1, 'surface2+text_dim')

        lh2 = tk.Label(list_hdr, text="STATUS", font=FONT_LABEL,
                        bg=Theme.surface2(), fg=Theme.text_dim())
        lh2.pack(side='right', padx=12, pady=4)
        self._reg_theme(lh2, 'surface2+text_dim')

        scroll_y = ttk.Scrollbar(list_frame, orient='vertical')
        scroll_y.pack(side='right', fill='y')

        self.file_list = tk.Listbox(
            list_frame, bg=Theme.surface(), fg=Theme.text(), font=FONT_MONO,
            selectbackground=Theme.list_sel(), selectforeground=Theme.accent(),
            borderwidth=0, highlightthickness=0,
            yscrollcommand=scroll_y.set, activestyle='none')
        self.file_list.pack(fill='both', expand=True, padx=4, pady=4)
        scroll_y.config(command=self.file_list.yview)
        self._reg_theme(self.file_list, 'listbox')

        self.file_count_var = tk.StringVar(value="0 files loaded")
        fc_lbl = tk.Label(self.frame_standalone, textvariable=self.file_count_var,
                          font=FONT_MONO, bg=Theme.bg(), fg=Theme.text_dim())
        fc_lbl.pack(anchor='w')
        self._reg_theme(fc_lbl, 'bg+text_dim')

        # ══════════════════════════════════════════════════════════════════════
        #  SEGMENT 2: SPECIAL ORDER
        # ══════════════════════════════════════════════════════════════════════

        self.frame_special = tk.Frame(parent, bg=Theme.bg())
        self._reg_theme(self.frame_special, 'bg')

        # ── Special Order file selector ──
        self._section(self.frame_special, "S2  //  SPECIAL ORDER (Product Broadcast)")

        so_frame = tk.Frame(self.frame_special, bg=Theme.surface(),
                            highlightthickness=1, highlightbackground=Theme.border())
        so_frame.pack(fill='x', pady=(0, 10))
        self._reg_theme(so_frame, 'surface')
        self._reg_theme(so_frame, 'border')

        so_inner = tk.Frame(so_frame, bg=Theme.surface())
        so_inner.pack(fill='x', padx=12, pady=10)
        self._reg_theme(so_inner, 'surface')

        self.so_var = tk.StringVar(value="No Special Order file loaded")
        self.so_count_var = tk.StringVar(value="")

        so_top = tk.Frame(so_inner, bg=Theme.surface())
        so_top.pack(fill='x')
        self._reg_theme(so_top, 'surface')
        self._btn(so_top, "📂 SELECT SPECIAL ORDER",
                  self._select_special_order, PINK).pack(side='left')

        so_c = tk.Label(so_top, textvariable=self.so_count_var,
                        font=FONT_MONO, bg=Theme.surface(), fg=GREEN)
        so_c.pack(side='right')
        self._reg_theme(so_c, 'surface+green')

        so_l = tk.Label(so_inner, textvariable=self.so_var,
                        font=FONT_MONO, bg=Theme.surface(), fg=Theme.text_dim(),
                        wraplength=400, anchor='w')
        so_l.pack(anchor='w', pady=(6, 0))
        self._reg_theme(so_l, 'surface+text_dim')

        # Template buttons
        so_tmpl_row = tk.Frame(self.frame_special, bg=Theme.bg())
        so_tmpl_row.pack(fill='x', pady=(0, 6))
        self._reg_theme(so_tmpl_row, 'bg')
        self._btn(so_tmpl_row, "📋 EKA_DATA TEMPLATE",
                  self._download_eka_template, GREEN).pack(side='left', padx=4)
        self._btn(so_tmpl_row, "📋 SPECIAL ORDER TEMPLATE",
                  self._download_so_template, GREEN).pack(side='left', padx=4)

        # Info panel
        info_frame = tk.Frame(self.frame_special, bg=Theme.surface2(),
                              highlightthickness=1, highlightbackground=Theme.border())
        info_frame.pack(fill='both', expand=True, pady=(0, 8))
        self._reg_theme(info_frame, 'surface2')

        info_txt = tk.Label(info_frame, text=(
            "Special Order broadcasts the SAME products to ALL locations.\n\n"
            "• EKA_DATA provides: locations with TO pattern, Transfer Code, Posting Group\n"
            "• Special Order provides: Products with per-type qty (EBO / Airport / Kiosk / Tester)\n"
            "• Items_March provides: Item No, MRP, GST Code (shared master)\n\n"
            "Output: Final Data with TO, Transfer-to, Posting Group PRE-FILLED"
        ), font=FONT_MONO, bg=Theme.surface2(), fg=Theme.text_dim(),
           justify='left', anchor='nw', wraplength=450, padx=12, pady=10)
        info_txt.pack(fill='both', expand=True)
        self._reg_theme(info_txt, 'surface2+text_dim')

    def _build_right(self, parent):
        """Build right panel: output config, stats, log."""

        # ── Output ──
        self._section(parent, "02  //  OUTPUT")

        out_frame = tk.Frame(parent, bg=Theme.surface(),
                             highlightthickness=1, highlightbackground=Theme.border())
        out_frame.pack(fill='x', pady=(0, 12))
        self._reg_theme(out_frame, 'surface')
        self._reg_theme(out_frame, 'border')

        ol1 = tk.Label(out_frame, text="Output folder:", font=FONT_LABEL,
                        bg=Theme.surface(), fg=Theme.text_dim())
        ol1.pack(anchor='w', padx=10, pady=(8, 2))
        self._reg_theme(ol1, 'surface+text_dim')

        ol2 = tk.Label(out_frame, text="  eka_output/  (auto-created next to uploaded files)",
                        font=FONT_MONO, bg=Theme.surface(), fg=Theme.accent2())
        ol2.pack(anchor='w', padx=10)
        self._reg_theme(ol2, 'surface+accent2')

        self.last_path_var = tk.StringVar(value="No run yet")
        ol3 = tk.Label(out_frame, text="Last saved:", font=FONT_LABEL,
                        bg=Theme.surface(), fg=Theme.text_dim())
        ol3.pack(anchor='w', padx=10, pady=(6, 2))
        self._reg_theme(ol3, 'surface+text_dim')

        ol4 = tk.Label(out_frame, textvariable=self.last_path_var,
                        font=FONT_MONO, bg=Theme.surface(), fg=GREEN,
                        wraplength=280, justify='left')
        ol4.pack(anchor='w', padx=10, pady=(0, 10))
        self._reg_theme(ol4, 'surface+green')

        # ── Stats ──
        self._section(parent, "03  //  LAST RUN STATS")

        stats_frame = tk.Frame(parent, bg=Theme.surface(),
                               highlightthickness=1, highlightbackground=Theme.border())
        stats_frame.pack(fill='x', pady=(0, 12))
        self._reg_theme(stats_frame, 'surface')
        self._reg_theme(stats_frame, 'border')

        stat_defs = [
            ("Locations",     "locations",  ACCENT),
            ("Total PO Qty",  "po_qty",     GREEN),
            ("PO Line Items", "po_items",   ACCENT2),
            ("Tester Qty",    "tester_qty", PINK),
            ("PWP Qty",       "pwp_qty",    AMBER),
            ("GWP Qty",       "gwp_qty",    '#00BCD4'),
            ("Non-Stock Qty", "ns_qty",     '#795548'),
            ("Grand Total",   "grand",      Theme.text()),
            ("Unmatched",     "unmatched",  RED),
        ]

        self.stat_vars = {}
        for i, (label, key, color) in enumerate(stat_defs):
            row_bg = Theme.surface2() if i % 2 == 0 else Theme.surface()
            row = tk.Frame(stats_frame, bg=row_bg)
            row.pack(fill='x')
            self._reg_theme(row, 'surface2' if i % 2 == 0 else 'surface')

            sl = tk.Label(row, text=f"  {label}", font=FONT_MONO,
                          bg=row_bg, fg=Theme.text_dim(), width=20, anchor='w')
            sl.pack(side='left', pady=3, padx=4)
            self._reg_theme(sl, 'surface2+text_dim' if i % 2 == 0 else 'surface+text_dim')

            var = tk.StringVar(value="—")
            self.stat_vars[key] = var
            sv = tk.Label(row, textvariable=var,
                          font=("Aptos Display", 11, "bold"), bg=row_bg, fg=color)
            sv.pack(side='right', padx=12, pady=3)
            self._reg_theme(sv, 'surface2' if i % 2 == 0 else 'surface')

        # ── Log ──
        self._section(parent, "04  //  LOG")

        log_frame = tk.Frame(parent, bg=Theme.surface(),
                             highlightthickness=1, highlightbackground=Theme.border())
        log_frame.pack(fill='both', expand=True)
        self._reg_theme(log_frame, 'surface')
        self._reg_theme(log_frame, 'border')

        scroll_log = ttk.Scrollbar(log_frame, orient='vertical')
        scroll_log.pack(side='right', fill='y')

        self.log_text = tk.Text(
            log_frame, bg=Theme.surface(), fg=Theme.text_dim(), font=FONT_MONO,
            height=6, wrap='word', state='disabled',
            borderwidth=0, highlightthickness=0,
            yscrollcommand=scroll_log.set)
        self.log_text.pack(fill='both', expand=True, padx=6, pady=6)
        scroll_log.config(command=self.log_text.yview)
        self._reg_theme(self.log_text, 'log')

        self.log_text.tag_config('ok',   foreground=GREEN)
        self.log_text.tag_config('err',  foreground=RED)
        self.log_text.tag_config('inf',  foreground=ACCENT)
        self.log_text.tag_config('dim',  foreground=Theme.text_dim())
        self.log_text.tag_config('warn', foreground=AMBER)

    def _build_bottom(self):
        """Build bottom bar: progress, status, action buttons."""
        bottom = tk.Frame(self, bg=Theme.surface(), height=64)
        bottom.pack(fill='x', side='bottom')
        bottom.pack_propagate(False)
        self._reg_theme(bottom, 'surface')

        pb_frame = tk.Frame(bottom, bg=Theme.surface())
        pb_frame.pack(fill='x', padx=16, pady=(8, 0))
        self._reg_theme(pb_frame, 'surface')

        self.progress_canvas = tk.Canvas(pb_frame, height=4, bg=Theme.surface2(),
                                         highlightthickness=0)
        self.progress_canvas.pack(fill='x')
        self._reg_theme(self.progress_canvas, 'progress')

        ctrl = tk.Frame(bottom, bg=Theme.surface())
        ctrl.pack(fill='x', padx=16, pady=(4, 8))
        self._reg_theme(ctrl, 'surface')

        self.status_label = tk.Label(ctrl,
            text="READY  //  Load master + EKA_DATA + add PO files",
            font=FONT_MONO, bg=Theme.surface(), fg=Theme.text_dim())
        self.status_label.pack(side='left')
        self._reg_theme(self.status_label, 'surface+text_dim')

        # Action buttons (right side)
        self.run_btn = self._btn(ctrl, "▶  PROCESS ALL", self._run, GREEN, large=True)
        self.run_btn.pack(side='right')

        # v1.2: D365 TO export button
        self._btn(ctrl, "📤 D365 TO PACKAGE",
                  self._export_d365_to, ACCENT2).pack(side='right', padx=8)

        self._btn(ctrl, "📂 OPEN OUTPUT",
                  self._open_output, Theme.text_dim()).pack(side='right', padx=8)

    # ── HELPERS ────────────────────────────────────────────────────────────────

    def _section(self, parent, title):
        """Create a section header with label + separator line."""
        f = tk.Frame(parent, bg=Theme.bg())
        f.pack(fill='x', pady=(6, 4))
        self._reg_theme(f, 'bg')

        sl = tk.Label(f, text=title, font=FONT_LABEL,
                      bg=Theme.bg(), fg=Theme.accent())
        sl.pack(side='left')
        self._reg_theme(sl, 'bg+accent')

        sep = tk.Frame(f, bg=Theme.border(), height=1)
        sep.pack(side='left', fill='x', expand=True, padx=8)
        self._reg_theme(sep, 'border')

    def _btn(self, parent, text, cmd, color, large=False):
        """Create a themed button (Label with click/hover bindings)."""
        padx = 16 if large else 10
        pady = 6 if large else 4
        btn = tk.Label(parent, text=text, font=FONT_BTN,
                       bg=Theme.surface2(), fg=color,
                       cursor='hand2', padx=padx, pady=pady,
                       relief='flat', bd=0,
                       highlightthickness=1, highlightbackground=color)
        btn.bind('<Button-1>', lambda e: cmd())
        btn.bind('<Enter>', lambda e: btn.config(bg=color, fg=Theme.bg()))
        btn.bind('<Leave>', lambda e: btn.config(bg=Theme.surface2(), fg=color))
        self._reg_theme(btn, 'surface2')
        return btn

    def _log(self, msg, tag='dim'):
        """Append a timestamped message to the log panel."""
        self.log_text.config(state='normal')
        ts = time.strftime("%H:%M:%S")
        self.log_text.insert('end', f"[{ts}] {msg}\n", tag)
        self.log_text.see('end')
        self.log_text.config(state='disabled')

    def _set_status(self, msg, color=None):
        """Update the bottom status label."""
        self.status_label.config(text=msg, fg=color or Theme.text_dim())

    def _set_progress(self, pct):
        """Update the progress bar (0-100)."""
        self.progress_canvas.update_idletasks()
        w = self.progress_canvas.winfo_width()
        self.progress_canvas.delete('all')
        self.progress_canvas.create_rectangle(0, 0, w, 4, fill=Theme.surface2(), outline='')
        if pct > 0:
            bar_w = int(w * pct / 100)
            self.progress_canvas.create_rectangle(0, 0, bar_w, 4, fill=PINK, outline='')

    def _refresh_file_list(self):
        """Refresh the file listbox from self.po_files."""
        self.file_list.delete(0, 'end')
        for p in self.po_files:
            name = os.path.basename(p)
            display = name if len(name) <= 50 else name[:47] + '...'
            self.file_list.insert('end', f"  {display}")
        self.file_count_var.set(f"{len(self.po_files)} file(s) loaded")

    # ── v1.2: FILENAME → EKA LOCATION LOOKUP ──────────────────────────────────

    def _lookup_location_from_filename(self, filepath: str) -> Optional[Dict]:
        """
        Match PO filename against EKA_DATA 'location' field.

        Filename convention: rename PO file to the ERP Location code.
        Example: EBO_AMD01.xlsx → extract 'EBO_AMD01' → match against
        EKA_DATA 'Location' column.

        Args:
            filepath: Full path to the PO file

        Returns:
            EKA location dict if matched, None if no match.
        """
        if not self.eka_locations:
            return None

        # Extract Location code from filename
        fname = os.path.basename(filepath)
        loc_code = fname.replace('.xlsx', '').replace('.xlsm', '').replace('.xls', '')

        # Exact match on 'location' field
        for loc in self.eka_locations:
            if loc.get('location', '') == loc_code:
                return loc

        # Partial match: filename starts with location code
        for loc in self.eka_locations:
            loc_val = loc.get('location', '')
            if loc_val and loc_code.startswith(loc_val):
                return loc

        return None

    # ── ACTIONS ────────────────────────────────────────────────────────────────

    def _select_master(self):
        """Open file dialog and load Items_March master file."""
        path = filedialog.askopenfilename(
            title="Select Items_March.xlsx (Master File)",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")])
        if not path:
            return
        try:
            count = self.engine.load_master(path)
            self.master_path = path
            self.master_var.set(os.path.basename(path))
            self.master_count_var.set(f"✓ {count:,} items loaded")
            self._log(f"Master loaded: {os.path.basename(path)} → {count:,} items", 'ok')
        except Exception as e:
            self.master_var.set(f"ERROR: {e}")
            self.master_count_var.set("")
            self._log(f"Master load failed: {e}", 'err')

    def _select_eka(self):
        """
        v1.2: Load EKA_DATA (shared across both modes).
        Parses locations immediately and stores in self.eka_locations.
        """
        path = filedialog.askopenfilename(
            title="Select EKA_DATA.xlsx (Location Registry)",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")])
        if not path:
            return
        self.eka_path = path

        # Parse immediately to populate eka_locations
        try:
            so_engine = SpecialOrderEngine(
                self.engine.master if self.engine.master else {})
            logs = []
            count = so_engine.load_eka_data(path, logs)
            self.eka_locations = so_engine.locations

            self.eka_var.set(os.path.basename(path))
            self.eka_count_var.set(f"✓ {count} locations loaded")
            self._log(f"EKA_DATA: {os.path.basename(path)} → {count} locations", 'ok')

            for level, msg in logs:
                if level != 'info':
                    tag = {'warn': 'warn', 'error': 'err', 'alert': 'warn'}.get(level, 'dim')
                    self._log(f"  {msg}", tag)
        except Exception as e:
            self.eka_var.set(f"ERROR: {e}")
            self._log(f"EKA_DATA load failed: {e}", 'err')

    def _select_special_order(self):
        """Open file dialog to select Special_Order.xlsx."""
        path = filedialog.askopenfilename(
            title="Select Special Order File",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")])
        if not path:
            return
        self.so_path = path
        self.so_var.set(os.path.basename(path))
        self.so_count_var.set("✓ Selected")
        self._log(f"Special Order selected: {os.path.basename(path)}", 'ok')

    def _add_files(self):
        """Open file dialog to add PO files."""
        files = filedialog.askopenfilenames(
            title="Select PO Excel Files",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")])
        added = 0
        for f in files:
            if f not in self.po_files and 'Items_March' not in f and 'PO_Output' not in f:
                self.po_files.append(f)
                added += 1
        if added:
            self._refresh_file_list()
            self._log(f"Added {added} PO file(s)", 'inf')

    def _add_folder(self):
        """Scan folder for .xlsx files and add them."""
        folder = filedialog.askdirectory(title="Select folder containing PO files")
        if not folder:
            return
        import glob
        files = sorted(glob.glob(os.path.join(folder, "*.xlsx")))
        added = 0
        for f in files:
            bname = os.path.basename(f)
            if (f not in self.po_files
                    and 'Items_March' not in bname
                    and 'PO_Output' not in bname
                    and not bname.startswith('~')):
                self.po_files.append(f)
                added += 1
        self._refresh_file_list()
        self._log(f"Scanned folder → added {added} PO file(s)", 'inf')

    def _clear_files(self):
        """Clear all PO files from the list."""
        self.po_files.clear()
        self._refresh_file_list()
        self._log("File list cleared", 'dim')

    # ── MODE SWITCHING ─────────────────────────────────────────────────────────

    def _switch_mode(self, mode: str):
        """Switch between 'standalone' and 'special' mode."""
        self.active_mode = mode
        if mode == 'standalone':
            self.frame_special.pack_forget()
            self.frame_standalone.pack(fill='both', expand=True)
            self.btn_standalone.config(
                bg=Theme.accent(), fg=Theme.bg(),
                highlightbackground=Theme.accent())
            self.btn_special.config(
                bg=Theme.surface2(), fg=Theme.text_dim(),
                highlightbackground=Theme.border())
        else:
            self.frame_standalone.pack_forget()
            self.frame_special.pack(fill='both', expand=True)
            self.btn_special.config(
                bg=PINK, fg=Theme.bg(), highlightbackground=PINK)
            self.btn_standalone.config(
                bg=Theme.surface2(), fg=Theme.text_dim(),
                highlightbackground=Theme.border())

    # ── v1.2: D365 TO EXPORT ──────────────────────────────────────────────────

    def _export_d365_to(self):
        """Export D365 Transfer Order package using last processing results."""
        if not self.last_results:
            messagebox.showwarning("No Data",
                "Run processing first to generate data for D365 export.")
            return

        template = filedialog.askopenfilename(
            title="Select D365 TO Template (EKA Sample Package)",
            filetypes=[("Excel files", "*.xlsx")])
        if not template:
            return

        # v1.2: Output goes in eka_output/ subfolder next to source files
        if self.po_files:
            output_dir = os.path.join(
                os.path.dirname(os.path.abspath(self.po_files[0])), "eka_output")
        elif self.so_path:
            output_dir = os.path.join(
                os.path.dirname(os.path.abspath(self.so_path)), "eka_output")
        else:
            output_dir = os.path.join(
                os.path.dirname(os.path.abspath(__file__)), "eka_output")
        os.makedirs(output_dir, exist_ok=True)
        output = os.path.join(output_dir,
            f"D365_TO_Package_{time.strftime('%d%m%Y_%H%M%S')}.xlsx")

        try:
            D365TOExporter.export(self.last_results, template, output)
            self._log(f"D365 TO package saved → {output}", 'ok')

            if messagebox.askyesno("D365 Exported",
                    f"D365 TO package saved!\n\n"
                    f"File: {os.path.basename(output)}\n\n"
                    f"Open file?"):
                if os.name == 'nt':
                    os.startfile(output)
                elif os.name == 'posix':
                    import subprocess
                    opener = 'xdg-open' if not os.uname().sysname == 'Darwin' else 'open'
                    subprocess.Popen([opener, output])
        except Exception as e:
            self._log(f"D365 TO export failed: {e}", 'err')
            messagebox.showerror("Error", f"D365 export failed:\n{e}")

    # ── TEMPLATE DOWNLOADS ─────────────────────────────────────────────────────

    def _download_template(self):
        """Generate and save a blank PO template with all 5 sheets and correct headers."""
        save_path = filedialog.asksaveasfilename(
            title="Save Blank PO Template", defaultextension=".xlsx",
            initialfile="PO_Template_Blank.xlsx",
            filetypes=[("Excel files", "*.xlsx")])
        if not save_path:
            return
        try:
            wb = Workbook()
            wb.remove(wb.active)
            hdr_fill = PatternFill('solid', fgColor='1A237E')
            hdr_font = Font(bold=True, color='FFFFFF', name='Aptos Display', size=11)

            def make_header(ws, headers, widths=None):
                for c, h in enumerate(headers, 1):
                    cell = ws.cell(row=1, column=c, value=h)
                    cell.font = hdr_font
                    cell.fill = hdr_fill
                    cell.alignment = Alignment(horizontal='center')
                if widths:
                    for c, w in enumerate(widths, 1):
                        ws.column_dimensions[get_column_letter(c)].width = w
                ws.freeze_panes = 'A2'

            # ── PO Sheet ──
            ws = wb.create_sheet('PO')
            make_header(ws,
                ['Rank', 'Category', 'EAN', 'SKU Code', 'Product Name',
                 'Brand', 'MRP', 'Available', 'Order Qty', 'Tester Qty'],
                [8, 20, 18, 14, 50, 14, 10, 12, 12, 12])
            # Sample row as guide
            sample = [1, 'Eyes', '8906121646979', '06D19087', 'SAMPLE PRODUCT NAME',
                      'RENEE', 450, 0, '', '']
            for c, v in enumerate(sample, 1):
                cell = ws.cell(row=2, column=c, value=v)
                cell.font = Font(name='Aptos Display', size=11, color='999999', italic=True)
            # Note row
            ws.cell(row=3, column=1,
                value='← Delete this sample row. Fill EAN, Order Qty, Tester Qty columns.').font = \
                Font(name='Aptos Display', size=11, color='FF6600', italic=True)

            # ── PWP Sheet ──
            ws = wb.create_sheet('PWP')
            make_header(ws, ['Sr. No.', 'Product Name', 'Avail.Qty', 'Req.Qty'], [10, 30, 12, 12])
            pwp_items = [
                (1, 'Stay With Me - Mini', '', ''),
                (2, 'Perfume', '', ''),
                (3, 'Crème Mini', '', ''),
            ]
            for row_data in pwp_items:
                ws.append(row_data)
            ws.append(('Total', None, 0, 0))

            # ── GWP Sheet ──
            ws = wb.create_sheet('GWP')
            make_header(ws, ['Sr. No.', 'EAN', 'Product Name', 'Avail.Qty', 'Req.Qty'],
                        [10, 18, 45, 12, 12])
            gwp_items = [
                (1, 8904473101658, 'RENEE Lunar Luxe Trousseau box – Silver', '', ''),
                (2, 8904473101672, 'RENEE Red Velvet Trousseau Box - Red', '', ''),
                (3, 8904473101665, 'RENEE Rose Glow Trousseau Box - Pink_', '', ''),
                (4, 8904473101009, 'RENEE Pink Puffer Pouch', '', ''),
                (5, 8904473101023, 'RENEE Red Puffer Pouch', '', ''),
                (6, 8904473101016, 'RENEE Silver Puffer Pouch', '', ''),
                (7, 8904473101733, 'RENEE LUNAR LUXE TROUSSEAU BOX SMALL Pink', '', ''),
                (8, 8904473101740, 'RENEE LUNAR LUXE TROUSSEAU BOX SMALL Red', '', ''),
                (9, 8904473101726, 'RENEE LUNAR LUXE TROUSSEAU BOX SMALL Silver', '', ''),
            ]
            for row_data in gwp_items:
                ws.append(row_data)
            ws.append(('Total', None, None, None, 0))

            # ── Non Stock Sheet ──
            ws = wb.create_sheet('Non Stock')
            make_header(ws, ['Sr. No.', 'Product Name', 'QTY'], [10, 30, 10])
            ns_items = [
                (1, 'Cotton Rolls'), (2, 'Mirrors'), (3, 'Carry Bag (Small)'),
                (4, 'Carry Bag (Big)'), (5, 'Cleansers'), (6, 'Calculator'),
                (7, 'Blotters'), (8, 'Swabs'), (9, 'Bill Roll'),
                (10, 'Renee Notebook'), (11, 'Pen'),
            ]
            for sr, name in ns_items:
                ws.append((sr, name, ''))
            ws.append(('Total', None, 0))

            # ── Summary Sheet ──
            ws = wb.create_sheet('Summary')
            ws.cell(row=3, column=2, value='[Location Name]')
            for c, h in enumerate(['PO', 'Tester', 'PWP', 'GWP', 'Non-Stock Requirement', 'Total'], 6):
                ws.cell(row=3, column=c, value=h)
            for c, h in enumerate(['Sr No', 'Order No', 'Order date', 'Order Email Date'], 2):
                ws.cell(row=4, column=c, value=h)
            for c in range(6, 12):
                ws.cell(row=4, column=c, value='Qty')

            wb.save(save_path)
            self._log(f"Template saved → {save_path}", 'ok')
            messagebox.showinfo("Template Saved",
                f"Blank PO template saved to:\n{save_path}\n\n"
                "Sheets: PO, PWP, GWP, Non Stock, Summary\n"
                "Fill in EAN, Order Qty, Tester Qty in PO sheet.\n"
                "Fill Req.Qty in PWP/GWP, QTY in Non Stock.")
        except Exception as e:
            self._log(f"Template save failed: {e}", 'err')
            messagebox.showerror("Error", f"Failed:\n{e}")

    def _download_eka_template(self):
        """
        Generate and save a blank EKA_DATA template with correct column headers.

        EKA_DATA is the Location Registry used by both modes (v1.2).
        Contains one row per store/outlet with routing metadata.
        """
        save_path = filedialog.asksaveasfilename(
            title="Save EKA_DATA Template", defaultextension=".xlsx",
            initialfile="EKA_DATA_Template.xlsx",
            filetypes=[("Excel files", "*.xlsx")])
        if not save_path:
            return
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = 'EKA_DATA'

            headers = ['Desc', 'Bill to', 'Ship to', 'Location',
                       'Gen. Biz.  Posting Group', 'Short Name',
                       'Prefix', 'Short Code', 'Transfer Code', 'Type',
                       'Example Regular', 'Example Tester', 'Status']
            widths = [40, 10, 12, 15, 22, 22, 8, 12, 15, 10, 25, 25, 10]

            hdr_fill = PatternFill('solid', fgColor='E65100')
            hdr_font = Font(bold=True, color='FFFFFF', name='Aptos Display', size=11)

            for c, h in enumerate(headers, 1):
                cell = ws.cell(row=1, column=c, value=h)
                cell.font = hdr_font
                cell.fill = hdr_fill
                cell.alignment = Alignment(horizontal='center')

            for c, w in enumerate(widths, 1):
                ws.column_dimensions[get_column_letter(c)].width = w

            # Sample rows (2 TO + 1 SO example)
            samples = [
                ('RENEE COSMETICS-ISCON ARCADE', '20329', '20329_1', 'EBO_AMD01',
                 'OFF-EBO', 'Ahmedabad EBO', 'TO', 'AHDEB', 'EBO_AMD01', 'EBO',
                 'TO/AHDEB/04/18426', 'TO/AHDEB/TT/18427', 'Active'),
                ('RENEE COSMETICS-CHENNAI AIRPORT', '20342', '20342_1', 'AP_CHEN01',
                 'OFF-AIRPORT', 'Chennai Airport', 'TO', 'CHNAP', 'AP_CHEN01', 'Airport',
                 'TO/CHNAP/04/18426', 'TO/CHNAP/TT/18427', 'Active'),
                ('OG BEAUTY PRIVATE LIMITED', '20395', '20395_1', 'EBO_PUNE02',
                 'OFF-EBO', 'Pune EBO', 'SO', 'PUNEB', '20395_1', 'EBO',
                 'SO/PUNEB/04/18426', 'SO/PUNEB/TT/18427', 'Active'),
            ]

            sample_font = Font(name='Aptos Display', size=11, color='666666', italic=True)
            for r, row_data in enumerate(samples, 2):
                for c, v in enumerate(row_data, 1):
                    cell = ws.cell(row=r, column=c, value=v)
                    cell.font = sample_font

            # Note row
            note_row = len(samples) + 2
            ws.cell(row=note_row, column=1,
                    value='← Delete sample rows. Add one row per location. '
                          'Prefix: TO for Transfer Order, SO for Sales Order. '
                          'Status: Active or Inactive.').font = \
                Font(name='Aptos Display', size=11, color='FF6600', italic=True)

            ws.freeze_panes = 'A2'
            wb.save(save_path)
            self._log(f"EKA_DATA template → {save_path}", 'ok')
            messagebox.showinfo("Template Saved",
                f"EKA_DATA template saved to:\n{save_path}\n\n"
                "Columns: Desc, Bill to, Ship to, Location, Posting Group,\n"
                "Short Name, Prefix (TO/SO), Short Code, Transfer Code,\n"
                "Type (EBO/Airport/Kiosk), Status (Active/Inactive)\n\n"
                "Delete sample rows and add your locations.")
        except Exception as e:
            self._log(f"EKA template failed: {e}", 'err')
            messagebox.showerror("Error", f"Failed:\n{e}")

    def _download_so_template(self):
        """
        Generate and save a blank Special Order template with correct column headers.

        Special Order is the product broadcast file — same products shipped to
        ALL locations with type-based quantities (EBO / Airport / Kiosk / Tester).
        """
        save_path = filedialog.asksaveasfilename(
            title="Save Special Order Template", defaultextension=".xlsx",
            initialfile="Special_Order_Template.xlsx",
            filetypes=[("Excel files", "*.xlsx")])
        if not save_path:
            return
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = 'Special Order'

            headers = ['Description', 'EAN', 'Item Category Code', 'MRP',
                       'EBO Qty', 'Airport Qty', 'Kiosk Qty', 'Tester Qty']
            widths = [45, 18, 18, 10, 12, 12, 12, 12]

            hdr_fill = PatternFill('solid', fgColor='1A237E')
            hdr_font = Font(bold=True, color='FFFFFF', name='Aptos Display', size=11)

            for c, h in enumerate(headers, 1):
                cell = ws.cell(row=1, column=c, value=h)
                cell.font = hdr_font
                cell.fill = hdr_fill
                cell.alignment = Alignment(horizontal='center')

            for c, w in enumerate(widths, 1):
                ws.column_dimensions[get_column_letter(c)].width = w

            # Sample rows
            samples = [
                ('RENEE PRO HD 3-IN-1 - AMANDE_9 GM', '8906121648515', 'POWDER', 650, 18, 18, 10, 1),
                ('RENEE PRO HD CONCEALER - BUFF_8 ML', '8906121648317', 'CONCEALER', 750, 18, 18, 10, 1),
                ('RENEE PAPER BAG BIG MULTICOLOR', '8904473105984', 'PAPER BAG', 750, '', '', '', 50),
            ]

            sample_font = Font(name='Aptos Display', size=11, color='666666', italic=True)
            for r, row_data in enumerate(samples, 2):
                for c, v in enumerate(row_data, 1):
                    cell = ws.cell(row=r, column=c, value=v)
                    cell.font = sample_font

            # Note row
            note_row = len(samples) + 2
            ws.cell(row=note_row, column=1,
                    value='← Delete sample rows. Add one row per product. '
                          'EAN must match GTIN in Items_March. '
                          'Leave qty blank or 0 if not applicable for that type. '
                          'Tester Qty goes to ALL locations at ₹0.54.').font = \
                Font(name='Aptos Display', size=11, color='FF6600', italic=True)

            ws.freeze_panes = 'A2'
            wb.save(save_path)
            self._log(f"SO template → {save_path}", 'ok')
            messagebox.showinfo("Template Saved",
                f"Special Order template saved to:\n{save_path}\n\n"
                "Columns: Description, EAN, Item Category Code, MRP,\n"
                "EBO Qty, Airport Qty, Kiosk Qty, Tester Qty\n\n"
                "• EAN must match GTIN in Items_March\n"
                "• Leave qty blank or 0 if not applicable\n"
                "• Tester Qty goes to ALL locations at ₹0.54\n"
                "• Paper bags / operational items: put qty in Tester only")
        except Exception as e:
            self._log(f"SO template failed: {e}", 'err')
            messagebox.showerror("Error", f"Failed:\n{e}")

    def _open_output(self):
        """Open the last output file or the source folder."""
        if self.last_output and os.path.exists(self.last_output):
            if os.name == 'nt':
                os.startfile(self.last_output)
            elif os.name == 'posix':
                import subprocess
                opener = 'xdg-open' if not os.uname().sysname == 'Darwin' else 'open'
                subprocess.Popen([opener, self.last_output])
        else:
            # v1.2: Open the eka_output folder next to source files
            if self.po_files:
                folder = os.path.join(
                    os.path.dirname(os.path.abspath(self.po_files[0])), "eka_output")
            elif self.so_path:
                folder = os.path.join(
                    os.path.dirname(os.path.abspath(self.so_path)), "eka_output")
            else:
                folder = os.path.join(
                    os.path.dirname(os.path.abspath(__file__)), "eka_output")

            if os.path.exists(folder):
                if os.name == 'nt':
                    os.startfile(folder)
            else:
                messagebox.showwarning("Not Found", "No output file yet. Run processing first.")

    # ── PROCESSING ─────────────────────────────────────────────────────────────

    def _run(self):
        """Entry point for processing. Routes to standalone or special order."""
        if self.is_running:
            return
        if not self.master_path:
            messagebox.showwarning("No Master", "Please load Items_March first.")
            return
        if self.active_mode == 'special':
            self._run_special()
        else:
            self._run_standalone()

    def _run_standalone(self):
        """Launch standalone PO processing in background thread."""
        if not self.po_files:
            messagebox.showwarning("No Files", "Please add at least one PO file.")
            return

        # v1.2: Warn if EKA_DATA not loaded (auto-fill won't work)
        if not self.eka_locations:
            proceed = messagebox.askyesno("EKA_DATA Not Loaded",
                "EKA_DATA is not loaded.\n\n"
                "Without EKA_DATA, the following fields will be EMPTY:\n"
                "  • TO number\n"
                "  • Transfer-to Code\n"
                "  • Gen. Bus. Posting Group\n"
                "  • Headers (TO) sheet\n\n"
                "Load EKA_DATA first for auto-fill from filename.\n\n"
                "Continue anyway?")
            if not proceed:
                return

        # v1.2: Output goes in eka_output/ subfolder next to the PO files
        output_dir = os.path.join(
            os.path.dirname(os.path.abspath(self.po_files[0])), "eka_output")
        os.makedirs(output_dir, exist_ok=True)
        timestamp = time.strftime("%d%m%Y_%H%M%S")
        output = os.path.join(output_dir, f"PO_Output_{timestamp}.xlsx")

        self.is_running = True
        self.run_btn.config(fg=Theme.text_dim())
        self._set_status("PROCESSING...", AMBER)

        threading.Thread(
            target=self._process_worker, args=(output,), daemon=True).start()

    def _run_special(self):
        """Launch Special Order processing in background thread."""
        if not self.eka_path:
            messagebox.showwarning("No EKA_DATA", "Please select EKA_DATA.xlsx.")
            return
        if not self.so_path:
            messagebox.showwarning("No Special Order", "Please select the Special Order file.")
            return

        # v1.2: Output goes in eka_output/ subfolder next to the SO file
        output_dir = os.path.join(
            os.path.dirname(os.path.abspath(self.so_path)), "eka_output")
        os.makedirs(output_dir, exist_ok=True)
        timestamp = time.strftime("%d%m%Y_%H%M%S")
        output = os.path.join(output_dir, f"SO_Output_{timestamp}.xlsx")

        self.is_running = True
        self.run_btn.config(fg=Theme.text_dim())
        self._set_status("PROCESSING SPECIAL ORDER...", AMBER)

        threading.Thread(
            target=self._process_special_worker, args=(output,), daemon=True).start()

    # ── STANDALONE WORKER ──────────────────────────────────────────────────────

    def _process_worker(self, output: str):
        """
        Background worker for standalone PO processing.

        v1.2: After extraction, auto-fills TO/Transfer/Posting from EKA_DATA
        if filename matches a Location code.
        """
        total = len(self.po_files)

        # ── PHASE 1: VALIDATION ──
        self.after(0, self._log, "─── PHASE 1: VALIDATING FILES ───", 'inf')
        self.after(0, self._set_status, "VALIDATING...", AMBER)

        # v1.2: Warn if EKA_DATA not loaded
        if not self.eka_locations:
            self.after(0, self._log,
                "⚠ EKA_DATA not loaded — TO number, Transfer-to Code, "
                "and Gen. Bus. Posting Group will be EMPTY. "
                "Load EKA_DATA first for auto-fill.", 'warn')

        validation_results = {}
        files_with_errors = []
        alert_messages = []

        for i, po_path in enumerate(self.po_files):
            fname = os.path.basename(po_path)
            loc = fname.replace('.xlsx', '')
            self.after(0, self._set_status,
                f"Validating {i+1}/{total}: {fname[:45]}", AMBER)
            self.after(0, self._set_progress, int((i / total) * 30))

            try:
                vlogs = self.engine.validate_file(po_path)
                has_err = any(l[0] == 'error' for l in vlogs)
                validation_results[po_path] = (vlogs, has_err)

                for level, msg in vlogs:
                    tag = {'info': 'inf', 'warn': 'warn', 'error': 'err',
                           'alert': 'warn'}.get(level, 'dim')
                    self.after(0, self._log, f"  [{loc}] {msg}", tag)
                    if level == 'alert':
                        alert_messages.append(f"• {loc}: {msg}")

                if has_err:
                    files_with_errors.append(fname)
                    self.after(0, self._log, f"✗ {loc}  →  VALIDATION FAILED", 'err')
                else:
                    self.after(0, self._log, f"✓ {loc}  →  OK", 'ok')

            except Exception as e:
                validation_results[po_path] = ([('error', str(e))], True)
                files_with_errors.append(fname)
                self.after(0, self._log, f"✗ {loc}  →  {e}", 'err')

        # Show auto-fix popup
        if alert_messages:
            alert_text = ("Column names were auto-fixed:\n\n"
                         + "\n".join(alert_messages[:20]))
            self.after(0, lambda: messagebox.showinfo("Auto-Fix Applied", alert_text))

        # Filter processable files
        processable = [f for f in self.po_files
                       if not validation_results.get(f, ([], True))[1]]
        if not processable:
            self.after(0, self._done, None,
                f"ABORTED — all {total} files failed validation", RED)
            return

        skipped = total - len(processable)
        if skipped > 0:
            self.after(0, self._log,
                f"⚠ Skipping {skipped} file(s) with errors", 'warn')

        # ── PHASE 2: EXTRACTION ──
        self.after(0, self._log,
            f"─── PHASE 2: EXTRACTING ({len(processable)} files) ───", 'inf')

        results: List[LocationResult] = []
        total_po = total_tester = total_pwp = total_gwp = total_ns = total_unmatched = 0

        for i, po_path in enumerate(processable):
            fname = os.path.basename(po_path)
            loc = fname.replace('.xlsx', '')
            self.after(0, self._set_status,
                f"Extracting {i+1}/{len(processable)}: {fname[:45]}", AMBER)
            self.after(0, self._set_progress,
                30 + int((i / len(processable)) * 55))

            try:
                res = self.engine.process_file(po_path)

                # ── v1.2: AUTO-FILL from EKA_DATA ──
                eka_loc = self._lookup_location_from_filename(po_path)

                if eka_loc:
                    to_regular = SpecialOrderEngine.generate_to_number(
                        eka_loc['prefix'], eka_loc['short_code'],
                        is_tester=False)
                    to_tester = SpecialOrderEngine.generate_to_number(
                        eka_loc['prefix'], eka_loc['short_code'],
                        is_tester=True)
                    tc = eka_loc['transfer_code']
                    pg = eka_loc['posting_group']

                    # Regular + PWP → regular TO
                    for item in res.regular_orders:
                        item.to = to_regular
                        item.transfer_to = tc
                        item.posting_group = pg

                    for item in res.pwp_orders:
                        item.to = to_regular
                        item.transfer_to = tc
                        item.posting_group = pg

                    # Tester + GWP + Non-Stock → tester TO
                    for item in res.tester_orders:
                        item.to = to_tester
                        item.transfer_to = tc
                        item.posting_group = pg

                    for item in res.gwp_orders:
                        item.to = to_tester
                        item.transfer_to = tc
                        item.posting_group = pg

                    for item in res.nonstock_orders:
                        item.to = to_tester
                        item.transfer_to = tc
                        item.posting_group = pg

                    res.logs.append(('info',
                        f"EKA auto-fill: {eka_loc['location']} → "
                        f"TO:{to_regular} / TT:{to_tester} / "
                        f"Transfer:{tc} / Posting:{pg}"))
                else:
                    if self.eka_locations:
                        res.logs.append(('warn',
                            f"EKA: '{loc}' not found in Location column — "
                            f"TO/Transfer/Posting left empty"))
                # ── END v1.2 auto-fill ──

                results.append(res)

                # Accumulate totals
                po_q = sum(r.qty for r in res.regular_orders)
                tt_q = sum(r.qty for r in res.tester_orders)
                pw_q = sum(r.qty for r in res.pwp_orders)
                gw_q = sum(r.qty for r in res.gwp_orders)
                ns_q = sum(r.qty for r in res.nonstock_orders)

                total_po += po_q
                total_tester += tt_q
                total_pwp += pw_q
                total_gwp += gw_q
                total_ns += ns_q
                total_unmatched += len(res.unmatched)

                # Flush extraction logs
                for level, msg in res.logs:
                    tag = {'info': 'inf', 'warn': 'warn', 'error': 'err'}.get(level, 'dim')
                    self.after(0, self._log, f"  [{loc}] {msg}", tag)

                parts = [f"PO:{po_q}"]
                if tt_q: parts.append(f"T:{tt_q}")
                if pw_q: parts.append(f"PWP:{pw_q}")
                if gw_q: parts.append(f"GWP:{gw_q}")
                if ns_q: parts.append(f"NS:{ns_q}")
                self.after(0, self._log,
                    f"✓ {loc}  →  {' | '.join(parts)}", 'ok')

            except Exception as e:
                self.after(0, self._log, f"✗ {fname}  →  {e}", 'err')

        if not results:
            self.after(0, self._done, None, "ERROR: No data processed", RED)
            return

        # ── Write output ──
        try:
            self.after(0, self._set_progress, 90)
            self.after(0, self._set_status, "Writing Excel...", AMBER)

            # v1.2: store results for D365 export
            self.last_results = results

            ExcelWriter.write(results, output)
            self.after(0, self._set_progress, 100)

            grand = total_po + total_tester + total_pwp + total_gwp + total_ns
            stats = {
                'locations':  str(len(results)),
                'po_qty':     f"{total_po:,}",
                'po_items':   str(sum(len(r.regular_orders) for r in results)),
                'tester_qty': f"{total_tester:,}",
                'pwp_qty':    f"{total_pwp:,}",
                'gwp_qty':    f"{total_gwp:,}",
                'ns_qty':     f"{total_ns:,}",
                'grand':      f"{grand:,}",
                'unmatched':  str(total_unmatched),
            }
            self.after(0, self._update_stats, stats)
            self.after(0, self._done, output,
                f"DONE  //  {len(results)} locations  |  {grand:,} total qty", GREEN)

        except Exception as e:
            self.after(0, self._done, None, f"ERROR: {e}", RED)

    # ── SPECIAL ORDER WORKER ──────────────────────────────────────────────────

    def _process_special_worker(self, output: str):
        """
        Background worker for Special Order processing.
        Loads EKA_DATA + Special Order, validates, generates output.
        """
        self.after(0, self._log, "═══ SPECIAL ORDER PROCESSING ═══", 'inf')

        # ── Load EKA_DATA ──
        self.after(0, self._set_status, "Loading EKA_DATA...", AMBER)
        self.after(0, self._set_progress, 5)

        so_engine = SpecialOrderEngine(self.engine.master)
        logs = []

        loc_count = so_engine.load_eka_data(self.eka_path, logs)
        for level, msg in logs:
            tag = {'info': 'inf', 'warn': 'warn', 'error': 'err',
                   'alert': 'warn'}.get(level, 'dim')
            self.after(0, self._log, f"  {msg}", tag)

        if loc_count == 0:
            self.after(0, self._done, None, "ABORTED — EKA_DATA loading failed", RED)
            return

        self.after(0, lambda: self.eka_count_var.set(f"✓ {loc_count} locations"))

        # ── Load Special Order ──
        self.after(0, self._set_status, "Loading Special Order...", AMBER)
        self.after(0, self._set_progress, 15)

        logs2 = []
        alert_messages = []
        prod_count = so_engine.load_special_order(self.so_path, logs2)
        for level, msg in logs2:
            tag = {'info': 'inf', 'warn': 'warn', 'error': 'err',
                   'alert': 'warn'}.get(level, 'dim')
            self.after(0, self._log, f"  {msg}", tag)
            if level == 'alert':
                alert_messages.append(f"• {msg}")

        if prod_count == 0:
            self.after(0, self._done, None, "ABORTED — Special Order loading failed", RED)
            return

        self.after(0, lambda: self.so_count_var.set(f"✓ {prod_count} products"))

        if alert_messages:
            alert_text = ("Column names auto-fixed in Special Order:\n\n"
                         + "\n".join(alert_messages))
            self.after(0, lambda: messagebox.showinfo("Auto-Fix Applied", alert_text))

        # ── Validate ──
        self.after(0, self._set_status, "Validating...", AMBER)
        self.after(0, self._set_progress, 25)

        logs3 = []
        if not so_engine.validate(logs3):
            for level, msg in logs3:
                tag = {'info': 'inf', 'warn': 'warn', 'error': 'err'}.get(level, 'dim')
                self.after(0, self._log, f"  {msg}", tag)
            self.after(0, self._done, None, "ABORTED — Validation failed", RED)
            return

        for level, msg in logs3:
            tag = {'info': 'inf', 'warn': 'warn', 'error': 'err'}.get(level, 'dim')
            self.after(0, self._log, f"  {msg}", tag)

        # ── Process ──
        self.after(0, self._log, "─── GENERATING OUTPUT ───", 'inf')
        self.after(0, self._set_status, "Generating output...", AMBER)
        self.after(0, self._set_progress, 40)

        logs4 = []
        results = so_engine.process(logs4)
        for level, msg in logs4:
            tag = {'info': 'inf', 'warn': 'warn', 'error': 'err'}.get(level, 'dim')
            self.after(0, self._log, f"  {msg}", tag)

        if not results:
            self.after(0, self._done, None, "ERROR — No results generated", RED)
            return

        # Log per-location summary
        total_po = total_tester = total_unmatched = 0
        for i, res in enumerate(results):
            po_q = sum(r.qty for r in res.regular_orders)
            tt_q = sum(r.qty for r in res.tester_orders)
            total_po += po_q
            total_tester += tt_q
            total_unmatched += len(res.unmatched)

            to_reg = res.regular_orders[0].to if res.regular_orders else '?'
            self.after(0, self._log,
                f"  ✓ {res.filename:<22} PO:{po_q} T:{tt_q}  [{to_reg}]", 'ok')
            self.after(0, self._set_progress,
                40 + int((i / len(results)) * 40))

        # ── Write output ──
        try:
            self.after(0, self._set_progress, 85)
            self.after(0, self._set_status, "Writing Excel...", AMBER)

            # v1.2: store results for D365 export
            self.last_results = results

            ExcelWriter.write(
                results, output,
                eka_locations=so_engine.locations,
                master=self.engine.master,
                so_products=so_engine.products)
            self.after(0, self._set_progress, 100)

            grand = total_po + total_tester
            stats = {
                'locations':  str(len(results)),
                'po_qty':     f"{total_po:,}",
                'po_items':   str(sum(len(r.regular_orders) for r in results)),
                'tester_qty': f"{total_tester:,}",
                'pwp_qty':    '0',
                'gwp_qty':    '0',
                'ns_qty':     '0',
                'grand':      f"{grand:,}",
                'unmatched':  str(total_unmatched),
            }
            self.after(0, self._update_stats, stats)
            self.after(0, self._done, output,
                f"DONE  //  {len(results)} locations × {prod_count} products  "
                f"|  {grand:,} total qty", GREEN)

        except Exception as e:
            self.after(0, self._done, None, f"ERROR: {e}", RED)

    # ── SHARED COMPLETION ──────────────────────────────────────────────────────

    def _update_stats(self, stats: Dict[str, str]):
        """Update the stats panel from a dict."""
        for key, val in stats.items():
            if key in self.stat_vars:
                self.stat_vars[key].set(val)

    def _done(self, output, msg, color):
        """Processing complete — update UI, optionally open output."""
        self.is_running = False
        self.run_btn.config(fg=GREEN)
        self._set_status(msg, color)

        if output:
            self.last_output = output
            self.last_path_var.set(os.path.basename(output))
            self._log(f"Saved → {output}", 'inf')
            if messagebox.askyesno("Done!",
                    f"Processing complete!\n\n{msg}\n\nOpen output file?"):
                self._open_output()


# ═══════════════════════════════════════════════════════════════════════════════
#  STYLE
# ═══════════════════════════════════════════════════════════════════════════════

def apply_style():
    """Configure ttk scrollbar style to match current theme."""
    style = ttk.Style()
    style.theme_use('default')
    style.configure('Vertical.TScrollbar',
                    background=Theme.surface2(),
                    troughcolor=Theme.surface(),
                    arrowcolor=Theme.text_dim(),
                    bordercolor=Theme.border(),
                    lightcolor=Theme.surface2(),
                    darkcolor=Theme.surface2())
    style.map('Vertical.TScrollbar',
              background=[('active', Theme.border())])


# ═══════════════════════════════════════════════════════════════════════════════
#  ENTRY POINT
# ═══════════════════════════════════════════════════════════════════════════════

if __name__ == '__main__':
    # ── Fix Windows DPI blurriness ──
    try:
        import ctypes
        ctypes.windll.shcore.SetProcessDpiAwareness(2)
    except Exception:
        try:
            ctypes.windll.shcore.SetProcessDpiAwareness(1)
        except Exception:
            try:
                ctypes.windll.user32.SetProcessDPIAware()
            except Exception:
                pass

    app = ReneePOApp()
    apply_style()
    app.mainloop()