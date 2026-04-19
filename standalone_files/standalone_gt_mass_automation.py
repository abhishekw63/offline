"""
╔═══════════════════════════════════════════════════════════════════════════════╗
║               GT MASS DUMP GENERATOR — v2.4                                  ║
║               Tkinter GUI Desktop Application                                ║
╠═══════════════════════════════════════════════════════════════════════════════╣
║  Author  : Agami AI / Vishal                                                ║
║  Version : 2.4 (Strict template validation + full file traceability)        ║
║  Purpose : Reads GT-Mass / Everyday PO Excel files from distributors,       ║
║            extracts meta info (SO Number, Distributor, City, State,          ║
║            Location) and ordered items (BC Code, Qty, Tester Qty),          ║
║            generates ERP-importable Sales Order sheets.                      ║
║  Stack   : Python 3.13, Tkinter, pandas, openpyxl                           ║
╚═══════════════════════════════════════════════════════════════════════════════╝

═══════════════════════════════════════════════════════════════════════════════
  CHANGELOG
═══════════════════════════════════════════════════════════════════════════════

  v2.4 — Strict Template Validation + Full File Traceability
  ═════════════════════════════════════════════════════════════════════════════

    ┌───────────────────────────────────────────────────────────────────────┐
    │  CHANGE                         │  IMPACT                             │
    ├─────────────────────────────────┼─────────────────────────────────────┤
    │  TemplateValidator class        │  Rejects non-standard files UP      │
    │  (new)                          │  FRONT with a clear reason, before  │
    │                                 │  any parsing is attempted.          │
    ├─────────────────────────────────┼─────────────────────────────────────┤
    │  attempted_files list on        │  Tracks EVERY file touched — even   │
    │  ProcessResult                  │  ones that fail validation before   │
    │                                 │  producing any OrderRow.            │
    ├─────────────────────────────────┼─────────────────────────────────────┤
    │  File → SO Mapping shows ALL    │  Every uploaded file gets an entry: │
    │  files (successful + failed +   │   • Success → SO Number             │
    │  warned)                        │   • Failed  → ❌ FAILED: reason     │
    │                                 │   • Warned  → SO + ⚠️ flag          │
    ├─────────────────────────────────┼─────────────────────────────────────┤
    │  Template-violation count in    │  "N file(s) don't meet template     │
    │  generate popup                 │  standard — see Warnings sheet"     │
    ├─────────────────────────────────┼─────────────────────────────────────┤
    │  Code formatting refactor       │  One statement per line, blank      │
    │                                 │  lines between logical blocks,      │
    │                                 │  comprehensive docstrings           │
    └─────────────────────────────────┴─────────────────────────────────────┘

  TEMPLATE COMPLIANCE RULES (strict — first sheet only)
  ──────────────────────────────────────────────────────────────────────────
  A file is REJECTED at validation stage if its FIRST sheet fails any of:

    1. Must contain a header row with 'BC Code' AND 'Order Qty'
    2. Must have 'PO Number' label with a value in meta rows

  Location is NOT a hard rejection — if missing, the file processes
  normally with Location Code left empty. A ❌ CRITICAL warning fires
  so you can enter it manually in the output.

  Multi-sheet files are allowed — extra sheets are simply ignored.

  ═════════════════════════════════════════════════════════════════════════════
  PREVIOUS VERSIONS
  ═════════════════════════════════════════════════════════════════════════════

  v2.3 — Code Quality Refactor
    • EmailBuilder + EmailSender split (SRP)
    • parse() broken into _find_header_row, _resolve_so_number, _extract_rows
    • Colors class (centralized palette)
    • Specific exception types throughout
    • safe_str_val() helper (DRY)
    • Renamed: ProcessResult, MetadataExtractor, SONumberFormatter

  v2.2 — File → SO Mapping, full SKU in email, source_file traceability
  v2.1 — SO from PO Number field, D365 export, email reports
  v2.0 — Initial ERP import format (Headers + Lines for Business Central)

═══════════════════════════════════════════════════════════════════════════════
  ARCHITECTURE
═══════════════════════════════════════════════════════════════════════════════

  ┌─────────────────────────────────────────────────────────────────────────┐
  │                       AutomationUI (Tkinter GUI)                        │
  │     Select files → Generate → D365 export → Email → Template dl        │
  └─────────────────────────┬───────────────────────────────────────────────┘
                            │
                            ▼
  ┌─────────────────────────────────────────────────────────────────────────┐
  │                    GTMassAutomation (Engine)                            │
  │                                                                          │
  │   for each file:                                                         │
  │     1. TemplateValidator.validate()   ← NEW in v2.4                    │
  │        └─► reject if fails compliance rules                             │
  │     2. ExcelParser.parse()                                               │
  │        ├─► _find_header_row()                                           │
  │        ├─► MetadataExtractor.extract() (meta + SO# + Location)         │
  │        ├─► _resolve_so_number() (file → filename → UNKNOWN)            │
  │        └─► _extract_rows() (data rows → OrderRow)                       │
  │     3. Collect rows/warnings/failures into ProcessResult                │
  └─────────────────────────┬───────────────────────────────────────────────┘
                            │
                            ▼
  ┌─────────────────────────────────────────────────────────────────────────┐
  │                    DumpExporter (Output)                                │
  │                                                                          │
  │   Produces 7 sheets:                                                     │
  │     Sheet 1: Headers (SO)       ← ERP Sales Order headers              │
  │     Sheet 2: Lines (SO)         ← ERP Sales Order lines                │
  │     Sheet 3: Sales Lines        ← Detailed flat reference              │
  │     Sheet 4: Sales Header       ← Grouped summary per SO               │
  │     Sheet 5: SKU Summary        ← Demand pivot per BC Code             │
  │     Sheet 6: File → SO Mapping  ← ALL files (success/fail/warn)        │
  │     Sheet 7: Warnings           ← Red-highlighted critical issues      │
  └─────────────────────────────────────────────────────────────────────────┘

  Email side (decoupled):
  ┌─────────────────────────┐    ┌─────────────────────────┐
  │  EmailBuilder           │    │  EmailSender            │
  │  ProcessResult → HTML   │ →  │  HTML → SMTP → delivery │
  │  (pure function)        │    │  (network I/O only)     │
  └─────────────────────────┘    └─────────────────────────┘

═══════════════════════════════════════════════════════════════════════════════
  DEPENDENCIES
═══════════════════════════════════════════════════════════════════════════════

  pip install pandas openpyxl
  pip install xlrd          # only for legacy .xls files

  Run:  python gt_mass_dump.py
"""


# ═══════════════════════════════════════════════════════════════════════════════
#  IMPORTS
# ═══════════════════════════════════════════════════════════════════════════════

from __future__ import annotations

import os
import sys
import platform
import time
import logging
import re
import smtplib

from email.message import EmailMessage
from dataclasses import dataclass, field
from pathlib import Path
from typing import List, Optional, Tuple, Dict
from datetime import datetime

import pandas as pd
import tkinter as tk
from tkinter import filedialog, messagebox

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ═══════════════════════════════════════════════════════════════════════════════
#  LOGGING CONFIGURATION
# ═══════════════════════════════════════════════════════════════════════════════
# All log messages go to console. Format: "2026-04-16 10:30:45 | INFO | ..."
# Change level to logging.DEBUG for verbose debugging.

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s | %(levelname)s | %(message)s"
)


# ═══════════════════════════════════════════════════════════════════════════════
#  EXPIRY CHECK
# ═══════════════════════════════════════════════════════════════════════════════

EXPIRY_DATE = "30-06-2026"


def check_expiry() -> None:
    """
    Check if the application has expired and act accordingly.

    Behavior:
        - Past expiry date  → error popup → exit application (sys.exit)
        - Within 7 days     → warning popup → continue normally
        - More than 7 days  → no popup, silent continue

    Called once at application startup in main().
    """
    expiry = datetime.strptime(EXPIRY_DATE, "%d-%m-%Y").date()
    today = datetime.now().date()

    # ── Expired: block usage entirely ──
    if today > expiry:
        root = tk.Tk()
        root.withdraw()
        messagebox.showerror(
            "Application Expired",
            f"This application expired on {EXPIRY_DATE}.\n"
            f"Please contact the administrator."
        )
        root.destroy()
        sys.exit(0)

    # ── Expiring soon: warn but allow ──
    days_remaining = (expiry - today).days
    if days_remaining <= 7:
        root = tk.Tk()
        root.withdraw()
        messagebox.showwarning(
            "Expiration Warning",
            f"⚠️ Expires in {days_remaining} day(s).\n"
            f"Expiry: {EXPIRY_DATE}"
        )
        root.destroy()


# ═══════════════════════════════════════════════════════════════════════════════
#  COLORS — centralized palette used across email HTML and Excel formatting
# ═══════════════════════════════════════════════════════════════════════════════

class Colors:
    """
    Centralized color palette.

    Change a color here and it updates everywhere — email HTML, Excel
    highlighting, and any future UI elements. Hex values are chosen
    for good contrast on both light and dark email clients.
    """

    NAVY   = '#1A237E'   # Primary brand — headers, SO table
    GREEN  = '#2E7D32'   # SKU demand section
    ORANGE = '#E65100'   # Order Qty accent
    PURPLE = '#6A1B9A'   # Tester Qty accent
    GOLD   = '#FFD600'   # Footer accent, branding highlight
    GRAY   = '#666666'   # Subtle text
    LTGRAY = '#f5f5f5'   # Light backgrounds


# ═══════════════════════════════════════════════════════════════════════════════
#  CONSTANTS
# ═══════════════════════════════════════════════════════════════════════════════

# ┌─────────────────────────────────────────────────────────────────────────┐
# │ LOCATION CODE MAPPING                                                    │
# │ Maps raw Location values from GT-Mass files → ERP Location Codes.       │
# │ Unknown values pass through unchanged. Empty values stay empty.         │
# └─────────────────────────────────────────────────────────────────────────┘

LOCATION_CODE_MAP: Dict[str, str] = {
    'AHD': 'PICK',           # Ahmedabad → PICK warehouse
    'BLR': 'DS_BL_OFF1',    # Bangalore → Dispatch office
}


# ┌─────────────────────────────────────────────────────────────────────────┐
# │ STATE-LIKE VALUES                                                        │
# │ Used as a safety check: if Distributor Name matches one of these,       │
# │ the source file has probably swapped rows → fire a warning.             │
# └─────────────────────────────────────────────────────────────────────────┘

STATE_LIKE_VALUES = {
    # Two-letter state codes
    "up", "mp", "ap", "hp", "uk", "jk", "wb", "tn", "kl", "ka",
    "gj", "rj", "hr", "pb", "br", "od", "as", "mh", "cg", "jh",

    # Zone names
    "north", "south", "east", "west", "central",

    # Full state names
    "uttar pradesh", "madhya pradesh", "rajasthan", "punjab",
    "maharashtra", "gujarat", "karnataka", "tamil nadu",
    "haryana", "delhi", "u.p", "u.p.", "m.p", "m.p.",
}


# ═══════════════════════════════════════════════════════════════════════════════
#  EMAIL CONFIGURATION
# ═══════════════════════════════════════════════════════════════════════════════

EMAIL_CONFIG = {
    # Gmail SMTP credentials (App Password, not regular password)
    'EMAIL_SENDER': 'abhishekwagh420@gmail.com',
    'EMAIL_PASSWORD': 'bomn ktfx jhct xexy',

    # SMTP server
    'SMTP_SERVER': 'smtp.gmail.com',
    'SMTP_PORT': 587,

    # Primary recipient
    'DEFAULT_RECIPIENT': 'abhishek.wagh@reneecosmetics.in',

    # CC recipients
    'CC_RECIPIENTS': [
        'offlineb2b@reneecosmetics.in',
        'kirpalsinh.bihola@reneecosmetics.in',
        'gtmassaccounts@reneecosmetics.in',
        'aritra.barmanray@reneecosmetics.in',
        'milan.nayak@reneecosmetics.in',
        'aashutosh.joshi@reneecosmetics.in',
        'ketan.jain@reneecosmetics.in'
    ],
}


# ═══════════════════════════════════════════════════════════════════════════════
#  UTILITY HELPERS
# ═══════════════════════════════════════════════════════════════════════════════

def safe_str_val(row_vals, idx: Optional[int], as_int_str: bool = False) -> str:
    """
    Safely extract a string value from a DataFrame row at the given index.

    Previously every field extraction repeated the same pattern:

        val = ''
        if idx is not None and pd.notna(row_vals[idx]):
            val = str(row_vals[idx]).strip()

    This helper collapses that into a single call.

    Args:
        row_vals   : numpy array representing one row from df.values
        idx        : Column index. Can be None if the column wasn't detected.
        as_int_str : If True, convert float → int → str (strips '.0' from
                     numeric IDs like EAN barcodes that Excel stores as floats).

    Returns:
        Cleaned string value, or '' if idx is None or the cell is blank/NaN.
    """
    if idx is None:
        return ''

    val = row_vals[idx]

    if pd.isna(val):
        return ''

    if as_int_str and isinstance(val, (int, float)):
        return str(int(val))

    return str(val).strip()


def format_indian(number) -> str:
    """
    Format a number using the Indian numbering system (lakhs, crores).

    Examples:
        1643      → "1,643"
        123456    → "1,23,456"
        1234567   → "12,34,567"
        12345.67  → "12,345.67"

    Args:
        number: int, float, or string-convertible numeric value

    Returns:
        Formatted string with Indian comma separators, or the raw
        string representation if conversion fails.
    """
    try:
        number = float(number)
    except (ValueError, TypeError):
        return str(number)

    sign = '-' if number < 0 else ''
    number = abs(number)

    # Split integer and decimal parts
    if number == int(number):
        int_part = str(int(number))
        dec_part = ''
    else:
        parts = f"{number:.2f}".split('.')
        int_part = parts[0]
        dec_part = '.' + parts[1]

    if len(int_part) <= 3:
        return sign + int_part + dec_part

    # Last 3 digits, then groups of 2
    result = int_part[-3:]
    remaining = int_part[:-3]

    while remaining:
        result = remaining[-2:] + ',' + result
        remaining = remaining[:-2]

    return sign + result + dec_part


# ═══════════════════════════════════════════════════════════════════════════════
#  DATA MODEL
# ═══════════════════════════════════════════════════════════════════════════════

@dataclass
class OrderRow:
    """
    Single ordered item extracted from a GT-Mass file.

    One file produces multiple OrderRows (one per SKU with qty > 0).

    Fields:
        so_number     : SO/GTM number (e.g., 'SO/GTM/6448')
        item_no       : BC Code / ERP Item No (e.g., '200163')
        ean           : EAN barcode (e.g., '8904473104307')
        category      : Product category (e.g., 'Eye', 'FACE')
        description   : Article Description
        qty           : Order Qty — regular stock
        tester_qty    : Tester Qty — samples/testers
        distributor   : Distributor Name from meta
        city          : City from meta
        state         : State from meta
        location      : Raw location value (e.g., 'AHD', 'BLR')
        location_code : Mapped ERP Location Code (e.g., 'PICK')
        source_file   : Original filename — for traceability via File → SO Mapping
    """
    so_number: str
    item_no: str
    ean: str
    category: str
    description: str
    qty: int
    tester_qty: int
    distributor: str
    city: str
    state: str
    location: str
    location_code: str
    source_file: str


@dataclass
class ProcessResult:
    """
    Aggregated result from processing all selected files.

    Populated by GTMassAutomation.process_files(), consumed by
    DumpExporter and EmailBuilder.

    Fields:
        rows            : All OrderRow objects across all files
        failed_files    : [(filename, reason)] — files that couldn't be parsed
        warned_files    : [(filename, warning_text)] — non-fatal issues
        attempted_files : ALL files processed, in selection order (NEW in v2.4).
                          Used by File → SO Mapping to show every uploaded
                          file — even ones that produced zero OrderRows.
        output_path     : Path to the generated reference Excel (set post-export)
    """
    rows: List[OrderRow] = field(default_factory=list)
    failed_files: List[Tuple[str, str]] = field(default_factory=list)
    warned_files: List[Tuple[str, str]] = field(default_factory=list)
    attempted_files: List[str] = field(default_factory=list)
    output_path: Optional[Path] = None


# ═══════════════════════════════════════════════════════════════════════════════
#  SO NUMBER FORMATTER — filename fallback
# ═══════════════════════════════════════════════════════════════════════════════

class SONumberFormatter:
    """
    Extracts an SO number from the filename as a FALLBACK only.

    The PRIMARY source is the file's PO Number field (Row 3/4, Col I).
    This class is used only when that field is empty.
    """

    @staticmethod
    def from_filename(filepath: Path) -> Optional[str]:
        """
        Extract the first digit sequence from the filename and format it.

        Example:
            "SOGTM6325.xlsx" → stem "SOGTM6325" → digits "6325" → "SO/GTM/6325"

        Args:
            filepath: Path to the Excel file

        Returns:
            "SO/GTM/####" string if digits were found, else None.
        """
        match = re.search(r"\d+", filepath.stem)

        if not match:
            logging.warning(f"No digits in filename: {filepath.name}")
            return None

        return f"SO/GTM/{match.group()}"


# ═══════════════════════════════════════════════════════════════════════════════
#  FILE READER — reads Excel into raw DataFrames (no header assumed)
# ═══════════════════════════════════════════════════════════════════════════════

class FileReader:
    """
    Reads Excel files into raw pandas DataFrames with NO header.

    The caller is responsible for finding the header row, because
    the real data header is buried below the meta rows (typically row 6)
    and its position can vary.
    """

    @staticmethod
    def read(file_path: Path) -> pd.DataFrame:
        """
        Read an Excel file using the correct engine for its extension.

        Engine selection:
            .xlsx / .xlsm → openpyxl (default)
            .xls          → xlrd (legacy format, requires `pip install xlrd`)

        Args:
            file_path: Path to the Excel file

        Returns:
            DataFrame with integer column indices and no header.

        Raises:
            RuntimeError: if file is corrupt, password-protected, has an
                          unsupported extension, or xlrd is missing for .xls.
        """
        ext = file_path.suffix.lower()

        # ── Modern Excel ──
        if ext in (".xlsx", ".xlsm"):
            try:
                df = pd.read_excel(file_path, header=None, engine="openpyxl")
                logging.info(f"{file_path.name} — openpyxl ({len(df)} rows)")
                return df
            except (ValueError, KeyError) as e:
                raise RuntimeError(f"Cannot read '{file_path.name}': {e}")

        # ── Legacy Excel ──
        if ext == ".xls":
            try:
                df = pd.read_excel(file_path, header=None, engine="xlrd")
                logging.info(f"{file_path.name} — xlrd ({len(df)} rows)")
                return df
            except ImportError:
                raise RuntimeError(
                    f"Cannot read '{file_path.name}' — xlrd not installed.\n"
                    f"Fix: pip install xlrd"
                )
            except (ValueError, KeyError) as e:
                raise RuntimeError(f"Cannot read '{file_path.name}': {e}")

        raise RuntimeError(
            f"Unsupported format: '{ext}'. Only .xlsx / .xlsm / .xls."
        )


# ═══════════════════════════════════════════════════════════════════════════════
#  TEMPLATE VALIDATOR — strict compliance check (NEW in v2.4)
# ═══════════════════════════════════════════════════════════════════════════════

class TemplateValidator:
    """
    Validates that an input file's FIRST sheet matches the GT-Mass PO template.

    Runs BEFORE parsing. A file that fails validation is rejected with
    a clear reason — the parser never touches it.

    Multi-sheet files are allowed — we always read the first sheet only
    (same as pd.read_excel default behavior). Extra sheets are ignored.

    Validation rules (all required):
        1. Header row must contain BOTH 'BC Code' AND 'Order Qty'
        2. Meta rows must contain 'PO Number' label with a value
        3. Meta rows must contain 'Location' label with a value
    """

    @staticmethod
    def validate(file_path: Path) -> Tuple[bool, Optional[str]]:
        """
        Run all template compliance checks against the first sheet of a file.

        Args:
            file_path: Path to the Excel file

        Returns:
            (is_valid, reason)
                is_valid = True   → reason is None, file passes
                is_valid = False  → reason contains human-readable message
                                    explaining what's non-compliant.
        """
        # Load the first sheet as raw data
        try:
            raw_df = FileReader.read(file_path)
        except RuntimeError as e:
            return False, f"Cannot read file: {e}"

        # ── Rule 1: Header row with BC Code + Order Qty ──
        header_ok, header_msg, header_row = TemplateValidator._check_header_row(raw_df)
        if not header_ok:
            return False, header_msg

        # ── Rules 2 & 3: Meta must have PO Number + Location ──
        meta_ok, meta_msg = TemplateValidator._check_required_meta(raw_df, header_row)
        if not meta_ok:
            return False, meta_msg

        return True, None

    @staticmethod
    def _check_header_row(raw_df: pd.DataFrame) -> Tuple[bool, Optional[str], int]:
        """
        Scan for the data header row containing 'BC Code' + 'Order Qty'.

        Args:
            raw_df: Raw DataFrame (no header) from FileReader.read()

        Returns:
            (ok, reason, header_row_index)
                header_row_index is -1 if not found.
        """
        for i, row_vals in enumerate(raw_df.values):
            vals_lower = [str(v).lower() for v in row_vals]

            has_bc_code = "bc code" in vals_lower
            has_order_qty = any("order qty" in v for v in vals_lower)

            if has_bc_code and has_order_qty:
                return True, None, i

        return False, (
            "Template violation: data header row not found. "
            "File must contain a row with BOTH 'BC Code' AND 'Order Qty' columns."
        ), -1

    @staticmethod
    def _check_required_meta(
        raw_df: pd.DataFrame,
        header_row: int
    ) -> Tuple[bool, Optional[str]]:
        """
        Ensure meta rows contain 'PO Number' label with a value.

        Location is NOT checked here — if missing, the file still processes
        and MetadataExtractor fires a ❌ CRITICAL warning. The user can
        enter the Location Code manually in the output.

        Only PO Number is a hard rejection — without it we can't create
        a valid SO number for the ERP import.

        Args:
            raw_df     : Raw DataFrame
            header_row : Index of the data header row (meta is above this)

        Returns:
            (ok, reason)
        """
        meta_df = raw_df.iloc[:header_row]

        po_number_found = False
        po_number_has_value = False

        for _, row in meta_df.iterrows():
            for col_idx in range(min(len(row) - 1, 10)):
                cell = row.iloc[col_idx]

                if pd.isna(cell):
                    continue

                cell_text = str(cell).strip().lower()

                # Check for PO Number label
                if cell_text == "po number":
                    po_number_found = True
                    if TemplateValidator._has_adjacent_value(row, col_idx):
                        po_number_has_value = True

        # Only PO Number is a hard requirement
        if not po_number_found:
            return False, (
                "Template violation: missing 'PO Number' label in meta rows. "
                "Fill this field and re-upload."
            )

        if not po_number_has_value:
            return False, (
                "Template violation: 'PO Number' label exists but value is empty. "
                "Fill the SO/GTM number and re-upload."
            )

        return True, None

    @staticmethod
    def _has_adjacent_value(row, col_idx: int) -> bool:
        """
        Check if the next 1 or 2 cells after col_idx contain a real value.

        Values are typically 2 columns to the right of their label
        (column G label, column I value). We scan the next two cells
        to handle slight layout variations.

        Args:
            row     : pandas Series (one row)
            col_idx : Column index of the label

        Returns:
            True if a non-empty, non-NaN value exists in cols
            col_idx+1 or col_idx+2.
        """
        for offset in range(1, 3):
            check_idx = col_idx + offset

            if check_idx >= len(row):
                return False

            val = row.iloc[check_idx]

            if pd.isna(val):
                continue

            val_str = str(val).strip()

            if val_str and val_str.lower() != 'nan':
                return True

        return False

# ═══════════════════════════════════════════════════════════════════════════════
#  METADATA EXTRACTOR
# ═══════════════════════════════════════════════════════════════════════════════

class MetadataExtractor:
    """
    Extracts meta fields (SO#, Distributor, City, State, Location) from the
    header region ABOVE the data table.

    Scanning strategy:
        LEFT SIDE  (Col A = label, Col B = value):
            - "Distributor Name"
            - "City"
            - "State"           (multiple possible — picks last non-blank)

        RIGHT SIDE (Col G = label, Col I = value):
            - "PO Number" → SO/GTM number
            - "Location"  → warehouse code (AHD, BLR, ...)

    Row positions vary slightly across files, so we scan by label matching
    rather than hardcoding row indices.
    """

    @staticmethod
    def extract(
        raw_df: pd.DataFrame,
        header_row: int
    ) -> Tuple[dict, List[str]]:
        """
        Scan rows 0..header_row-1 for meta field labels and values.

        Args:
            raw_df     : Full DataFrame (no header) from FileReader
            header_row : Row index of the data header

        Returns:
            Tuple of:
                meta_dict with keys: distributor, city, state, location,
                                     location_code, so_number
                warnings list (non-fatal issues found during extraction)
        """
        distributor = ""
        city = ""
        location = ""
        so_number = ""
        state_values: List[str] = []
        warnings: List[str] = []

        meta_df = raw_df.iloc[:header_row]

        for _, row in meta_df.iterrows():
            # ── LEFT SIDE: Col A (label) + Col B (value) ──
            label = ""
            if pd.notna(row.iloc[0]):
                label = str(row.iloc[0]).strip().lower()

            value = ""
            if pd.notna(row.iloc[1]):
                value = str(row.iloc[1]).strip()
                if value.lower() == "nan":
                    value = ""

            if label == "distributor name" and not distributor:
                distributor = value
                logging.info(f"Distributor: '{distributor}'")

            elif label == "city" and not city:
                city = value
                logging.info(f"City: '{city}'")

            elif label == "state":
                state_values.append(value)

            # ── RIGHT SIDE: scan cols 0-9 for "PO Number" / "Location" ──
            for col_idx in range(min(len(row) - 1, 10)):
                if pd.isna(row.iloc[col_idx]):
                    continue

                cell_text = str(row.iloc[col_idx]).strip().lower()

                if cell_text == "location":
                    for vi in range(col_idx + 1, min(col_idx + 3, len(row))):
                        lv = row.iloc[vi]
                        if pd.notna(lv) and str(lv).strip() and str(lv).strip().lower() != 'nan':
                            location = str(lv).strip()
                            logging.info(f"Location: '{location}'")
                            break

                elif cell_text == "po number" and not so_number:
                    for vi in range(col_idx + 1, min(col_idx + 3, len(row))):
                        pv = row.iloc[vi]
                        if pd.notna(pv) and str(pv).strip() and str(pv).strip().lower() != 'nan':
                            so_number = str(pv).strip()
                            logging.info(f"SO Number: '{so_number}'")
                            break

        # ── Resolve state: prefer the last non-blank value ──
        state = next((s for s in reversed(state_values) if s), "")
        logging.info(f"State: '{state}'")

        # ── Map raw Location → ERP Location Code ──
        location_code = ""
        if location:
            loc_upper = location.upper().strip()
            location_code = LOCATION_CODE_MAP.get(loc_upper, location)
            logging.info(f"Location code: '{location_code}'")

        # ── Collect meta-related warnings ──
        if not distributor:
            warnings.append("Distributor Name is blank.")

        if not city:
            warnings.append("City is blank.")

        if not state:
            warnings.append("State is blank.")

        if not location_code:
            warnings.append(
                "❌ CRITICAL: Location Code is EMPTY — "
                "ERP import will fail without Location Code. "
                "Fix the source file immediately."
            )

        # Safety check: if Distributor value looks like a state, rows are swapped
        if distributor and distributor.strip().lower() in STATE_LIKE_VALUES:
            warnings.append(
                f"Distributor '{distributor}' looks like a state/zone — verify."
            )

        meta = {
            "distributor": distributor,
            "city": city,
            "state": state,
            "location": location,
            "location_code": location_code,
            "so_number": so_number,
        }

        return meta, warnings


# ═══════════════════════════════════════════════════════════════════════════════
#  EXCEL PARSER
# ═══════════════════════════════════════════════════════════════════════════════

class ExcelParser:
    """
    Parses a single GT-Mass / Everyday PO Excel file into OrderRow objects.

    Processing pipeline:
        1. Read raw file (FileReader.read)
        2. Find header row (_find_header_row)
        3. Extract meta fields (MetadataExtractor)
        4. Resolve SO number (_resolve_so_number)
        5. Build data table (rows below header with proper column names)
        6. Detect column positions (_detect_columns)
        7. Extract ordered items into OrderRow objects (_extract_rows)

    Note: Callers should run TemplateValidator first. This parser assumes
    the file has already passed basic structural checks.
    """

    # Column name constants (lowercase for case-insensitive matching)
    BC_COLUMN     = "bc code"       # Item No from ERP
    QTY_COLUMN    = "order qty"     # Regular stock quantity
    TESTER_COLUMN = "tester qty"    # Sample/tester quantity

    def parse(self, file_path: Path) -> Tuple[List[OrderRow], List[str]]:
        """
        Parse one GT-Mass file end-to-end.

        Args:
            file_path: Path to the Excel file

        Returns:
            Tuple of (list of OrderRow, list of warning strings)

        Raises:
            RuntimeError: if file has broken structure (e.g., no BC Code column
                          despite passing template validation — should be rare).
        """
        logging.info(f"Parsing: {file_path.name}")

        warnings: List[str] = []

        # Step 1: Read raw file
        raw_df = FileReader.read(file_path)

        # Step 2: Find header row
        header_row = self._find_header_row(raw_df)

        # Step 3: Extract meta fields
        meta, meta_warnings = MetadataExtractor.extract(raw_df, header_row)
        warnings.extend(meta_warnings)

        # Step 4: Resolve SO number (file → filename → UNKNOWN)
        so_number, so_warnings = self._resolve_so_number(meta, file_path)
        warnings.extend(so_warnings)

        # Step 5: Build data table
        df = raw_df.iloc[header_row + 1:].copy()
        df.columns = raw_df.iloc[header_row].values
        df = df.reset_index(drop=True)

        # Steps 6 & 7: Detect columns + extract rows
        rows, extract_warnings = self._extract_rows(
            df, so_number, meta, file_path.name
        )
        warnings.extend(extract_warnings)

        return rows, warnings

    def _find_header_row(self, raw_df: pd.DataFrame) -> int:
        """
        Scan the DataFrame for the row containing 'BC Code' + 'Order Qty'.

        Args:
            raw_df: Raw DataFrame (no header)

        Returns:
            Integer row index of the header row.

        Raises:
            RuntimeError: if header row is not found.
        """
        for i, row_vals in enumerate(raw_df.values):
            vals_lower = [str(v).lower() for v in row_vals]

            has_bc_code = "bc code" in vals_lower
            has_order_qty = any("order qty" in v for v in vals_lower)

            if has_bc_code and has_order_qty:
                return i

        raise RuntimeError(
            "Header row not found — no 'BC Code' + 'Order Qty' row."
        )

    def _resolve_so_number(
        self,
        meta: dict,
        file_path: Path
    ) -> Tuple[str, List[str]]:
        """
        Resolve the SO number using priority: file → filename → UNKNOWN.

        Args:
            meta      : Meta dict from MetadataExtractor
            file_path : Path to the Excel file

        Returns:
            (so_number_string, warnings_list)
        """
        warnings: List[str] = []

        so = meta.get("so_number", "")

        # Priority 1: from file's PO Number field
        if so:
            logging.info(f"SO from file: '{so}'")
            return so, warnings

        # Priority 2: from filename digits (fallback)
        so = SONumberFormatter.from_filename(file_path)

        if so:
            warnings.append(
                f"SO from filename: '{so}'. Fill the PO Number field in source."
            )
            return so, warnings

        # Priority 3: unknown
        warnings.append("SO not found in file or filename — using 'SO/GTM/UNKNOWN'.")
        return "SO/GTM/UNKNOWN", warnings

    def _extract_rows(
        self,
        df: pd.DataFrame,
        so_number: str,
        meta: dict,
        filename: str
    ) -> Tuple[List[OrderRow], List[str]]:
        """
        Iterate the data table and build OrderRow objects for items with qty > 0.

        Args:
            df        : Data DataFrame with column names set from header row
            so_number : Resolved SO number to stamp on every row
            meta      : Meta dict (distributor, city, state, location, etc.)
            filename  : Source filename for traceability

        Returns:
            (list of OrderRow, list of warnings)

        Raises:
            RuntimeError: if BC Code or Order Qty columns are missing.
        """
        warnings: List[str] = []

        # Detect required columns
        bc_col, qty_col, tester_col, ean_col, cat_col, desc_col = (
            self._detect_columns(df)
        )

        if bc_col is None:
            raise RuntimeError("'BC Code' column not found in data table.")

        if qty_col is None:
            raise RuntimeError("'Order Qty' column not found in data table.")

        if tester_col is None:
            warnings.append("'Tester Qty' column not found — defaulting to 0.")

        # Pre-compute column indices for fast tuple access
        bc_idx = df.columns.get_loc(bc_col)
        qty_idx = df.columns.get_loc(qty_col)
        tester_idx = df.columns.get_loc(tester_col) if tester_col else None
        ean_idx = df.columns.get_loc(ean_col) if ean_col else None
        cat_idx = df.columns.get_loc(cat_col) if cat_col else None
        desc_idx = df.columns.get_loc(desc_col) if desc_col else None

        rows: List[OrderRow] = []

        for row_vals in df.values:
            # Read BC Code — must be a real integer
            bc = row_vals[bc_idx]

            if pd.isna(bc):
                continue

            try:
                bc = int(bc)
            except (ValueError, TypeError):
                continue

            # Read quantities
            qty = self._clean_qty(row_vals[qty_idx])
            tqty = self._clean_qty(row_vals[tester_idx]) if tester_idx is not None else 0

            # Skip rows with no demand
            if qty <= 0 and tqty <= 0:
                continue

            # Build the OrderRow
            rows.append(OrderRow(
                so_number=so_number,
                item_no=str(bc),
                ean=safe_str_val(row_vals, ean_idx, as_int_str=True),
                category=safe_str_val(row_vals, cat_idx),
                description=safe_str_val(row_vals, desc_idx),
                qty=qty,
                tester_qty=tqty,
                distributor=meta["distributor"],
                city=meta["city"],
                state=meta["state"],
                location=meta["location"],
                location_code=meta["location_code"],
                source_file=filename,
            ))

        if not rows:
            warnings.append(
                "No ordered rows — all Order Qty and Tester Qty values are 0."
            )

        return rows, warnings

    def _detect_columns(self, df) -> Tuple[Optional[str], ...]:
        """
        Find required and optional columns by name matching.

        Match rules:
            bc_col     : exact match 'bc code'
            qty_col    : substring match 'order qty'
            tester_col : substring match 'tester qty'
            ean_col    : exact match 'ean'
            cat_col    : exact match 'category'
            desc_col   : substring 'article description' or exact 'description'

        Args:
            df: DataFrame with column names set from header row

        Returns:
            Tuple of 6 optional column names (None if not found):
                (bc_col, qty_col, tester_col, ean_col, cat_col, desc_col)
        """
        bc_col = qty_col = tester_col = None
        ean_col = cat_col = desc_col = None

        for col in df.columns:
            name = str(col).strip().lower()

            if name == self.BC_COLUMN:
                bc_col = col

            if self.QTY_COLUMN in name:
                qty_col = col

            if self.TESTER_COLUMN in name:
                tester_col = col

            if name == 'ean' and not ean_col:
                ean_col = col

            if name == 'category' and not cat_col:
                cat_col = col

            if 'article description' in name:
                desc_col = col
            elif name == 'description' and not desc_col:
                desc_col = col

        return bc_col, qty_col, tester_col, ean_col, cat_col, desc_col

    @staticmethod
    def _clean_qty(value) -> int:
        """
        Clean a quantity cell value and convert to int.

        Handles messy Excel values:
            NaN / None           → 0
            "" / "-"             → 0
            "1,234"              → 1234  (strips commas)
            "12.0"               → 12    (float → int)
            "abc" (non-numeric)  → 0

        Args:
            value: Raw cell value

        Returns:
            Integer quantity (0 if invalid or empty).
        """
        if pd.isna(value):
            return 0

        value = str(value).strip()

        if value in ("", "-"):
            return 0

        value = value.replace(",", "")

        try:
            return int(float(value))
        except (ValueError, TypeError):
            return 0


# ═══════════════════════════════════════════════════════════════════════════════
#  EMAIL — EmailBuilder (HTML) + EmailSender (SMTP)
# ═══════════════════════════════════════════════════════════════════════════════

class EmailBuilder:
    """
    Pure transformation: ProcessResult → HTML string.

    No network I/O. Testable without SMTP credentials. Separated from
    EmailSender to follow the Single Responsibility Principle.
    """

    @staticmethod
    def _aggregate(result: 'ProcessResult') -> dict:
        """
        Aggregate OrderRows into email-ready summaries.

        Args:
            result: Full ProcessResult from the engine

        Returns:
            Dict with keys:
                unique_sos     : List of representative OrderRows (one per SO)
                so_groups      : {so_number: {'order', 'tester'}}
                total_items    : Total number of line items
                total_order    : Sum of all Order Qty
                total_tester   : Sum of all Tester Qty
                sorted_skus    : [(item_no, info)] sorted by total demand desc
        """
        unique_sos = list({r.so_number: r for r in result.rows}.values())
        total_order = sum(r.qty for r in result.rows)
        total_tester = sum(r.tester_qty for r in result.rows)

        # Aggregate per SKU (item_no)
        sku_groups: Dict[str, dict] = {}

        for r in result.rows:
            if r.item_no not in sku_groups:
                sku_groups[r.item_no] = {
                    'desc': r.description,
                    'cat': r.category,
                    'order': 0,
                    'tester': 0,
                }

            sku_groups[r.item_no]['order'] += r.qty
            sku_groups[r.item_no]['tester'] += r.tester_qty

            # Fill in description/category if previously blank
            if not sku_groups[r.item_no]['desc'] and r.description:
                sku_groups[r.item_no]['desc'] = r.description

        # Aggregate per SO
        so_groups: Dict[str, dict] = {}

        for r in result.rows:
            if r.so_number not in so_groups:
                so_groups[r.so_number] = {'order': 0, 'tester': 0}

            so_groups[r.so_number]['order'] += r.qty
            so_groups[r.so_number]['tester'] += r.tester_qty

        sorted_skus = sorted(
            sku_groups.items(),
            key=lambda x: x[1]['order'] + x[1]['tester'],
            reverse=True,
        )

        return {
            'unique_sos': unique_sos,
            'so_groups': so_groups,
            'total_items': len(result.rows),
            'total_order': total_order,
            'total_tester': total_tester,
            'sorted_skus': sorted_skus,
        }

    @staticmethod
    def build_subject(result: 'ProcessResult') -> str:
        """Build the email subject line."""
        ts = datetime.now().strftime('%d-%m-%Y %H:%M')
        so_count = len({r.so_number for r in result.rows})
        item_count = len(result.rows)

        return (
            f"📊 GT Mass SO Report: {so_count} SOs, "
            f"{item_count} Items — {ts}"
        )

    @staticmethod
    def build_html(result: 'ProcessResult', elapsed_str: str) -> str:
        """
        Build the full HTML email body.

        Args:
            result      : Full ProcessResult
            elapsed_str : Processing time for the footer

        Returns:
            Complete HTML string ready to attach to EmailMessage.
        """
        d = EmailBuilder._aggregate(result)
        C = Colors

        total_qty = d['total_order'] + d['total_tester']
        ts = datetime.now().strftime('%d-%m-%Y %H:%M:%S')

        html = f"""<html><body style="margin:0;padding:0;font-family:Arial,sans-serif;background:#f0f2f5;">
<table width="100%" cellpadding="0" cellspacing="0" style="background:#f0f2f5;">
<tr><td align="center" style="padding:20px 10px;">
<table width="800" cellpadding="0" cellspacing="0" style="background:#fff;border-radius:8px;overflow:hidden;border:1px solid #ddd;">
<tr><td style="background:{C.NAVY};padding:25px 30px;text-align:center;">
    <p style="margin:0;font-size:22px;font-weight:bold;color:white;">📊 GT Mass — Sales Order Report</p>
    <p style="margin:8px 0 0;font-size:12px;color:#9fa8da;">Generated: {ts} | Processing: {elapsed_str}</p>
    <table style="margin:10px auto 0;"><tr><td style="background:#283593;padding:5px 15px;border-radius:15px;">
        <span style="font-size:10px;color:#9fa8da;letter-spacing:1px;">⚡ GT MASS DUMP GENERATOR v2.4</span>
    </td></tr></table>
</td></tr>
<tr><td style="height:4px;font-size:0;"><table width="100%" cellpadding="0" cellspacing="0"><tr>
    <td width="25%" style="background:{C.ORANGE};height:4px;"></td>
    <td width="25%" style="background:{C.GOLD};height:4px;"></td>
    <td width="25%" style="background:#00E676;height:4px;"></td>
    <td width="25%" style="background:#2979FF;height:4px;"></td>
</tr></table></td></tr>
<tr><td style="padding:0;border-bottom:1px solid #eee;"><table width="100%" cellpadding="0" cellspacing="0"><tr>
    <td width="25%" style="text-align:center;padding:20px 10px;border-right:1px solid #f0f0f0;">
        <p style="margin:0;font-size:32px;font-weight:bold;color:{C.NAVY};">{len(d['unique_sos'])}</p>
        <p style="margin:5px 0 0;font-size:10px;color:#999;text-transform:uppercase;letter-spacing:1px;">Sales Orders</p>
    </td>
    <td width="25%" style="text-align:center;padding:20px 10px;border-right:1px solid #f0f0f0;">
        <p style="margin:0;font-size:32px;font-weight:bold;color:{C.GREEN};">{format_indian(d['total_items'])}</p>
        <p style="margin:5px 0 0;font-size:10px;color:#999;text-transform:uppercase;letter-spacing:1px;">Line Items</p>
    </td>
    <td width="25%" style="text-align:center;padding:20px 10px;border-right:1px solid #f0f0f0;">
        <p style="margin:0;font-size:32px;font-weight:bold;color:{C.ORANGE};">{format_indian(d['total_order'])}</p>
        <p style="margin:5px 0 0;font-size:10px;color:#999;text-transform:uppercase;letter-spacing:1px;">Order Qty</p>
    </td>
    <td width="25%" style="text-align:center;padding:20px 10px;">
        <p style="margin:0;font-size:32px;font-weight:bold;color:{C.PURPLE};">{format_indian(d['total_tester'])}</p>
        <p style="margin:5px 0 0;font-size:10px;color:#999;text-transform:uppercase;letter-spacing:1px;">Tester Qty</p>
    </td>
</tr></table></td></tr>
<tr><td style="padding:12px 20px;background:#f8f9fa;"><table width="100%" cellpadding="0" cellspacing="0"><tr>
    <td width="33%" style="height:2px;background:{C.NAVY};font-size:0;">&nbsp;</td>
    <td width="34%" style="height:2px;background:{C.GREEN};font-size:0;">&nbsp;</td>
    <td width="33%" style="height:2px;background:{C.ORANGE};font-size:0;">&nbsp;</td>
</tr></table></td></tr>
<tr><td style="padding:14px 20px;font-weight:bold;font-size:14px;color:{C.NAVY};border-left:5px solid {C.NAVY};background:#E8EAF6;">📋 Sales Order Details</td></tr>
<tr><td style="padding:0;"><table width="100%" cellpadding="0" cellspacing="0" style="border-collapse:collapse;">
<tr>
    <th style="background:{C.NAVY};color:white;padding:10px 8px;font-size:11px;text-transform:uppercase;">SO Number</th>
    <th style="background:{C.NAVY};color:white;padding:10px 8px;font-size:11px;text-transform:uppercase;">Distributor</th>
    <th style="background:{C.NAVY};color:white;padding:10px 8px;font-size:11px;text-transform:uppercase;">City</th>
    <th style="background:{C.NAVY};color:white;padding:10px 8px;font-size:11px;text-transform:uppercase;">State</th>
    <th style="background:{C.NAVY};color:white;padding:10px 8px;font-size:11px;text-transform:uppercase;">Location</th>
    <th style="background:{C.NAVY};color:white;padding:10px 8px;font-size:11px;text-transform:uppercase;">Order Qty</th>
    <th style="background:{C.NAVY};color:white;padding:10px 8px;font-size:11px;text-transform:uppercase;">Tester Qty</th>
    <th style="background:{C.NAVY};color:white;padding:10px 8px;font-size:11px;text-transform:uppercase;">Total</th>
</tr>"""

        for i, so_row in enumerate(d['unique_sos']):
            si = d['so_groups'].get(so_row.so_number, {'order': 0, 'tester': 0})
            bg = '#f9f9f9' if i % 2 == 1 else '#ffffff'

            html += f'''<tr style="background:{bg};">
    <td style="padding:9px 8px;text-align:center;font-size:12px;border-bottom:1px solid #eee;font-weight:bold;">{so_row.so_number}</td>
    <td style="padding:9px 8px;text-align:left;font-size:12px;border-bottom:1px solid #eee;">{so_row.distributor or "—"}</td>
    <td style="padding:9px 8px;text-align:center;font-size:12px;border-bottom:1px solid #eee;">{so_row.city or "—"}</td>
    <td style="padding:9px 8px;text-align:center;font-size:12px;border-bottom:1px solid #eee;">{so_row.state or "—"}</td>
    <td style="padding:9px 8px;text-align:center;font-size:12px;border-bottom:1px solid #eee;">{so_row.location_code or "—"}</td>
    <td style="padding:9px 8px;text-align:center;font-size:12px;border-bottom:1px solid #eee;">{format_indian(si['order'])}</td>
    <td style="padding:9px 8px;text-align:center;font-size:12px;border-bottom:1px solid #eee;">{format_indian(si['tester'])}</td>
    <td style="padding:9px 8px;text-align:center;font-size:12px;border-bottom:1px solid #eee;font-weight:bold;">{format_indian(si['order']+si['tester'])}</td>
</tr>'''

        html += f'''<tr style="background:#E8EAF6;font-weight:bold;">
    <td style="padding:10px 8px;text-align:center;font-size:12px;">TOTAL</td>
    <td colspan="4" style="padding:10px 8px;text-align:left;font-size:12px;">{len(d['unique_sos'])} Sales Orders</td>
    <td style="padding:10px 8px;text-align:center;font-size:12px;">{format_indian(d['total_order'])}</td>
    <td style="padding:10px 8px;text-align:center;font-size:12px;">{format_indian(d['total_tester'])}</td>
    <td style="padding:10px 8px;text-align:center;font-size:12px;">{format_indian(total_qty)}</td>
</tr></table></td></tr>
<tr><td style="padding:12px 20px;background:#f8f9fa;"><table width="100%" cellpadding="0" cellspacing="0"><tr>
    <td width="33%" style="height:2px;background:{C.NAVY};font-size:0;">&nbsp;</td>
    <td width="34%" style="height:2px;background:{C.GREEN};font-size:0;">&nbsp;</td>
    <td width="33%" style="height:2px;background:{C.ORANGE};font-size:0;">&nbsp;</td>
</tr></table></td></tr>
<tr><td style="padding:14px 20px;font-weight:bold;font-size:14px;color:{C.GREEN};border-left:5px solid {C.GREEN};background:#E8F5E9;">📦 SKU Demand Summary</td></tr>
<tr><td style="padding:0;"><table width="100%" cellpadding="0" cellspacing="0" style="border-collapse:collapse;">
<tr>
    <th style="background:{C.GREEN};color:white;padding:10px 6px;font-size:11px;">#</th>
    <th style="background:{C.GREEN};color:white;padding:10px 6px;font-size:11px;">BC CODE</th>
    <th style="background:{C.GREEN};color:white;padding:10px 6px;font-size:11px;">DESCRIPTION</th>
    <th style="background:{C.GREEN};color:white;padding:10px 6px;font-size:11px;">CATEGORY</th>
    <th style="background:{C.GREEN};color:white;padding:10px 6px;font-size:11px;">ORDER</th>
    <th style="background:{C.GREEN};color:white;padding:10px 6px;font-size:11px;">TESTER</th>
    <th style="background:{C.GREEN};color:white;padding:10px 6px;font-size:11px;">TOTAL</th>
</tr>'''

        for rank, (item_no, info) in enumerate(d['sorted_skus'], 1):
            total = info['order'] + info['tester']
            desc = info['desc'][:45] + '...' if len(info['desc']) > 45 else info['desc']
            bg = '#f1f8e9' if rank % 2 == 0 else '#ffffff'

            html += f'''<tr style="background:{bg};">
    <td style="padding:8px 6px;text-align:center;font-size:12px;color:#999;border-bottom:1px solid #eee;">{rank}</td>
    <td style="padding:8px 6px;text-align:center;font-size:12px;font-weight:bold;border-bottom:1px solid #eee;">{item_no}</td>
    <td style="padding:8px 6px;text-align:left;font-size:12px;border-bottom:1px solid #eee;">{desc or "—"}</td>
    <td style="padding:8px 6px;text-align:center;font-size:12px;border-bottom:1px solid #eee;">{info["cat"] or "—"}</td>
    <td style="padding:8px 6px;text-align:center;font-size:12px;border-bottom:1px solid #eee;">{format_indian(info["order"])}</td>
    <td style="padding:8px 6px;text-align:center;font-size:12px;border-bottom:1px solid #eee;">{format_indian(info["tester"])}</td>
    <td style="padding:8px 6px;text-align:center;font-size:12px;font-weight:bold;border-bottom:1px solid #eee;">{format_indian(total)}</td>
</tr>'''

        html += f'''<tr style="background:#E8F5E9;font-weight:bold;">
    <td style="padding:10px 6px;text-align:center;font-size:12px;"></td>
    <td style="padding:10px 6px;text-align:center;font-size:12px;">GRAND TOTAL</td>
    <td style="padding:10px 6px;text-align:left;font-size:12px;">{len(d['sorted_skus'])} unique SKUs</td>
    <td style="padding:10px 6px;text-align:center;font-size:12px;"></td>
    <td style="padding:10px 6px;text-align:center;font-size:12px;">{format_indian(d['total_order'])}</td>
    <td style="padding:10px 6px;text-align:center;font-size:12px;">{format_indian(d['total_tester'])}</td>
    <td style="padding:10px 6px;text-align:center;font-size:12px;">{format_indian(total_qty)}</td>
</tr></table></td></tr>
<tr><td style="padding:12px 20px;background:#f8f9fa;"><table width="100%" cellpadding="0" cellspacing="0"><tr>
    <td width="33%" style="height:2px;background:{C.NAVY};font-size:0;">&nbsp;</td>
    <td width="34%" style="height:2px;background:{C.GREEN};font-size:0;">&nbsp;</td>
    <td width="33%" style="height:2px;background:{C.ORANGE};font-size:0;">&nbsp;</td>
</tr></table></td></tr>
<tr><td style="background:{C.NAVY};padding:30px;text-align:center;">
    <p style="margin:0 0 5px;font-size:16px;font-weight:bold;color:{C.GOLD};letter-spacing:1px;">⚡ GT MASS DUMP GENERATOR v2.4</p>
    <p style="margin:0 0 18px;font-size:11px;color:#7986CB;">Warehouse Automation Suite — One-click PO Intelligence</p>
    <table style="margin:0 auto;max-width:400px;"><tr><td style="background:rgba(255,255,255,0.08);border:1px solid rgba(255,255,255,0.15);padding:18px;border-radius:10px;text-align:center;">
        <p style="margin:0 0 3px;font-size:10px;color:#7986CB;text-transform:uppercase;letter-spacing:2px;">🚀 Engineered by</p>
        <p style="margin:0 0 5px;font-size:18px;font-weight:bold;color:white;">Abhishek Wagh</p>
        <p style="margin:0 0 3px;font-size:11px;color:#9FA8DA;">Order Management and Automation</p>
        <p style="margin:0;font-size:10px;color:#7986CB;">📧 abhishek.wagh@reneecosmetics.in</p>
    </td></tr></table>
    <table style="margin:15px auto 0;max-width:450px;"><tr><td style="background:rgba(255,255,255,0.05);padding:12px 20px;border-radius:8px;border-left:3px solid {C.GOLD};text-align:left;">
        <p style="margin:0;font-size:11px;font-style:italic;color:#C5CAE9;">🏆 "Automation isn't just about saving time — it's about building systems that <span style="color:{C.GOLD};font-weight:bold;">sell while you sleep.</span>"</p>
    </td></tr></table>
    <p style="margin:18px 0 0;font-size:9px;color:#5C6BC0;">© 2026 RENEE Cosmetics Pvt. Ltd. | Warehouse Automation Division | Confidential</p>
</td></tr></table></td></tr></table></body></html>'''

        return html


class EmailSender:
    """
    Sends HTML emails via SMTP.

    No data aggregation or HTML building — receives a prebuilt HTML
    string from EmailBuilder and pushes it to Gmail SMTP.
    """

    @staticmethod
    def send_report(
        result: 'ProcessResult',
        elapsed_str: str
    ) -> Tuple[bool, str]:
        """
        Send the email report.

        Args:
            result      : ProcessResult with all data
            elapsed_str : Processing time for the footer

        Returns:
            (success_bool, error_message_if_failed)
        """
        config = EMAIL_CONFIG

        if not config['EMAIL_SENDER'] or not config['DEFAULT_RECIPIENT']:
            return False, "Email not configured."

        try:
            html = EmailBuilder.build_html(result, elapsed_str)
            subject = EmailBuilder.build_subject(result)

            msg = EmailMessage()
            msg['From'] = config['EMAIL_SENDER']
            msg['To'] = config['DEFAULT_RECIPIENT']

            if config['CC_RECIPIENTS']:
                msg['Cc'] = ', '.join(config['CC_RECIPIENTS'])

            msg['Subject'] = subject
            msg.set_content("Please view in HTML-compatible client.")
            msg.add_alternative(html, subtype='html')

            server = smtplib.SMTP(config['SMTP_SERVER'], config['SMTP_PORT'])
            server.starttls()
            server.login(config['EMAIL_SENDER'], config['EMAIL_PASSWORD'])

            recipients = [config['DEFAULT_RECIPIENT']] + config['CC_RECIPIENTS']
            server.send_message(msg, to_addrs=recipients)
            server.quit()

            logging.info(
                f"Email sent to {config['DEFAULT_RECIPIENT']} "
                f"+ {len(config['CC_RECIPIENTS'])} CC"
            )
            return True, ""

        except smtplib.SMTPAuthenticationError as e:
            logging.error(f"Email auth failed: {e}")
            return False, f"Auth failed — check EMAIL_PASSWORD. ({e})"

        except smtplib.SMTPException as e:
            logging.error(f"SMTP error: {e}")
            return False, f"SMTP error: {e}"

        except (ConnectionError, OSError) as e:
            logging.error(f"Network error: {e}")
            return False, f"Network error: {e}"

        except (ValueError, KeyError) as e:
            logging.error(f"Config error: {e}")
            return False, f"Config error: {e}"

# ═══════════════════════════════════════════════════════════════════════════════
#  DUMP EXPORTER — writes the 7-sheet output workbook
# ═══════════════════════════════════════════════════════════════════════════════

class DumpExporter:
    """
    Writes the final output Excel workbook.

    Seven sheets:
        1. Headers (SO)       — ERP Sales Order headers (one per SO)
        2. Lines (SO)         — ERP Sales Order lines (one per item)
        3. Sales Lines        — Detailed flat list with product details
        4. Sales Header       — Grouped summary per SO
        5. SKU Summary        — Demand pivot per BC Code across all SOs
        6. File → SO Mapping  — Every uploaded file + its SO status (v2.4)
        7. Warnings           — Non-fatal issues (red-highlighted criticals)

    Consistent formatting across all sheets:
        Headers: navy blue bg, white bold text (Aptos Display 11pt)
        Data:    light grey borders, Aptos Display 11pt
        Auto column widths based on content.
    """

    # ── Shared Excel formatting constants ──
    HEADER_FILL = PatternFill('solid', fgColor='1A237E')
    HEADER_FONT = Font(bold=True, color='FFFFFF', name='Aptos Display', size=11)
    DATA_FONT   = Font(name='Aptos Display', size=11)
    THIN_SIDE   = Side(style='thin', color='CCCCCC')
    BORDER      = Border(left=THIN_SIDE, right=THIN_SIDE, top=THIN_SIDE, bottom=THIN_SIDE)

    # ─────────────────────────────────────────────────────────────────────────
    #  CELL FORMATTING HELPERS
    # ─────────────────────────────────────────────────────────────────────────

    def _hdr_cell(self, ws, row, col, value):
        """Create a formatted header cell (navy bg, white bold text)."""
        cell = ws.cell(row=row, column=col, value=value)
        cell.font = self.HEADER_FONT
        cell.fill = self.HEADER_FILL
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = self.BORDER
        return cell

    def _data_cell(self, ws, row, col, value, fmt=None):
        """Create a formatted data cell (standard font, light borders)."""
        cell = ws.cell(row=row, column=col, value=value)
        cell.font = self.DATA_FONT
        cell.border = self.BORDER

        if fmt:
            cell.number_format = fmt

        return cell

    def _auto_width(self, ws, max_w=50):
        """Auto-size column widths based on the longest value per column."""
        for col in ws.columns:
            letter = col[0].column_letter
            width = max((len(str(c.value or '')) for c in col), default=8)
            ws.column_dimensions[letter].width = min(width + 3, max_w)

    # ─────────────────────────────────────────────────────────────────────────
    #  MAIN EXPORT ENTRY POINT
    # ─────────────────────────────────────────────────────────────────────────

    def export(self, result: ProcessResult) -> Optional[Path]:
        """
        Write the reference output Excel workbook (7 sheets).

        Args:
            result: ProcessResult from the engine

        Returns:
            Path to the saved file, or None if no data was available.
        """
        # Show error popup for files that couldn't be read at all
        if result.failed_files:
            msg = "Files skipped:\n\n"
            for fname, reason in result.failed_files:
                msg += f"  • {fname}: {reason}\n"
            messagebox.showerror("Files Failed", msg)

        # Nothing to export — abort
        if not result.rows:
            messagebox.showwarning(
                "No Data",
                "No valid rows found across all files. Nothing to export."
            )
            # Still return a path? No — if we have no data, but we DO have
            # attempted_files, we should still produce a File→SO Mapping
            # sheet so the user can see why everything failed.
            if not result.attempted_files:
                return None

        # Prepare output path
        Path("output").mkdir(exist_ok=True)
        timestamp = datetime.now().strftime('%d-%m-%Y_%H%M%S')
        ref_path = Path("output") / f"gt_mass_dump_{timestamp}.xlsx"

        wb = Workbook()
        wb.remove(wb.active)

        # Write all sheets in order
        if result.rows:
            self._write_headers_so(wb, result)
            self._write_lines_so(wb, result)
            self._write_sales_lines(wb, result)
            self._write_sales_header(wb, result)
            self._write_sku_summary(wb, result)

        # File → SO Mapping is always written if we attempted anything
        self._write_file_so_mapping(wb, result)
        self._write_warnings(wb, result)

        # Safety: workbook must have at least 1 sheet
        if not wb.sheetnames:
            wb.create_sheet('Empty')

        wb.save(str(ref_path))
        logging.info(f"Saved: {ref_path}")

        return ref_path

    # ─────────────────────────────────────────────────────────────────────────
    #  D365 TEMPLATE EXPORT
    # ─────────────────────────────────────────────────────────────────────────

    def export_d365(
        self,
        result: ProcessResult,
        template_path: str
    ) -> Optional[Path]:
        """
        Fill the D365 sample package template with processed data.

        The template has pre-formatted empty rows. We replace empty cells
        with filled ones via regex. If data exceeds template capacity,
        we inject new <row> XML elements before filling.

        Args:
            result        : ProcessResult with data to export
            template_path : Path to the D365 template .xlsx

        Returns:
            Path to the generated D365 import file, or None on failure.
        """
        if not result.rows:
            messagebox.showwarning("No Data", "Generate dump first.")
            return None

        try:
            import shutil
            import zipfile
            import re as re_mod

            Path("output").mkdir(exist_ok=True)
            timestamp = datetime.now().strftime('%d-%m-%Y_%H%M%S')
            d365_path = Path("output") / f"d365_import_{timestamp}.xlsx"
            shutil.copy2(template_path, str(d365_path))

            today_str = datetime.now().strftime("%d-%m-%Y")

            # Collect unique SOs in order
            seen = set()
            unique_sos = []
            for row in result.rows:
                if row.so_number not in seen:
                    seen.add(row.so_number)
                    unique_sos.append(row)

            # Read all ZIP contents
            zip_contents = {}
            with zipfile.ZipFile(str(d365_path), 'r') as z:
                for item in z.namelist():
                    zip_contents[item] = z.read(item)

            # ── Extend sharedStrings.xml with our new strings ──
            ss_xml = zip_contents['xl/sharedStrings.xml'].decode('utf-8')
            existing = re_mod.findall(r'<t[^>]*>([^<]*)</t>', ss_xml)
            string_map = {s: i for i, s in enumerate(existing)}

            new_strings = {'Order', 'Item', 'B2B', today_str}
            for r in unique_sos:
                new_strings.add(r.so_number)
                if r.location_code:
                    new_strings.add(r.location_code)
            for r in result.rows:
                new_strings.add(r.so_number)
                if r.location_code:
                    new_strings.add(r.location_code)

            next_idx = len(existing)
            for s in sorted(new_strings):
                if s not in string_map:
                    string_map[s] = next_idx
                    next_idx += 1

            total_count = next_idx
            si_items = [''] * total_count
            for s, idx in string_map.items():
                esc = s.replace('&', '&amp;').replace('<', '&lt;').replace('>', '&gt;')
                si_items[idx] = f'<si><t>{esc}</t></si>'

            new_ss = (
                '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>\r\n'
                f'<sst xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main" '
                f'count="{total_count}" uniqueCount="{total_count}">'
                + ''.join(si_items) + '</sst>'
            )
            zip_contents['xl/sharedStrings.xml'] = new_ss.encode('utf-8')

            # ── Helper: fill an empty <c r="A4" s="11"/> with a value ──
            def fill_cell(xml, col_letter, row_num, value, is_string=True):
                """Replace empty pre-formatted cell with filled one."""
                ref = f"{col_letter}{row_num}"
                pattern = f'<c r="{ref}" s="(\\d+)"\\s*/>'

                if is_string:
                    idx = string_map.get(str(value), 0)
                    replacement = f'<c r="{ref}" s="\\1" t="s"><v>{idx}</v></c>'
                else:
                    replacement = f'<c r="{ref}" s="\\1"><v>{value}</v></c>'

                return re_mod.sub(pattern, replacement, xml, count=1)

            # ── Helper: inject new empty rows when template runs out ──
            def inject_empty_row(xml, row_num, columns, style_id):
                """
                Inject a new empty <row> with pre-formatted <c> cells.

                The D365 template has a fixed number of pre-formatted rows.
                When our data exceeds that, fill_cell's regex matches nothing
                and the data is silently dropped. This function adds the missing
                rows before filling so every SO/item gets a slot.
                """
                cells = ''.join(
                    f'<c r="{c}{row_num}" s="{style_id}"/>'
                    for c in columns
                )
                new_row = (
                    f'<row r="{row_num}" spans="1:{len(columns)}" '
                    f'x14ac:dyDescent="0.3">{cells}</row>'
                )
                return xml.replace('</sheetData>', new_row + '</sheetData>')

            # ── Count template capacity vs actual data ──
            s1 = zip_contents['xl/worksheets/sheet1.xml'].decode('utf-8')
            s1_template_rows = len(re_mod.findall(r'<row r="(\d+)"', s1)) - 2
            hdr_cols = list('ABCDEFGHIJKLMNOPQR')

            if len(unique_sos) > s1_template_rows:
                logging.info(
                    f"D365: Sheet1 template has {s1_template_rows} rows, "
                    f"need {len(unique_sos)} — injecting extras"
                )
                for extra in range(s1_template_rows + 4, len(unique_sos) + 4):
                    s1 = inject_empty_row(s1, extra, hdr_cols, '11')

            s2 = zip_contents['xl/worksheets/sheet2.xml'].decode('utf-8')
            s2_template_rows = len(re_mod.findall(r'<row r="(\d+)"', s2)) - 3
            line_cols = list('ABCDEFGH')

            if len(result.rows) > s2_template_rows:
                logging.info(
                    f"D365: Sheet2 template has {s2_template_rows} rows, "
                    f"need {len(result.rows)} — injecting extras"
                )
                for extra in range(s2_template_rows + 4, len(result.rows) + 4):
                    s2 = inject_empty_row(s2, extra, line_cols, '8')

            # ── Fill Sales Header (sheet1) ──
            for i, row in enumerate(unique_sos):
                r = i + 4
                s1 = fill_cell(s1, 'A', r, 'Order')
                s1 = fill_cell(s1, 'B', r, row.so_number)

                for c in 'EFGHI':
                    s1 = fill_cell(s1, c, r, today_str)

                s1 = fill_cell(s1, 'J', r, row.so_number)

                if row.location_code:
                    s1 = fill_cell(s1, 'K', r, row.location_code)

                s1 = fill_cell(s1, 'M', r, 'B2B')

            zip_contents['xl/worksheets/sheet1.xml'] = s1.encode('utf-8')

            # ── Fill Sales Line (sheet2) ──
            current_so = None
            line_no = 0

            for i, row in enumerate(result.rows):
                if row.so_number != current_so:
                    current_so = row.so_number
                    line_no = 0

                line_no += 10000
                r = i + 4

                s2 = fill_cell(s2, 'A', r, 'Order')
                s2 = fill_cell(s2, 'B', r, row.so_number)
                s2 = fill_cell(s2, 'C', r, line_no, is_string=False)
                s2 = fill_cell(s2, 'D', r, 'Item')

                try:
                    s2 = fill_cell(s2, 'E', r, int(row.item_no), is_string=False)
                except (ValueError, TypeError):
                    s2 = fill_cell(s2, 'E', r, row.item_no)

                if row.location_code:
                    s2 = fill_cell(s2, 'F', r, row.location_code)

                s2 = fill_cell(s2, 'G', r, row.qty, is_string=False)

            zip_contents['xl/worksheets/sheet2.xml'] = s2.encode('utf-8')

            # ── Remove unused rows & update dimension/table refs ──
            last_hdr = 3 + len(unique_sos)
            last_line = 3 + len(result.rows)

            s1_clean = zip_contents['xl/worksheets/sheet1.xml'].decode('utf-8')
            for r in range(last_hdr + 1, 37):
                s1_clean = re_mod.sub(
                    rf'<row r="{r}"[^>]*>.*?</row>',
                    '',
                    s1_clean,
                    flags=re_mod.DOTALL,
                )
            s1_clean = re_mod.sub(
                r'<dimension ref="[^"]*"/>',
                f'<dimension ref="A1:R{last_hdr}"/>',
                s1_clean,
            )
            zip_contents['xl/worksheets/sheet1.xml'] = s1_clean.encode('utf-8')

            s2_clean = zip_contents['xl/worksheets/sheet2.xml'].decode('utf-8')
            for r in range(last_line + 1, 500):
                s2_clean = re_mod.sub(
                    rf'<row r="{r}"[^>]*>.*?</row>',
                    '',
                    s2_clean,
                    flags=re_mod.DOTALL,
                )
            s2_clean = re_mod.sub(
                r'<dimension ref="[^"]*"/>',
                f'<dimension ref="A1:H{last_line}"/>',
                s2_clean,
            )
            zip_contents['xl/worksheets/sheet2.xml'] = s2_clean.encode('utf-8')

            t1 = zip_contents['xl/tables/table1.xml'].decode('utf-8')
            t1 = re_mod.sub(r'ref="A3:[A-Z]+\d+"', f'ref="A3:R{last_hdr}"', t1)
            zip_contents['xl/tables/table1.xml'] = t1.encode('utf-8')

            t2 = zip_contents['xl/tables/table2.xml'].decode('utf-8')
            t2 = re_mod.sub(r'ref="A3:[A-Z]+\d+"', f'ref="A3:H{last_line}"', t2)
            zip_contents['xl/tables/table2.xml'] = t2.encode('utf-8')

            # Write the modified ZIP back
            with zipfile.ZipFile(str(d365_path), 'w', zipfile.ZIP_DEFLATED) as zo:
                for name, data in zip_contents.items():
                    zo.writestr(name, data)

            logging.info(f"D365 saved: {d365_path}")
            return d365_path

        except (FileNotFoundError, PermissionError) as e:
            messagebox.showerror("D365 Error", f"File error: {e}")
            return None

        except (KeyError, ValueError) as e:
            messagebox.showerror("D365 Error", f"Template error: {e}")
            return None

    # ─────────────────────────────────────────────────────────────────────────
    #  SHEET WRITERS
    # ─────────────────────────────────────────────────────────────────────────

    def _write_headers_so(self, wb, result: ProcessResult):
        """
        Sheet 1: 'Headers (SO)' — one row per unique SO for ERP header import.
        """
        ws = wb.create_sheet('Headers (SO)')

        headers = [
            'Document Type', 'No.', 'Sell-to Customer No.', 'Ship-to Code',
            'Posting Date', 'Order Date', 'Document Date',
            'Invoice From Date', 'Invoice To Date',
            'External Document No.', 'Location Code', 'Dimension Set ID',
            'Supply Type', 'Voucher Narration',
            'Brand Code (Dimension)', 'Channel Code (Dimension)',
            'Catagory (Dimension)', 'Geography Code (Dimension)',
        ]

        for c, h in enumerate(headers, 1):
            self._hdr_cell(ws, 1, c, h)

        today_str = datetime.now().strftime("%d-%m-%Y")

        # Deduplicate while preserving order
        seen = set()
        unique_sos = []
        for row in result.rows:
            if row.so_number not in seen:
                seen.add(row.so_number)
                unique_sos.append(row)

        r = 2
        for row in unique_sos:
            self._data_cell(ws, r, 1, 'Order')
            self._data_cell(ws, r, 2, row.so_number)
            self._data_cell(ws, r, 3, '')
            self._data_cell(ws, r, 4, '')

            for c in range(5, 10):
                self._data_cell(ws, r, c, today_str)

            self._data_cell(ws, r, 10, row.so_number)
            self._data_cell(ws, r, 11, row.location_code)
            self._data_cell(ws, r, 12, '')
            self._data_cell(ws, r, 13, 'B2B')
            r += 1

        self._auto_width(ws)

    def _write_lines_so(self, wb, result: ProcessResult):
        """
        Sheet 2: 'Lines (SO)' — one row per item for ERP line import.
        Line numbers increment by 10000 per SO, reset when SO changes.
        """
        ws = wb.create_sheet('Lines (SO)')

        headers = [
            'Document Type', 'Document No.', 'Line No.', 'Type',
            'No.', 'Location Code', 'Quantity', 'Unit Price',
        ]

        for c, h in enumerate(headers, 1):
            self._hdr_cell(ws, 1, c, h)

        r = 2
        current_so = None
        line_no = 0

        for row in result.rows:
            if row.so_number != current_so:
                current_so = row.so_number
                line_no = 0

            line_no += 10000

            self._data_cell(ws, r, 1, 'Order')
            self._data_cell(ws, r, 2, row.so_number)
            self._data_cell(ws, r, 3, line_no)
            self._data_cell(ws, r, 4, 'Item')
            self._data_cell(ws, r, 5, row.item_no)
            self._data_cell(ws, r, 6, row.location_code)
            self._data_cell(ws, r, 7, row.qty)
            self._data_cell(ws, r, 8, '')
            r += 1

        self._auto_width(ws)

    def _write_sales_lines(self, wb, result: ProcessResult):
        """
        Sheet 3: 'Sales Lines' — detailed flat list with product details.
        Reference sheet for cross-verification against source files.
        """
        ws = wb.create_sheet('Sales Lines')

        headers = [
            'SO Number', 'EAN', 'BC Code', 'Category',
            'Article Description', 'Order Qty', 'Tester Qty',
        ]

        for c, h in enumerate(headers, 1):
            self._hdr_cell(ws, 1, c, h)

        for r, row in enumerate(result.rows, 2):
            self._data_cell(ws, r, 1, row.so_number)
            self._data_cell(ws, r, 2, row.ean)
            self._data_cell(ws, r, 3, row.item_no)
            self._data_cell(ws, r, 4, row.category)
            self._data_cell(ws, r, 5, row.description)
            self._data_cell(ws, r, 6, row.qty)
            self._data_cell(ws, r, 7, row.tester_qty)

        self._auto_width(ws)

    def _write_sales_header(self, wb, result: ProcessResult):
        """
        Sheet 4: 'Sales Header' — grouped summary per SO (quantities + meta).
        """
        ws = wb.create_sheet('Sales Header')

        headers = [
            'SO Number', 'Order Qty', 'Tester Qty', 'Total Qty',
            'Distributor', 'City', 'State', 'Location',
        ]

        for c, h in enumerate(headers, 1):
            self._hdr_cell(ws, 1, c, h)

        # Group rows by SO, aggregating quantities
        so_groups: Dict[str, dict] = {}

        for row in result.rows:
            if row.so_number not in so_groups:
                so_groups[row.so_number] = {
                    'order_qty': 0,
                    'tester_qty': 0,
                    'distributor': row.distributor,
                    'city': row.city,
                    'state': row.state,
                    'location': row.location,
                }

            so_groups[row.so_number]['order_qty'] += row.qty
            so_groups[row.so_number]['tester_qty'] += row.tester_qty

        r = 2
        for so_num, info in so_groups.items():
            total = info['order_qty'] + info['tester_qty']

            self._data_cell(ws, r, 1, so_num)
            self._data_cell(ws, r, 2, info['order_qty'])
            self._data_cell(ws, r, 3, info['tester_qty'])
            self._data_cell(ws, r, 4, total)
            self._data_cell(ws, r, 5, info['distributor'])
            self._data_cell(ws, r, 6, info['city'])
            self._data_cell(ws, r, 7, info['state'])
            self._data_cell(ws, r, 8, info['location'])
            r += 1

        self._auto_width(ws)

    def _write_sku_summary(self, wb, result: ProcessResult):
        """
        Sheet 5: 'SKU Summary' — demand pivot per BC Code across all SOs.
        Sorted by total demand descending.
        """
        ws = wb.create_sheet('SKU Summary')

        headers = [
            'BC Code', 'Description', 'Category',
            'Order Qty', 'Tester Qty', 'Total Qty',
        ]

        for c, h in enumerate(headers, 1):
            self._hdr_cell(ws, 1, c, h)

        # Aggregate per BC Code
        sku_groups: Dict[str, dict] = {}

        for row in result.rows:
            if row.item_no not in sku_groups:
                sku_groups[row.item_no] = {
                    'desc': row.description,
                    'cat': row.category,
                    'order': 0,
                    'tester': 0,
                }

            sku_groups[row.item_no]['order'] += row.qty
            sku_groups[row.item_no]['tester'] += row.tester_qty

            if not sku_groups[row.item_no]['desc'] and row.description:
                sku_groups[row.item_no]['desc'] = row.description

            if not sku_groups[row.item_no]['cat'] and row.category:
                sku_groups[row.item_no]['cat'] = row.category

        sorted_skus = sorted(
            sku_groups.items(),
            key=lambda x: x[1]['order'] + x[1]['tester'],
            reverse=True,
        )

        r = 2
        grand_order = 0
        grand_tester = 0

        for item_no, info in sorted_skus:
            total = info['order'] + info['tester']
            grand_order += info['order']
            grand_tester += info['tester']

            self._data_cell(ws, r, 1, item_no)
            self._data_cell(ws, r, 2, info['desc'])
            self._data_cell(ws, r, 3, info['cat'])
            self._data_cell(ws, r, 4, info['order'])
            self._data_cell(ws, r, 5, info['tester'])
            self._data_cell(ws, r, 6, total)
            r += 1

        # Grand total row
        bold = Font(name='Aptos Display', size=11, bold=True)

        ws.cell(row=r, column=1, value='GRAND TOTAL').font = bold
        ws.cell(row=r, column=2, value=f'{len(sorted_skus)} unique SKUs').font = bold
        ws.cell(row=r, column=4, value=grand_order).font = bold
        ws.cell(row=r, column=5, value=grand_tester).font = bold
        ws.cell(row=r, column=6, value=grand_order + grand_tester).font = bold

        for c in range(1, 7):
            ws.cell(row=r, column=c).border = self.BORDER

        self._auto_width(ws)

    def _write_file_so_mapping(self, wb, result: ProcessResult):
        """
        Sheet 6: 'File → SO Mapping' — complete traceability for every file.

        v2.4: Shows ALL attempted files, not just successful ones.
            - Successful files  → SO Number (e.g., 'SO/GTM/6448')
            - Failed files      → '❌ FAILED: reason'
            - Warned files      → SO Number + '⚠️' flag

        Use case: if a file doesn't appear in the other sheets, look here
        to find out whether it was rejected and why.

        Status column (v2.4 addition):
            ✅ OK       — parsed cleanly
            ⚠️ WARNING  — parsed but with warnings
            ❌ FAILED   — rejected at validation or parsing stage
        """
        ws = wb.create_sheet('File → SO Mapping')

        headers = ['Sr No', 'Filename', 'SO Number', 'Status']
        for c, h in enumerate(headers, 1):
            self._hdr_cell(ws, 1, c, h)

        # Build quick-lookup maps from result
        file_to_so: Dict[str, str] = {}
        for row in result.rows:
            if row.source_file not in file_to_so:
                file_to_so[row.source_file] = row.so_number

        failed_map: Dict[str, str] = {
            fname: reason for fname, reason in result.failed_files
        }

        # Files with warnings (even if they succeeded)
        warned_set = {fname for fname, _ in result.warned_files}

        # Red styling for failed rows
        red_fill = PatternFill('solid', fgColor='FFCDD2')
        red_font = Font(name='Aptos Display', size=11, bold=True, color='D32F2F')

        # Yellow styling for warned rows
        yellow_fill = PatternFill('solid', fgColor='FFF9C4')

        # Iterate ALL attempted files in original selection order
        sr_no = 1
        failed_count = 0
        warned_count = 0
        success_count = 0

        for filename in result.attempted_files:
            r = sr_no + 1  # Row 2 onwards (row 1 is header)

            # Determine status for this file
            if filename in failed_map:
                so_display = f"❌ FAILED: {failed_map[filename]}"
                status = "❌ FAILED"
                failed_count += 1

                self._data_cell(ws, r, 1, sr_no)
                self._data_cell(ws, r, 2, filename)
                self._data_cell(ws, r, 3, so_display)
                self._data_cell(ws, r, 4, status)

                # Red highlight for failed rows
                for c in range(1, 5):
                    cell = ws.cell(row=r, column=c)
                    cell.fill = red_fill
                    cell.font = red_font

            elif filename in file_to_so:
                so_num = file_to_so[filename]

                if filename in warned_set:
                    status = "⚠️ WARNING"
                    warned_count += 1

                    self._data_cell(ws, r, 1, sr_no)
                    self._data_cell(ws, r, 2, filename)
                    self._data_cell(ws, r, 3, so_num)
                    self._data_cell(ws, r, 4, status)

                    # Yellow highlight for warned rows
                    for c in range(1, 5):
                        ws.cell(row=r, column=c).fill = yellow_fill

                else:
                    status = "✅ OK"
                    success_count += 1

                    self._data_cell(ws, r, 1, sr_no)
                    self._data_cell(ws, r, 2, filename)
                    self._data_cell(ws, r, 3, so_num)
                    self._data_cell(ws, r, 4, status)

            else:
                # Attempted but produced no rows and no failure —
                # shouldn't normally happen but handle gracefully
                so_display = "(no data extracted)"
                status = "❌ FAILED"
                failed_count += 1

                self._data_cell(ws, r, 1, sr_no)
                self._data_cell(ws, r, 2, filename)
                self._data_cell(ws, r, 3, so_display)
                self._data_cell(ws, r, 4, status)

                for c in range(1, 5):
                    cell = ws.cell(row=r, column=c)
                    cell.fill = red_fill
                    cell.font = red_font

            sr_no += 1

        # Summary row at the bottom
        summary_row = sr_no + 1
        bold = Font(name='Aptos Display', size=11, bold=True)

        ws.cell(row=summary_row, column=1, value='TOTAL').font = bold
        ws.cell(
            row=summary_row, column=2,
            value=f'{len(result.attempted_files)} file(s) attempted'
        ).font = bold
        ws.cell(
            row=summary_row, column=3,
            value=f'✅ {success_count} OK  |  ⚠️ {warned_count} warn  |  ❌ {failed_count} failed'
        ).font = bold
        ws.cell(row=summary_row, column=4, value='').font = bold

        for c in range(1, 5):
            ws.cell(row=summary_row, column=c).border = self.BORDER

        self._auto_width(ws)

    def _write_warnings(self, wb, result: ProcessResult):
        """
        Sheet 7: 'Warnings' — non-fatal issues. Only created if any exist.

        Critical warnings (containing '❌ CRITICAL') are highlighted
        with a red fill and bold red text so they stand out.
        """
        if not result.warned_files and not result.failed_files:
            return

        ws = wb.create_sheet('Warnings')

        headers = ['File', 'Type', 'Message']
        for c, h in enumerate(headers, 1):
            self._hdr_cell(ws, 1, c, h)

        red_fill = PatternFill('solid', fgColor='FFCDD2')
        red_font = Font(name='Aptos Display', size=11, bold=True, color='D32F2F')

        r = 2

        # First, write all failures (always critical red)
        for fname, reason in result.failed_files:
            self._data_cell(ws, r, 1, fname)
            self._data_cell(ws, r, 2, '❌ FAILED')
            self._data_cell(ws, r, 3, reason)

            for c in range(1, 4):
                ws.cell(row=r, column=c).fill = red_fill
                ws.cell(row=r, column=c).font = red_font

            r += 1

        # Then, write warnings
        for fname, warning in result.warned_files:
            is_critical = '❌ CRITICAL' in warning

            self._data_cell(ws, r, 1, fname)
            self._data_cell(
                ws, r, 2,
                '❌ CRITICAL' if is_critical else '⚠️ WARNING'
            )
            self._data_cell(ws, r, 3, warning)

            if is_critical:
                for c in range(1, 4):
                    ws.cell(row=r, column=c).fill = red_fill
                    ws.cell(row=r, column=c).font = red_font

            r += 1

        self._auto_width(ws)


# ═══════════════════════════════════════════════════════════════════════════════
#  FILE OPENER (cross-platform)
# ═══════════════════════════════════════════════════════════════════════════════

def open_file(file_path: Path):
    """
    Open a file using the OS default application.

    Windows: os.startfile (opens Excel for .xlsx)
    macOS:   'open' command
    Linux:   'xdg-open' command
    """
    try:
        system = platform.system()

        if system == "Windows":
            os.startfile(str(file_path))
        elif system == "Darwin":
            import subprocess as sp
            sp.Popen(["open", str(file_path)])
        else:
            import subprocess as sp
            sp.Popen(["xdg-open", str(file_path)])

    except (FileNotFoundError, OSError) as e:
        messagebox.showerror("Open File Error", f"Could not open:\n{e}")


# ═══════════════════════════════════════════════════════════════════════════════
#  MAIN AUTOMATION ENGINE
# ═══════════════════════════════════════════════════════════════════════════════

class GTMassAutomation:
    """
    Orchestrates file validation, parsing, and export.

    Processing flow:
        1. For each selected file:
           a. Record in attempted_files (v2.4)
           b. Validate against template (v2.4) — skip parsing if invalid
           c. Parse into OrderRows
           d. Collect warnings, rows, failures
        2. Return aggregated ProcessResult

    Error handling strategy:
        Specific exception types are caught first (expected failures
        like RuntimeError for format issues, data errors, OS errors).
        A final broad Exception handler exists only as a safety net
        so one bad file cannot crash the entire batch.
    """

    def __init__(self):
        """Initialize with fresh validator, parser, and exporter instances."""
        self.validator = TemplateValidator()
        self.parser = ExcelParser()
        self.exporter = DumpExporter()

    def process_files(self, files: List[Path]) -> ProcessResult:
        """
        Process all selected files and return an aggregated result.

        Pipeline per file:
            1. Record the filename in attempted_files
            2. TemplateValidator.validate() — skip rest if non-compliant
            3. ExcelParser.parse() — extract rows + warnings
            4. Append rows and warnings to the result

        Any exception during validation/parsing moves the file to
        failed_files with a descriptive reason. The batch continues.

        Args:
            files: List of Path objects pointing to selected Excel files

        Returns:
            ProcessResult with rows, warnings, failures, and attempted_files
        """
        result = ProcessResult()

        for file_path in files:
            fname = file_path.name

            # Record EVERY attempted file (v2.4)
            result.attempted_files.append(fname)

            # ── Template compliance check (v2.4) ──
            is_valid, reason = self.validator.validate(file_path)

            if not is_valid:
                result.failed_files.append((fname, reason))
                logging.error(f"{fname} REJECTED: {reason}")
                continue

            # ── Parse the file ──
            try:
                rows, warnings = self.parser.parse(file_path)
                result.rows.extend(rows)

                for w in warnings:
                    result.warned_files.append((fname, w))
                    logging.warning(f"{fname}: {w}")

            except RuntimeError as e:
                result.failed_files.append((fname, str(e)))
                logging.error(f"{fname} FAILED: {e}")

            except (ValueError, KeyError, TypeError) as e:
                result.failed_files.append((fname, f"Data error: {e}"))
                logging.error(f"{fname} DATA: {e}")

            except OSError as e:
                result.failed_files.append((fname, f"File error: {e}"))
                logging.error(f"{fname} FILE: {e}")

            except Exception as e:
                # Safety net — one bad file must not crash the batch
                result.failed_files.append((fname, f"Unexpected: {e}"))
                logging.error(f"{fname} UNEXPECTED: {e}")

        # ── Log summary ──
        unique_so_count = len({r.so_number for r in result.rows})
        logging.info(
            f"Done — {len(result.attempted_files)} attempted | "
            f"{len(result.rows)} rows | "
            f"{unique_so_count} SOs | "
            f"{len(result.failed_files)} failed | "
            f"{len(result.warned_files)} warnings"
        )

        return result


# ═══════════════════════════════════════════════════════════════════════════════
#  TKINTER UI
# ═══════════════════════════════════════════════════════════════════════════════

class AutomationUI:
    """
    Tkinter GUI for the GT Mass Dump Generator.

    Six buttons:
        1. Select Excel Files     → file chooser
        2. Generate Dump          → run pipeline + export
        3. Open Last Output File  → opens last generated file
        4. Download PO Template   → saves a blank template
        5. Export D365 Package    → fills D365 template
        6. Send Email Report      → sends HTML summary email

    Status line shows: selected count, processing state, time elapsed.
    """

    def __init__(self, automation: GTMassAutomation):
        """
        Build the UI window and all widgets.

        Args:
            automation: Shared engine instance for processing
        """
        self.automation = automation
        self.files: List[Path] = []
        self.last_output_path: Optional[Path] = None
        self.last_result: Optional[ProcessResult] = None
        self.last_elapsed: str = ""

        # ── Main window ──
        self.root = tk.Tk()
        self.root.title("GT Mass Dump Generator v2.4")
        self.root.geometry("460x520")
        self.root.resizable(False, False)

        # ── Title + subtitle ──
        tk.Label(
            self.root,
            text="GT Mass Dump Generator",
            font=("Arial", 14, "bold"),
        ).pack(pady=10)

        tk.Label(
            self.root,
            text="GT-Mass / Everyday PO Files → ERP Import",
            font=("Arial", 9),
            fg="gray",
        ).pack(pady=0)

        # ── File count display ──
        self.label = tk.Label(
            self.root,
            text="Selected Files: 0",
            font=("Arial", 10),
        )
        self.label.pack(pady=6)

        # ── Action buttons ──
        tk.Button(
            self.root, text="Select Excel Files", width=22,
            command=self.select_files,
        ).pack(pady=6)

        tk.Button(
            self.root, text="Generate Dump", width=22,
            command=self.generate_dump,
        ).pack(pady=6)

        self.open_button = tk.Button(
            self.root, text="Open Last Output File", width=22,
            state=tk.DISABLED,
            command=self.open_last_file,
        )
        self.open_button.pack(pady=6)

        tk.Button(
            self.root, text="📋 Download PO Template", width=22,
            command=self._download_template,
        ).pack(pady=6)

        tk.Button(
            self.root, text="📤 Export D365 Package", width=22,
            command=self._export_d365,
        ).pack(pady=6)

        tk.Button(
            self.root, text="📧 Send Email Report", width=22,
            command=self._send_email,
        ).pack(pady=6)

        # ── Status labels ──
        self.status = tk.Label(
            self.root,
            text="Status: Waiting",
            font=("Arial", 10),
            fg="gray",
        )
        self.status.pack(pady=6)

        self.time_label = tk.Label(
            self.root,
            text="",
            font=("Arial", 9),
            fg="darkgreen",
        )
        self.time_label.pack(pady=2)

    # ─────────────────────────────────────────────────────────────────────────
    #  ACTION HANDLERS
    # ─────────────────────────────────────────────────────────────────────────

    def select_files(self):
        """Open file chooser to select Excel files."""
        files = filedialog.askopenfilenames(
            title="Select Sales Order Files",
            filetypes=[("Excel Files", "*.xlsx *.xls *.xlsm")],
        )
        self.files = [Path(f) for f in files]
        self.label.config(text=f"Selected Files: {len(self.files)}")
        self.time_label.config(text="")
        self.status.config(text="Status: Files selected", fg="gray")

    def generate_dump(self):
        """Run validation + parsing + export pipeline and show result popup."""
        if not self.files:
            messagebox.showwarning("Warning", "Select files first.")
            return

        # ── Start ──
        start = time.time()
        self.status.config(text="Processing...", fg="blue")
        self.time_label.config(text="")
        self.root.update()

        # ── Process ──
        result = self.automation.process_files(self.files)
        self.last_result = result

        # ── Export ──
        output_path = self.automation.exporter.export(result)

        elapsed_str = f"{time.time() - start:.2f} seconds"
        self.last_elapsed = elapsed_str

        # ── Compute summary counts ──
        failed = len(result.failed_files)
        warned = len(result.warned_files)
        rows = len(result.rows)
        sos = len({r.so_number for r in result.rows}) if result.rows else 0
        attempted = len(result.attempted_files)

        # Count template violations specifically
        template_violations = sum(
            1 for _, reason in result.failed_files
            if 'Template violation' in reason
        )

        # Count SOs with missing Location Code
        missing_loc_sos = len({
            r.so_number for r in result.rows if not r.location_code
        })

        # ── Update status line ──
        if output_path:
            self.last_output_path = output_path
            self.open_button.config(state=tk.NORMAL)

            if failed > 0 or warned > 0:
                status_text = (
                    f"Done — {rows} rows | {failed} failed | {warned} warn"
                )
                status_color = "orange"
            else:
                status_text = f"Done — {rows} rows, {sos} SOs"
                status_color = "darkgreen"

            self.status.config(text=status_text, fg=status_color)
            self.time_label.config(text=f"⏱ {elapsed_str}")

            # ── Build popup message ──
            lines = [
                f"File   : {output_path.name}",
                f"Attempted : {attempted} file(s)",
                f"Rows      : {rows}",
                f"SOs       : {sos}",
                f"Time      : {elapsed_str}",
            ]

            if template_violations:
                lines.append(
                    f"\n🚫 {template_violations} file(s) don't meet template "
                    f"standard — see Warnings sheet"
                )

            if failed and failed > template_violations:
                other_failures = failed - template_violations
                lines.append(f"❌ {other_failures} other failure(s)")

            if warned:
                lines.append(f"⚠️ {warned} warning(s) — check Warnings sheet")

            if missing_loc_sos:
                lines.append(
                    f"🔴 {missing_loc_sos} SO(s) have EMPTY Location Code "
                    f"— fix source files!"
                )

            lines.append("\n📋 Check 'File → SO Mapping' for full traceability.")
            lines.append("\nOpen file?")

            answer = messagebox.askyesno("Done", '\n'.join(lines))

            if answer:
                open_file(output_path)

        else:
            self.status.config(text="No data", fg="red")
            self.time_label.config(text=f"⏱ {elapsed_str}")

    def open_last_file(self):
        """Open the most recently generated output file."""
        if self.last_output_path and self.last_output_path.exists():
            open_file(self.last_output_path)
        else:
            messagebox.showwarning(
                "Not Found",
                "File gone. Generate a new dump.",
            )

    def _export_d365(self):
        """Prompt for D365 template and fill it with processed data."""
        if not self.last_result or not self.last_result.rows:
            messagebox.showwarning("No Data", "Generate dump first.")
            return

        # Warn about missing Location Codes
        missing_loc_sos = sorted({
            r.so_number
            for r in self.last_result.rows
            if not r.location_code
        })

        if missing_loc_sos:
            proceed = messagebox.askyesno(
                "⚠️ Missing Location Codes",
                f"The following {len(missing_loc_sos)} SO(s) have EMPTY "
                f"Location Code:\n\n"
                + "\n".join(f"  • {s}" for s in missing_loc_sos) +
                f"\n\nD365 import may fail for these SOs.\n\n"
                f"Continue with export anyway?"
            )

            if not proceed:
                return

        template_path = filedialog.askopenfilename(
            title="Select D365 Template",
            filetypes=[("Excel", "*.xlsx")],
        )

        if not template_path:
            return

        d365 = self.automation.exporter.export_d365(
            self.last_result, template_path
        )

        if d365:
            sos = len({r.so_number for r in self.last_result.rows})
            items = len(self.last_result.rows)

            answer = messagebox.askyesno(
                "D365 Package Exported",
                f"D365 import file created successfully!\n\n"
                f"File   : {d365.name}\n"
                f"SOs    : {sos}\n"
                f"Items  : {items}\n\n"
                f"Do you want to open the exported file?"
            )

            if answer:
                open_file(d365)

    def _send_email(self):
        """Send the email report via SMTP."""
        if not self.last_result or not self.last_result.rows:
            messagebox.showwarning("No Data", "Generate dump first.")
            return

        self.status.config(text="Sending email...", fg="blue")
        self.root.update()

        ok, err = EmailSender.send_report(
            self.last_result, self.last_elapsed
        )

        if ok:
            self.status.config(text="Email sent ✓", fg="darkgreen")

            cc_list = ', '.join(EMAIL_CONFIG['CC_RECIPIENTS']) or 'none'

            messagebox.showinfo(
                "Email Sent",
                f"Report sent successfully!\n\n"
                f"To : {EMAIL_CONFIG['DEFAULT_RECIPIENT']}\n"
                f"CC : {cc_list}"
            )
        else:
            self.status.config(text="Email failed ✗", fg="red")
            messagebox.showerror("Failed", f"{err}")

    def _download_template(self):
        """Generate and save a blank GT-Mass PO template Excel file."""
        save_path = filedialog.asksaveasfilename(
            title="Save Template",
            defaultextension=".xlsx",
            initialfile="GT-Mass_PO_Template.xlsx",
            filetypes=[("Excel", "*.xlsx")],
        )

        if not save_path:
            return

        try:
            wb = Workbook()
            ws = wb.active
            ws.title = 'PO Template'

            # Font styles
            title_font = Font(name='Aptos Display', size=14, bold=True, color='1A237E')
            label_font = Font(name='Aptos Display', size=11, bold=True)
            value_font = Font(name='Aptos Display', size=11, color='0000CC')
            note_font = Font(name='Aptos Display', size=10, italic=True, color='FF6600')
            sample_font = Font(name='Aptos Display', size=11, color='888888', italic=True)
            hdr_font = Font(bold=True, color='FFFFFF', name='Aptos Display', size=11)
            crit_font = Font(name='Aptos Display', size=11, bold=True, color='D32F2F')
            crit_hdr_font = Font(bold=True, color='FFFFFF', name='Aptos Display', size=11)

            # Fills
            hdr_fill = PatternFill('solid', fgColor='1A237E')
            meta_fill = PatternFill('solid', fgColor='E3F2FD')
            crit_fill = PatternFill('solid', fgColor='FFCDD2')
            crit_hdr_fill = PatternFill('solid', fgColor='D32F2F')

            # ── Row 1: Title ──
            ws.cell(row=1, column=1, value='Purchase Order GT-Mass (Template)').font = title_font

            # ── Meta rows ──
            meta_layout = [
                (2, 'Distributor Name', '<Enter Distributor Name>'),
                (3, 'DB Code', '<DB Code>'),
                (5, 'City', '<City>'),
                (6, 'State', '<State>'),
            ]

            for r, label, value in meta_layout:
                lc = ws.cell(row=r, column=1, value=label)
                lc.font = label_font
                lc.fill = meta_fill

                vc = ws.cell(row=r, column=2, value=value)
                vc.font = value_font

            # Right-side meta (non-critical)
            ws.cell(row=2, column=7, value='ASM').font = label_font
            ws.cell(row=2, column=7).fill = meta_fill
            ws.cell(row=2, column=9, value='<ASM Name>').font = value_font

            ws.cell(row=3, column=7, value='RSM').font = label_font
            ws.cell(row=3, column=7).fill = meta_fill
            ws.cell(row=3, column=9, value='<RSM Name>').font = value_font

            ws.cell(row=4, column=1, value='BDE Name').font = label_font
            ws.cell(row=4, column=1).fill = meta_fill
            ws.cell(row=4, column=2, value='<BDE Name>').font = value_font

            # ── Critical fields: PO Number + Location (red) ──
            ws.cell(row=4, column=7, value='PO Number').font = crit_font
            ws.cell(row=4, column=7).fill = crit_fill
            ws.cell(row=4, column=9, value='SO/GTM/0000').font = crit_font
            ws.cell(row=4, column=9).fill = crit_fill

            ws.cell(row=5, column=7, value='Date of PO').font = label_font
            ws.cell(row=5, column=7).fill = meta_fill
            ws.cell(row=5, column=9, value='DD.MM.YYYY').font = value_font

            ws.cell(row=6, column=7, value='Location').font = crit_font
            ws.cell(row=6, column=7).fill = crit_fill
            ws.cell(row=6, column=9, value='AHD').font = crit_font
            ws.cell(row=6, column=9).fill = crit_fill

            # ── Row 7: Data header ──
            data_headers = [
                'EAN', 'BC Code', 'Category', 'Article Description ',
                'Nail Paint Shade Number ', 'Product Classification',
                'HSN Code\n8 Digit', 'MRP', 'Retiler Margin',
                'Trade & Display Scheme', 'Ullage', 'QPS',
                'Qty In Case', 'Rate @ RLP', 'Amount @ RLP',
                'Order Qty', 'Order Amount', 'Tester Qty',
            ]

            critical_cols = {'EAN', 'BC Code', 'Order Qty', 'Tester Qty'}

            for ci, h in enumerate(data_headers, 1):
                cell = ws.cell(row=7, column=ci, value=h)
                cell.alignment = Alignment(
                    horizontal='center', vertical='center', wrap_text=True
                )

                if h.strip() in critical_cols:
                    cell.font = crit_hdr_font
                    cell.fill = crit_hdr_fill
                else:
                    cell.font = hdr_font
                    cell.fill = hdr_fill

            # ── Row 8: Sample data (grey italic) ──
            sample_data = [
                8904473104307, 201238, 'Eye',
                'RENEE PURE BROWN KAJAL PEN WITH SHARPENER, 0.35GM',
                '-', 'Cosmetics', 33049990, 199, 1.2,
                '16.67% on RLP', '1.66 % on RLP', '4.81% on RLP',
                '', '', '', 72, '', 6,
            ]

            for ci, v in enumerate(sample_data, 1):
                ws.cell(row=8, column=ci, value=v).font = sample_font

            # ── Row 10: Instructions ──
            ws.cell(row=10, column=1, value='⚠ INSTRUCTIONS:').font = Font(
                name='Aptos Display', size=11, bold=True, color='D32F2F'
            )

            instructions = [
                '1. Fill PO Number (Row 4, Col I) SO/GTM/####',
                '2. Fill Location (Row 6, Col I) AHD/BLR',
                '3. Fill Distributor, City, State',
                '4. Data from Row 8, delete sample',
                '5. BC Code numeric, Qty numeric',
                '6. RED = critical fields',
                '7. Save .xlsx → load into generator',
            ]

            for i, ins in enumerate(instructions):
                ws.cell(row=11 + i, column=1, value=ins).font = note_font

            # Column widths
            widths = {
                'A': 16, 'B': 12, 'C': 12, 'D': 50, 'E': 12, 'F': 18,
                'G': 14, 'H': 8, 'I': 14, 'J': 20, 'K': 16, 'L': 14,
                'M': 12, 'N': 12, 'O': 14, 'P': 12, 'Q': 14, 'R': 12,
            }

            for col_letter, w in widths.items():
                ws.column_dimensions[col_letter].width = w

            ws.freeze_panes = 'A8'
            wb.save(save_path)

            messagebox.showinfo("Saved", f"Template: {save_path}")

        except (PermissionError, OSError) as e:
            messagebox.showerror("Error", f"Save failed: {e}")

    def run(self):
        """Start the Tkinter event loop (blocks until window closes)."""
        self.root.mainloop()


# ═══════════════════════════════════════════════════════════════════════════════
#  ENTRY POINT
# ═══════════════════════════════════════════════════════════════════════════════

def main():
    """
    Application entry point.

    Sequence:
        1. check_expiry()     — block if expired, warn if within 7 days
        2. GTMassAutomation() — build processing engine
        3. AutomationUI()     — build GUI window
        4. ui.run()           — start event loop (blocks)
    """
    check_expiry()
    automation = GTMassAutomation()
    ui = AutomationUI(automation)
    ui.run()


if __name__ == "__main__":
    main()