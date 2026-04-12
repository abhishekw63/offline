"""
Marketplace GRN / Return Note Parser — Tkinter GUI
====================================================
A scalable desktop app to extract SKU-level data from marketplace PDFs.

Supported Marketplaces:
    1. Blinkit  — GRN (Goods Receipt Note) PDFs
    2. Flipkart — Return Note PDFs

Architecture:
    - BaseMarketplaceParser (ABC) defines the interface
    - Each marketplace has its own parser subclass
    - MarketplaceRegistry auto-discovers and registers parsers
    - Adding a new marketplace = adding one new class

Requirements:
    pip install pdfplumber pandas openpyxl

Run:
    python marketplace_grn_parser.py
"""

import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import threading
import os
import glob
import re
import time
import abc
from dataclasses import dataclass, field
from typing import List, Dict, Any, Optional, Tuple, Type

import pdfplumber
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ═══════════════════════════════════════════════════════════════════════════════
#  CORE DATA STRUCTURES
# ═══════════════════════════════════════════════════════════════════════════════

@dataclass
class ParsedResult:
    """Standardized output from any marketplace parser."""
    marketplace: str
    header: Dict[str, Any]
    items: pd.DataFrame
    summary: Dict[str, Any]
    raw_text: str = ""


# ═══════════════════════════════════════════════════════════════════════════════
#  BASE PARSER (Abstract)
# ═══════════════════════════════════════════════════════════════════════════════

class BaseMarketplaceParser(abc.ABC):
    """
    Abstract base class for all marketplace parsers.
    To add a new marketplace:
        1. Subclass this
        2. Implement all abstract methods/properties
        3. The registry auto-discovers it
    """

    @property
    @abc.abstractmethod
    def marketplace_name(self) -> str:
        """Display name, e.g. 'Blinkit', 'Flipkart'."""
        ...

    @property
    @abc.abstractmethod
    def document_type(self) -> str:
        """Document type label, e.g. 'GRN', 'Return Note'."""
        ...

    @property
    @abc.abstractmethod
    def item_sheet_name(self) -> str:
        """Excel sheet name for line items."""
        ...

    @property
    @abc.abstractmethod
    def summary_sheet_name(self) -> str:
        """Excel sheet name for summary."""
        ...

    @property
    @abc.abstractmethod
    def column_widths(self) -> List[int]:
        """Column widths for Excel formatting."""
        ...

    @property
    def status_column_name(self) -> Optional[str]:
        """Column with status values for color-coding. None to skip."""
        return None

    @property
    def status_colors(self) -> Dict[str, str]:
        """Map status value → hex color for Excel cell fill."""
        return {}

    @abc.abstractmethod
    def parse(self, pdf_path: str) -> ParsedResult:
        """Parse one PDF file → ParsedResult."""
        ...

    @abc.abstractmethod
    def get_stats(self, combined_df: pd.DataFrame) -> Dict[str, Any]:
        """Compute stats from the combined DataFrame for the GUI panel."""
        ...

    @abc.abstractmethod
    def get_summary_columns(self) -> List[str]:
        """Ordered list of columns for the summary sheet."""
        ...

    @classmethod
    def detect(cls, pdf_path: str) -> bool:
        """Auto-detect if a PDF belongs to this marketplace. Override for auto-detect."""
        return False


# ═══════════════════════════════════════════════════════════════════════════════
#  MARKETPLACE REGISTRY
# ═══════════════════════════════════════════════════════════════════════════════

class MarketplaceRegistry:
    """Auto-discovers all BaseMarketplaceParser subclasses."""

    _parsers: Dict[str, Type[BaseMarketplaceParser]] = {}

    @classmethod
    def register(cls, parser_class: Type[BaseMarketplaceParser]):
        instance = parser_class()
        cls._parsers[instance.marketplace_name] = parser_class
        return parser_class

    @classmethod
    def get_names(cls) -> List[str]:
        return sorted(cls._parsers.keys())

    @classmethod
    def get_parser(cls, name: str) -> BaseMarketplaceParser:
        return cls._parsers[name]()

    @classmethod
    def auto_detect(cls, pdf_path: str) -> Optional[str]:
        for name, klass in cls._parsers.items():
            if klass.detect(pdf_path):
                return name
        return None


# ═══════════════════════════════════════════════════════════════════════════════
#  UTILITY FUNCTIONS
# ═══════════════════════════════════════════════════════════════════════════════

def clean_upc(raw):
    return re.sub(r'\s+', '', str(raw))


def clean_number(val):
    if val is None or str(val).strip() in ('-', '', 'None'):
        return 0.0
    try:
        return float(str(val).replace('\n', '').replace(',', '').strip())
    except (ValueError, TypeError):
        return 0.0


# ═══════════════════════════════════════════════════════════════════════════════
#  BLINKIT PARSER
# ═══════════════════════════════════════════════════════════════════════════════

@MarketplaceRegistry.register
class BlinkitParser(BaseMarketplaceParser):

    @property
    def marketplace_name(self) -> str:
        return "Blinkit"

    @property
    def document_type(self) -> str:
        return "GRN"

    @property
    def item_sheet_name(self) -> str:
        return "GRN Line Items"

    @property
    def summary_sheet_name(self) -> str:
        return "PO Summary"

    @property
    def column_widths(self) -> List[int]:
        return [16, 10, 8, 10, 14, 16, 52, 10, 14, 10, 10, 12, 14, 12, 16, 28]

    @property
    def status_column_name(self) -> str:
        return "Line GRN Status"

    @property
    def status_colors(self) -> Dict[str, str]:
        return {'Full GRN': '00C853', 'Partial GRN': 'FFB300', 'Not GRNed': 'D50000'}

    def get_summary_columns(self) -> List[str]:
        return [
            'PO Number', 'PO Date', 'Facility',
            'Total PO Qty', 'Total GRN Qty', 'Fill Rate %',
            'Articles in PO', 'Articles in GRN',
            'Total PO Amount', 'Net GRN Amount', 'GMV Loss',
        ]

    def get_stats(self, df: pd.DataFrame) -> Dict[str, Any]:
        vc = df['Line GRN Status'].value_counts()
        return {
            'pos':     str(df['PO Number'].nunique()),
            'skus':    str(len(df)),
            'full':    str(vc.get('Full GRN', 0)),
            'partial': str(vc.get('Partial GRN', 0)),
            'not_grn': str(vc.get('Not GRNed', 0)),
            'grn_qty': f"{int(df['GRN Qty'].sum()):,}",
        }

    @classmethod
    def detect(cls, pdf_path: str) -> bool:
        try:
            with pdfplumber.open(pdf_path) as pdf:
                text = (pdf.pages[0].extract_text() or "").lower()
                return 'blinkit' in text or 'bcpl' in text or 'p.o. number' in text
        except Exception:
            return False

    def _extract_header(self, pdf) -> Dict[str, Any]:
        text = pdf.pages[0].extract_text() or ""
        po  = re.search(r'P\.O\.\s*Number\s*[:\s]+(\d+)', text)
        dt  = re.search(r'Date\s*[:\s]+([\w.]+\s+\d+,\s+\d{4})', text)
        fac = re.search(r'BCPL\s*-\s*(.+?)(?:\n|Contact)', text)
        return {
            'PO Number': po.group(1).strip()  if po  else 'UNKNOWN',
            'PO Date':   dt.group(1).strip()  if dt  else '',
            'Facility':  fac.group(1).strip() if fac else '',
        }

    def _extract_summary(self, all_text: str) -> Dict[str, Any]:
        def find(pattern):
            m = re.search(pattern, all_text, re.IGNORECASE)
            return clean_number(m.group(1)) if m else 0.0
        fr = re.search(r'Fill rate:\s*([\d.]+)%', all_text)
        return {
            'Total PO Qty':    find(r'Total Quantity in PO:\s*([\d,]+)'),
            'Total GRN Qty':   find(r'Total Quantity in GRN\(s\):\s*([\d,]+)'),
            'Fill Rate %':     float(fr.group(1)) if fr else 0.0,
            'Articles in PO':  find(r'Articles in PO:\s*([\d,]+)'),
            'Articles in GRN': find(r'Articles in GRN\(s\):\s*([\d,]+)'),
            'Total PO Amount': find(r'Total Amount in PO\s+([\d,\.]+)'),
            'Net GRN Amount':  find(r'Net amt\. by GRN\s+([\d,\.]+)'),
            'GMV Loss':        find(r'Potential GMV Loss \(in INR\)\s+([\d,\.]+)'),
        }

    def parse(self, pdf_path: str) -> ParsedResult:
        rows = []
        all_text = ""
        with pdfplumber.open(pdf_path) as pdf:
            header = self._extract_header(pdf)
            for page in pdf.pages:
                all_text += (page.extract_text() or "") + "\n"
                for table in page.extract_tables():
                    for row in table:
                        if not row or row[0] is None:
                            continue
                        if not re.match(r'^\d+$', str(row[0]).strip()):
                            continue
                        try:
                            upc     = clean_upc(row[2] if len(row) > 2 else '')
                            desc    = str(row[3] or '').replace('\n', ' ').strip()
                            po_qty  = int(clean_number(row[8]))  if len(row) > 8  else 0
                            grn_qty = int(clean_number(row[9]))  if len(row) > 9  else 0
                            mrp     = clean_number(row[4])        if len(row) > 4  else 0.0
                            lr      = clean_number(row[6])        if len(row) > 6  else 0.0
                            fr_raw  = str(row[10] or '').strip()  if len(row) > 10 else '-'
                            fr      = clean_number(fr_raw) if fr_raw != '-' else 0.0
                            grn_amt = clean_number(row[11]) if len(row) > 11 else 0.0
                            gmv     = clean_number(row[12]) if len(row) > 12 else 0.0

                            if grn_qty == 0:
                                status = 'Not GRNed'
                            elif grn_qty < po_qty:
                                status = 'Partial GRN'
                            else:
                                status = 'Full GRN'

                            rows.append({
                                'PO Number':       header['PO Number'],
                                'PO Date':         header['PO Date'],
                                'Facility':        header['Facility'],
                                'Sr No':           int(row[0]),
                                'Item Code':       str(row[1] or '').strip(),
                                'UPC / GTIN':      upc,
                                'Description':     desc,
                                'MRP':             mrp,
                                'Landing Rate':    lr,
                                'PO Qty':          po_qty,
                                'GRN Qty':         grn_qty,
                                'Fill Rate %':     fr,
                                'GRN Amount':      grn_amt,
                                'GMV Loss':        gmv,
                                'Line GRN Status': status,
                                'PO<>EAN':         f"{header['PO Number']}<>{upc}",
                            })
                        except Exception:
                            pass

        summary = self._extract_summary(all_text)
        header.update(summary)
        return ParsedResult(
            marketplace=self.marketplace_name,
            header=header,
            items=pd.DataFrame(rows),
            summary=summary,
            raw_text=all_text,
        )


# ═══════════════════════════════════════════════════════════════════════════════
#  FLIPKART PARSER
# ═══════════════════════════════════════════════════════════════════════════════

@MarketplaceRegistry.register
class FlipkartParser(BaseMarketplaceParser):

    @property
    def marketplace_name(self) -> str:
        return "Flipkart"

    @property
    def document_type(self) -> str:
        return "Return Note"

    @property
    def item_sheet_name(self) -> str:
        return "Return Note Items"

    @property
    def summary_sheet_name(self) -> str:
        return "Return Summary"

    @property
    def column_widths(self) -> List[int]:
        return [14, 22, 20, 10, 14, 20, 50, 16, 8, 10, 10, 14, 10, 10, 10, 10, 10, 12, 14, 28]

    @property
    def status_column_name(self) -> str:
        return "Section"

    @property
    def status_colors(self) -> Dict[str, str]:
        return {
            'A': 'FF5252',    # Debit (short supply etc.) — Red
            'B': 'FFB300',    # Consignment return — Amber
            'C': '7B61FF',    # Unbilled rejected — Purple
            'D': '00C853',    # Credit — Green
        }

    def get_summary_columns(self) -> List[str]:
        return [
            'Return Note No', 'Return Note Date', 'Return Order Date',
            'PO Number', 'Vendor Invoice No', 'Vendor Invoice Date',
            'IRN Id', 'Debit Note Ids', 'Vendor GSTIN',
            'Total Items', 'Total Qty', 'Net Debit Amount',
            'Net Consignment Amount',
        ]

    def get_stats(self, df: pd.DataFrame) -> Dict[str, Any]:
        sec_counts = df['Section'].value_counts()
        return {
            'pos':     str(df['Return Note No'].nunique()),
            'skus':    str(len(df)),
            'full':    str(sec_counts.get('A', 0)),     # Section A (Debit)
            'partial': str(sec_counts.get('B', 0)),     # Section B (Consignment)
            'not_grn': str(sec_counts.get('C', 0)),     # Section C (Unbilled)
            'grn_qty': f"{int(df['Qty'].sum()):,}",
        }

    @classmethod
    def detect(cls, pdf_path: str) -> bool:
        try:
            with pdfplumber.open(pdf_path) as pdf:
                text = (pdf.pages[0].extract_text() or "").lower()
                return 'flipkart' in text and 'return note' in text
        except Exception:
            return False

    def _extract_header(self, pdf) -> Dict[str, Any]:
        text = pdf.pages[0].extract_text() or ""

        def find(pattern, default=''):
            m = re.search(pattern, text, re.IGNORECASE)
            return m.group(1).strip() if m else default

        return {
            'Return Note No':    find(r'Flipkart Return Note No:\s*(\S+)'),
            'Return Note Date':  find(r'Flipkart Return Note Date:\s*(\S+)'),
            'Return Order Date': find(r'Return Order Creation Date:\s*(\S+)'),
            'PO Number':         find(r'PO Number:\s*(\S+)'),
            'Vendor Invoice No': find(r'Vendor Invoice Number:\s*(\S+)'),
            'Vendor Invoice Date': find(r'Vendor Invoice Date:\s*(\S+)'),
            'IRN Id':            find(r'IRN Id:\s*(\S+)'),
            'Debit Note Ids':    find(r'Debit Note Ids:\s*(\S+)'),
            'Vendor GSTIN':      find(r'Vendor Gstin\s*-\s*(\S+)'),
            'Vendor Name':       find(r'^(.*?)\n.*HK ESTATE', ''),
        }

    def _parse_fsn_and_description(self, particulars_cell: str) -> Tuple[str, str, str]:
        """
        Parse the merged FSN + HSN + Description from the Particulars column.

        Full item cell looks like:
            PERHFY4JVY4BGH        ← FSN part 1
            H8                    ← FSN part 2 (line-wrapped)
            33030050              ← HSN code (pure digits)
            Renee Florl Fest ...  ← Description

        Section C (Unbilled) cells may only have FSN:
            PERGGNX6XPHZU
            YH9

        Returns: (fsn, hsn_code, description)
        """
        if not particulars_cell:
            return ('', '', '')

        lines = [l.strip() for l in particulars_cell.split('\n') if l.strip()]

        if not lines:
            return ('', '', '')

        # Strategy:
        # 1. Collect leading lines that are pure uppercase alphanumeric → FSN parts
        # 2. If we hit a pure-digit line (4-8 digits) → HSN code
        # 3. Everything after HSN → description
        # 4. If all lines are alphanumeric (no HSN, no desc) → it's FSN-only

        fsn_parts = []
        hsn_code = ''
        desc_start_idx = len(lines)  # default: no description

        for i, line in enumerate(lines):
            # Check if this line is the HSN code (pure digits, 4-8 chars)
            if re.match(r'^\d{4,8}$', line):
                hsn_code = line
                desc_start_idx = i + 1
                break
            elif re.match(r'^[A-Z0-9]+$', line):
                # Part of FSN (uppercase alphanumeric)
                fsn_parts.append(line)
            else:
                # Reached a description line (mixed case, spaces, etc.)
                desc_start_idx = i
                break

        fsn = ''.join(fsn_parts)
        description = ' '.join(lines[desc_start_idx:]).strip()

        # Fallback: if nothing parsed as FSN, use first line
        if not fsn and lines:
            fsn = re.sub(r'\s+', '', lines[0])

        return (fsn, hsn_code, description)

    def _extract_totals(self, all_text: str) -> Dict[str, Any]:
        """Extract grand totals from the Return Note."""
        net_debit = 0.0
        net_consignment = 0.0

        # Net Debit Note Detail (A+B-D)
        m = re.search(r'Net Debit Note Detail.*?(\d[\d,.]+)\s*INR', all_text, re.DOTALL)
        if m:
            net_debit = clean_number(m.group(1))

        # Net Consignment Detail (B+C-D)
        m = re.search(r'Net Consignment Detail.*?(\d[\d,.]+)\s*INR', all_text, re.DOTALL)
        if m:
            net_consignment = clean_number(m.group(1))

        return {
            'Net Debit Amount': net_debit,
            'Net Consignment Amount': net_consignment,
        }

    def parse(self, pdf_path: str) -> ParsedResult:
        rows = []
        all_text = ""

        with pdfplumber.open(pdf_path) as pdf:
            header = self._extract_header(pdf)

            for page in pdf.pages:
                page_text = page.extract_text() or ""
                all_text += page_text + "\n"

                for table in page.extract_tables():
                    for row in table:
                        if not row or len(row) < 6:
                            continue

                        section = str(row[0] or '').strip()
                        sno     = str(row[1] or '').strip()

                        # Skip header rows and non-data rows
                        if section not in ('A', 'B', 'C', 'D'):
                            continue
                        if not re.match(r'^\d+$', sno):
                            continue

                        try:
                            debit_note_id = str(row[2] or '').replace('\n', '').strip()
                            particulars_raw = str(row[3] or '')
                            reason = str(row[4] or '').replace('\n', ' ').strip()
                            qty = int(clean_number(row[5]))
                            currency = str(row[6] or '').strip() if len(row) > 6 else 'INR'
                            unit_price = clean_number(row[7]) if len(row) > 7 else 0.0
                            taxable = clean_number(row[8]) if len(row) > 8 else 0.0

                            # Tax fields — may be empty for Section C
                            igst_pct = clean_number(row[9])   if len(row) > 9  else 0.0
                            cgst_pct = clean_number(row[10])  if len(row) > 10 else 0.0
                            sgst_pct = clean_number(row[11])  if len(row) > 11 else 0.0
                            cess     = clean_number(row[12])  if len(row) > 12 else 0.0
                            tcs      = clean_number(row[13])  if len(row) > 13 else 0.0
                            tax_amt  = clean_number(row[14])  if len(row) > 14 else 0.0
                            total    = clean_number(row[15])  if len(row) > 15 else 0.0

                            # Parse FSN, HSN, Description from merged Particulars
                            fsn, hsn_code, description = self._parse_fsn_and_description(particulars_raw)

                            rows.append({
                                'Return Note No':   header.get('Return Note No', ''),
                                'Return Note Date': header.get('Return Note Date', ''),
                                'PO Number':        header.get('PO Number', ''),
                                'Section':          section,
                                'Sr No':            int(sno),
                                'Debit Note ID':    debit_note_id,
                                'FSN':              fsn,
                                'HSN Code':         hsn_code,
                                'Description':      description,
                                'Reason':           reason,
                                'Qty':              qty,
                                'Currency':         currency,
                                'Unit Price':       unit_price,
                                'Taxable Amount':   taxable,
                                'IGST %':           igst_pct,
                                'CGST %':           cgst_pct,
                                'SGST %':           sgst_pct,
                                'Cess':             cess,
                                'TCS':              tcs,
                                'Tax Amount':       tax_amt,
                                'Total':            total,
                            })
                        except Exception:
                            pass

        totals = self._extract_totals(all_text)
        summary = {**header}
        summary['Total Items'] = len(rows)
        summary['Total Qty'] = sum(r['Qty'] for r in rows)
        summary.update(totals)

        return ParsedResult(
            marketplace=self.marketplace_name,
            header=summary,
            items=pd.DataFrame(rows),
            summary=totals,
            raw_text=all_text,
        )


# ═══════════════════════════════════════════════════════════════════════════════
#  EXCEL FORMATTER
# ═══════════════════════════════════════════════════════════════════════════════

class ExcelFormatter:
    """Marketplace-agnostic Excel formatter."""

    HEADER_FILL = PatternFill('solid', start_color='1A237E')
    HEADER_FONT = Font(bold=True, color='FFFFFF', name='Courier New', size=9)
    THIN_SIDE   = Side(style='thin', color='CCCCCC')
    BORDER      = Border(left=THIN_SIDE, right=THIN_SIDE, top=THIN_SIDE, bottom=THIN_SIDE)

    @classmethod
    def format(cls, output_path: str, parser: BaseMarketplaceParser):
        wb = load_workbook(output_path)

        # Format line items sheet
        ws = wb[parser.item_sheet_name]
        cls._format_header_row(ws)
        cls._apply_column_widths(ws, parser.column_widths)
        cls._apply_status_colors(ws, parser.status_column_name, parser.status_colors)
        cls._format_data_rows(ws)
        ws.freeze_panes = 'A2'
        ws.auto_filter.ref = ws.dimensions

        # Format summary sheet
        ws2 = wb[parser.summary_sheet_name]
        cls._format_header_row(ws2)
        cls._format_data_rows(ws2)
        ws2.freeze_panes = 'A2'
        for i in range(1, ws2.max_column + 1):
            ws2.column_dimensions[get_column_letter(i)].width = 22

        wb.save(output_path)

    @classmethod
    def _format_header_row(cls, ws):
        for cell in ws[1]:
            cell.fill = cls.HEADER_FILL
            cell.font = cls.HEADER_FONT
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = cls.BORDER
        ws.row_dimensions[1].height = 28

    @classmethod
    def _format_data_rows(cls, ws):
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            for cell in row:
                cell.border = cls.BORDER
                cell.font = Font(name='Courier New', size=9)
                cell.alignment = Alignment(vertical='center')

    @classmethod
    def _apply_column_widths(cls, ws, widths: List[int]):
        for i, w in enumerate(widths[:ws.max_column], 1):
            ws.column_dimensions[get_column_letter(i)].width = w

    @classmethod
    def _apply_status_colors(cls, ws, col_name: Optional[str], colors: Dict[str, str]):
        if not col_name or not colors:
            return
        sc = None
        for cell in ws[1]:
            if cell.value == col_name:
                sc = cell.column
                break
        if not sc:
            return
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            sv = str(row[sc - 1].value or '')
            c = colors.get(sv)
            if c:
                row[sc - 1].fill = PatternFill('solid', start_color=c)
                row[sc - 1].font = Font(
                    name='Courier New', size=9, bold=True,
                    color='FFFFFF' if sv in ('Not GRNed', 'A') else '000000'
                )
                row[sc - 1].alignment = Alignment(horizontal='center', vertical='center')


# ═══════════════════════════════════════════════════════════════════════════════
#  COLORS & FONTS (GUI)
# ═══════════════════════════════════════════════════════════════════════════════

BG       = "#0F1117"
SURFACE  = "#1A1D27"
SURFACE2 = "#22263A"
ACCENT   = "#00D4FF"
ACCENT2  = "#7B61FF"
GREEN    = "#00E676"
RED      = "#FF5252"
AMBER    = "#FFB300"
TEXT     = "#E8EAF6"
TEXT_DIM = "#6B7280"
BORDER_C = "#2D3250"

FONT_TITLE = ("Courier New", 18, "bold")
FONT_SUB   = ("Courier New", 10)
FONT_LABEL = ("Courier New", 9, "bold")
FONT_MONO  = ("Courier New", 9)
FONT_BTN   = ("Courier New", 10, "bold")


# ═══════════════════════════════════════════════════════════════════════════════
#  GUI APPLICATION
# ═══════════════════════════════════════════════════════════════════════════════

class MarketplaceParserApp(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Marketplace GRN / Return Note Parser")
        self.geometry("960x720")
        self.resizable(True, True)
        self.configure(bg=BG)
        self.minsize(800, 580)

        self.pdf_files: List[str]  = []
        self.last_output: Optional[str] = None
        self.is_running = False

        self._build_ui()

    # ── UI CONSTRUCTION ────────────────────────────────────────────────────────

    def _build_ui(self):
        # Header
        hdr = tk.Frame(self, bg=SURFACE, height=60)
        hdr.pack(fill='x')
        hdr.pack_propagate(False)
        tk.Label(hdr, text="▶ MARKETPLACE PARSER", font=FONT_TITLE,
                 bg=SURFACE, fg=ACCENT).pack(side='left', padx=20, pady=12)
        tk.Label(hdr, text="PDF → Excel  //  Multi-Marketplace Extraction",
                 font=FONT_SUB, bg=SURFACE, fg=TEXT_DIM).pack(side='left', padx=4)

        # Body
        body = tk.Frame(self, bg=BG)
        body.pack(fill='both', expand=True, padx=16, pady=12)
        body.columnconfigure(0, weight=3)
        body.columnconfigure(1, weight=2)
        body.rowconfigure(0, weight=1)

        left = tk.Frame(body, bg=BG)
        right = tk.Frame(body, bg=BG)
        left.grid(row=0, column=0, sticky='nsew', padx=(0, 8))
        right.grid(row=0, column=1, sticky='nsew')

        self._build_left(left)
        self._build_right(right)
        self._build_bottom()

    def _build_left(self, parent):
        # ── Marketplace Selector ──
        self._section(parent, "00  //  MARKETPLACE")
        sel_frame = tk.Frame(parent, bg=SURFACE, highlightthickness=1,
                             highlightbackground=BORDER_C)
        sel_frame.pack(fill='x', pady=(0, 10))

        sel_inner = tk.Frame(sel_frame, bg=SURFACE)
        sel_inner.pack(fill='x', padx=12, pady=10)

        tk.Label(sel_inner, text="SELECT MARKETPLACE:", font=FONT_LABEL,
                 bg=SURFACE, fg=TEXT_DIM).pack(anchor='w')

        self.marketplace_var = tk.StringVar(value=MarketplaceRegistry.get_names()[0])
        mp_row = tk.Frame(sel_inner, bg=SURFACE)
        mp_row.pack(fill='x', pady=(4, 0))

        for name in MarketplaceRegistry.get_names():
            rb = tk.Radiobutton(
                mp_row, text=name.upper(), variable=self.marketplace_var,
                value=name, font=FONT_BTN, bg=SURFACE, fg=ACCENT2,
                selectcolor=SURFACE2, activebackground=SURFACE,
                activeforeground=ACCENT, indicatoron=True,
                highlightthickness=0, bd=0
            )
            rb.pack(side='left', padx=(0, 20))

        # Auto-detect label
        self.detect_label = tk.Label(
            sel_inner, text="TIP: Parser auto-detects marketplace if you're unsure",
            font=("Courier New", 8), bg=SURFACE, fg=TEXT_DIM
        )
        self.detect_label.pack(anchor='w', pady=(4, 0))

        # ── PDF Files ──
        self._section(parent, "01  //  PDF FILES")

        dz = tk.Frame(parent, bg=SURFACE2, relief='flat', bd=0,
                      highlightthickness=1, highlightbackground=BORDER_C)
        dz.pack(fill='x', pady=(0, 8))
        dz_inner = tk.Frame(dz, bg=SURFACE2)
        dz_inner.pack(fill='x', padx=1, pady=1)

        tk.Label(dz_inner, text="DROP ZONE", font=("Courier New", 9, "bold"),
                 bg=SURFACE2, fg=ACCENT2).pack(pady=(10, 2))
        tk.Label(dz_inner, text="Add GRN / Return Note PDFs",
                 font=FONT_MONO, bg=SURFACE2, fg=TEXT_DIM).pack(pady=(0, 10))

        btn_row = tk.Frame(dz_inner, bg=SURFACE2)
        btn_row.pack(pady=(0, 10))
        self._btn(btn_row, "+ ADD FILES", self._add_files, ACCENT).pack(side='left', padx=4)
        self._btn(btn_row, "+ ADD FOLDER", self._add_folder, ACCENT2).pack(side='left', padx=4)
        self._btn(btn_row, "✕ CLEAR ALL", self._clear_files, RED).pack(side='left', padx=4)

        # File list
        list_frame = tk.Frame(parent, bg=SURFACE,
                              highlightthickness=1, highlightbackground=BORDER_C)
        list_frame.pack(fill='both', expand=True, pady=(0, 8))

        list_hdr = tk.Frame(list_frame, bg=SURFACE2)
        list_hdr.pack(fill='x')
        tk.Label(list_hdr, text="  FILE", font=FONT_LABEL, bg=SURFACE2, fg=TEXT_DIM,
                 width=38, anchor='w').pack(side='left', padx=4, pady=4)
        tk.Label(list_hdr, text="STATUS", font=FONT_LABEL, bg=SURFACE2, fg=TEXT_DIM
                 ).pack(side='right', padx=12, pady=4)

        scroll_y = ttk.Scrollbar(list_frame, orient='vertical')
        scroll_y.pack(side='right', fill='y')
        self.file_list = tk.Listbox(
            list_frame, bg=SURFACE, fg=TEXT, font=FONT_MONO,
            selectbackground=SURFACE2, selectforeground=ACCENT,
            borderwidth=0, highlightthickness=0,
            yscrollcommand=scroll_y.set, activestyle='none'
        )
        self.file_list.pack(fill='both', expand=True, padx=4, pady=4)
        scroll_y.config(command=self.file_list.yview)

        self.file_count_var = tk.StringVar(value="0 files loaded")
        tk.Label(parent, textvariable=self.file_count_var,
                 font=FONT_MONO, bg=BG, fg=TEXT_DIM).pack(anchor='w')

    def _build_right(self, parent):
        # Output
        self._section(parent, "02  //  OUTPUT")
        out_frame = tk.Frame(parent, bg=SURFACE,
                             highlightthickness=1, highlightbackground=BORDER_C)
        out_frame.pack(fill='x', pady=(0, 12))

        tk.Label(out_frame, text="Output folder:", font=FONT_LABEL,
                 bg=SURFACE, fg=TEXT_DIM).pack(anchor='w', padx=10, pady=(8, 2))
        tk.Label(out_frame, text="  output_parsed/  (auto-created next to script)",
                 font=FONT_MONO, bg=SURFACE, fg=ACCENT2).pack(anchor='w', padx=10)
        tk.Label(out_frame, text="Filename:", font=FONT_LABEL,
                 bg=SURFACE, fg=TEXT_DIM).pack(anchor='w', padx=10, pady=(6, 2))
        tk.Label(out_frame, text="  {marketplace}_ddmmyyyy_hhmmss.xlsx",
                 font=FONT_MONO, bg=SURFACE, fg=TEXT_DIM).pack(anchor='w', padx=10)

        self.last_path_var = tk.StringVar(value="No run yet")
        tk.Label(out_frame, text="Last saved:", font=FONT_LABEL,
                 bg=SURFACE, fg=TEXT_DIM).pack(anchor='w', padx=10, pady=(6, 2))
        tk.Label(out_frame, textvariable=self.last_path_var,
                 font=FONT_MONO, bg=SURFACE, fg=GREEN,
                 wraplength=280, justify='left').pack(anchor='w', padx=10, pady=(0, 10))

        # Stats
        self._section(parent, "03  //  LAST RUN STATS")
        stats_frame = tk.Frame(parent, bg=SURFACE,
                               highlightthickness=1, highlightbackground=BORDER_C)
        stats_frame.pack(fill='x', pady=(0, 12))

        self.stat_labels_config = {
            'Blinkit': [
                ("POs Processed", "pos"),
                ("Total SKUs",    "skus"),
                ("Full GRN",      "full"),
                ("Partial GRN",   "partial"),
                ("Not GRNed",     "not_grn"),
                ("Total GRN Qty", "grn_qty"),
            ],
            'Flipkart': [
                ("Return Notes",    "pos"),
                ("Total Items",     "skus"),
                ("Section A (Debit)", "full"),
                ("Section B (Cons.)", "partial"),
                ("Section C (Unbill)", "not_grn"),
                ("Total Qty",       "grn_qty"),
            ],
        }

        # We use the same 6 stat slots, just rename labels per marketplace
        self.stat_vars = {}
        self.stat_label_widgets = {}
        stat_colors = {
            "full": GREEN, "partial": AMBER, "not_grn": RED,
            "pos": ACCENT, "skus": ACCENT2, "grn_qty": TEXT
        }

        # Default labels
        default_labels = self.stat_labels_config['Blinkit']
        for i, (label, key) in enumerate(default_labels):
            row = tk.Frame(stats_frame, bg=SURFACE2 if i % 2 == 0 else SURFACE)
            row.pack(fill='x')
            lbl_widget = tk.Label(row, text=f"  {label}", font=FONT_MONO,
                                  bg=row['bg'], fg=TEXT_DIM, width=20, anchor='w')
            lbl_widget.pack(side='left', pady=4, padx=4)
            self.stat_label_widgets[key] = lbl_widget

            var = tk.StringVar(value="—")
            self.stat_vars[key] = var
            tk.Label(row, textvariable=var, font=("Courier New", 11, "bold"),
                     bg=row['bg'], fg=stat_colors.get(key, TEXT)).pack(side='right', padx=12, pady=4)

        # Update labels when marketplace changes
        self.marketplace_var.trace_add('write', self._on_marketplace_change)

        # Log
        self._section(parent, "04  //  LOG")
        log_frame = tk.Frame(parent, bg=SURFACE,
                             highlightthickness=1, highlightbackground=BORDER_C)
        log_frame.pack(fill='both', expand=True)
        scroll_log = ttk.Scrollbar(log_frame, orient='vertical')
        scroll_log.pack(side='right', fill='y')
        self.log_text = tk.Text(
            log_frame, bg=SURFACE, fg=TEXT_DIM, font=FONT_MONO,
            height=6, wrap='word', state='disabled',
            borderwidth=0, highlightthickness=0,
            yscrollcommand=scroll_log.set
        )
        self.log_text.pack(fill='both', expand=True, padx=6, pady=6)
        scroll_log.config(command=self.log_text.yview)
        self.log_text.tag_config('ok',  foreground=GREEN)
        self.log_text.tag_config('err', foreground=RED)
        self.log_text.tag_config('inf', foreground=ACCENT)
        self.log_text.tag_config('dim', foreground=TEXT_DIM)

    def _build_bottom(self):
        bottom = tk.Frame(self, bg=SURFACE, height=64)
        bottom.pack(fill='x', side='bottom')
        bottom.pack_propagate(False)

        pb_frame = tk.Frame(bottom, bg=SURFACE)
        pb_frame.pack(fill='x', padx=16, pady=(8, 0))
        self.progress_canvas = tk.Canvas(pb_frame, height=4, bg=SURFACE2,
                                         highlightthickness=0)
        self.progress_canvas.pack(fill='x')

        ctrl = tk.Frame(bottom, bg=SURFACE)
        ctrl.pack(fill='x', padx=16, pady=(4, 8))

        self.status_label = tk.Label(ctrl, text="READY  //  Select marketplace & add PDFs",
                                     font=FONT_MONO, bg=SURFACE, fg=TEXT_DIM)
        self.status_label.pack(side='left')

        self.run_btn = self._btn(ctrl, "▶  EXTRACT DATA", self._run, ACCENT, large=True)
        self.run_btn.pack(side='right')

        self._btn(ctrl, "📂 OPEN OUTPUT", self._open_output, TEXT_DIM).pack(side='right', padx=8)

    # ── HELPERS ────────────────────────────────────────────────────────────────

    def _section(self, parent, title):
        f = tk.Frame(parent, bg=BG)
        f.pack(fill='x', pady=(6, 4))
        tk.Label(f, text=title, font=FONT_LABEL, bg=BG, fg=ACCENT).pack(side='left')
        tk.Frame(f, bg=BORDER_C, height=1).pack(side='left', fill='x', expand=True, padx=8)

    def _btn(self, parent, text, cmd, color, large=False):
        font = ("Courier New", 10, "bold") if large else ("Courier New", 9, "bold")
        padx = 16 if large else 10
        pady = 6  if large else 4
        btn = tk.Label(parent, text=text, font=font, bg=SURFACE2, fg=color,
                       cursor='hand2', padx=padx, pady=pady, relief='flat', bd=0,
                       highlightthickness=1, highlightbackground=color)
        btn.bind('<Button-1>', lambda e: cmd())
        btn.bind('<Enter>', lambda e: btn.config(bg=color, fg=BG))
        btn.bind('<Leave>', lambda e: btn.config(bg=SURFACE2, fg=color))
        return btn

    def _log(self, msg, tag='dim'):
        self.log_text.config(state='normal')
        ts = time.strftime("%H:%M:%S")
        self.log_text.insert('end', f"[{ts}] {msg}\n", tag)
        self.log_text.see('end')
        self.log_text.config(state='disabled')

    def _set_status(self, msg, color=TEXT_DIM):
        self.status_label.config(text=msg, fg=color)

    def _set_progress(self, pct):
        self.progress_canvas.update_idletasks()
        w = self.progress_canvas.winfo_width()
        self.progress_canvas.delete('all')
        self.progress_canvas.create_rectangle(0, 0, w, 4, fill=SURFACE2, outline='')
        if pct > 0:
            bar_w = int(w * pct / 100)
            self.progress_canvas.create_rectangle(0, 0, bar_w, 4, fill=ACCENT, outline='')

    def _refresh_file_list(self):
        self.file_list.delete(0, 'end')
        for p in self.pdf_files:
            name = os.path.basename(p)
            display = name if len(name) <= 44 else name[:41] + '...'
            self.file_list.insert('end', f"  {display}")
        self.file_count_var.set(f"{len(self.pdf_files)} file(s) loaded")

    def _on_marketplace_change(self, *_):
        mp = self.marketplace_var.get()
        labels = self.stat_labels_config.get(mp, self.stat_labels_config['Blinkit'])
        for label_text, key in labels:
            if key in self.stat_label_widgets:
                self.stat_label_widgets[key].config(text=f"  {label_text}")

    # ── ACTIONS ────────────────────────────────────────────────────────────────

    def _add_files(self):
        files = filedialog.askopenfilenames(
            title="Select PDFs",
            filetypes=[("PDF files", "*.pdf"), ("All files", "*.*")]
        )
        added = 0
        for f in files:
            if f not in self.pdf_files:
                self.pdf_files.append(f)
                added += 1
        if added:
            self._refresh_file_list()
            self._log(f"Added {added} file(s)", 'inf')

    def _add_folder(self):
        folder = filedialog.askdirectory(title="Select folder containing PDFs")
        if not folder:
            return
        files = sorted(glob.glob(os.path.join(folder, "*.pdf")))
        added = 0
        for f in files:
            if f not in self.pdf_files:
                self.pdf_files.append(f)
                added += 1
        self._refresh_file_list()
        self._log(f"Scanned folder → added {added} PDF(s)", 'inf')

    def _clear_files(self):
        self.pdf_files.clear()
        self._refresh_file_list()
        self._log("File list cleared", 'dim')

    def _open_output(self):
        if self.last_output and os.path.exists(self.last_output):
            if os.name == 'nt':
                os.startfile(self.last_output)
            else:
                os.system(f'open "{self.last_output}"')
        else:
            folder = os.path.join(os.path.dirname(os.path.abspath(__file__)), "output_parsed")
            if os.path.exists(folder):
                if os.name == 'nt':
                    os.startfile(folder)
                else:
                    os.system(f'open "{folder}"')
            else:
                messagebox.showwarning("Not Found", "No output file yet. Run extraction first.")

    def _run(self):
        if self.is_running:
            return
        if not self.pdf_files:
            messagebox.showwarning("No Files", "Please add at least one PDF.")
            return

        mp_name = self.marketplace_var.get()
        parser = MarketplaceRegistry.get_parser(mp_name)

        script_dir = os.path.dirname(os.path.abspath(__file__))
        output_dir = os.path.join(script_dir, "output_parsed")
        os.makedirs(output_dir, exist_ok=True)
        timestamp = time.strftime("%d%m%Y_%H%M%S")
        output = os.path.join(output_dir, f"{mp_name}_{timestamp}.xlsx")

        self.is_running = True
        self.run_btn.config(fg=TEXT_DIM)
        self._set_status("PROCESSING...", AMBER)
        threading.Thread(
            target=self._extract_worker,
            args=(parser, output),
            daemon=True
        ).start()

    def _extract_worker(self, parser: BaseMarketplaceParser, output: str):
        all_items = []
        all_summaries = []
        total = len(self.pdf_files)

        for i, pdf_path in enumerate(self.pdf_files):
            fname = os.path.basename(pdf_path)
            self.after(0, self._set_status,
                       f"Processing {i+1}/{total}:  {fname[:40]}", AMBER)
            self.after(0, self._set_progress, int((i / total) * 90))
            try:
                result = parser.parse(pdf_path)
                all_items.append(result.items)
                all_summaries.append(result.header)
                msg = (f"✓ {fname}  →  {parser.document_type}  |  "
                       f"{len(result.items)} items")
                self.after(0, self._log, msg, 'ok')
            except Exception as e:
                self.after(0, self._log, f"✗ {fname}  →  {e}", 'err')

        if not all_items:
            self.after(0, self._done, None, None, parser,
                       "ERROR: No data extracted", RED)
            return

        try:
            combined = pd.concat(all_items, ignore_index=True)
            summary = pd.DataFrame(all_summaries)

            sc_cols = parser.get_summary_columns()
            summary = summary[[c for c in sc_cols if c in summary.columns]]

            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                combined.to_excel(writer, sheet_name=parser.item_sheet_name, index=False)
                summary.to_excel(writer, sheet_name=parser.summary_sheet_name, index=False)

            self.after(0, self._set_progress, 95)
            ExcelFormatter.format(output, parser)
            self.after(0, self._set_progress, 100)

            # Update stats
            stats = parser.get_stats(combined)
            self.after(0, self._update_stats, stats)

            self.after(0, self._done, combined, output, parser,
                       f"DONE  //  {len(combined)} rows saved", GREEN)
        except Exception as e:
            self.after(0, self._done, None, None, parser,
                       f"ERROR: {e}", RED)

    def _update_stats(self, stats: Dict[str, str]):
        for key, val in stats.items():
            if key in self.stat_vars:
                self.stat_vars[key].set(val)

    def _done(self, df, output, parser, msg, color):
        self.is_running = False
        self.run_btn.config(fg=ACCENT)
        self._set_status(msg, color)
        if df is not None:
            self.last_output = output
            self.last_path_var.set(os.path.basename(output))
            self._log(f"Saved → {output}", 'inf')
            if messagebox.askyesno("Done!",
                                   f"Extraction complete!\n\n{msg}\n\nOpen output file?"):
                self._open_output()


# ═══════════════════════════════════════════════════════════════════════════════
#  STYLE
# ═══════════════════════════════════════════════════════════════════════════════

def apply_style():
    style = ttk.Style()
    style.theme_use('default')
    style.configure('Vertical.TScrollbar',
                    background=SURFACE2, troughcolor=SURFACE,
                    arrowcolor=TEXT_DIM, bordercolor=BORDER_C,
                    lightcolor=SURFACE2, darkcolor=SURFACE2)
    style.map('Vertical.TScrollbar', background=[('active', BORDER_C)])


# ═══════════════════════════════════════════════════════════════════════════════
#  ENTRY POINT
# ═══════════════════════════════════════════════════════════════════════════════

if __name__ == '__main__':
    app = MarketplaceParserApp()
    apply_style()
    app.mainloop()